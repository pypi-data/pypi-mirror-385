from typing import Any, TypeAlias

import openpyxl as op
import pandas as pd

from ut_aod.aod import AoD
from ut_dic.dic import Dic

from ut_xls.op.iocwb import IocWb
from ut_xls.op.ws import Ws, SheetNms

TyCe: TypeAlias = op.cell.cell.Cell
TyWb: TypeAlias = op.workbook.workbook.Workbook
TyWs: TypeAlias = op.worksheet.worksheet.Worksheet
TyCs: TypeAlias = op.chartsheet.chartsheet.Chartsheet
TyPdDf: TypeAlias = pd.DataFrame

TyArr = list[Any]
TyAoA = list[TyArr]
TyAoAoA = list[TyAoA]
TyDic = dict[Any, Any]
TyAoD = list[TyDic]
TyAoS = list[str]
TyAoWs = list[TyWs]
TyDoD = dict[Any, TyDic]
TyDoAoA = dict[Any, TyAoA]
TyDoAoD = dict[Any, TyAoD]
TyDoWs = dict[Any, TyWs]
TyDoPdDf = dict[Any, TyPdDf]
TyAoD_DoAoD = TyAoD | TyDoAoD
TySheetIx = int
TySheetNm = str
TySheet = TySheetIx | TySheetNm
TyAoSheet = list[TySheet]
TySheets = TySheet | TyAoSheet
TySheetNms = list[TySheetNm]
TyStrArr = str | TyArr
TyToCe = tuple[TyCe, ...]
TyWsCs = TyWs | TyCs

TnArr = None | TyArr
TnAoA = None | TyAoA
TnAoD = None | TyAoD
TnDic = None | TyDic
TnDoAoA = None | TyDoAoA
TnAoD_DoAoD = None | TyAoD_DoAoD
TnAoWs = None | TyAoWs
TnDoWs = None | TyDoWs
TnSheet = None | TySheet
TnAoSheet = None | TyAoSheet
TnSheets = None | TySheets
TnSheetNm = None | TySheetNm
TnWb = None | TyWb
TnCs = None | TyCs
TnWs = None | TyWs
TnWsCs = None | TyWsCs


class DoAoA:

    @staticmethod
    def create_wb(doaoa: TnDoAoA) -> TyWb:
        # def create_wb_with_doaoa(doaoa: TnDoAoA) -> TyWb:
        wb: TyWb = IocWb.get(write_only=True)
        if not doaoa:
            ws: TnWsCs = wb.active
            if ws is not None:
                wb.remove(ws)
            return wb
        for ws_id, aoa in doaoa.items():
            _ws: TnWs = wb.create_sheet()
            if _ws is None:
                continue
            _ws.title = ws_id
            Ws.append_rows(_ws, aoa)
        return wb


class DoAoD:

    @staticmethod
    def create_wb(doaod: TyDoAoD) -> TyWb:
        wb: TyWb = IocWb.get(write_only=True)
        if not doaod:
            return wb
        for ws_id, aod in doaod.items():
            a_header = [list(aod[0].keys())]
            a_data = [list(d.values()) for d in aod]
            a_row = a_header + a_data
            ws: TyWs = wb.create_sheet()
            ws.title = ws_id
            Ws.append_rows(ws, a_row)
        return wb


class Wb:

    @staticmethod
    def iter_sheet_names(wb: TyWb, **kwargs):
        cols_count = kwargs.get('cols_count', 0)
        sheet_names: TyArr = kwargs.get('sheet_names', [])
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            if sheet.max_column == cols_count:
                yield sheet_name

    @staticmethod
    def iter_sheet(wb: TyWb, sheet_count):
        for _ii in range(0, sheet_count):
            _ws = wb.create_sheet()
            yield _ws

    @classmethod
    def sh_sheetnm(cls, wb: TnWb, sheet: TnSheet) -> TnSheetNm:
        if wb is None or sheet is None:
            return None
        _sheetnm: TnSheetNm = SheetNms.sh_sheetnm(wb.sheetnames, sheet)
        return _sheetnm

    @classmethod
    def sh_sheetnms(cls, wb: TnWb, sheets: TnSheets) -> TySheetNms:
        _sheetnms: TySheetNms = []
        if wb is None or not sheets:
            return _sheetnms
        if isinstance(sheets, (int, str)):
            _sheetnm: TnSheetNm = SheetNms.sh_sheetnm(wb.sheetnames, sheets)
            if _sheetnm:
                _sheetnms = [_sheetnm]
            return _sheetnms
        if isinstance(sheets, (list, tuple)):
            _sheetnms = SheetNms.sh_sheetnms(wb.sheetnames, sheets)
            return _sheetnms
        return _sheetnms

    @classmethod
    def sh_sheet_by_sheetnm(
            cls, wb: TnWb, sheetnm: TnSheetNm) -> TnWs:
        if wb is None:
            return None
        if not sheetnm:
            return None
        return wb[sheetnm]

    @classmethod
    def sh_sheet(cls, wb: TnWb, sheet: TySheet) -> TnWs:
        _ws: TnWs = cls.sh_sheet_by_sheetnm(wb, cls.sh_sheetnm(wb, sheet))
        return _ws

    @classmethod
    def sh_chartsheet_by_sheetnm(cls, wb: TnWb, sheet_name: TnSheetNm) -> TnCs:
        _cs: TnCs = Ws.sh_chartsheet(cls.sh_sheet_by_sheetnm(wb, sheet_name))
        return _cs

    @classmethod
    def sh_worksheet_by_sheetnm(cls, wb: TnWb, sheet_name: TnSheetNm) -> TnWs:
        _ws: TnWs = Ws.sh_worksheet(cls.sh_sheet_by_sheetnm(wb, sheet_name))
        return _ws

    @classmethod
    def sh_chartsheet(cls, wb: TnWb, sheet: TnSheet) -> TnCs:
        return cls.sh_chartsheet_by_sheetnm(wb, cls.sh_sheetnm(wb, sheet))

    @classmethod
    def sh_worksheet(cls, wb: TnWb, sheet: TnSheet) -> TnWs:
        return cls.sh_worksheet_by_sheetnm(wb, cls.sh_sheetnm(wb, sheet))

    @classmethod
    def to_aod(cls, wb: TnWb, sheet: TnSheet) -> TyAoD:
        if wb is None:
            return []
        _ws: TnWs = cls.sh_worksheet(wb, sheet)
        _aod: TyAoD = Ws.to_aod(_ws)
        return _aod

    @classmethod
    def to_doaod(cls, wb: TnWb, sheet: TnSheets) -> TyDoAoD:
        if wb is None:
            return {}
        doaod: TyDoAoD = {}
        if wb is None:
            return doaod
        _sheetnms: TySheetNms = cls.sh_sheetnms(wb, sheet)
        if not _sheetnms:
            return doaod
        for _sheetnm in _sheetnms:
            _ws: TnWs = cls.sh_worksheet_by_sheetnm(wb, _sheetnm)
            Dic.set_by_key(doaod, _sheetnm, Ws.to_aod(_ws))
        return doaod

    @classmethod
    def to_aod_or_doaod(
            cls, wb: TyWb, sheet: TnSheets) -> TyAoD_DoAoD:
        doaod: TyDoAoD = {}
        _sheetnms: TySheetNms = cls.sh_sheetnms(wb, sheet)
        if not _sheetnms:
            return doaod
        if len(_sheetnms) == 1:
            _sheetnm = _sheetnms[0]
            _ws: TnWs = Wb.sh_worksheet_by_sheetnm(wb, _sheetnm)
            _aod: TyAoD = Ws.to_aod(_ws)
            return _aod
        for _sheetnm in _sheetnms:
            _ws = Wb.sh_worksheet_by_sheetnm(wb, _sheetnm)
            Dic.set_by_key(doaod, _sheetnm, Ws.to_aod(_ws))
        return doaod

    @classmethod
    def createupdate_wb_with_doaoa(cls, wb: TnWb, doaoa: TnDoAoA) -> None:
        if not doaoa:
            return
        if wb is None:
            DoAoA.create_wb(doaoa)
        else:
            cls.update_wb_with_doaoa(wb, doaoa)

    @classmethod
    def update_wb_with_aoa(cls, wb: TnWb, aoa: TnAoA, sheet: TySheet) -> None:
        if wb is None:
            return
        if not aoa:
            return
        _sheetnm: TnSheetNm = cls.sh_sheetnm(wb, sheet)
        _ws: TnWs = cls.sh_worksheet_by_sheetnm(wb, _sheetnm)
        Ws.append_rows(_ws, aoa)

    @classmethod
    def update_wb_with_aod(cls, wb: TnWb, aod: TnAoD, sheet: TySheet) -> None:
        if wb is None:
            return
        _aoa: TnAoA = AoD.to_aoa(aod, sw_keys=False)
        cls.update_wb_with_aoa(wb, _aoa, sheet)

    @classmethod
    def update_wb_with_doaoa(cls, wb: TnWb, doaoa: TnDoAoA) -> None:
        if wb is None:
            return
        if not doaoa:
            return
        a_ws_id: TyArr = Dic.sh_keys(doaoa, wb.sheetnames)
        for ws_id in a_ws_id:
            aoa: TyAoA = doaoa[ws_id]
            ws: TnWs = cls.sh_worksheet(wb, ws_id)
            Ws.append_rows(ws, aoa)

    @staticmethod
    def update_wb_with_dodf(wb: TnWb, dodf: TyDoPdDf, **kwargs) -> TnWb:
        if wb is None:
            return None
        _d_update: TyDic = kwargs.get('d_update', {})
        _d_head: TyDic = _d_update.get('d_head', {})
        _a_key: TyArr = Dic.show_sorted_keys(dodf)
        for _key in _a_key:
            _df = dodf[_key]
            _ws_tmpl: TyWs = wb['TMPL']
            _ws_new: TyWs = wb.copy_worksheet(_ws_tmpl)
            _ws_new.title = _key
            _d_head['title']['value'] = _key
            Ws.update_ws_cell_from_df_with_d_body(_ws_new, _df, _d_update)
            Ws.update_ws_cell_with_d_head(_ws_new, _d_head)
        return wb
