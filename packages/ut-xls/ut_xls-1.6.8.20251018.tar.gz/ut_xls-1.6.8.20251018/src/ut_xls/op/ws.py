from collections.abc import Callable
from typing import Any, TypeAlias

import openpyxl as op
import pandas as pd

from ut_log.log import LogEq
from ut_arr.arr import Arr
from ut_dic.dic import Dic

from ut_xls.op.rw import Rw

TyWs: TypeAlias = op.worksheet.worksheet.Worksheet
TyCs: TypeAlias = op.chartsheet.chartsheet.Chartsheet
TyPdDf: TypeAlias = pd.DataFrame

TyArr = list[Any]
TyAoA = list[TyArr]
TyCall = Callable
TyDic = dict[Any, Any]
TyAoD = list[TyDic]
TyInt = int
TyStrArr = str | TyArr
TySheet = str | int
TyAoSheet = list[TySheet]
TySheetNm = str
TySheetNms = list[TySheetNm]
TyWsCs = TyWs | TyCs

TnSheet = None | TySheet
TnAoSheet = None | TyAoSheet
TnSheetNm = None | TySheetNm
TnSheetNms = None | TySheetNms
TnCs = None | TyCs
TnWsCs = None | TyWsCs
TnWs = None | TyWs
TnInt = None | TyInt
TnStr = None | str


class SheetNms:

    @staticmethod
    def sh_sheetnm(sheetnms: TySheetNms, sheet: TnSheet) -> TnSheetNm:
        if sheet is None:
            return None
        if isinstance(sheet, int):
            if sheet < len(sheetnms):
                return sheetnms[sheet]
            return None
        if isinstance(sheet, str):
            if sheet in sheetnms:
                return sheet
            return None
        return None

    @classmethod
    def sh_sheetnms(cls, sheetnms: TySheetNms, sheets: TnAoSheet) -> TySheetNms:
        _sheetnms: TySheetNms = []
        if sheets is None:
            return _sheetnms
        for _sheet in sheets:
            _sheetnm = cls.sh_sheetnm(sheetnms, _sheet)
            if _sheetnm:
                _sheetnms.append(_sheetnm)
        return _sheetnms


class Ws:

    class Headers:

        @staticmethod
        def iter_column(ws: TyWs, **kwargs):
            row_header = kwargs.get('headers_start', 1)
            col_name_prefix = kwargs.get('col_name_prefix', '')
            col_start = 1
            col_end = ws.max_column+1
            sw_add_sheet_name = kwargs.get('sw_add_sheet_name', False)
            if sw_add_sheet_name:
                header_sheet_name = kwargs.get(
                    'header_sheet_name', 'sheet_name')
                col_name: str = f"{col_name_prefix}{header_sheet_name}"
                yield col_name
            for col in range(col_start, col_end):
                cell = ws.cell(column=col, row=row_header)
                # col_name: str = cell.value
                col_name = f"{col_name_prefix}{cell.value}"
                yield col_name

    @staticmethod
    def append_rows(ws: TnWs, a_row: TyArr) -> TnWsCs:
        if ws is None:
            return ws
        for _row in a_row:
            ws.append(_row)
        return ws

    @staticmethod
    def filter_rows(ws: TnWs) -> TyArr:
        if ws is None:
            return []
        a_row = list(ws.iter_rows(values_only=True))
        a_row_new = []
        for row in a_row[1:]:
            if row is None or len(row) < 3:
                continue
            if row[0] is not None and row[1] is not None:
                a_row_new.append(row)
        return a_row_new

    @staticmethod
    def iter_sheet_lst(ws: TyWs, **kwargs):
        row_start = kwargs.get('row_start', 1)
        row_end = kwargs.get('row_end', ws.max_row)
        col_start = kwargs.get('col_start', 1)
        col_end = kwargs.get('col_end', ws.max_column)

        for row in ws.iter_rows(
                     min_row=row_start, min_col=col_start,
                     max_row=row_end, max_col=col_end):
            yield list(Rw.iter_cell_value(row, **kwargs))

    @classmethod
    def sh_headers(cls, ws: TyWs, **kwargs) -> TyArr:
        return list(cls.Headers.iter_column(ws, **kwargs))

    @classmethod
    def sh_aoa(cls, ws: TyWs, **kwargs) -> TyAoA:
        return list(cls.iter_sheet_lst(ws, **kwargs))

    @staticmethod
    def sh_id(sheet_name: None | str = None, sheet_index: int = 0) -> TnSheet:
        if sheet_name is not None:
            return sheet_name
        return sheet_index

    @classmethod
    def sh_chartsheet(cls, ws: TnWsCs) -> TnCs:
        # if cls.is_type(ws, TnCs):
        if isinstance(ws, TnCs):
            return ws
        return None

    @classmethod
    def sh_worksheet(cls, ws: TnWsCs) -> TnWs:
        # if cls.is_type(ws, TnWs):
        if isinstance(ws, TnWs):
            return ws
        return None

    @classmethod
    def to_aod_apply_fnc_to_value(cls, ws: TyWs, fnc: Callable, **kwargs) -> TyAoD:
        aod: TyAoD = []
        if ws is None:
            return aod
        i_row = ws.iter_rows(values_only=True)
        try:
            header = next(i_row)
        except StopIteration:
            return aod
        for row in i_row:
            dic = Arr.sh_dic_zip(header, Arr.apply_function(row, fnc, **kwargs))
            aod.append(dic)
        LogEq.debug("aod", aod)
        return aod

    @classmethod
    def to_aod_apply_str_to_value(cls, ws: TyWs) -> TyAoD:
        return cls.to_aod_apply_fnc_to_value(ws, str)

    @classmethod
    def to_aod(cls, ws: TnWs) -> TyAoD:
        aod: TyAoD = []
        if ws is None:
            return aod
        i_row = ws.iter_rows(values_only=True)
        try:
            header = next(i_row)
        except StopIteration:
            return aod
        for row in i_row:
            dic = Arr.sh_dic_zip(header, row)
            aod.append(dic)
        LogEq.debug("aod", aod)
        return aod

    @classmethod
    def to_rows(cls, ws: TyWs) -> TyArr:
        return list(ws.iter_rows())

    @classmethod
    def to_row_values(cls, ws: TyWs) -> TyArr:
        return list(ws.iter_rows(values_only=True))

    @classmethod
    def to_dic(cls, ws: TnWs) -> TyDic:
        dic: TyDic = {}
        if ws is None:
            return dic
        a_row: TyArr = list(ws.iter_rows(values_only=True))
        header = a_row[0]
        if len(header) == 2:
            for row in a_row[1:]:
                dic[row[0]] = row[1]
        else:
            for row in a_row[1:]:
                dic[row[0]] = row[1]
        LogEq.debug("dic", dic)
        return dic

    @staticmethod
    def update_ws_cell_from_df_with_d_body(
            ws: TyWs, df: TyPdDf, d_update: TyDic) -> None:
        # def update_xls_cell_with_d_body(
        _pv_indexes: TyStrArr = d_update.get('pv_indexes', '')
        _pv_a_nm_col: TyDic = d_update.get('pv_a_nm_col', [])

        _d_body: TyDic = d_update.get('d_body', {})
        _a_nm_row: TyArr = _d_body.get('a_row', [])
        _a_nm_col: TyArr = _d_body.get('a_col', [])
        _xls_off_row: TyInt = Dic.locate(_d_body, ['offset', 'row'])
        _xls_off_col: TyInt = Dic.locate(_d_body, ['offset', 'col'])

        df_new = df.set_index(_pv_indexes)
        _a_df_nm_row = df_new.index.tolist()

        for _nm_row in _a_df_nm_row:
            _ix_row = _a_nm_row.index(_nm_row)+_xls_off_row
            for _nm_col in _pv_a_nm_col:
                _ix_col = _a_nm_col.index(_nm_col)+_xls_off_col
                _df_cell_value = df_new.loc[_nm_row, _nm_col]
                ws.cell(_ix_row, _ix_col).value = _df_cell_value

    @staticmethod
    def update_ws_cell_with_d_head(
            ws: TyWs, d_head: TyDic) -> None:
        # def update_xls_cell_with_d_head(ws: TyWs, d_head: TyDic) -> None:
        for _key in d_head.keys():
            _ix_row = d_head[_key]['row']
            _ix_col = d_head[_key]['col']
            _value = d_head[_key]['value']
            ws.cell(_ix_row, _ix_col).value = _value
