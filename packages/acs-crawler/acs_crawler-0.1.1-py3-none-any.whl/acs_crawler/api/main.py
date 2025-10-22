"""FastAPI application for ACS Crawler web interface."""

import asyncio
import json
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, HTTPException, Request, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

from ..config import setup_logging, get_logger, DATA_DIR
from ..crawler_service import CrawlerService
from ..models.paper import CrawlJob, Paper, JobStatus

# Initialize logging
setup_logging()
logger = get_logger(__name__)

# Initialize FastAPI app
app = FastAPI(
    title="ACS Paper Crawler",
    description="A professional web interface for crawling ACS papers",
    version="0.1.0",
)

# Setup templates and static files
TEMPLATES_DIR = Path(__file__).parent / "templates"
STATIC_DIR = Path(__file__).parent / "static"

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# Mount static files
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# Initialize crawler service
crawler_service = CrawlerService()

# Thread pool for background tasks
executor = ThreadPoolExecutor(max_workers=3)


# Pydantic models for API
class JobCreate(BaseModel):
    """Request model for creating a new crawl job."""

    journal_url: str
    max_results: Optional[int] = None


class JobResponse(BaseModel):
    """Response model for job information."""

    job_id: str
    journal_url: str
    status: str
    created_at: str
    started_at: Optional[str]
    completed_at: Optional[str]
    total_papers: int
    crawled_papers: int
    failed_papers: int
    error_message: Optional[str]


class PaperResponse(BaseModel):
    """Response model for paper information."""

    paper_id: str
    title: str
    doi: str
    url: str
    authors: List[str]
    abstract: Optional[str]
    journal: Optional[str]
    publication_date: Optional[str]
    crawled_at: str


# Helper functions
def job_to_response(job: CrawlJob) -> JobResponse:
    """Convert CrawlJob to JobResponse."""
    return JobResponse(
        job_id=job.job_id,
        journal_url=job.journal_url,
        status=job.status.value,
        created_at=job.created_at.isoformat(),
        started_at=job.started_at.isoformat() if job.started_at else None,
        completed_at=job.completed_at.isoformat() if job.completed_at else None,
        total_papers=job.total_papers,
        crawled_papers=job.crawled_papers,
        failed_papers=job.failed_papers,
        error_message=job.error_message,
    )


def paper_to_response(paper: Paper) -> PaperResponse:
    """Convert Paper to PaperResponse."""
    return PaperResponse(
        paper_id=paper.paper_id,
        title=paper.metadata.title,
        doi=paper.metadata.doi,
        url=paper.metadata.url,
        authors=[author.name for author in paper.metadata.authors],
        abstract=paper.metadata.abstract,
        journal=paper.metadata.journal,
        publication_date=paper.metadata.publication_date,
        crawled_at=paper.crawled_at.isoformat(),
    )


def run_job_in_background(job_id: str) -> None:
    """Run a crawl job in the background."""
    try:
        logger.info(f"Starting background job: {job_id}")
        crawler_service.run_job(job_id)
        logger.info(f"Background job completed: {job_id}")
    except Exception as e:
        logger.error(f"Background job failed: {job_id}, error: {e}")


# Web routes (HTML)
@app.get("/", response_class=HTMLResponse)
async def index(request: Request) -> HTMLResponse:
    """Render the main dashboard page."""
    jobs = crawler_service.list_all_jobs()
    papers = crawler_service.list_all_papers()

    # Get statistics
    total_papers = len(papers)
    total_jobs = len(jobs)
    running_jobs = sum(1 for job in jobs if job.status == JobStatus.RUNNING)
    completed_jobs = sum(1 for job in jobs if job.status == JobStatus.COMPLETED)

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "total_papers": total_papers,
            "total_jobs": total_jobs,
            "running_jobs": running_jobs,
            "completed_jobs": completed_jobs,
            "recent_jobs": jobs[:10],
            "recent_papers": papers[:10],
        },
    )


@app.get("/jobs", response_class=HTMLResponse)
async def jobs_page(request: Request) -> HTMLResponse:
    """Render the jobs management page."""
    jobs = crawler_service.list_all_jobs()
    return templates.TemplateResponse("jobs.html", {"request": request, "jobs": jobs})


@app.get("/papers", response_class=HTMLResponse)
async def papers_page(request: Request) -> HTMLResponse:
    """Render the papers list page."""
    papers = crawler_service.list_all_papers()
    return templates.TemplateResponse("papers.html", {"request": request, "papers": papers})


@app.get("/papers/{paper_id:path}", response_class=HTMLResponse)
async def paper_detail(request: Request, paper_id: str) -> HTMLResponse:
    """Render paper detail page."""
    paper = crawler_service.get_paper(paper_id)
    if not paper:
        raise HTTPException(status_code=404, detail="Paper not found")

    return templates.TemplateResponse("paper_detail.html", {"request": request, "paper": paper})


# API routes (JSON)
@app.post("/api/jobs", response_model=JobResponse)
async def create_job(job_create: JobCreate, background_tasks: BackgroundTasks) -> JobResponse:
    """Create a new crawl job and start it in the background.

    Args:
        job_create: Job creation request with journal URL and optional max_results
        background_tasks: FastAPI background tasks manager

    Returns:
        Created job information

    Raises:
        HTTPException: If job creation fails
    """
    try:
        job = crawler_service.create_job(job_create.journal_url, max_results=job_create.max_results)
        # Start job in background
        background_tasks.add_task(run_job_in_background, job.job_id)
        logger.info(f"Created and started job {job.job_id} with max_results={job_create.max_results}")
        return job_to_response(job)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to create job: {e}")
        raise HTTPException(status_code=500, detail="Failed to create job")


@app.get("/api/jobs", response_model=List[JobResponse])
async def list_jobs() -> List[JobResponse]:
    """List all crawl jobs.

    Returns:
        List of all jobs
    """
    jobs = crawler_service.list_all_jobs()
    return [job_to_response(job) for job in jobs]


@app.get("/api/jobs/{job_id}", response_model=JobResponse)
async def get_job(job_id: str) -> JobResponse:
    """Get a specific job by ID.

    Args:
        job_id: Job ID

    Returns:
        Job information

    Raises:
        HTTPException: If job not found
    """
    job = crawler_service.get_job_status(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job_to_response(job)


@app.delete("/api/jobs/{job_id}")
async def cancel_job(job_id: str) -> dict:
    """Cancel a running job.

    Args:
        job_id: Job ID

    Returns:
        Success message

    Raises:
        HTTPException: If job not found or cannot be cancelled
    """
    job = crawler_service.get_job_status(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if job.status not in [JobStatus.PENDING, JobStatus.RUNNING]:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot cancel job in {job.status} status"
        )

    # Mark job as failed/cancelled
    job.status = JobStatus.FAILED
    job.error_message = "Cancelled by user"
    job.completed_at = datetime.now()
    crawler_service.storage.save_job(job)

    logger.info(f"Job {job_id} cancelled by user")
    return {"message": f"Job {job_id} cancelled successfully"}


@app.get("/api/papers", response_model=List[PaperResponse])
async def list_papers() -> List[PaperResponse]:
    """List all crawled papers.

    Returns:
        List of all papers
    """
    papers = crawler_service.list_all_papers()
    return [paper_to_response(paper) for paper in papers]


@app.get("/api/papers/export/xlsx")
async def export_papers_xlsx():
    """Export all papers to Excel (XLSX) format.

    Returns:
        Excel file with all paper metadata
    """
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        papers = crawler_service.list_all_papers()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "ACS Papers"

        # Define header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header
        headers = [
            'DOI', 'Title', 'Authors', 'Journal', 'Volume', 'Issue', 'Pages',
            'Publication Date', 'Open Access', 'PDF URL', 'Abstract', 'Keywords', 'URL', 'Crawled At'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, paper in enumerate(papers, 2):
            # Join authors with semicolon + space
            authors_str = '; '.join([author.name for author in paper.metadata.authors])
            # Join keywords with semicolon + space
            keywords_str = '; '.join(paper.metadata.keywords) if paper.metadata.keywords else ''
            # Open Access status
            oa_status = 'Yes' if paper.metadata.is_open_access else 'No'

            ws.cell(row=row_num, column=1, value=paper.metadata.doi or '')
            ws.cell(row=row_num, column=2, value=paper.metadata.title or '')
            ws.cell(row=row_num, column=3, value=authors_str)
            ws.cell(row=row_num, column=4, value=paper.metadata.journal or '')
            ws.cell(row=row_num, column=5, value=paper.metadata.volume or '')
            ws.cell(row=row_num, column=6, value=paper.metadata.issue or '')
            ws.cell(row=row_num, column=7, value=paper.metadata.pages or '')
            ws.cell(row=row_num, column=8, value=paper.metadata.publication_date or '')
            ws.cell(row=row_num, column=9, value=oa_status)
            ws.cell(row=row_num, column=10, value=paper.metadata.oa_pdf_url or '')
            ws.cell(row=row_num, column=11, value=paper.metadata.abstract or '')
            ws.cell(row=row_num, column=12, value=keywords_str)
            ws.cell(row=row_num, column=13, value=paper.metadata.url or '')
            ws.cell(row=row_num, column=14, value=paper.crawled_at.isoformat())

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)  # Cap at 80 for readability
            ws.column_dimensions[column].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=acs_papers.xlsx"
            }
        )
    except Exception as e:
        logger.error(f"Failed to export papers to Excel: {e}")
        raise HTTPException(status_code=500, detail="Failed to export papers to Excel")


@app.get("/api/papers/{paper_id:path}", response_model=PaperResponse)
async def get_paper(paper_id: str) -> PaperResponse:
    """Get a specific paper by ID.

    Args:
        paper_id: Paper ID (DOI)

    Returns:
        Paper information

    Raises:
        HTTPException: If paper not found
    """
    paper = crawler_service.get_paper(paper_id)
    if not paper:
        raise HTTPException(status_code=404, detail="Paper not found")
    return paper_to_response(paper)


@app.get("/api/journals")
async def get_journals() -> JSONResponse:
    """Get the list of available ACS journals.

    Returns:
        JSON with journals list
    """
    journals_file = Path(__file__).parent.parent / "data" / "acs_journals.json"
    try:
        with open(journals_file, 'r', encoding='utf-8') as f:
            journals_data = json.load(f)
        return JSONResponse(content=journals_data)
    except FileNotFoundError:
        logger.error(f"Journals file not found: {journals_file}")
        return JSONResponse(content={"journals": []}, status_code=500)
    except Exception as e:
        logger.error(f"Failed to load journals: {e}")
        return JSONResponse(content={"journals": []}, status_code=500)


@app.get("/api/statistics")
async def get_statistics() -> JSONResponse:
    """Get database statistics for charts and visualizations.

    Returns:
        JSON with various statistics
    """
    try:
        stats = crawler_service.storage.get_statistics()

        # Get all papers for time-based analysis
        papers = crawler_service.list_all_papers()

        # Papers by month
        from collections import defaultdict
        papers_by_month = defaultdict(int)
        papers_by_year = defaultdict(int)

        for paper in papers:
            if paper.crawled_at:
                month_key = paper.crawled_at.strftime('%Y-%m')
                year_key = paper.crawled_at.strftime('%Y')
                papers_by_month[month_key] += 1
                papers_by_year[year_key] += 1

        # Sort by date
        sorted_months = sorted(papers_by_month.items())

        # Get top authors
        from collections import Counter
        all_authors = []
        for paper in papers:
            for author in paper.metadata.authors:
                all_authors.append(author.name)

        top_authors = Counter(all_authors).most_common(10)

        # Publication years from metadata
        pub_years = defaultdict(int)
        for paper in papers:
            if paper.metadata.publication_date:
                try:
                    # Try to extract year from publication date
                    import re
                    year_match = re.search(r'20\d{2}', paper.metadata.publication_date)
                    if year_match:
                        pub_years[year_match.group()] += 1
                except:
                    pass

        response_data = {
            "total_papers": stats.get('total_papers', 0),
            "papers_with_abstracts": stats.get('papers_with_abstracts', 0),
            "papers_by_journal": stats.get('papers_by_journal', {}),
            "papers_by_month": dict(sorted_months),
            "papers_by_year": dict(sorted(papers_by_year.items())),
            "publication_years": dict(sorted(pub_years.items())),
            "top_authors": [{"name": name, "count": count} for name, count in top_authors],
            "total_jobs": stats.get('total_jobs', 0),
            "jobs_completed": stats.get('jobs_completed', 0),
            "jobs_failed": stats.get('jobs_failed', 0),
        }

        return JSONResponse(content=response_data)
    except Exception as e:
        logger.error(f"Failed to get statistics: {e}")
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.on_event("shutdown")
async def shutdown_event() -> None:
    """Cleanup on shutdown."""
    crawler_service.cleanup()
    executor.shutdown(wait=True)
    logger.info("Application shutdown complete")
