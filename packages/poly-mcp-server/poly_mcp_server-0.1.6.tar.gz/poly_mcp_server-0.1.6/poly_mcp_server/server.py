#!/usr/bin/env python3
"""
Poly MCP Server - 다양한 유틸리티 도구를 제공하는 MCP 서버
"""

import asyncio
import hashlib
import os
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent, CallToolResult, INVALID_PARAMS, INTERNAL_ERROR


# 서버 인스턴스 생성
app = Server("poly-mcp-server")


@app.list_tools()
async def list_tools() -> list[Tool]:
    """사용 가능한 도구 목록을 반환합니다."""
    return [
        Tool(
            name="calculator",
            description="기본 수학 계산을 수행합니다 (덧셈, 뺄셈, 곱셈, 나눗셈)",
            inputSchema={
                "type": "object",
                "properties": {
                    "expression": {
                        "type": "string",
                        "description": "계산할 수학 표현식 (예: '2 + 3 * 4')",
                    }
                },
                "required": ["expression"],
            },
        ),
        Tool(
            name="text_analyzer",
            description="텍스트 분석 도구 (글자 수, 단어 수, 문장 수 등)",
            inputSchema={
                "type": "object",
                "properties": {
                    "text": {
                        "type": "string",
                        "description": "분석할 텍스트",
                    },
                    "language": {
                        "type": "string",
                        "description": "텍스트 언어 (ko, en 등)",
                        "default": "ko",
                    },
                },
                "required": ["text"],
            },
        ),
        Tool(
            name="uuid_generator",
            description="UUID를 생성합니다",
            inputSchema={
                "type": "object",
                "properties": {
                    "version": {
                        "type": "string",
                        "description": "UUID 버전 (v1, v4)",
                        "default": "v4",
                    },
                    "count": {
                        "type": "integer",
                        "description": "생성할 UUID 개수",
                        "default": 1,
                    },
                },
            },
        ),
        Tool(
            name="timestamp_converter",
            description="타임스탬프를 다양한 형식으로 변환합니다",
            inputSchema={
                "type": "object",
                "properties": {
                    "timestamp": {
                        "type": "string",
                        "description": "변환할 타임스탬프 또는 날짜 문자열",
                    },
                    "format": {
                        "type": "string",
                        "description": "출력 형식 (iso, unix, korean 등)",
                        "default": "iso",
                    },
                },
                "required": ["timestamp"],
            },
        ),
        Tool(
            name="hash_generator",
            description="문자열의 해시값을 생성합니다",
            inputSchema={
                "type": "object",
                "properties": {
                    "text": {
                        "type": "string",
                        "description": "해시할 텍스트",
                    },
                    "algorithm": {
                        "type": "string",
                        "description": "해시 알고리즘 (md5, sha1, sha256, sha512)",
                        "default": "sha256",
                    },
                },
                "required": ["text"],
            },
        ),
        Tool(
            name="postgres_query",
            description="PostgreSQL 데이터베이스에 쿼리를 실행합니다 (환경변수로 연결 정보 설정 필요)",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "실행할 SQL 쿼리 (SELECT만 허용)",
                    },
                    "limit": {
                        "type": "integer",
                        "description": "반환할 최대 행 수",
                        "default": 100,
                    },
                },
                "required": ["query"],
            },
        ),
        Tool(
            name="wordit_edutem_usage",
            description="Wordit 에듀템 월별 사용 통계를 조회하고 엑셀 파일로 저장합니다 (예: '2024년 8월')",
            inputSchema={
                "type": "object",
                "properties": {
                    "period": {
                        "type": "string",
                        "description": "조회 기간 (예: '2024년 8월', '2024-08', '8월')",
                    },
                    "output_path": {
                        "type": "string",
                        "description": "엑셀 파일 저장 경로 (기본값: 현재 디렉토리)",
                        "default": ".",
                    },
                },
                "required": ["period"],
            },
        ),
    ]


@app.call_tool()
async def call_tool(name: str, arguments: Any) -> list[TextContent]:
    """도구를 호출하고 결과를 반환합니다."""
    try:
        if name == "calculator":
            return await handle_calculator(arguments)
        elif name == "text_analyzer":
            return await handle_text_analyzer(arguments)
        elif name == "uuid_generator":
            return await handle_uuid_generator(arguments)
        elif name == "timestamp_converter":
            return await handle_timestamp_converter(arguments)
        elif name == "hash_generator":
            return await handle_hash_generator(arguments)
        elif name == "postgres_query":
            return await handle_postgres_query(arguments)
        elif name == "wordit_edutem_usage":
            return await handle_wordit_edutem_usage(arguments)
        else:
            raise ValueError(f"Unknown tool: {name}")
    except Exception as e:
        raise RuntimeError(f"Tool execution failed: {str(e)}")


async def handle_calculator(args: dict) -> list[TextContent]:
    """계산기 도구를 처리합니다."""
    expression = args.get("expression", "")
    
    # 안전한 수학 표현식만 허용
    allowed_chars = set("0123456789+-*/()%. ")
    if not all(c in allowed_chars for c in expression):
        raise ValueError(f"잘못된 수학 표현식입니다: {expression}")
    
    try:
        result = eval(expression, {"__builtins__": {}}, {})
        return [
            TextContent(
                type="text",
                text=f"계산 결과: {expression} = {result}"
            )
        ]
    except Exception as e:
        raise ValueError(f"계산 오류: {str(e)}")


async def handle_text_analyzer(args: dict) -> list[TextContent]:
    """텍스트 분석 도구를 처리합니다."""
    text = args.get("text", "")
    language = args.get("language", "ko")
    
    char_count = len(text)
    char_count_no_spaces = len(text.replace(" ", "").replace("\n", "").replace("\t", ""))
    
    if language == "ko":
        # 한국어: 공백, 구두점으로 단어 분리
        import re
        words = [w for w in re.split(r'[\s,，.。!！?？;；:：]+', text) if w]
        sentences = [s for s in re.split(r'[.。!！?？]+', text) if s.strip()]
    else:
        # 영어: 공백으로 단어 분리
        words = [w for w in text.split() if w]
        sentences = [s for s in text.replace('!', '.').replace('?', '.').split('.') if s.strip()]
    
    word_count = len(words)
    sentence_count = len(sentences)
    avg_words = round(word_count / sentence_count, 1) if sentence_count > 0 else 0
    
    result = f"""텍스트 분석 결과:
📊 기본 통계:
- 전체 글자 수: {char_count}자
- 공백 제외 글자 수: {char_count_no_spaces}자
- 단어 수: {word_count}개
- 문장 수: {sentence_count}개
- 문장당 평균 단어 수: {avg_words}개

📝 언어: {'한국어' if language == 'ko' else '영어'}"""
    
    return [TextContent(type="text", text=result)]


async def handle_uuid_generator(args: dict) -> list[TextContent]:
    """UUID 생성기 도구를 처리합니다."""
    version = args.get("version", "v4")
    count = args.get("count", 1)
    
    uuids = []
    for _ in range(count):
        if version == "v1":
            uuids.append(str(uuid.uuid1()))
        else:  # v4
            uuids.append(str(uuid.uuid4()))
    
    result = f"생성된 UUID ({version}):\n" + "\n".join(uuids)
    return [TextContent(type="text", text=result)]


async def handle_timestamp_converter(args: dict) -> list[TextContent]:
    """타임스탬프 변환기 도구를 처리합니다."""
    timestamp_str = args.get("timestamp", "")
    format_type = args.get("format", "iso")
    
    # 타임스탬프 파싱
    try:
        if timestamp_str.isdigit():
            # Unix 타임스탬프
            ts_int = int(timestamp_str)
            if ts_int > 10000000000:  # 밀리초
                dt = datetime.fromtimestamp(ts_int / 1000, tz=timezone.utc)
            else:  # 초
                dt = datetime.fromtimestamp(ts_int, tz=timezone.utc)
        else:
            # ISO 형식 또는 일반 날짜 문자열
            from dateutil import parser
            dt = parser.parse(timestamp_str)
    except Exception as e:
        raise ValueError(f"잘못된 타임스탬프 형식입니다: {timestamp_str}")
    
    # 형식에 따라 변환
    if format_type == "iso":
        formatted = dt.isoformat()
    elif format_type == "unix":
        formatted = str(int(dt.timestamp()))
    elif format_type == "korean":
        formatted = dt.strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
    else:
        formatted = str(dt)
    
    # 현재 시간과의 차이 계산
    now = datetime.now(timezone.utc)
    diff_seconds = (now - dt.replace(tzinfo=timezone.utc)).total_seconds()
    diff_days = int(diff_seconds / 86400)
    
    if diff_days > 0:
        time_diff = f"{diff_days}일 전"
    elif diff_days < 0:
        time_diff = f"{abs(diff_days)}일 후"
    else:
        time_diff = "오늘"
    
    result = f"""타임스탬프 변환 결과:
🕒 입력: {timestamp_str}
📅 변환된 시간: {formatted}
⏰ Unix 타임스탬프: {int(dt.timestamp())}
🌏 ISO 형식: {dt.isoformat()}
📊 현재로부터: {time_diff}"""
    
    return [TextContent(type="text", text=result)]


async def handle_hash_generator(args: dict) -> list[TextContent]:
    """해시 생성기 도구를 처리합니다."""
    text = args.get("text", "")
    algorithm = args.get("algorithm", "sha256").lower()
    
    valid_algorithms = ["md5", "sha1", "sha256", "sha512"]
    if algorithm not in valid_algorithms:
        raise ValueError(
            f"지원되지 않는 해시 알고리즘입니다. "
            f"지원 알고리즘: {', '.join(valid_algorithms)}"
        )
    
    # 해시 생성
    hash_obj = hashlib.new(algorithm)
    hash_obj.update(text.encode('utf-8'))
    hash_value = hash_obj.hexdigest()
    
    result = f"""해시 생성 결과:
📝 원본 텍스트: "{text}"
🔒 알고리즘: {algorithm.upper()}
🔑 해시값: {hash_value}
📏 길이: {len(hash_value)}자"""
    
    return [TextContent(type="text", text=result)]


async def handle_postgres_query(args: dict) -> list[TextContent]:
    """PostgreSQL 쿼리 실행 도구를 처리합니다."""
    query = args.get("query", "").strip()
    limit = args.get("limit", 100)
    
    # SELECT 쿼리만 허용 (보안)
    if not query.upper().startswith("SELECT"):
        raise ValueError("보안상 SELECT 쿼리만 허용됩니다.")
    
    # 환경변수에서 연결 정보 가져오기
    db_config = {
        "host": os.getenv("POSTGRES_HOST", "localhost"),
        "port": int(os.getenv("POSTGRES_PORT", "5432")),
        "database": os.getenv("POSTGRES_DB"),
        "user": os.getenv("POSTGRES_USER"),
        "password": os.getenv("POSTGRES_PASSWORD"),
    }
    
    # 필수 연결 정보 확인
    if not all([db_config["database"], db_config["user"], db_config["password"]]):
        raise ValueError(
            "PostgreSQL 연결 정보가 설정되지 않았습니다.\n"
            "다음 환경변수를 설정하세요:\n"
            "- POSTGRES_HOST (기본값: localhost)\n"
            "- POSTGRES_PORT (기본값: 5432)\n"
            "- POSTGRES_DB (필수)\n"
            "- POSTGRES_USER (필수)\n"
            "- POSTGRES_PASSWORD (필수)"
        )
    
    conn = None
    try:
        # PostgreSQL 연결
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()
        
        # 쿼리 실행
        cursor.execute(query)
        
        # 결과 가져오기
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchmany(limit)
        
        if not rows:
            result = "🔍 쿼리 결과: 데이터가 없습니다."
        else:
            # 테이블 형식으로 결과 포맷팅
            result_lines = [
                f"🔍 쿼리 결과 ({len(rows)}개 행):",
                "",
                "| " + " | ".join(columns) + " |",
                "|" + "|".join(["---"] * len(columns)) + "|",
            ]
            
            for row in rows:
                formatted_row = [str(val) if val is not None else "NULL" for val in row]
                result_lines.append("| " + " | ".join(formatted_row) + " |")
            
            if len(rows) == limit:
                result_lines.append("")
                result_lines.append(f"⚠️ 결과가 {limit}개로 제한되었습니다.")
            
            result = "\n".join(result_lines)
        
        cursor.close()
        
        return [TextContent(type="text", text=result)]
        
    except Exception as e:
        raise ValueError(f"PostgreSQL 오류: {str(e)}")
    finally:
        if conn:
            conn.close()


def parse_period(period: str) -> tuple[str, str]:
    """기간 문자열을 파싱하여 시작일과 종료일을 반환합니다.
    
    Args:
        period: '2024년 8월', '2024-08', '8월' 등
        
    Returns:
        (start_date, end_date) 튜플 (YYYY-MM-DD 형식)
    """
    current_year = datetime.now().year
    
    # 패턴 매칭
    # 1. "2024년 8월" 또는 "2024년08월"
    match = re.search(r'(\d{4})\s*년\s*(\d{1,2})\s*월', period)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
    # 2. "2024-08" 또는 "2024/08"
    elif re.search(r'(\d{4})[-/](\d{1,2})', period):
        match = re.search(r'(\d{4})[-/](\d{1,2})', period)
        year = int(match.group(1))
        month = int(match.group(2))
    # 3. "8월" (현재 연도 사용)
    elif re.search(r'(\d{1,2})\s*월', period):
        match = re.search(r'(\d{1,2})\s*월', period)
        year = current_year
        month = int(match.group(1))
    else:
        raise ValueError(f"기간 형식을 인식할 수 없습니다: {period}\n지원 형식: '2024년 8월', '2024-08', '8월'")
    
    # 날짜 계산
    from dateutil.relativedelta import relativedelta
    start_date = datetime(year, month, 1)
    end_date = start_date + relativedelta(months=1)
    
    return start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')


async def handle_wordit_edutem_usage(args: dict) -> list[TextContent]:
    """Wordit 에듀템 월별 사용 통계를 조회하고 엑셀로 저장합니다."""
    period = args.get("period", "")
    output_path = args.get("output_path", ".")
    
    if not period:
        raise ValueError("조회 기간을 입력해주세요 (예: '2024년 8월', '2024-08')")
    
    # 기간 파싱
    try:
        start_date, end_date = parse_period(period)
    except Exception as e:
        raise ValueError(str(e))
    
    # 환경변수에서 연결 정보 가져오기
    db_config = {
        "host": os.getenv("POSTGRES_HOST", "localhost"),
        "port": int(os.getenv("POSTGRES_PORT", "5432")),
        "database": os.getenv("POSTGRES_DB"),
        "user": os.getenv("POSTGRES_USER"),
        "password": os.getenv("POSTGRES_PASSWORD"),
    }
    
    # 필수 연결 정보 확인
    if not all([db_config["database"], db_config["user"], db_config["password"]]):
        raise ValueError("PostgreSQL 연결 정보가 설정되지 않았습니다.")
    
    conn = None
    try:
        # PostgreSQL 연결
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()
        
        # 쿼리 실행
        query = f"""
        SELECT tdc.course_name
            , tms.member_name_kr
            , tms.member_code
            , count(trmr.created_at) as mission_count
            , min(trmr.created_at) as first_complete
            , max(trmr.created_at) as last_complete
        FROM voca_wordit.tb_rea_mission_result trmr
            JOIN voca_wordit.tb_mba_student tms ON trmr.member_code = tms.member_code
            JOIN voca_wordit.tb_dfa_course tdc ON tdc.course_code = tms.course_code
            JOIN voca_wordit.tb_cla_client tcc ON tcc.client_code = tms.client_code
        WHERE trmr.mission_type = 'WB'
            AND tms.client_code NOT IN ('0508003', '1301003')
            AND trmr.created_at >= '{start_date}'::timestamp
            AND trmr.created_at < '{end_date}'::timestamp
        GROUP BY tdc.course_name, tms.member_code, tms.member_name_kr
        ORDER BY tdc.course_name, tms.member_name_kr, tms.member_code
        """
        
        cursor.execute(query)
        
        # 결과 가져오기
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        
        cursor.close()
        
        if not rows:
            return [TextContent(type="text", text=f"📊 {period} 기간의 데이터가 없습니다.")]
        
        # 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.title = f"Wordit 사용통계_{start_date[:7]}"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        for col_idx, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        filename = f"Wordit_에듀템_사용통계_{start_date[:7]}.xlsx"
        filepath = output_dir / filename
        
        wb.save(filepath)
        
        result = f"""📊 Wordit 에듀템 사용 통계 조회 완료!

📅 조회 기간: {start_date} ~ {end_date}
👥 조회 결과: {len(rows)}명
💾 저장 위치: {filepath.absolute()}

📈 요약:
- 코스별 학생 수
- 미션 완료 횟수
- 첫 완료일 / 마지막 완료일

✅ 엑셀 파일이 성공적으로 저장되었습니다."""
        
        return [TextContent(type="text", text=result)]
        
    except Exception as e:
        raise ValueError(f"Wordit 에듀템 사용 통계 조회 오류: {str(e)}")
    finally:
        if conn:
            conn.close()


def main():
    """메인 함수 - stdio로 서버를 실행합니다."""
    import sys
    from mcp.server.stdio import stdio_server
    
    # stderr로 로그 출력
    print("Poly MCP Server running on stdio", file=sys.stderr)
    sys.stderr.flush()
    
    # stdio 서버 실행
    async def run():
        async with stdio_server() as (read_stream, write_stream):
            await app.run(
                read_stream,
                write_stream,
                app.create_initialization_options()
            )
    
    asyncio.run(run())


if __name__ == "__main__":
    main()
