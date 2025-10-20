#!/usr/bin/env python3
"""
Poly MCP Server - 다양한 유틸리티 도구를 제공하는 MCP 서버
"""

import asyncio
import calendar
import os
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent, CallToolResult, INVALID_PARAMS, INTERNAL_ERROR


# 서버 인스턴스 생성
app = Server("poly-mcp-server")


@app.list_tools()
async def list_tools() -> list[Tool]:
    """사용 가능한 도구 목록을 반환합니다."""
    return [
        Tool(
            name="postgres_query",
            description="PostgreSQL 데이터베이스에 쿼리를 실행합니다 (환경변수로 연결 정보 설정 필요)",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "실행할 SQL 쿼리 (SELECT만 허용)",
                    },
                    "limit": {
                        "type": "integer",
                        "description": "반환할 최대 행 수",
                        "default": 100,
                    },
                },
                "required": ["query"],
            },
        ),
        Tool(
            name="wordit_edutem_usage",
            description="Wordit 에듀템 사용 통계를 조회하고 엑셀 파일로 저장합니다",
            inputSchema={
                "type": "object",
                "properties": {
                    "start_date": {
                        "type": "string",
                        "description": "시작일 (YYYY-MM-DD 형식, 예: '2025-09-01')",
                    },
                    "end_date": {
                        "type": "string",
                        "description": "종료일 (YYYY-MM-DD 형식, 예: '2025-09-07'). 생략시 시작일이 속한 월의 말일",
                    },
                    "campus_name": {
                        "type": "string",
                        "description": "캠퍼스명 필터 (예: '강남캠퍼스'). 생략시 전체 캠퍼스 조회",
                    },
                    "output_path": {
                        "type": "string",
                        "description": "엑셀 파일 저장 경로 (기본값: ./output)",
                        "default": "./output",
                    },
                },
                "required": ["start_date"],
            },
        ),
    ]


@app.call_tool()
async def call_tool(name: str, arguments: Any) -> list[TextContent]:
    """도구를 호출하고 결과를 반환합니다."""
    try:
        if name == "postgres_query":
            return await handle_postgres_query(arguments)
        elif name == "wordit_edutem_usage":
            return await handle_wordit_edutem_usage(arguments)
        else:
            raise ValueError(f"Unknown tool: {name}")
    except Exception as e:
        raise RuntimeError(f"Tool execution failed: {str(e)}")


async def handle_postgres_query(args: dict) -> list[TextContent]:
    """PostgreSQL 쿼리 실행 도구를 처리합니다."""
    query = args.get("query", "").strip()
    limit = args.get("limit", 100)
    
    # SELECT 쿼리만 허용 (보안)
    if not query.upper().startswith("SELECT"):
        raise ValueError("보안상 SELECT 쿼리만 허용됩니다.")
    
    # 환경변수에서 연결 정보 가져오기
    db_config = {
        "host": os.getenv("POSTGRES_HOST", "localhost"),
        "port": int(os.getenv("POSTGRES_PORT", "5432")),
        "database": os.getenv("POSTGRES_DB"),
        "user": os.getenv("POSTGRES_USER"),
        "password": os.getenv("POSTGRES_PASSWORD"),
    }
    
    # 필수 연결 정보 확인
    if not all([db_config["database"], db_config["user"], db_config["password"]]):
        raise ValueError(
            "PostgreSQL 연결 정보가 설정되지 않았습니다.\n"
            "다음 환경변수를 설정하세요:\n"
            "- POSTGRES_HOST (기본값: localhost)\n"
            "- POSTGRES_PORT (기본값: 5432)\n"
            "- POSTGRES_DB (필수)\n"
            "- POSTGRES_USER (필수)\n"
            "- POSTGRES_PASSWORD (필수)"
        )
    
    conn = None
    try:
        # PostgreSQL 연결
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()
        
        # 쿼리 실행
        cursor.execute(query)
        
        # 결과 가져오기
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchmany(limit)
        
        if not rows:
            result = "🔍 쿼리 결과: 데이터가 없습니다."
        else:
            # 테이블 형식으로 결과 포맷팅
            result_lines = [
                f"🔍 쿼리 결과 ({len(rows)}개 행):",
                "",
                "| " + " | ".join(columns) + " |",
                "|" + "|".join(["---"] * len(columns)) + "|",
            ]
            
            for row in rows:
                formatted_row = [str(val) if val is not None else "NULL" for val in row]
                result_lines.append("| " + " | ".join(formatted_row) + " |")
            
            if len(rows) == limit:
                result_lines.append("")
                result_lines.append(f"⚠️ 결과가 {limit}개로 제한되었습니다.")
            
            result = "\n".join(result_lines)
        
        cursor.close()
        
        return [TextContent(type="text", text=result)]
        
    except Exception as e:
        raise ValueError(f"PostgreSQL 오류: {str(e)}")
    finally:
        if conn:
            conn.close()


async def handle_wordit_edutem_usage(args: dict) -> list[TextContent]:
    """Wordit 에듀템 사용 통계를 조회하고 엑셀로 저장합니다."""
    start_date = args.get("start_date", "")
    end_date = args.get("end_date", "")
    campus_name = args.get("campus_name", "")
    output_path = args.get("output_path", "./output")
    
    if not start_date:
        raise ValueError("시작일을 입력해주세요 (예: '2025-09-01')")
    
    # 날짜 형식 검증 및 파싱
    try:
        from dateutil import parser
        from dateutil.relativedelta import relativedelta
        import calendar
        
        start_dt = parser.parse(start_date)
        start_date_str = start_dt.strftime('%Y-%m-%d')
        
        if end_date:
            end_dt = parser.parse(end_date)
            # end_date를 포함하기 위해 +1일 추가
            end_dt = end_dt + relativedelta(days=1)
            end_date_str = end_dt.strftime('%Y-%m-%d')
        else:
            # end_date가 없으면 해당 월의 말일까지 (+1일)
            last_day = calendar.monthrange(start_dt.year, start_dt.month)[1]
            end_dt = datetime(start_dt.year, start_dt.month, last_day) + relativedelta(days=1)
            end_date_str = end_dt.strftime('%Y-%m-%d')
    except Exception as e:
        raise ValueError(f"날짜 형식 오류: {str(e)}\n형식: YYYY-MM-DD (예: 2025-09-01)")
    
    # 환경변수에서 연결 정보 가져오기
    db_config = {
        "host": os.getenv("POSTGRES_HOST", "localhost"),
        "port": int(os.getenv("POSTGRES_PORT", "5432")),
        "database": os.getenv("POSTGRES_DB"),
        "user": os.getenv("POSTGRES_USER"),
        "password": os.getenv("POSTGRES_PASSWORD"),
    }
    
    # 필수 연결 정보 확인
    if not all([db_config["database"], db_config["user"], db_config["password"]]):
        raise ValueError("PostgreSQL 연결 정보가 설정되지 않았습니다.")
    
    conn = None
    try:
        # PostgreSQL 연결
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()
        
        # 캠퍼스 필터 조건 추가
        campus_filter = ""
        if campus_name:
            campus_filter = f"AND tcc.client_name_kr LIKE '%{campus_name}%'"
        
        # 쿼리 실행
        query = f"""
        SELECT tdc.course_name
            , tms.member_name_kr
            , tms.member_code
            , tcc.client_name_kr as campus_name
            , count(trmr.created_at) as mission_count
            , min(trmr.created_at) as first_complete
            , max(trmr.created_at) as last_complete
        FROM voca_wordit.tb_rea_mission_result trmr
            JOIN voca_wordit.tb_mba_student tms ON trmr.member_code = tms.member_code
            JOIN voca_wordit.tb_dfa_course tdc ON tdc.course_code = tms.course_code
            JOIN voca_wordit.tb_cla_client tcc ON tcc.client_code = tms.client_code
        WHERE trmr.mission_type = 'WB'
            AND tms.client_code NOT IN ('0508003', '1301003')
            {campus_filter}
            AND trmr.created_at >= '{start_date_str}'::timestamp
            AND trmr.created_at < '{end_date_str}'::timestamp
        GROUP BY tdc.course_name, tms.member_code, tms.member_name_kr, tcc.client_name_kr
        ORDER BY tcc.client_name_kr, tdc.course_name, tms.member_name_kr, tms.member_code
        """
        
        cursor.execute(query)
        
        # 결과 가져오기
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        
        cursor.close()
        
        if not rows:
            campus_msg = f" (캠퍼스: {campus_name})" if campus_name else ""
            return [TextContent(type="text", text=f"📊 {start_date_str} ~ {end_date_str} 기간{campus_msg}의 데이터가 없습니다.")]
        
        # 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        
        # 파일명 생성
        period_str = f"{start_date_str}_to_{end_date_str}".replace('-', '')
        campus_suffix = f"_{campus_name}" if campus_name else ""
        ws.title = f"Wordit_{period_str[:8]}"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        for col_idx, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        filename = f"Wordit_에듀템_사용통계_{period_str}{campus_suffix}.xlsx"
        filepath = output_dir / filename
        
        wb.save(filepath)
        
        campus_info = f"\n🏫 캠퍼스: {campus_name}" if campus_name else "\n🏫 캠퍼스: 전체"
        
        result = f"""📊 Wordit 에듀템 사용 통계 조회 완료!

📅 조회 기간: {start_date_str} ~ {end_date_str}{campus_info}
👥 조회 결과: {len(rows)}명
💾 저장 위치: {filepath.absolute()}

📈 포함 정보:
- 캠퍼스명
- 코스명
- 학생명/코드
- 미션 완료 횟수
- 첫 완료일 / 마지막 완료일

✅ 엑셀 파일이 성공적으로 저장되었습니다."""
        
        return [TextContent(type="text", text=result)]
        
    except Exception as e:
        raise ValueError(f"Wordit 에듀템 사용 통계 조회 오류: {str(e)}")
    finally:
        if conn:
            conn.close()


def main():
    """메인 함수 - stdio로 서버를 실행합니다."""
    import sys
    from mcp.server.stdio import stdio_server
    
    # stderr로 로그 출력
    print("Poly MCP Server running on stdio", file=sys.stderr)
    sys.stderr.flush()
    
    # stdio 서버 실행
    async def run():
        async with stdio_server() as (read_stream, write_stream):
            await app.run(
                read_stream,
                write_stream,
                app.create_initialization_options()
            )
    
    asyncio.run(run())


if __name__ == "__main__":
    main()
