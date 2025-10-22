# SPDX-FileCopyrightText: 2025 German Aerospace Center (DLR)
#
# SPDX-License-Identifier: BSD-3-Clause

import os
import pathlib
from io import BytesIO
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from flowstrider.converters import metadata_xsxl_converter
from flowstrider.models.common_models import Cluster, Edge, Node
from flowstrider.models.dataflowdiagram import DataflowDiagram
from flowstrider.rules import attributes_dict

__location__ = os.path.dirname(__file__)


@pytest.mark.parametrize(
    "tags, expected_type",
    [
        (Node("id1", "", ["STRIDE:Interactor"]), "Node: Interactor"),
        (Node("id2", "", ["STRIDE:DataStore"]), "Node: DataStore"),
        (Node("id3", "", ["STRIDE:Process"]), "Node: Process"),
        (Edge("id4", "", ["STRIDE:Dataflow"]), "Edge: Dataflow"),
        (Cluster("id5", [], "", ["STRIDE:TrustBoundary"]), "TrustBoundary"),
        ([], "Not defined"),
        (["UnknownTag"], "Not defined"),
    ],
)
def test_determine_type(tags, expected_type):
    assert metadata_xsxl_converter.determine_type(tags) == expected_type


def test_metadata_check():
    attributes = {}
    attributes["key1"] = attributes_dict.Attribute(
        "key1name", "explanation1", ["Node: Process"], [], ["test_rules"]
    )
    attributes["key2"] = attributes_dict.Attribute(
        "key2name", "explanation2", ["Node: Process"], [], ["test_rules"]
    )
    attributes["key3"] = attributes_dict.Attribute(
        "key3name", "explanation3", ["Edge"], [], ["test_rules"]
    )
    attributes["key4"] = attributes_dict.Attribute(
        "key4name", "explanation4", ["Cluster"], [], ["test_rules"]
    )

    node1 = Node(
        id="node1",
        name="Node 1",
        tags=["STRIDE:Process"],
        attributes={"key1": "value1"},
    )
    node2 = Node(id="node2", name="Node 2", tags=[], attributes={})

    dfd = DataflowDiagram(
        id="test_diagram",
        clusters={},
        nodes={"node1": node1, "node2": node2},
        edges={},
        tags=["test_rules"],
    )

    with (
        patch(
            "flowstrider.rules.attributes_dict.attributes",
            attributes,
        ),
        patch(
            "flowstrider.converters.metadata_xsxl_converter.Workbook", autospec=True
        ) as MockWorkbook,
    ):
        mock_workbook = MockWorkbook.return_value
        mock_worksheet = mock_workbook.active

        metadata_xsxl_converter.metadata_check(
            dfd, pathlib.Path("output/test_diagram_metadata_overview.xlsx")
        )

        MockWorkbook.assert_called_once()
        mock_workbook.save.assert_called_once_with(
            pathlib.Path("output/test_diagram_metadata_overview.xlsx")
        )

        expected_header = [
            "Entity Type",
            "Entity ID",
            "Entity Name",
            "key1name",
            "key2name",
            "key3name",
            "key4name",
        ]
        for col_num, value in enumerate(expected_header, start=1):
            assert mock_worksheet.cell.call_args_list[col_num - 1][1]["value"] == value

        expected_explanations = [
            "",
            "",
            "",
            "explanation1",
            "explanation2",
            "explanation3",
            "explanation4",
        ]
        for col_num, value in enumerate(expected_explanations, start=1):
            assert (
                mock_worksheet.cell.call_args_list[
                    len(expected_header) * 2 + col_num - 1
                ][1]["value"]
                == value
            )

        # Filter out .cell calls that don't have 3 arguments
        # ...(the .cell(x,y).border() calls)
        cell_write_call_args_list = [
            entry for entry in mock_worksheet.cell.call_args_list if len(entry[1]) == 3
        ]

        expected_row = ["Node: Process", "node1", "Node 1", "value1", "Missing", "", ""]
        for col_num, value in enumerate(expected_row, start=1):
            list_index = len(expected_header) * 3 + col_num - 1
            assert cell_write_call_args_list[list_index][1]["value"] == value


def test_metadata_check_conditional_formating():
    node1 = Node(
        id="node1",
        name="Node 1",
        tags=["STRIDE:DataStore"],
        attributes={"encryption_method": "AES_256"},
    )
    edge1 = Edge(
        id="edge1",
        name="Edge 1",
        source_id="node1",
        sink_id="node1",
        tags=["STRIDE:Dataflow"],
        attributes={},
    )

    dfd = DataflowDiagram(
        id="test_diagram",
        clusters={},
        nodes={"node1": node1},
        edges={"edge1": edge1},
        tags={"stride", "bsi_rules", "linddun_rules"},
    )

    metadata_xsxl_converter.metadata_check(
        dfd, pathlib.Path("output/test_diagram_metadata_overview.xlsx")
    )

    test_diagram_path = os.path.join(
        __location__, "../../output/test_diagram_metadata_overview.xlsx"
    )
    assert os.path.exists(test_diagram_path)

    workbook = load_workbook(filename=test_diagram_path)
    worksheet = workbook.active

    # Check that all entities exist in the metadata overview
    assert worksheet["B4"].value == node1.id
    assert worksheet["B5"].value == edge1.id

    # Get all conditional formating rules
    cond_formating_rules = []
    for range_string in worksheet.conditional_formatting._cf_rules:
        for cfRule in worksheet.conditional_formatting._cf_rules[range_string]:
            cond_formating_rules.append(cfRule.formula)

    # Find columns
    pos_encryption_method = -1
    pos_handles_confidential_data = -1
    for i in range(len(attributes_dict.attributes)):
        if worksheet[get_column_letter(i + 4) + "1"].value == "Encryption method":
            pos_encryption_method = i + 4

        elif (
            worksheet[get_column_letter(i + 4) + "1"].value
            == "Handles confidential data"
        ):
            pos_handles_confidential_data = i + 4

    assert pos_encryption_method > 3
    assert pos_handles_confidential_data > 3

    # Check content of cells
    assert worksheet[get_column_letter(pos_encryption_method) + "4"].value == "AES_256"
    assert (
        worksheet[get_column_letter(pos_handles_confidential_data) + "4"].value
        in "Missing"
    )

    # Check that attribute cells have formatting
    found_rule_encryption_method = False
    found_rule_handles_confidential_data = False
    for i in range(len(cond_formating_rules)):
        if (
            not found_rule_encryption_method
            and get_column_letter(pos_encryption_method) + "4"
            in cond_formating_rules[i][0]
        ):
            found_rule_encryption_method = True
            assert ('=OR(ISNUMBER(SEARCH("missing",{col}4)),ISBLANK({col}4))').format(
                col=get_column_letter(pos_encryption_method)
            ) in cond_formating_rules[i][0]
            assert ('"AES-256",{col}4').format(
                col=get_column_letter(pos_encryption_method)
            ) in cond_formating_rules[i + 1][0]
            assert ('"AES256",{col}4').format(
                col=get_column_letter(pos_encryption_method)
            ) in cond_formating_rules[i + 2][0]

        if (
            not found_rule_handles_confidential_data
            and get_column_letter(pos_handles_confidential_data) + "5"
            in cond_formating_rules[i][0]
        ):
            found_rule_handles_confidential_data = True
            assert ('=OR(ISNUMBER(SEARCH("missing",{col}5)),ISBLANK({col}5))').format(
                col=get_column_letter(pos_handles_confidential_data)
            ) in cond_formating_rules[i][0]
            assert ('"True",{col}5').format(
                col=get_column_letter(pos_handles_confidential_data)
            ) in cond_formating_rules[i + 1][0]
            assert ('"no",{col}5').format(
                col=get_column_letter(pos_handles_confidential_data)
            ) in cond_formating_rules[i + 2][0]

    assert found_rule_encryption_method
    assert found_rule_handles_confidential_data


def test_parse_value():
    # Test boolean parsing
    assert metadata_xsxl_converter.parse_value("true") is True
    assert metadata_xsxl_converter.parse_value("yes") is True
    assert metadata_xsxl_converter.parse_value("1") is True
    assert metadata_xsxl_converter.parse_value("false") is False
    assert metadata_xsxl_converter.parse_value("no") is False
    assert metadata_xsxl_converter.parse_value("0") is False

    # Test JSON parsing
    assert metadata_xsxl_converter.parse_value('{"key": "value"}') == {"key": "value"}
    assert metadata_xsxl_converter.parse_value("[1, 2, 3]") == [1, 2, 3]

    # Test list parsing
    assert metadata_xsxl_converter.parse_value("item1, item2, item3") == [
        "item1",
        "item2",
        "item3",
    ]

    # Test non-string input (returned as is)
    assert metadata_xsxl_converter.parse_value(42) == 42
    assert metadata_xsxl_converter.parse_value(None) is None

    # Test normal string (returned as is)
    assert (
        metadata_xsxl_converter.parse_value("unparseable string")
        == "unparseable string"
    )


def test_update_dfd_json_from_xlsx():
    node1 = Node(id="node1", name="Node 1", tags=["STRIDE:Process"], attributes={})

    dfd = DataflowDiagram(
        id="test_diagram",
        clusters={},
        nodes={"node1": node1},
        edges={},
    )

    # Create an in-memory Excel file
    workbook = Workbook()
    sheet = workbook.active

    # Add headers
    headers = ["Entity Type", "Entity ID", "Entity Name", "Attribute 1", "Attribute 2"]
    keys = ["", "", "", "attribute_1_key", "attribute_2_key"]
    explanations = ["", "", "", "explanation1", "explanation2"]
    sheet.append(headers)
    sheet.append(keys)
    sheet.append(explanations)

    # Add data rows
    sheet.append(["Node", "node1", "Node 1", "value1", "value2"])

    # Save workbook to in-memory file
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Call function
    updated_dfd = metadata_xsxl_converter.update_dfd_json_from_xlsx(dfd, excel_file)

    # Validate updates
    assert updated_dfd.nodes["node1"].attributes == {
        "attribute_1_key": "value1",
        "attribute_2_key": "value2",
    }
