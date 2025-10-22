import argparse
import polars as pl
import re
import pyarrow
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from pathlib import Path
import yaml
from dbqt.tools.utils import Timer, setup_logging
import logging

logger = logging.getLogger(__name__)

# Define default type mappings for equivalent data types
DEFAULT_TYPE_MAPPINGS = {
    "INTEGER": ["INT", "INTEGER", "BIGINT", "SMALLINT", "TINYINT", "NUMBER"],
    "VARCHAR": ["VARCHAR", "TEXT", "CHAR", "STRING", "NVARCHAR", "VARCHAR2"],
    "DECIMAL": ["DECIMAL", "NUMERIC", "NUMBER"],
    "FLOAT": ["FLOAT", "REAL", "DOUBLE", "DOUBLE PRECISION"],
    "TIMESTAMP": ["TIMESTAMP", "DATETIME"],
    "DATE": ["DATE", "TIMESTAMP"],
    "BOOLEAN": ["BOOLEAN", "BOOL", "BIT"],
}


def load_type_mappings(config_path=None):
    """Load type mappings from YAML config file or return defaults"""
    if config_path and Path(config_path).exists():
        logger.info(f"Loading type mappings from {config_path}")
        with open(config_path, "r") as f:
            config = yaml.safe_load(f)
            return config.get("type_mappings", DEFAULT_TYPE_MAPPINGS)
    return DEFAULT_TYPE_MAPPINGS


def generate_config_file(output_path="colcompare_config.yaml"):
    """Generate a default configuration file with type mappings"""
    config = {
        "type_mappings": DEFAULT_TYPE_MAPPINGS,
        "description": "Column comparison type mappings configuration. "
        "Each key represents a type group, and the list contains equivalent types.",
    }

    output_file = Path(output_path)
    if output_file.exists():
        logger.warning(f"Config file already exists at {output_path}")
        response = input("Overwrite? (y/n): ")
        if response.lower() != "y":
            logger.info("Config generation cancelled")
            return

    with open(output_path, "w") as f:
        yaml.dump(config, f, default_flow_style=False, sort_keys=False)

    logger.info(f"Generated default config file at {output_path}")
    print(f"\nDefault configuration saved to: {output_path}")
    print("You can now edit this file to customize type mappings.")


def are_types_compatible(type1, type2, type_mappings=None):
    """Check if two data types are considered compatible"""
    if type_mappings is None:
        type_mappings = DEFAULT_TYPE_MAPPINGS

    type1, type2 = type1.upper(), type2.upper()

    # Strip length specifications like VARCHAR(50) to VARCHAR
    type1 = type1.split("(")[0].strip()
    type2 = type2.split("(")[0].strip()

    # If types are exactly the same, they're compatible
    if type1 == type2:
        return True

    # Special handling for TIMESTAMP variations
    if re.match(r"^TIMESTAMP.*", type1) and re.match(r"^TIMESTAMP.*", type2):
        return True

    # Check if types belong to the same group
    for type_group in type_mappings.values():
        if type1 in type_group and type2 in type_group:
            return True

    return False


def _process_nested_type(field_type, parent_name="", processed_fields=None):
    """Recursively process nested types in Parquet schema"""
    if processed_fields is None:
        processed_fields = []

    # Handle list types
    if isinstance(field_type, (pyarrow.lib.ListType, pyarrow.lib.LargeListType)):
        element_type = field_type.value_type
        if isinstance(element_type, pyarrow.lib.StructType):
            for nested_field in element_type:
                full_name = (
                    f"{parent_name}__{nested_field.name}"
                    if parent_name
                    else nested_field.name
                )
                if isinstance(
                    nested_field.type,
                    (
                        pyarrow.lib.ListType,
                        pyarrow.lib.LargeListType,
                        pyarrow.lib.StructType,
                        pyarrow.lib.MapType,
                    ),
                ):
                    _process_nested_type(nested_field.type, full_name, processed_fields)
                else:
                    processed_fields.append(
                        {"col_name": full_name, "type": str(nested_field.type)}
                    )
        else:
            processed_fields.append({"col_name": parent_name, "type": str(field_type)})

    # Handle struct types
    elif isinstance(field_type, pyarrow.lib.StructType):
        for nested_field in field_type:
            full_name = (
                f"{parent_name}__{nested_field.name}"
                if parent_name
                else nested_field.name
            )
            if isinstance(
                nested_field.type,
                (
                    pyarrow.lib.ListType,
                    pyarrow.lib.LargeListType,
                    pyarrow.lib.StructType,
                    pyarrow.lib.MapType,
                ),
            ):
                _process_nested_type(nested_field.type, full_name, processed_fields)
            else:
                processed_fields.append(
                    {"col_name": full_name, "type": str(nested_field.type)}
                )

    # Handle map types
    elif isinstance(field_type, pyarrow.lib.MapType):
        processed_fields.append({"col_name": parent_name, "type": str(field_type)})

    return processed_fields


def compare_and_unnest_parquet_schema(source_path, target_path):
    """Compare schemas of two Parquet files without loading full dataset"""
    schema1 = pq.read_schema(source_path)
    schema2 = pq.read_schema(target_path)

    # Process both schemas to expand nested fields
    processed_fields1 = []
    processed_fields2 = []

    for field in schema1:
        if isinstance(
            field.type,
            (
                pyarrow.lib.ListType,
                pyarrow.lib.LargeListType,
                pyarrow.lib.StructType,
                pyarrow.lib.MapType,
            ),
        ):
            processed_fields1.extend(_process_nested_type(field.type, field.name))
        else:
            processed_fields1.append({"col_name": field.name, "type": str(field.type)})

    for field in schema2:
        if isinstance(
            field.type,
            (
                pyarrow.lib.ListType,
                pyarrow.lib.LargeListType,
                pyarrow.lib.StructType,
                pyarrow.lib.MapType,
            ),
        ):
            processed_fields2.extend(_process_nested_type(field.type, field.name))
        else:
            processed_fields2.append({"col_name": field.name, "type": str(field.type)})

    # Convert processed fields to polars DataFrame
    schema1_df = pl.DataFrame(
        {
            "SCH_TABLE": ["pq"] * len(processed_fields1),
            "COL_NAME": [f["col_name"] for f in processed_fields1],
            "DATA_TYPE": [f["type"] for f in processed_fields1],
        }
    )

    schema2_df = pl.DataFrame(
        {
            "SCH_TABLE": ["pq"] * len(processed_fields2),
            "COL_NAME": [f["col_name"] for f in processed_fields2],
            "DATA_TYPE": [f["type"] for f in processed_fields2],
        }
    )

    return schema1_df, schema2_df


def read_files(source_path, target_path):
    """Read source and target files using Polars"""
    # Check if files are Parquet
    if source_path.endswith(".parquet") and target_path.endswith(".parquet"):
        # For Parquet files, only load the schema information
        source_df, target_df = compare_and_unnest_parquet_schema(
            source_path, target_path
        )
    else:
        source_df = pl.read_csv(source_path)
        target_df = pl.read_csv(target_path)

        # Handle missing DATA_TYPE column
        if "DATA_TYPE" not in source_df.columns:
            source_df = source_df.with_columns(pl.lit("N/A").alias("DATA_TYPE"))
        if "DATA_TYPE" not in target_df.columns:
            target_df = target_df.with_columns(pl.lit("N/A").alias("DATA_TYPE"))

        # Create SCH_TABLE column - concatenate SCH and NAME if SCH exists, otherwise use NAME
        if "SCH" in source_df.columns:
            source_df = source_df.with_columns(
                pl.concat_str([pl.col("SCH"), pl.lit("."), pl.col("TABLE_NAME")]).alias(
                    "SCH_TABLE"
                )
            )
        else:
            source_df = source_df.with_columns(pl.col("TABLE_NAME").alias("SCH_TABLE"))

        if "SCH" in target_df.columns:
            target_df = target_df.with_columns(
                pl.concat_str([pl.col("SCH"), pl.lit("."), pl.col("TABLE_NAME")]).alias(
                    "SCH_TABLE"
                )
            )
        else:
            target_df = target_df.with_columns(pl.col("TABLE_NAME").alias("SCH_TABLE"))

    return source_df, target_df


def compare_tables(source_df, target_df):
    """Compare tables between source and target"""
    source_tables = set(source_df["SCH_TABLE"].unique())
    target_tables = set(target_df["SCH_TABLE"].unique())

    common_tables = source_tables.intersection(target_tables)
    source_only = source_tables - target_tables
    target_only = target_tables - source_tables

    return {
        "common": sorted(list(common_tables)),
        "source_only": sorted(list(source_only)),
        "target_only": sorted(list(target_only)),
    }


def compare_columns(source_df, target_df, table_name, type_mappings=None):
    """Compare columns for a specific table"""
    source_cols = source_df.filter(pl.col("SCH_TABLE") == table_name).select(
        ["COL_NAME", "DATA_TYPE"]
    )
    target_cols = target_df.filter(pl.col("SCH_TABLE") == table_name).select(
        ["COL_NAME", "DATA_TYPE"]
    )

    source_cols_set = set(source_cols["COL_NAME"].to_list())
    target_cols_set = set(target_cols["COL_NAME"].to_list())

    common_cols = source_cols_set.intersection(target_cols_set)
    source_only = source_cols_set - target_cols_set
    target_only = target_cols_set - source_cols_set

    # Compare data types for common columns
    datatype_mismatches = []
    for col in common_cols:
        source_type = source_cols.filter(pl.col("COL_NAME") == col)["DATA_TYPE"].item()
        target_type = target_cols.filter(pl.col("COL_NAME") == col)["DATA_TYPE"].item()
        if not are_types_compatible(source_type, target_type, type_mappings):
            datatype_mismatches.append(
                {"column": col, "source_type": source_type, "target_type": target_type}
            )

    return {
        "common": sorted(list(common_cols)),
        "source_only": sorted(list(source_only)),
        "target_only": sorted(list(target_only)),
        "datatype_mismatches": datatype_mismatches,
    }


def format_worksheet(ws):
    """Apply formatting to worksheet"""
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Calculate width: add 2 for padding, cap at 21.6 (3 inches)
        # If content is smaller, use the smaller width
        adjusted_width = min(max_length + 2, 21.6)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


def create_excel_report(
    comparison_results, source_df, target_df, file_name, type_mappings=None
):
    """Create formatted Excel report"""
    wb = Workbook()

    # Table Comparison Sheet
    ws_tables = wb.active
    ws_tables.title = "Table Comparison"
    ws_tables.append(["Category", "Table Name"])

    for table in comparison_results["tables"]["common"]:
        ws_tables.append(["Common", table])
    for table in comparison_results["tables"]["source_only"]:
        ws_tables.append(["Source Only", table])
    for table in comparison_results["tables"]["target_only"]:
        ws_tables.append(["Target Only", table])

    format_worksheet(ws_tables)

    # Column Comparison Sheet
    ws_columns = wb.create_sheet("Column Comparison")
    ws_columns.append(
        ["Table Name", "Column Name", "Status", "Source Type", "Target Type"]
    )

    for table in comparison_results["columns"]:
        # Get all columns from source and target for this table
        table_name = table["table_name"]
        source_cols = source_df.filter(pl.col("SCH_TABLE") == table_name).select(
            ["COL_NAME", "DATA_TYPE"]
        )
        target_cols = target_df.filter(pl.col("SCH_TABLE") == table_name).select(
            ["COL_NAME", "DATA_TYPE"]
        )

        # Create dictionaries for easy lookup
        source_types = dict(zip(source_cols["COL_NAME"], source_cols["DATA_TYPE"]))
        target_types = dict(zip(target_cols["COL_NAME"], target_cols["DATA_TYPE"]))

        # Process all columns
        all_columns = sorted(set(list(source_types.keys()) + list(target_types.keys())))

        for col in all_columns:
            source_type = source_types.get(col, "N/A")
            target_type = target_types.get(col, "N/A")

            if col in table["source_only"]:
                status = "Source Only"
            elif col in table["target_only"]:
                status = "Target Only"
            else:  # Column exists in both
                if are_types_compatible(source_type, target_type, type_mappings):
                    status = "Matching"
                else:
                    status = "Different Types"

            ws_columns.append(
                [table["table_name"], col, status, source_type, target_type]
            )

    format_worksheet(ws_columns)

    # Datatype Mismatches Sheet
    ws_datatypes = wb.create_sheet("Datatype Mismatches")
    ws_datatypes.append(["Table Name", "Column Name", "Source Type", "Target Type"])

    for table in comparison_results["columns"]:
        for mismatch in table["datatype_mismatches"]:
            ws_datatypes.append(
                [
                    table["table_name"],
                    mismatch["column"],
                    mismatch["source_type"],
                    mismatch["target_type"],
                ]
            )

    format_worksheet(ws_datatypes)

    # Save the workbook with timestamp
    os.makedirs("results", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if "/" in file_name:
        file_name = file_name.split("/")[-1]
    wb.save(f"results/{file_name}_{timestamp}.xlsx")


def colcompare(args=None):
    if isinstance(args, (list, type(None))):
        # Called from command line
        parser = argparse.ArgumentParser(
            description="Compare column schemas between two CSV or Parquet files",
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog="""
Generates an Excel report with three sheets:
- Table Comparison: Lists tables present in source/target
- Column Comparison: Details of column presence and type matching
- Datatype Mismatches: Highlights columns with incompatible types

The report is saved to ./results/ with a timestamp in the filename.

To generate a default configuration file:
  dbqt colcompare --generate-config [--output PATH]
            """,
        )

        parser.add_argument(
            "--generate-config",
            action="store_true",
            help="Generate a default type mappings configuration file",
        )
        parser.add_argument(
            "--output",
            "-o",
            default="colcompare_config.yaml",
            help="Output path for config file (used with --generate-config)",
        )
        parser.add_argument(
            "source", nargs="?", help="Path to the source CSV/Parquet file"
        )
        parser.add_argument(
            "target", nargs="?", help="Path to the target CSV/Parquet file"
        )
        parser.add_argument(
            "--config", "-c", help="Path to type mappings configuration file"
        )
        parser.add_argument(
            "--verbose", "-v", action="store_true", help="Verbose logging"
        )

        args = parser.parse_args(args)

        # Setup logging
        setup_logging(args.verbose)

        # Handle generate-config command
        if args.generate_config:
            generate_config_file(args.output)
            return

        # Validate that source and target are provided for comparison
        if not args.source or not args.target:
            parser.error("source and target arguments are required for comparison")

    # Load type mappings
    type_mappings = load_type_mappings(getattr(args, "config", None))

    with Timer("Column comparison"):
        # Read source and target files
        source_df, target_df = read_files(args.source, args.target)

        # Compare tables
        table_comparison = compare_tables(source_df, target_df)

        # Compare columns for common tables
        column_comparisons = []
        for table in table_comparison["common"]:
            column_comparison = compare_columns(
                source_df, target_df, table, type_mappings
            )
            column_comparisons.append({"table_name": table, **column_comparison})

        # Create comparison results dictionary
        comparison_results = {"tables": table_comparison, "columns": column_comparisons}
        target_file_name = args.target.split(".")[0]
        # Generate Excel report
        create_excel_report(
            comparison_results, source_df, target_df, target_file_name, type_mappings
        )


def main(args=None):
    """Entry point for the colcompare tool"""
    colcompare(args)


if __name__ == "__main__":
    main()
