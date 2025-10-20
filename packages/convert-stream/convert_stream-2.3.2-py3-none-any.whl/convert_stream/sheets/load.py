#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
from abc import ABC, abstractmethod
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from pandas.io.parsers import TextFileReader
import pandas as pd
import csv
from soup_files import File, ProgressBarAdapter, CreatePbar


class ABCSheetReader(ABC):
    """Classe abstrata para leitura de planilhas CSV e EXCEL"""

    def __init__(
            self,
            file: File, *,
            pbar: ProgressBarAdapter = CreatePbar().get()
    ):
        self.pbar: ProgressBarAdapter = pbar
        self.file: File = file
        self.df = None
        self.isLoading: bool = False

    @property
    def is_running(self) -> bool:
        return self.isLoading

    @abstractmethod
    def read(self, sheet_name: str = None, *, progress: bool = True) -> None:
        pass

    @abstractmethod
    def get_sheet_names(self) -> list[str]:
        pass

    @abstractmethod
    def get_total_rows(self, sheet_name: str = None, *, progress: bool = False) -> int:
        pass

    @abstractmethod
    def get_total_columns(self, sheet_name: str = None, *, progress: bool = False) -> int:
        pass

    @abstractmethod
    def get_columns(self, sheet_name: str = None, *, progress: bool = False) -> list[str]:
        pass

    @abstractmethod
    def get_dataframe(self, sheet_name: str = None, *, progress: bool = True) -> pd.DataFrame:
        pass


class ReaderCsv(ABCSheetReader):
    def __init__(self, file: File, *, pbar: ProgressBarAdapter = None, separator: str = '\t'):
        super().__init__(file, pbar=pbar)
        self.separator: str = separator
        self.df = pd.DataFrame()

    def get_sheet_names(self) -> list[str]:
        return [self.file.name()]

    def get_total_rows(self, sheet_name: str = None, *, progress: bool = False) -> int:
        try:
            with open(self.file.absolute(), newline='', encoding='utf-8') as f:
                _reader_csv = csv.reader(f, delimiter=self.separator)
                total_lines = sum(1 for _ in _reader_csv)
        except FileNotFoundError:
            self.pbar.update_text(f'Erro: o arquivo não existe: {self.file.basename()}')
            return 0
        except Exception as err:
            self.pbar.update_text(f'{err}')
            return 0
        else:
            return total_lines - 1

    def get_total_columns(self, sheet_name: str = None, *, progress: bool = False) -> int:
        try:
            with open(self.file.absolute(), 'r', newline='', encoding='utf-8') as f:
                leitor = csv.reader(f)
                # Lê a primeira linha do arquivo
                _first_line = next(leitor)
                # O número de colunas é o tamanho da primeira linha
                num_cols = len(_first_line)
        except FileNotFoundError:
            print(f"Erro: O arquivo '{self.file.basename()}' não foi encontrado.")
            return 0
        except StopIteration:
            print(f"Erro: O arquivo '{self.file.basename()}' está vazio.")
            return 0
        except Exception as e:
            print(f"Ocorreu um erro: {e}")
            return 0
        else:
            return num_cols

    def get_columns(self, sheet_name: str = None, *, progress: bool = False) -> list[str]:
        try:
            with open(self.file.absolute(), newline='', encoding='utf-8') as f:
                leitor = csv.reader(f, delimiter=self.separator)
                head_csv = next(leitor)  # lê a primeira linha
        except Exception as err:
            self.pbar.update_text(f'{__class__.__name__}: {err}')
            return []
        else:
            return head_csv

    def read(self, sheet_name: str = None, *, progress: bool = True) -> None:
        self.isLoading = True
        self.pbar.start()
        self.pbar.update(0, "Iniciando leitura do CSV")

        total_lines = self.get_total_rows()
        chunks_text: list[TextFileReader] = []
        for num, chunk in enumerate(
                pd.read_csv(self.file.absolute(), chunksize=1000, sep=self.separator),
        ):
            percent: float = (((num + 1) * 1000) / total_lines) * 100
            self.pbar.update(percent, f"Lendo CSV [{self.file.basename()}]")
            chunks_text.append(chunk)

        if len(chunks_text) > 0:
            self.df = pd.concat(chunks_text, ignore_index=True)
        self.pbar.update(100, "Leitura finalizada!", )
        self.isLoading = False
        self.pbar.stop()

    def get_dataframe(self, sheet_name: str = None, *, progress: bool = True) -> pd.DataFrame:
        if self.df.empty:
            self.read()
        return self.df


class ReaderExcel(ABCSheetReader):

    def __init__(
            self, file: File, *,
            pbar: ProgressBarAdapter = CreatePbar().get()
    ):
        super().__init__(file, pbar=pbar)
        self.df = pd.DataFrame()

    def get_columns(self, sheet_name: str = None, *, progress: bool = False) -> list[str]:
        """Cabeçalho original da planilha"""
        ws = self.get_work_sheet(sheet_name)
        if ws is None:
            return list()
        return list(next(ws.iter_rows(values_only=True)))

    def get_total_rows(self, sheet_name: str = None, progress: bool = False) -> int:
        return self.get_work_sheet(sheet_name, progress).max_row

    def get_total_columns(self, sheet_name: str = None, progress: bool = False) -> int:
        return self.get_work_sheet(sheet_name, progress).max_column

    def get_sheet_names(self) -> list[str]:
        try:
            return self.get_workbook().sheetnames
        except Exception as e:
            print(e)
            return []

    def get_workbook(self, progress: bool = True) -> Workbook | None:
        self.pbar.start()
        if progress:
            self.pbar.update_text(f'Lendo Excel: {self.file.basename()}')
        try:
            excel_wb: Workbook = load_workbook(self.file.absolute(), read_only=True)
        except Exception as e:
            if progress:
                self.pbar.update_text(f'{__class__.__name__} -> {e}\n')
            self.pbar.stop()
            return None
        else:
            self.pbar.stop()
            return excel_wb

    def get_active_sheet(self, progress: bool = True) -> ReadOnlyWorksheet | None:
        return self.get_workbook(progress).active

    def get_work_sheet(self, sheet_name: str = None, progress: bool = True) -> ReadOnlyWorksheet | None:
        file_workbook = self.get_workbook(progress)
        if file_workbook is None:
            return None

        if sheet_name is None:
            return file_workbook.active
        else:
            _names = self.get_sheet_names()
            if not sheet_name in _names:
                return None
            return file_workbook[sheet_name]

    def read(self, sheet_name: str = None, progress: bool = True) -> None:
        if progress:
            return self._read_yes_progress(sheet_name)
        return self._read_no_progress(sheet_name)

    def _read_yes_progress(self, sheet_name: str = None) -> None:
        self.isLoading = True
        self.pbar.start()
        self.pbar.update(0, "Iniciando leitura do Excel")
        ws: ReadOnlyWorksheet = self.get_work_sheet(sheet_name)
        if ws is None:
            self.pbar.update(0, "Falha na leitura")
            self.isLoading = False
            self.pbar.stop()
            return

        list_data: list[tuple] = []
        _rows = ws.iter_rows(values_only=True)
        max_num: int = self.get_total_rows()
        for num, row in enumerate(_rows, 1):
            self.pbar.update(
                ((num + 1) / max_num) * 100,
                f'Lendo planilha: {self.get_work_sheet().title} | Linhas [{num + 1} de {max_num}]'
            )
            list_data.append(row)

        if len(list_data) > 0:
            self.df = pd.DataFrame(list_data)
        if (self.get_columns() is not None) and (self.get_columns() != []):
            self.df.columns = self.get_columns(sheet_name)
        print()
        self.pbar.update(100, 'Operação finalizada!!!')
        self.isLoading = False
        self.pbar.stop()

    def _read_no_progress(self, sheet_name: str = None) -> None:
        self.isLoading = True
        self.pbar.start()
        try:
            if sheet_name is None:
                self.df = pd.read_excel(self.file.absolute())
            else:
                self.df = pd.read_excel(self.file.absolute(), sheet_name=sheet_name)
        except Exception as e:
            print(__class__.__name__, e)
        finally:
            self.isLoading = False
            self.pbar.stop()

    def get_dataframe(self, sheet_name: str = None, *, progress: bool = True) -> pd.DataFrame:
        if self.df.empty:
            self.read(sheet_name, progress=progress)
        return self.df


class ReaderOds(ABCSheetReader):

    def __init__(self, file: File, *, pbar: ProgressBarAdapter = CreatePbar().get()):
        super().__init__(file, pbar=pbar)
        self.df = pd.DataFrame()

    def get_columns(self, sheet_name: str = None, *, progress: bool = False) -> list[str]:
        self.read(sheet_name, progress=progress)
        try:
            return self.df.columns.to_list()
        except Exception as err:
            print(f'Erro ao ler cabeçalho: {err}')
            return []

    def get_total_rows(self, sheet_name: str = None, *, progress: bool = False) -> int:
        self.read(sheet_name, progress=progress)
        try:
            return self.df.shape[0]
        except Exception as err:
            print(f'Erro ao contar linhas: {err}')
            return 0

    def get_total_columns(self, sheet_name: str = None, *, progress: bool = False) -> int:
        cols = self.get_columns(sheet_name, progress=progress)
        return len(cols)

    def get_sheet_names(self) -> list[str]:
        try:
            xls = pd.ExcelFile(self.file.absolute(), engine='odf')
            return xls.sheet_names
        except Exception as err:
            print(f'Erro ao obter nomes das abas: {err}')
            return []

    def read(self, sheet_name: str = None, *, progress: bool = False) -> None:
        if not self.df.empty:
            return

        self.isLoading = True
        self.pbar.start()
        if progress:
            self.pbar.update(0, f'Lendo ODS: {self.file.basename()}')

        try:
            if sheet_name is None:
                self.df = pd.read_excel(self.file.absolute(), engine='odf')
            else:
                self.df = pd.read_excel(self.file.absolute(), sheet_name=sheet_name, engine='odf')
        except Exception as err:
            self.pbar.update_text(f'Erro ao ler ODS com progresso: {err}')
        finally:
            print()
            if progress:
                self.pbar.update(100, 'Operação finalizada!')
            self.pbar.stop()
            self.isLoading = False

    def get_dataframe(self, sheet_name: str = None, *, progress: bool = True) -> pd.DataFrame:
        self.read(progress=progress)
        return self.df


class ReadFileSheet(object):

    def __init__(
            self, file: File, *,
            separator: str = '\t',
            pbar: ProgressBarAdapter = CreatePbar().get()
    ):
        if file.is_csv():
            self.sheet_reader: ABCSheetReader = ReaderCsv(file, pbar=pbar, separator=separator)
        elif file.is_excel():
            self.sheet_reader: ABCSheetReader = ReaderExcel(file, pbar=pbar)
        elif file.is_ods():
            self.sheet_reader: ABCSheetReader = ReaderOds(file, pbar=pbar)
        else:
            raise NotImplementedError(f'Tipo de planilha desconhecida: {file.basename()}')

    def read(self, sheet_name: str = None, *, progress: bool = False) -> None:
        return self.sheet_reader.read(sheet_name, progress=progress)

    def get_sheet_names(self) -> list[str]:
        return self.sheet_reader.get_sheet_names()

    def get_total_rows(self, sheet_name: str = None, *, progress: bool = False) -> int:
        return self.sheet_reader.get_total_rows(sheet_name, progress=progress)

    def get_total_columns(self, sheet_name: str = None, *, progress: bool = False) -> int:
        return self.sheet_reader.get_total_columns(sheet_name, progress=progress)

    def get_columns(self, sheet_name: str = None, *, progress: bool = False) -> list[str]:
        return self.sheet_reader.get_columns(sheet_name, progress=progress)

    def get_dataframe(self, sheet_name: str = None, *, progress: bool = True) -> pd.DataFrame:
        return self.sheet_reader.get_dataframe(sheet_name, progress=progress)
