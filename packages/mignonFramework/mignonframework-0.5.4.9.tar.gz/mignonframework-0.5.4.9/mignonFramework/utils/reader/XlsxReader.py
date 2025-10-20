import os
import random
from typing import List, Iterator, Tuple, Dict, Any

from mignonFramework.utils.reader.BaseReader import BaseReader
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')
try:
    import openpyxl
except ImportError:
    openpyxl = None


class XlsxReader(BaseReader):
    """
    一个具体的Reader实现，用于读取XLSX格式的Excel文件（.xlsx）。
    - 默认读取工作簿中的活动工作表 (active sheet)。
    - 允许指定标题行（header row）的位置。
    """

    def __init__(self, path: str, header_row: int = 1):
        """
        初始化XlsxReader。

        Args:
            path (str): 要读取的 .xlsx 文件路径或包含 .xlsx 文件的目录路径。
            header_row (int): 标题所在的行号（从1开始）。默认为1。

        Raises:
            ImportError: 如果 'openpyxl' 库未安装。
            ValueError: 如果 header_row 小于1。
        """
        if openpyxl is None:
            raise ImportError("使用 XlsxReader 需要 'openpyxl' 库。请通过 'pip install openpyxl' 安装。")

        if header_row < 1:
            raise ValueError("header_row 必须是大于等于1的正整数。")

        self.header_row = header_row
        super().__init__(path)

    def _discover_files(self) -> List[str]:
        """
        在指定路径下发现所有以 .xlsx 结尾的文件。
        """
        if os.path.isdir(self.path):
            return [os.path.join(self.path, f) for f in os.listdir(self.path)
                    if os.path.isfile(os.path.join(self.path, f)) and f.lower().endswith('.xlsx')]
        elif os.path.isfile(self.path) and self.path.lower().endswith('.xlsx'):
            return [self.path]
        return []

    def get_total_items(self, file_path: str) -> int:
        """
        获取单个 .xlsx 文件中的总数据行数（不包括标题行及之前的行）。

        Args:
            file_path (str): 文件的完整路径。

        Returns:
            int: 文件中的总数据行数。
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active
            # 总行数减去标题行及之前的所有行
            return max(0, sheet.max_row - self.header_row)
        except Exception:
            return 0

    def read_file(self, file_path: str, start_line: int = 1) -> Iterator[Tuple[Dict[str, Any], int]]:
        """
        逐行读取单个 .xlsx 文件并将其解析为字典。

        Args:
            file_path (str): 要读取的文件的完整路径。
            start_line (int): 从指定的行号开始读取。注意：行号对应Excel中的实际行号。
                              如果 start_line 小于等于标题行，将从第一个数据行开始处理。

        Yields:
            Iterator[Tuple[Dict[str, Any], int]]: 包含解析后数据和对应行号的元组。
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active

            # 从指定的 header_row 读取标题
            if sheet.max_row < self.header_row:
                raise ValueError(f"文件 '{os.path.basename(file_path)}' 的总行数 ({sheet.max_row}) 小于指定的标题行号 ({self.header_row})。")
            headers = [cell.value for cell in sheet[self.header_row]]
            if not any(headers):
                raise ValueError(f"文件 '{os.path.basename(file_path)}' 的第 {self.header_row} 行（标题行）无效或为空。")

            # 数据行从标题行的下一行开始
            data_start_row = self.header_row + 1

            # 从数据行的第一行开始遍历
            for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=data_start_row), start=data_start_row):
                if row_idx < start_line:
                    continue

                if all(cell.value is None for cell in row_cells):
                    continue

                row_data = {headers[i]: cell.value for i, cell in enumerate(row_cells) if i < len(headers)}
                yield row_data, row_idx

        except Exception as e:
            raise IOError(f"读取 XLSX 文件 '{file_path}' 时出错: {e}") from e

    def get_samples(self, sample_size: int) -> List[Dict[str, Any]]:
        """
        从第一个可用的 .xlsx 文件中随机抽取指定数量的样本行。

        Args:
            sample_size (int): 要抽取的样本数量。

        Returns:
            List[Dict[str, Any]]: 包含样本数据的字典列表。
        """
        if not self.files_to_process:
            return []

        file_to_sample = self.files_to_process[0]
        samples = []
        try:
            total_items = self.get_total_items(file_to_sample)
            if total_items <= 0:
                return []

            num_samples_to_take = min(sample_size, total_items)
            if num_samples_to_take == 0:
                return []

            # 数据行号的范围是从 header_row + 1 开始
            data_start_row = self.header_row + 1
            data_end_row = self.header_row + total_items
            data_row_indices = range(data_start_row, data_end_row + 1)

            target_row_nums = sorted(random.sample(data_row_indices, num_samples_to_take))

            workbook = openpyxl.load_workbook(file_to_sample, read_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[self.header_row]]

            target_pointer = 0
            for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=data_start_row), start=data_start_row):
                if target_pointer >= len(target_row_nums):
                    break

                if row_idx == target_row_nums[target_pointer]:
                    if not all(cell.value is None for cell in row_cells):
                        row_data = {headers[i]: cell.value for i, cell in enumerate(row_cells) if i < len(headers)}
                        samples.append(row_data)
                    target_pointer += 1

        except Exception as e:
            print(f"[WARNING] 从文件 '{os.path.basename(file_to_sample)}' 随机抽样时出错: {e}")

        return samples