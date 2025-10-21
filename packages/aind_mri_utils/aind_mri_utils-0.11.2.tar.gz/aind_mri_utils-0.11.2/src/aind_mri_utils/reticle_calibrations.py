"""
Functions to read reticle calibration data, find a transformation between
coordinate frames, and apply the transformation.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from pathlib import Path
from typing import TYPE_CHECKING, Any

import numpy as np
from openpyxl import load_workbook
from scipy.spatial.transform import Rotation

if TYPE_CHECKING:
    from numpy.typing import NDArray

from aind_mri_utils.arc_angles import vector_to_arc_angles
from aind_mri_utils.rotations import (
    apply_affine,
    apply_inverse_affine,
    compose_transforms,
)

logger = logging.getLogger(__name__)


def _extract_calibration_metadata(
    ws: Any,
) -> tuple[float, float, float, NDArray[np.floating[Any]], str]:
    """
    Extract calibration metadata from an Excel worksheet.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Worksheet object from which to extract the calibration metadata.

    Returns
    -------
    tuple
        Tuple containing:
        - global_factor : float
            The global scale value.
        - global_rotation_degrees : float
            The global rotation in degrees.
        - manipulator_factor : float
            The manipulator scale value.
        - global_offset : numpy.ndarray
            The global offset as a 3-element array.
        - reticle_name : str
            The name of the reticle.
    """
    row_iter = ws.iter_rows(min_row=1, max_row=2, values_only=True)
    col_name_lookup = {k: i for i, k in enumerate(next(row_iter))}
    metadata_values = next(row_iter)
    global_factor = metadata_values[col_name_lookup["GlobalFactor"]]
    global_rotation_degrees = metadata_values[
        col_name_lookup["GlobalRotationDegrees"]
    ]
    manipulator_factor = metadata_values[col_name_lookup["ManipulatorFactor"]]
    reticle_name = metadata_values[col_name_lookup["Reticule"]]
    offset_x_pos = col_name_lookup["GlobalOffsetX"]
    global_offset = np.array(
        metadata_values[offset_x_pos : offset_x_pos + 3],
        dtype=float,
    )
    return (
        global_factor,
        global_rotation_degrees,
        manipulator_factor,
        global_offset,
        reticle_name,
    )


def find_similarity(
    X: NDArray[np.floating[Any]], Y: NDArray[np.floating[Any]]
) -> tuple[NDArray, NDArray, NDArray, int]:
    """
    Estimate the similarity transform between two sets of points.

    Finds the rotation matrix R, translation vector t, and the reflection
    correction matrix F. For data arrays with points in rows (transposed), the
    transformation is:

        Y = (X @ F) @ R.T + t

    Parameters
    ----------
    X : array_like, shape (N, M)
        M-D Points in the source frame.
    Y : array_like, shape (N, M)
        M-D Points in the target frame.

    Returns
    -------
    F : numpy.ndarray, shape (M, M)
        Reflection correction matrix.
    R : numpy.ndarray, shape (M, M)
        Rotation matrix.
    translation_vector : numpy.ndarray, shape (M,)
        Translation vector.
    rank : int
        Rank of the cross-covariance matrix.
    """
    X, Y = np.asarray(X), np.asarray(Y)
    Xm, Ym = X.mean(0), Y.mean(0)
    Xc, Yc = X - Xm, Y - Ym

    H = Yc.T @ Xc
    ndim = H.shape[0]
    H_det = np.linalg.det(H)
    U, S, Vt = np.linalg.svd(H, full_matrices=False)
    tol = max(H.shape) * np.finfo(S.dtype).eps * S[0]
    rank: int = np.sum(S > tol)
    # ---- handedness correction ---------------------------------------------
    R = U @ Vt
    d = np.sign(np.linalg.det(R))
    D = np.eye(ndim)
    if d < 0:
        U[:, -1] *= -1  # Flip last column of U
        R = U @ Vt
        D[-1, -1] = -1  # Flip last diagonal element of D
    # Generate the reflection correction matrix F
    # If Hdet is close to zero, use the identity matrix
    if np.isclose(H_det, 0):
        F = np.eye(ndim)
    else:
        F = Vt.T @ D @ Vt

    # ---- translation --------------------------------------------------------
    translation_vector = Ym - (R @ (F @ Xm))

    return F, R, translation_vector, int(rank)


def reticle_metadata_transform(
    global_rotation_degrees: float,
) -> NDArray[np.floating[Any]]:
    """
    Calculate the rotation matrix for the global rotation.

    Parameters
    ----------
    global_rotation_degrees : float
        The global rotation in degrees.

    Returns
    -------
    numpy.ndarray
        The rotation matrix.
    """
    R = (
        Rotation.from_euler("z", global_rotation_degrees, degrees=True)
        .as_matrix()
        .squeeze()
    )
    return R


def _contains_none(arr: NDArray[Any]) -> bool:
    """
    Check if any element in the iterable is None.

    Parameters
    ----------
    arr : iterable
        Iterable of elements to check.

    Returns
    -------
    bool
        True if any element is None, False otherwise.
    """
    return any(x is None for x in arr)


def _combine_pairs(
    list_of_pairs: list[
        tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]
    ],
) -> tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]:
    """
    Combine lists of pairs into separate global and manipulator points arrays.

    Parameters
    ----------
    list_of_pairs : list of tuple
        List of tuples, each containing a reticle point and a probe point as
        numpy arrays.

    Returns
    -------
    tuple of numpy.ndarray
        Two arrays: one for global points, one for manipulator points.
    """
    global_pts, manipulator_pts = [np.vstack(x) for x in zip(*list_of_pairs)]
    return global_pts, manipulator_pts


def _extract_calibration_pairs(
    ws: Any,
) -> dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]]:
    """
    Extract calibration pairs from an Excel worksheet.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Worksheet object from which to extract the calibration pairs.

    Returns
    -------
    dict
        Keys are probe names, values are tuples of arrays (global_pts,
        manipulator_pts).
    """
    pair_lists_by_probe: dict[
        int, list[tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]]
    ] = dict()
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        probe_name = row[0]
        if probe_name is None:
            continue
        reticle_pt = np.array(row[1:4])
        probe_pt = np.array(row[4:7])
        if _contains_none(reticle_pt) or _contains_none(probe_pt):
            continue
        if probe_name not in pair_lists_by_probe:
            pair_lists_by_probe[probe_name] = []
        pair_lists_by_probe[probe_name].append((reticle_pt, probe_pt))
    pair_mats_by_probe = {
        k: _combine_pairs(v) for k, v in pair_lists_by_probe.items()
    }
    return pair_mats_by_probe


def _apply_metadata_to_pair_mats(
    global_pts: NDArray[np.floating[Any]],
    manipulator_pts: NDArray[np.floating[Any]],
    global_factor: float,
    global_rotation_degrees: float,
    global_offset: NDArray[np.floating[Any]],
    manipulator_factor: float,
) -> tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]:
    """
    Apply calibration metadata to global and manipulator points arrays.

    Parameters
    ----------
    global_pts : numpy.ndarray
        Global points array.
    manipulator_pts : numpy.ndarray
        Manipulator points array.
    global_factor : float
        Global scale factor.
    global_rotation_degrees : float
        Global rotation in degrees.
    global_offset : numpy.ndarray
        Global offset as a 3-element array.
    manipulator_factor : float
        Manipulator scale factor.

    Returns
    -------
    tuple of numpy.ndarray
        Adjusted global points and manipulator points arrays.
    """
    if global_rotation_degrees != 0:
        rot_mat = reticle_metadata_transform(global_rotation_degrees)
        # Transposed because points are row vectors
        global_pts = global_pts @ rot_mat.T
    global_pts = global_pts * global_factor + global_offset
    manipulator_pts = manipulator_pts * manipulator_factor
    return global_pts, manipulator_pts


def read_manual_reticle_calibration(
    filename: str,
    points_sheet_name: str = "points",
    metadata_sheet_name: str = "metadata",
) -> tuple[
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    NDArray[np.floating[Any]],
    float,
    str,
]:
    """
    Read reticle calibration data from an Excel file.

    Parameters
    ----------
    filename : str
        Path to the Excel file containing the calibration data.
    points_sheet_name : str, optional
        Name of the sheet containing the calibration points. Default is
        "points".
    metadata_sheet_name : str, optional
        Name of the sheet containing the calibration metadata. Default is
        "metadata".

    Returns
    -------
    adjusted_pairs_by_probe : dict
        Adjusted calibration pairs by probe name.
    global_offset : numpy.ndarray
        The global offset as a 3-element array.
    global_rotation_degrees : float
        The global rotation in degrees.
    reticle_name : str
        The name of the reticle.

    Raises
    ------
    ValueError
        If the specified sheets are not found in the Excel file.
    """
    in_mem_file = None
    with open(filename, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    wb = load_workbook(in_mem_file, read_only=True, data_only=True)
    if points_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {points_sheet_name} not found in {filename}")
    if metadata_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet {metadata_sheet_name} not found in {filename}"
        )
    (
        global_factor,
        global_rotation_degrees,
        manipulator_factor,
        global_offset,
        reticle_name,
    ) = _extract_calibration_metadata(wb[metadata_sheet_name])
    pairs_by_probe = _extract_calibration_pairs(wb["points"])
    adjusted_pairs_by_probe = {
        k: _apply_metadata_to_pair_mats(
            *v,
            global_factor,
            global_rotation_degrees,
            global_offset,
            manipulator_factor,
        )
        for k, v in pairs_by_probe.items()
    }
    return (
        adjusted_pairs_by_probe,
        global_offset,
        global_rotation_degrees,
        reticle_name,
    )


def read_parallax_calibration_dir(
    parallax_points_dir: str,
    sn_filename_regexp: re.Pattern[str] = re.compile(
        r"(?i)points_SN\d+(?:_.*)?.csv$"
    ),
    *args: Any,
    **kwargs: Any,
) -> dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]]:
    """
    Read parallax calibration data from a directory of CSV files.

    Parameters
    ----------
    parallax_points_dir : str
        Directory containing the parallax calibration CSV files.
    sn_filename_regexp : re.Pattern, optional
        Regular expression pattern to match filenames.
    *args
        Additional arguments to pass to the calibration file reader.
    **kwargs
        Additional keyword arguments to pass to the calibration file reader.

    Returns
    -------
    dict
        Keys are controller numbers, values are tuples of arrays (global_pts,
        manipulator_pts).
    """
    pairs_by_controller: dict[
        int,
        list[
            tuple[
                NDArray[np.floating[Any]],
                NDArray[np.floating[Any]],
            ]
        ],
    ] = {}
    p_path = Path(parallax_points_dir)
    for filename in p_path.iterdir():
        if filename.is_file() and re.search(sn_filename_regexp, filename.name):
            _append_parallax_calibration_file(
                pairs_by_controller, filename, *args, **kwargs
            )
    mats_by_controller = {
        k: _combine_pairs(v) for k, v in pairs_by_controller.items()
    }
    return mats_by_controller


def read_parallax_calibration_file(
    parallax_points_filename: str, *args: Any, **kwargs: Any
) -> dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]]:
    """
    Read parallax calibration data from a single CSV file.

    Parameters
    ----------
    parallax_points_filename : str
        Path to the CSV file containing the parallax points data.
    *args
        Additional arguments to pass to the calibration file reader.
    **kwargs
        Additional keyword arguments to pass to the calibration file reader.

    Returns
    -------
    dict
        Keys are controller numbers, values are tuples of arrays (global_pts,
        manipulator_pts).
    """
    pairs_by_controller: dict[
        int,
        list[
            tuple[
                NDArray[np.floating[Any]],
                NDArray[np.floating[Any]],
            ]
        ],
    ] = {}
    _append_parallax_calibration_file(
        pairs_by_controller, parallax_points_filename, *args, **kwargs
    )
    mats_by_controller: dict[
        int,
        tuple[
            NDArray[np.floating[Any]],
            NDArray[np.floating[Any]],
        ],
    ] = {k: _combine_pairs(v) for k, v in pairs_by_controller.items()}
    return mats_by_controller


def read_parallax_calibration_dir_and_correct(
    parallax_calibration_dir: str,
    reticle_offset: NDArray[np.floating[Any]],
    reticle_rotation: float,
    local_scale_factor: float = 1 / 1000,
    global_scale_factor: float = 1 / 1000,
) -> dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]]:
    """
    Read and correct parallax calibration data from a directory of CSV files.

    Parameters
    ----------
    parallax_calibration_dir : str
        Directory containing parallax calibration data.
    reticle_offset : numpy.ndarray
        Offset of the reticle.
    reticle_rotation : float
        Rotation of the reticle in degrees.
    local_scale_factor : float, optional
        Local scale factor for calibration. Default is 1/1000.
    global_scale_factor : float, optional
        Global scale factor for calibration. Default is 1/1000.

    Returns
    -------
    dict
        Keys are controller numbers, values are tuples of arrays (global_pts,
        manipulator_pts).
    """
    pairs_by_probe = read_parallax_calibration_dir(parallax_calibration_dir)
    corrected_pairs_by_probe = {}
    for controller, pairs in pairs_by_probe.items():
        reticle_pts, manip_pts = _apply_metadata_to_pair_mats(
            *pairs,
            global_scale_factor,
            reticle_rotation,
            reticle_offset,
            local_scale_factor,
        )
        corrected_pairs_by_probe[controller] = (reticle_pts, manip_pts)
    return corrected_pairs_by_probe


def _append_parallax_calibration_file(
    pairs_by_controller: dict[
        int, list[tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]]
    ],
    parallax_points_filename: str | Path,
    sn_colname: str = "sn",
    sn_regexp: re.Pattern[str] = re.compile(r"(\d+)$"),
) -> dict[
    int, list[tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]]
]:
    """
    Read parallax calibration data from a CSV file and organize by controller
    number.

    Parameters
    ----------
    pairs_by_controller : dict
        Dictionary to append results to.
    parallax_points_filename : str
        Path to the CSV file containing parallax points data.
    sn_colname : str, optional
        Column name for the serial number. Default is "sn".
    sn_regexp : re.Pattern, optional
        Regular expression pattern to extract the controller number.

    Returns
    -------
    dict
        Keys are controller numbers, values are lists of tuples (reticle_pt,
        manip_pt).

    Notes
    -----
    The CSV file is expected to have columns named "global_x", "global_y",
    "global_z" for reticle points and "local_x", "local_y", "local_z" for
    manipulator points.
    """
    dims = ["x", "y", "z"]
    reticle_colnames = [f"global_{dim}" for dim in dims]
    manipulator_colnames = [f"local_{dim}" for dim in dims]
    with open(parallax_points_filename, newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        # Map each column name to its first index
        index_map = {}
        for i, col in enumerate(header):
            if col not in index_map:
                index_map[col] = i
        # Find indices for the columns we care about
        sn_col_ndx = index_map[sn_colname]
        reticle_pt_ndxs = [index_map[col] for col in reticle_colnames]
        manip_pt_ndxs = [index_map[col] for col in manipulator_colnames]
        # Read the data
        for row in reader:
            res = re.search(sn_regexp, row[sn_col_ndx])
            if res is None:
                raise ValueError(
                    f"Could not extract controller number from "
                    f"{row[sn_col_ndx]} using regexp {sn_regexp.pattern}"
                )
            controller_no = int(res.group(1))
            ret_pt = np.array([float(row[ndx]) for ndx in reticle_pt_ndxs])
            manip_pt = np.array([float(row[ndx]) for ndx in manip_pt_ndxs])
            # Append to this manipulator's list of pairs, creating a new list
            # if needed
            pairs_by_controller.setdefault(controller_no, []).append(
                (ret_pt, manip_pt)
            )
    return pairs_by_controller


def fit_rotation_params_interpretable(
    bregma_pts: NDArray[np.floating[Any]],
    probe_pts: NDArray[np.floating[Any]],
) -> tuple[
    NDArray[np.floating[Any]],
    NDArray[np.floating[Any]],
    NDArray[np.floating[Any]],
    int,
]:
    """
    Fit rotation parameters to align bregma points with probe points using
    least squares optimization. The rotation matrix and translation vector
    are the solution for the equation

    probe_pts = R @ bregma_pts + translation

    where each point is a column vector.

    Because numpy is row-major, points are often stored as row vectors. In this
    case, you should use the transpose of this equation:

    probe_pts = bregma_pts @ R.T + translation

    Parameters
    ----------
    bregma_pts : numpy.ndarray, shape (N, 3)
        Bregma points to be transformed.
    probe_pts : numpy.ndarray, shape (N, 3)
        Probe points to align with.

    Returns
    -------
    F : numpy.ndarray, shape (3, 3)
        Handedness correction matrix.
    R : numpy.ndarray, shape (3, 3)
        Rotation matrix.
    t : numpy.ndarray, shape (3,)
        Translation vector.
    rank : int
        Rank of the cross-covariance matrix.
    """
    if bregma_pts.shape != probe_pts.shape:
        raise ValueError("bregma_pts and probe_pts must have the same shape")
    if bregma_pts.shape[1] != 3:
        raise ValueError("bregma_pts and probe_pts must have 3 columns")

    F, R, t, rank = find_similarity(bregma_pts, probe_pts)
    return F, R, t, rank


def fit_rotation_params(
    bregma_pts: NDArray[np.floating[Any]],
    probe_pts: NDArray[np.floating[Any]],
    **kwargs: Any,
) -> tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]:
    """
    Fit rotation parameters to align bregma points with probe points using
    least squares optimization. The rotation matrix and translation vector
    are the solution for the equation

    probe_pts = R @ bregma_pts + translation

    where each point is a column vector.

    Because numpy is row-major, points are often stored as row vectors. In this
    case, you should use the transpose of this equation:

    probe_pts = bregma_pts @ R.T + translation

    Parameters
    ----------
    bregma_pts : numpy.ndarray, shape (N, 3)
        Bregma points to be transformed.
    probe_pts : numpy.ndarray, shape (N, 3)
        Probe points to align with.
    **kwargs
        Additional keyword arguments for fitting.

    Returns
    -------
    Rcomb : numpy.ndarray, shape (3, 3)
        Combined transformation matrix.
    t : numpy.ndarray, shape (3,)
        Translation vector.
    """
    F, R, t, rank = fit_rotation_params_interpretable(
        bregma_pts, probe_pts, **kwargs
    )
    R_comb = R @ F
    return R_comb, t


def _fit_by_probe(
    pairs_by_probe: dict[
        int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]
    ],
    *args: Any,
    **kwargs: Any,
) -> dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]]:
    """
    Fit rotation parameters for each probe.

    Parameters
    ----------
    pairs_by_probe : dict
        Keys are probe names, values are tuples of arrays (reticle_pts,
        probe_pts).
    *args
        Additional arguments for fitting.
    **kwargs
        Additional keyword arguments for fitting.

    Returns
    -------
    dict
        Keys are probe names, values are tuples (R, t).
    """
    cal_by_probe = {
        k: fit_rotation_params(*v, *args, **kwargs)
        for k, v in pairs_by_probe.items()
    }
    return cal_by_probe


def fit_rotation_params_from_manual_calibration(
    filename: str, *args: Any, **kwargs: Any
) -> tuple[
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    NDArray[np.floating[Any]],
    NDArray[np.floating[Any]],
]:
    """
    Fit rotation parameters from manual calibration data.

    Parameters
    ----------
    filename : str
        Path to the `.xlsx` file containing calibration data.
    *args
        Additional arguments for fitting.
    **kwargs
        Additional keyword arguments for fitting.

    Returns
    -------
    cal_by_probe : dict
        Calibration parameters by probe. Each value is a tuple containing the
        rotation matrix, translation vector, and scaling factor. These
        parameters transform bregma points to probe points, both in
        millimeters.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma coordinates.
    t_reticle_to_bregma : numpy.ndarray
        Translation vector from reticle to bregma coordinates.
    """
    adjusted_pairs_by_probe, global_offset, global_rotation_degrees, _ = (
        read_manual_reticle_calibration(filename)
    )
    R_reticle_to_bregma = reticle_metadata_transform(global_rotation_degrees)
    cal_by_probe = _fit_by_probe(adjusted_pairs_by_probe, *args, **kwargs)
    return cal_by_probe, R_reticle_to_bregma, global_offset


def fit_rotation_params_from_parallax(
    parallax_calibration_dir: str,
    reticle_offset: NDArray[np.floating[Any]],
    reticle_rotation: float,
    local_scale_factor: float = 1 / 1000,
    global_scale_factor: float = 1 / 1000,
    *args: Any,
    **kwargs: Any,
) -> tuple[
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    NDArray[np.floating[Any]],
]:
    """
    Fit rotation parameters from parallax calibration data.

    Parameters
    ----------
    parallax_calibration_dir : str
        Directory containing parallax calibration data.
    reticle_offset : numpy.ndarray
        Offset of the reticle in bregma-relative coordinates.
    reticle_rotation : float
        Rotation of the reticle. i.e. the angle about the z-axis of the reticle
        coordinate system and the bregma coordinate system in degrees. The sign
        of the angle should be positive if the reticle is rotated clockwise
        w.r.t bregma coordinate system when viewed from above.
    local_scale_factor : float, optional
        Local scale factor for calibration, by default 1/1000. Parallax stores
        points in microns, so this factor is used to convert to millimeters.
    global_scale_factor : float, optional
        Global scale factor for calibration, by default 1/1000. Parallax stores
        points in microns, so this factor is used to convert to millimeters.
    *args : tuple
        Additional arguments to pass to the fitting function. See
        `fit_rotation_params`.
    **kwargs : dict
        Additional keyword arguments to pass to the fitting function.
        See `fit_rotation_params`.

    Returns
    -------
    cal_by_probe : dict
        Calibration parameters by probe. Each value is a tuple containing the
        rotation matrix, translation vector, and scaling factor. These
        parameters transform bregma points to probe points, both in
        millimeters.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma coordinates, in millimeters.
    """
    adjusted_pairs_by_probe = read_parallax_calibration_dir_and_correct(
        parallax_calibration_dir,
        reticle_offset,
        reticle_rotation,
        local_scale_factor,
        global_scale_factor,
    )
    cal_by_probe = _fit_by_probe(adjusted_pairs_by_probe, *args, **kwargs)
    R_reticle_to_bregma = reticle_metadata_transform(reticle_rotation)
    return cal_by_probe, R_reticle_to_bregma


def _debug_print_pt_err(
    reticle: NDArray[np.floating[Any]],
    probe: NDArray[np.floating[Any]],
    predicted_probe: NDArray[np.floating[Any]],
    err: float,
    decimals: int = 3,
) -> None:
    """
    Print the error for a single point.

    Parameters
    ----------
    reticle : numpy.ndarray
        Reticle point.
    probe : numpy.ndarray
        Probe point.
    predicted_probe : numpy.ndarray
        Predicted probe point.
    err : float
        Error value.
    decimals : int, optional
        Number of decimal places to round to. Default is 3.
    """
    rounded_reticle = np.round(reticle, decimals=decimals)
    rounded_probe = np.round(probe, decimals=decimals)
    rounded_pred = np.round(predicted_probe, decimals=decimals)
    logger.debug(
        f"Reticle {rounded_reticle} -> "
        f"Probe {rounded_probe}: predicted {rounded_pred} "
        f"error {err:.2f} µm"
    )


def _debug_print_err_stats(errs: NDArray[np.floating[Any]]) -> None:
    """
    Print error statistics for a probe.

    Parameters
    ----------
    errs : numpy.ndarray
        Array of error values.
    """
    logger.debug(
        f"mean error {errs.mean():.2f} µm, max error {errs.max():.2f} µm"
    )


def _debug_fits(
    cal_by_probe: dict[
        int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]
    ],
    R_reticle_to_bregma: NDArray[np.floating[Any]],
    t_reticle_to_bregma: NDArray[np.floating[Any]],
    adjusted_pairs_by_probe: dict[
        int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]
    ],
) -> dict[int, NDArray[np.floating[Any]]]:
    """
    Debug the fits for each probe.

    Parameters
    ----------
    cal_by_probe : dict
        Calibration parameters by probe.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma coordinates.
    t_reticle_to_bregma : numpy.ndarray
        Translation vector from reticle to bregma coordinates.
    adjusted_pairs_by_probe : dict
        Adjusted calibration pairs by probe.

    Returns
    -------
    dict
        Keys are probe names, values are arrays of error values.
    """
    errs_by_probe = {}
    for probe, (bregma_pts, probe_pts) in adjusted_pairs_by_probe.items():
        R, t = cal_by_probe[probe]
        predicted_probe_pts = transform_bregma_to_probe(bregma_pts, R, t)
        # in mm
        errs = np.linalg.norm(predicted_probe_pts - probe_pts, axis=1)
        errs_by_probe[probe] = errs
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug(f"Probe {probe}:")
            logger.debug("rotation:")
            logger.debug(R)
            logger.debug(f"translation: {t}")
            _debug_print_err_stats(1000 * errs)
            reticle_pts = transform_probe_to_bregma(
                bregma_pts, R_reticle_to_bregma, t_reticle_to_bregma
            )
            for i in range(len(errs)):
                _debug_print_pt_err(
                    reticle_pts[i],
                    probe_pts[i],
                    predicted_probe_pts[i],
                    1000 * errs[i],
                )
    return errs_by_probe


def debug_manual_calibration(
    filename: str, *args: Any, **kwargs: Any
) -> tuple[
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    NDArray[np.floating[Any]],
    NDArray[np.floating[Any]],
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    dict[int, NDArray[np.floating[Any]]],
]:
    """
    Debugs the manual calibration process by fitting rotation parameters and
    reading manual reticle calibration data.

    Parameters
    ----------
    filename : str
        Path to the file containing manual calibration data.
    *args
        Additional arguments for calibration.
    **kwargs
        Additional keyword arguments for calibration.

    Returns
    -------
    cal_by_probe : dict
        Calibration data for each probe.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma.
    t_reticle_to_bregma : numpy.ndarray
        Translation vector from reticle to bregma.
    adjusted_pairs_by_probe : dict
        Adjusted pairs of calibration data by probe.
    errs_by_probe : dict
        Errors in the fits for each probe.
    """
    cal_by_probe, R_reticle_to_bregma, t_reticle_to_bregma = (
        fit_rotation_params_from_manual_calibration(filename, *args, **kwargs)
    )
    (
        adjusted_pairs_by_probe,
        global_offset,
        global_rotation_degrees,
        _,
    ) = read_manual_reticle_calibration(filename)
    errs_by_probe = _debug_fits(
        cal_by_probe,
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        adjusted_pairs_by_probe,
    )
    return (
        cal_by_probe,
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        adjusted_pairs_by_probe,
        errs_by_probe,
    )


def debug_parallax_calibration(
    parallax_calibration_dir: str,
    reticle_offset: NDArray[np.floating[Any]],
    reticle_rotation: float,
    local_scale_factor: float = 1 / 1000,
    global_scale_factor: float = 1 / 1000,
    *args: Any,
    **kwargs: Any,
) -> tuple[
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    NDArray[np.floating[Any]],
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    dict[int, NDArray[np.floating[Any]]],
]:
    """
    Debugs the parallax calibration process by reading calibration data,
    applying corrections, fitting the data, and calculating errors.

    Parameters
    ----------
    parallax_calibration_dir : str
        Directory containing the parallax calibration data.
    reticle_offset : numpy.ndarray
        Offset of the reticle in the calibration setup.
    reticle_rotation : float
        Rotation of the reticle in the calibration setup.
    local_scale_factor : float, optional
        Scale factor for local adjustments. Default is 1/1000.
    global_scale_factor : float, optional
        Scale factor for global adjustments. Default is 1/1000.
    *args
        Additional arguments for fitting.
    **kwargs
        Additional keyword arguments for fitting.

    Returns
    -------
    cal_by_probe : dict
        Calibration data organized by probe.
    R_reticle_to_bregma : numpy.ndarray
        Transformation matrix from reticle to bregma.
    adjusted_pairs_by_probe : dict
        Adjusted calibration pairs organized by probe.
    errs_by_probe : dict
        Errors in the calibration fits organized by probe.
    """
    adjusted_pairs_by_probe = read_parallax_calibration_dir_and_correct(
        parallax_calibration_dir,
        reticle_offset,
        reticle_rotation,
        local_scale_factor,
        global_scale_factor,
    )
    cal_by_probe = _fit_by_probe(adjusted_pairs_by_probe, *args, **kwargs)
    R_reticle_to_bregma = reticle_metadata_transform(reticle_rotation)
    errs_by_probe = _debug_fits(
        cal_by_probe,
        R_reticle_to_bregma,
        reticle_offset,
        adjusted_pairs_by_probe,
    )
    return (
        cal_by_probe,
        R_reticle_to_bregma,
        adjusted_pairs_by_probe,
        errs_by_probe,
    )


def transform_bregma_to_probe(
    bregma_pts: NDArray[np.floating[Any]],
    R: NDArray[np.floating[Any]],
    translation: NDArray[np.floating[Any]],
) -> NDArray[np.floating[Any]]:
    """
    Transform bregma points to probe points using rotation and translation.

    Parameters
    ----------
    bregma_pts : numpy.ndarray, shape (N, 3)
        Bregma points to transform.
    R : numpy.ndarray, shape (3, 3)
        Affine matrix as provided from fit functions in this module.
    translation : numpy.ndarray, shape (3,)
        Translation vector as provided from fit functions in this module.

    Returns
    -------
    numpy.ndarray, shape (N, 3)
        Transformed probe points.

    Notes
    -----
    Expects the affine and transform to take bregma points to probe points.
    """
    return apply_affine(bregma_pts, R, translation)


def transform_probe_to_bregma(
    probe_pts: NDArray[np.floating[Any]],
    R: NDArray[np.floating[Any]],
    translation: NDArray[np.floating[Any]],
) -> NDArray[np.floating[Any]]:
    """
    Transform probe points to bregma points using rotation and translation.

    Parameters
    ----------
    probe_pts : numpy.ndarray, shape (N, 3)
        Probe points to transform.
    R : numpy.ndarray, shape (3, 3)
        Affine matrix as provided from fit functions in this module.
    translation : numpy.ndarray, shape (3,)
        Translation vector as provided from fit functions in this module.

    Returns
    -------
    numpy.ndarray, shape (N, 3)
        Transformed bregma points.

    Notes
    -----
    Expects the affine and transform to take bregma points to probe points.
    """
    return apply_inverse_affine(probe_pts, R, translation)


def transform_bregma_to_reticle(
    bregma_pts: NDArray[np.floating[Any]],
    R: NDArray[np.floating[Any]],
    translation: NDArray[np.floating[Any]],
) -> NDArray[np.floating[Any]]:
    """
    Transform bregma points to reticle points using rotation and translation.

    Parameters
    ----------
    bregma_pts : numpy.ndarray, shape (N, 3)
        Bregma points to transform.
    R : numpy.ndarray, shape (3, 3)
        Affine matrix as provided from fit functions in this module.
    translation : numpy.ndarray, shape (3,)
        Translation vector as provided from fit functions in this module.

    Returns
    -------
    numpy.ndarray, shape (N, 3)
        Transformed reticle points.
    """
    return apply_inverse_affine(bregma_pts, R, translation)


def transform_reticle_to_bregma(
    reticle_pts: NDArray[np.floating[Any]],
    R: NDArray[np.floating[Any]],
    translation: NDArray[np.floating[Any]],
) -> NDArray[np.floating[Any]]:
    """
    Transform reticle points to bregma points using rotation and translation.

    Parameters
    ----------
    reticle_pts : numpy.ndarray, shape (N, 3)
        Reticle points to transform.
    R : numpy.ndarray, shape (3, 3)
        Affine matrix as provided from fit functions in this module.
    translation : numpy.ndarray, shape (3,)
        Translation vector as provided from fit functions in this module.

    Returns
    -------
    numpy.ndarray, shape (N, 3)
        Transformed bregma points.
    """
    return apply_affine(reticle_pts, R, translation)


def combine_reticle_to_probe_transforms(
    R_bregma_to_probe: NDArray[np.floating[Any]],
    t_bregma_to_probe: NDArray[np.floating[Any]],
    R_reticle_to_bregma: NDArray[np.floating[Any]],
    t_reticle_to_bregma: NDArray[np.floating[Any]],
) -> tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]:
    """
    Combine transformation matrices and translation vectors from reticle to
    bregma and bregma to probe.

    Parameters
    ----------
    R_bregma_to_probe : numpy.ndarray, shape (3, 3)
        Rotation matrix from bregma to probe.
    t_bregma_to_probe : numpy.ndarray, shape (3,)
        Translation vector from bregma to probe.
    R_reticle_to_bregma : numpy.ndarray, shape (3, 3)
        Rotation matrix from reticle to bregma.
    t_reticle_to_bregma : numpy.ndarray, shape (3,)
        Translation vector from reticle to bregma.

    Returns
    -------
    tuple
        Combined rotation matrix and translation vector from reticle to probe.
    """
    return compose_transforms(
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        R_bregma_to_probe,
        t_bregma_to_probe,
    )


def transform_reticle_to_probe(
    reticle_pts: NDArray[np.floating[Any]],
    R_bregma_to_probe: NDArray[np.floating[Any]],
    t_bregma_to_probe: NDArray[np.floating[Any]],
    R_reticle_to_bregma: NDArray[np.floating[Any]],
    t_reticle_to_bregma: NDArray[np.floating[Any]],
) -> NDArray[np.floating[Any]]:
    """
    Transform reticle points to probe points using rotation and translation.

    Parameters
    ----------
    reticle_pts : numpy.ndarray, shape (N, 3)
        Reticle points to transform.
    R_bregma_to_probe : numpy.ndarray, shape (3, 3)
        Rotation matrix from bregma to probe coordinates.
    t_bregma_to_probe : numpy.ndarray, shape (3,)
        Translation vector from bregma to probe coordinates.
    R_reticle_to_bregma : numpy.ndarray, shape (3, 3)
        Rotation matrix from reticle to bregma coordinates.
    t_reticle_to_bregma : numpy.ndarray, shape (3,)
        Translation vector from reticle to bregma coordinates.

    Returns
    -------
    numpy.ndarray, shape (N, 3)
        Transformed probe points.
    """
    R_reticle_to_probe, t_reticle_to_probe = (
        combine_reticle_to_probe_transforms(
            R_bregma_to_probe,
            t_bregma_to_probe,
            R_reticle_to_bregma,
            t_reticle_to_bregma,
        )
    )
    return apply_affine(reticle_pts, R_reticle_to_probe, t_reticle_to_probe)


def transform_probe_to_reticle(
    probe_pts: NDArray[np.floating[Any]],
    R_bregma_to_probe: NDArray[np.floating[Any]],
    t_bregma_to_probe: NDArray[np.floating[Any]],
    R_reticle_to_bregma: NDArray[np.floating[Any]],
    t_reticle_to_bregma: NDArray[np.floating[Any]],
) -> NDArray[np.floating[Any]]:
    """
    Transform probe points to reticle coordinates.

    Parameters
    ----------
    probe_pts : numpy.ndarray, shape (N, 3)
        Probe points to transform.
    R_bregma_to_probe : numpy.ndarray, shape (3, 3)
        Rotation matrix from bregma to probe coordinates.
    t_bregma_to_probe : numpy.ndarray, shape (3,)
        Translation vector from bregma to probe coordinates.
    R_reticle_to_bregma : numpy.ndarray, shape (3, 3)
        Rotation matrix from reticle to bregma coordinates.
    t_reticle_to_bregma : numpy.ndarray, shape (3,)
        Translation vector from reticle to bregma coordinates.

    Returns
    -------
    numpy.ndarray, shape (N, 3)
        Transformed points in reticle coordinates.
    """
    R_reticle_to_probe, t_reticle_to_probe = (
        combine_reticle_to_probe_transforms(
            R_bregma_to_probe,
            t_bregma_to_probe,
            R_reticle_to_bregma,
            t_reticle_to_bregma,
        )
    )
    return apply_inverse_affine(
        probe_pts, R_reticle_to_probe, t_reticle_to_probe
    )


def find_probe_insertion_vector(
    R: NDArray[np.floating[Any]],
    newscale_z_down: NDArray[np.floating[Any]] | None = None,
) -> NDArray[np.floating[Any]]:
    """
    Find the probe insertion vector from the rotation matrix.

    Parameters
    ----------
    R : numpy.ndarray, shape (3, 3)
        Rotation matrix.
    newscale_z_down : numpy.ndarray, shape (3,), optional
        Vector representing the z-axis in probe coordinates.
        Default is [0, 0, 1].

    Returns
    -------
    numpy.ndarray, shape (3,)
        Probe insertion vector in bregma coordinates.
    """
    if newscale_z_down is None:
        newscale_z_down = np.array([0, 0, 1])
    # Probe coordinate system has z-axis pointing down
    z_axis = transform_probe_to_bregma(newscale_z_down, R, np.zeros(3))
    return z_axis


def find_probe_angle(
    R: NDArray[np.floating[Any]],
    newscale_z_down: NDArray[np.floating[Any]] | None = None,
    **kwargs: Any,
) -> tuple[float, float] | None:
    """
    Find the probe angle from the calibration rotation matrix.

    Parameters
    ----------
    R : numpy.ndarray, shape (3, 3)
        Rotation matrix.
    newscale_z_down : numpy.ndarray, shape (3,), optional
        Vector representing the z-axis in probe coordinates.
        Default is [0, 0, 1].
    **kwargs
        Additional keyword arguments for arc angle calculation.

    Returns
    -------
    tuple of float
        The calculated arc angles in degrees. The first element is the angle
        around the x-axis, and the second element is the angle around the
        y-axis.  Returns None if the input vector is a zero vector.
    """
    if newscale_z_down is None:
        newscale_z_down = np.array([0, 0, 1])
    # Probe coordinate system has z-axis pointing down
    z_axis = find_probe_insertion_vector(R, newscale_z_down=newscale_z_down)
    return vector_to_arc_angles(z_axis, **kwargs)


def _validate_combined_calibration_inputs(
    manual_calibration_files: str | list[str],
    parallax_directories: str | list[str],
) -> tuple[list[str], list[str]]:
    """
    Validate and normalize input lists for combined calibration.

    Parameters
    ----------
    manual_calibration_files : str or list of str
        Manual calibration file(s).
    parallax_directories : str or list of str
        Parallax calibration directory(ies).

    Returns
    -------
    tuple
        (manual_calibration_files, parallax_directories) as lists.
    """
    if isinstance(manual_calibration_files, list):
        if len(manual_calibration_files) == 0:
            raise ValueError("No manual calibration files provided")
    else:
        manual_calibration_files = [manual_calibration_files]
    if not isinstance(parallax_directories, list):
        parallax_directories = [parallax_directories]
    return manual_calibration_files, parallax_directories


def combine_parallax_and_manual_calibrations(
    manual_calibration_files: list[str] | str,
    parallax_directories: list[str] | str,
    probes_to_ignore_manual: list[str] | None = None,
    *args: Any,
    **kwargs: Any,
) -> tuple[
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    NDArray[np.floating[Any]],
    NDArray[np.floating[Any]],
]:
    """
    Combine parallax and manual calibration data.

    In each list of calibrations, calibrations later in the list take priority.
    Manual calibrations take priority over parallax calibrations, unless
    specified.

    Parameters
    ----------
    manual_calibration_files : list of str
        List of files containing manual calibration data.
    parallax_directories : list of str
        List of directories containing parallax calibration data.
    probes_to_ignore_manual : list of str, optional
        List of probe names to ignore from the manual calibrations.
        Default is [].
    *args
        Additional arguments for fitting.
    **kwargs
        Additional keyword arguments for fitting.

    Returns
    -------
    cal_by_probe_combined : dict
        Combined calibration data by probe.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma.
    global_offset : numpy.ndarray
        Global offset applied to the calibration data.

    Raises
    ------
    ValueError
        If no manual calibration files or parallax directories are provided.
    """
    manual_calibration_files, parallax_directories = (
        _validate_combined_calibration_inputs(
            manual_calibration_files, parallax_directories
        )
    )
    # Read the first manual calibration to get the reticle metadata
    first_manual = manual_calibration_files[0]
    adjusted_pairs_by_probe, global_offset, global_rotation_degrees, _ = (
        read_manual_reticle_calibration(first_manual)
    )
    R_reticle_to_bregma = reticle_metadata_transform(global_rotation_degrees)

    # Fit the manual calibrations
    cal_by_probe_manual = _fit_by_probe(
        adjusted_pairs_by_probe, *args, **kwargs
    )
    for manual_calibration_file in manual_calibration_files[1:]:
        cal_by_probe, _, _ = fit_rotation_params_from_manual_calibration(
            manual_calibration_file, *args, **kwargs
        )
        cal_by_probe_manual.update(cal_by_probe)
    cal_by_probe_combined: dict[
        int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]
    ] = {}
    for parallax_dir in parallax_directories:
        cal_by_probe, _ = fit_rotation_params_from_parallax(
            parallax_dir,
            global_offset,
            global_rotation_degrees,
            *args,
            **kwargs,
        )
        cal_by_probe_combined.update(cal_by_probe)

    # Drop any probes from exclusion list from the manual calibrations
    if probes_to_ignore_manual is None:
        probes_to_ignore_manual = []
    for probe_name in probes_to_ignore_manual:
        cal_by_probe_manual.pop(probe_name, None)  # type: ignore
    # Add the first manual calibration to the combined calibrations
    # Because they are added last, these manual calibrations take priority
    cal_by_probe_combined.update(cal_by_probe_manual)

    return cal_by_probe_combined, R_reticle_to_bregma, global_offset


def debug_parallax_and_manual_calibrations(
    manual_calibration_files: list[str] | str,
    parallax_directories: list[str] | str,
    probes_to_ignore_manual: list[str] | None = None,
    local_scale_factor: float = 1 / 1000,
    global_scale_factor: float = 1 / 1000,
    *args: Any,
    **kwargs: Any,
) -> tuple[
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    NDArray[np.floating[Any]],
    NDArray[np.floating[Any]],
    dict[int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]],
    dict[int, NDArray[np.floating[Any]]],
]:
    """
    Debug combined parallax and manual calibrations.

    In each list of calibrations, calibrations later in the list take priority.
    Manual calibrations take priority over parallax calibrations, unless
    specified.

    Parameters
    ----------
    manual_calibration_files : list of str
        List of file paths to the manual calibration files.
    parallax_directories : list of str
        List of directories containing parallax calibration data.
    probes_to_ignore_manual : list of str, optional
        List of probe names to ignore from the manual calibration data.
        Default is [].
    local_scale_factor : float, optional
        Local scale factor to apply to the calibration data. Default is 1/1000.
    global_scale_factor : float, optional
        Global scale factor to apply to the calibration data.
        Default is 1/1000.
    *args
        Additional arguments for calibration.
    **kwargs
        Additional keyword arguments for calibration.

    Returns
    -------
    combined_cal_by_probe : dict
        Combined calibration data by probe.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma.
    t_reticle_to_bregma : numpy.ndarray
        Translation vector from reticle to bregma.
    combined_pairs_by_probe : dict
        Combined pairs of calibration data by probe.
    errs_by_probe : dict
        Errors by probe from the debug fits.
    """
    if probes_to_ignore_manual is None:
        probes_to_ignore_manual = []
    manual_calibration_files, parallax_directories = (
        _validate_combined_calibration_inputs(
            manual_calibration_files, parallax_directories
        )
    )
    # Read the first manual calibration to get the reticle metadata

    # Fit the manual calibrations
    manual_cal_by_probe = {}
    manual_pairs_by_probe = {}

    for filename in manual_calibration_files:
        cal_by_probe, R_reticle_to_bregma, t_reticle_to_bregma = (
            fit_rotation_params_from_manual_calibration(
                filename, *args, **kwargs
            )
        )
        manual_cal_by_probe.update(cal_by_probe)
        (
            adjusted_pairs_by_probe,
            global_offset,
            global_rotation_degrees,
            _,
        ) = read_manual_reticle_calibration(filename)
        manual_pairs_by_probe.update(adjusted_pairs_by_probe)

    combined_cal_by_probe: dict[
        int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]
    ] = {}
    combined_pairs_by_probe: dict[
        int, tuple[NDArray[np.floating[Any]], NDArray[np.floating[Any]]]
    ] = {}
    for parallax_dir in parallax_directories:
        adjusted_pairs_by_probe = read_parallax_calibration_dir_and_correct(
            parallax_dir,
            global_offset,
            global_rotation_degrees,
            local_scale_factor,
            global_scale_factor,
        )
        combined_cal_by_probe.update(
            _fit_by_probe(adjusted_pairs_by_probe, *args, **kwargs)
        )
        combined_pairs_by_probe.update(adjusted_pairs_by_probe)
    for probe_name in probes_to_ignore_manual:
        manual_cal_by_probe.pop(probe_name, None)  # type: ignore
        manual_pairs_by_probe.pop(probe_name, None)  # type: ignore
    combined_cal_by_probe.update(manual_cal_by_probe)
    combined_pairs_by_probe.update(manual_pairs_by_probe)
    errs_by_probe = _debug_fits(
        combined_cal_by_probe,
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        combined_pairs_by_probe,
    )
    return (
        combined_cal_by_probe,
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        combined_pairs_by_probe,
        errs_by_probe,
    )
