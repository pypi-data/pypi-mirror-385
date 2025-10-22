"""
SheetManager 单元测试
"""

import pytest
from openpyxl import load_workbook

from mcp_excel_supabase.excel.sheet_manager import SheetManager
from mcp_excel_supabase.utils.errors import (
    SheetNotFoundError,
    SheetAlreadyExistsError,
    ValidationError,
)


class TestSheetManager:
    """SheetManager 测试类"""

    def test_create_sheet_at_end(self, simple_excel_file):
        """测试在末尾创建工作表"""
        manager = SheetManager()

        # 创建工作表
        result = manager.create_sheet(simple_excel_file, "NewSheet")

        assert result["success"] is True

        # 验证工作表已创建
        wb = load_workbook(simple_excel_file)
        assert "NewSheet" in wb.sheetnames
        assert wb.sheetnames[-1] == "NewSheet"  # 在末尾

    def test_create_sheet_at_position(self, simple_excel_file):
        """测试在指定位置创建工作表"""
        manager = SheetManager()

        # 在位置 0 创建工作表
        result = manager.create_sheet(simple_excel_file, "FirstSheet", position=0)

        assert result["success"] is True

        # 验证工作表在第一个位置
        wb = load_workbook(simple_excel_file)
        assert "FirstSheet" in wb.sheetnames
        assert wb.sheetnames[0] == "FirstSheet"

    def test_create_sheet_already_exists(self, simple_excel_file):
        """测试创建已存在的工作表"""
        manager = SheetManager()

        # 获取现有工作表名称
        wb = load_workbook(simple_excel_file)
        existing_name = wb.sheetnames[0]

        # 尝试创建同名工作表
        with pytest.raises(SheetAlreadyExistsError):
            manager.create_sheet(simple_excel_file, existing_name)

    def test_create_sheet_invalid_name(self, simple_excel_file):
        """测试创建无效名称的工作表"""
        manager = SheetManager()

        # 工作表名称过长（>31字符）
        long_name = "A" * 32
        with pytest.raises(ValidationError):
            manager.create_sheet(simple_excel_file, long_name)

    def test_delete_sheet(self, multi_sheet_excel_file):
        """测试删除工作表"""
        manager = SheetManager()

        # 删除工作表
        result = manager.delete_sheet(multi_sheet_excel_file, "Expenses")

        assert result["success"] is True

        # 验证工作表已删除
        wb = load_workbook(multi_sheet_excel_file)
        assert "Expenses" not in wb.sheetnames

    def test_delete_sheet_not_found(self, simple_excel_file):
        """测试删除不存在的工作表"""
        manager = SheetManager()

        with pytest.raises(SheetNotFoundError):
            manager.delete_sheet(simple_excel_file, "NonExistent")

    def test_delete_last_sheet(self, simple_excel_file):
        """测试删除最后一个工作表"""
        manager = SheetManager()

        # 获取唯一的工作表名称
        wb = load_workbook(simple_excel_file)
        sheet_name = wb.sheetnames[0]

        # 尝试删除最后一个工作表
        with pytest.raises(ValidationError) as exc_info:
            manager.delete_sheet(simple_excel_file, sheet_name)

        assert "不能删除最后一个工作表" in str(exc_info.value)

    def test_rename_sheet(self, simple_excel_file):
        """测试重命名工作表"""
        manager = SheetManager()

        # 获取原工作表名称
        wb = load_workbook(simple_excel_file)
        old_name = wb.sheetnames[0]

        # 重命名工作表
        new_name = "RenamedSheet"
        result = manager.rename_sheet(simple_excel_file, old_name, new_name)

        assert result["success"] is True

        # 验证工作表已重命名
        wb = load_workbook(simple_excel_file)
        assert old_name not in wb.sheetnames
        assert new_name in wb.sheetnames

    def test_rename_sheet_not_found(self, simple_excel_file):
        """测试重命名不存在的工作表"""
        manager = SheetManager()

        with pytest.raises(SheetNotFoundError):
            manager.rename_sheet(simple_excel_file, "NonExistent", "NewName")

    def test_rename_sheet_name_exists(self, multi_sheet_excel_file):
        """测试重命名为已存在的名称"""
        manager = SheetManager()

        # 尝试重命名为已存在的名称
        with pytest.raises(SheetAlreadyExistsError):
            manager.rename_sheet(multi_sheet_excel_file, "Sales", "Expenses")

    def test_copy_sheet(self, simple_excel_file):
        """测试复制工作表"""
        manager = SheetManager()

        # 获取源工作表名称
        wb = load_workbook(simple_excel_file)
        source_name = wb.sheetnames[0]

        # 复制工作表
        target_name = "CopiedSheet"
        result = manager.copy_sheet(simple_excel_file, source_name, target_name)

        assert result["success"] is True

        # 验证工作表已复制
        wb = load_workbook(simple_excel_file)
        assert source_name in wb.sheetnames
        assert target_name in wb.sheetnames

    def test_copy_sheet_with_position(self, simple_excel_file):
        """测试复制工作表到指定位置"""
        manager = SheetManager()

        # 获取源工作表名称
        wb = load_workbook(simple_excel_file)
        source_name = wb.sheetnames[0]

        # 复制工作表到位置 0
        target_name = "CopiedSheet"
        result = manager.copy_sheet(simple_excel_file, source_name, target_name, position=0)

        assert result["success"] is True

        # 验证工作表在第一个位置
        wb = load_workbook(simple_excel_file)
        assert wb.sheetnames[0] == target_name

    def test_copy_sheet_not_found(self, simple_excel_file):
        """测试复制不存在的工作表"""
        manager = SheetManager()

        with pytest.raises(SheetNotFoundError):
            manager.copy_sheet(simple_excel_file, "NonExistent", "NewSheet")

    def test_copy_sheet_target_exists(self, multi_sheet_excel_file):
        """测试复制到已存在的名称"""
        manager = SheetManager()

        with pytest.raises(SheetAlreadyExistsError):
            manager.copy_sheet(multi_sheet_excel_file, "Sales", "Expenses")

    def test_move_sheet(self, multi_sheet_excel_file):
        """测试移动工作表"""
        manager = SheetManager()

        # 移动工作表到位置 0
        result = manager.move_sheet(multi_sheet_excel_file, "Summary", 0)

        assert result["success"] is True

        # 验证工作表已移动
        wb = load_workbook(multi_sheet_excel_file)
        assert wb.sheetnames[0] == "Summary"

    def test_move_sheet_not_found(self, simple_excel_file):
        """测试移动不存在的工作表"""
        manager = SheetManager()

        with pytest.raises(SheetNotFoundError):
            manager.move_sheet(simple_excel_file, "NonExistent", 0)

    def test_move_sheet_invalid_position(self, multi_sheet_excel_file):
        """测试移动到无效位置"""
        manager = SheetManager()

        # 位置超出范围
        with pytest.raises(ValidationError):
            manager.move_sheet(multi_sheet_excel_file, "Sales", 999)

    def test_copy_sheet_preserves_data(self, formatted_excel_file):
        """测试复制工作表保留数据"""
        manager = SheetManager()

        # 获取源工作表
        wb = load_workbook(formatted_excel_file)
        source_name = wb.sheetnames[0]
        source_ws = wb[source_name]
        source_value = source_ws["A1"].value

        # 复制工作表
        target_name = "CopiedSheet"
        manager.copy_sheet(formatted_excel_file, source_name, target_name)

        # 验证数据已复制
        wb = load_workbook(formatted_excel_file)
        target_ws = wb[target_name]
        assert target_ws["A1"].value == source_value
