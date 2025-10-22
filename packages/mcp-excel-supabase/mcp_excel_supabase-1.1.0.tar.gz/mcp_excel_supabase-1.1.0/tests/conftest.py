"""
Pytest 配置文件

提供测试所需的 fixtures 和配置，包括：
- 测试数据生成器
- 临时文件和目录管理
- Mock 对象
- 测试环境配置
"""

import os
import sys
import pytest
import tempfile
import shutil
from pathlib import Path
from typing import Generator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 将 src 目录添加到 Python 路径
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))


# ============================================================================
# 测试环境配置
# ============================================================================


@pytest.fixture(scope="session", autouse=True)
def setup_test_environment():
    """设置测试环境"""
    # 设置测试环境变量
    os.environ["LOG_LEVEL"] = "DEBUG"
    os.environ["SUPABASE_URL"] = "https://test.supabase.co"
    os.environ["SUPABASE_KEY"] = "test_key_12345"
    os.environ["DEFAULT_BUCKET"] = "test_bucket"

    yield

    # 清理（如果需要）
    pass


# ============================================================================
# 临时目录和文件 Fixtures
# ============================================================================


@pytest.fixture
def temp_dir() -> Generator[Path, None, None]:
    """
    创建临时目录

    Yields:
        临时目录的 Path 对象
    """
    temp_path = Path(tempfile.mkdtemp())
    yield temp_path
    # 测试结束后清理
    if temp_path.exists():
        shutil.rmtree(temp_path)


@pytest.fixture
def temp_file(temp_dir: Path) -> Generator[Path, None, None]:
    """
    创建临时文件

    Yields:
        临时文件的 Path 对象
    """
    temp_file_path = temp_dir / "test_file.txt"
    temp_file_path.write_text("test content")
    yield temp_file_path


# ============================================================================
# Excel 测试数据生成器
# ============================================================================


@pytest.fixture
def simple_excel_file(temp_dir: Path) -> Path:
    """
    创建简单的 Excel 文件（无格式）

    Returns:
        Excel 文件路径
    """
    file_path = temp_dir / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 添加简单数据
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["C1"] = "City"

    ws["A2"] = "Alice"
    ws["B2"] = 25
    ws["C2"] = "Beijing"

    ws["A3"] = "Bob"
    ws["B3"] = 30
    ws["C3"] = "Shanghai"

    ws["A4"] = "Charlie"
    ws["B4"] = 35
    ws["C4"] = "Guangzhou"

    wb.save(file_path)
    return file_path


@pytest.fixture
def formatted_excel_file(temp_dir: Path) -> Path:
    """
    创建带格式的 Excel 文件

    Returns:
        Excel 文件路径
    """
    file_path = temp_dir / "formatted.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "FormattedSheet"

    # 添加数据
    headers = ["Product", "Price", "Quantity", "Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        # 设置表头格式
        cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 添加数据行
    data = [
        ["Apple", 5.5, 10, 55.0],
        ["Banana", 3.2, 15, 48.0],
        ["Orange", 4.8, 12, 57.6],
    ]

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")

            # 为数字列添加边框
            if col_idx > 1:
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

    # 设置列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    # 设置行高
    ws.row_dimensions[1].height = 25

    wb.save(file_path)
    return file_path


@pytest.fixture
def multi_sheet_excel_file(temp_dir: Path) -> Path:
    """
    创建多 Sheet 的 Excel 文件

    Returns:
        Excel 文件路径
    """
    file_path = temp_dir / "multi_sheet.xlsx"
    wb = Workbook()

    # 第一个 Sheet
    ws1 = wb.active
    ws1.title = "Sales"
    ws1["A1"] = "Month"
    ws1["B1"] = "Revenue"
    ws1["A2"] = "January"
    ws1["B2"] = 10000
    ws1["A3"] = "February"
    ws1["B3"] = 12000

    # 第二个 Sheet
    ws2 = wb.create_sheet("Expenses")
    ws2["A1"] = "Category"
    ws2["B1"] = "Amount"
    ws2["A2"] = "Rent"
    ws2["B2"] = 5000
    ws2["A3"] = "Utilities"
    ws2["B3"] = 1000

    # 第三个 Sheet
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Total Revenue"
    ws3["B1"] = 22000
    ws3["A2"] = "Total Expenses"
    ws3["B2"] = 6000
    ws3["A3"] = "Net Profit"
    ws3["B3"] = 16000

    wb.save(file_path)
    return file_path


@pytest.fixture
def excel_with_formulas(temp_dir: Path) -> Path:
    """
    创建包含公式的 Excel 文件

    Returns:
        Excel 文件路径
    """
    file_path = temp_dir / "with_formulas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculations"

    # 添加数据
    ws["A1"] = "Value1"
    ws["B1"] = "Value2"
    ws["C1"] = "Sum"
    ws["D1"] = "Average"

    ws["A2"] = 10
    ws["B2"] = 20
    ws["C2"] = "=A2+B2"
    ws["D2"] = "=AVERAGE(A2:B2)"

    ws["A3"] = 15
    ws["B3"] = 25
    ws["C3"] = "=A3+B3"
    ws["D3"] = "=AVERAGE(A3:B3)"

    ws["A4"] = 30
    ws["B4"] = 40
    ws["C4"] = "=A4+B4"
    ws["D4"] = "=AVERAGE(A4:B4)"

    # 总计行
    ws["A5"] = "Total"
    ws["C5"] = "=SUM(C2:C4)"
    ws["D5"] = "=AVERAGE(D2:D4)"

    wb.save(file_path)
    return file_path


@pytest.fixture
def merged_cells_excel_file(temp_dir: Path) -> Path:
    """
    创建包含合并单元格的 Excel 文件

    Returns:
        Excel 文件路径
    """
    file_path = temp_dir / "merged_cells.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MergedCells"

    # 合并单元格
    ws.merge_cells("A1:C1")
    ws["A1"] = "Merged Header"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, size=14)

    ws.merge_cells("A2:A3")
    ws["A2"] = "Merged Vertical"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # 普通数据
    ws["B2"] = "Data1"
    ws["C2"] = "Data2"
    ws["B3"] = "Data3"
    ws["C3"] = "Data4"

    wb.save(file_path)
    return file_path


# ============================================================================
# 大文件生成器
# ============================================================================


@pytest.fixture
def large_excel_file(temp_dir: Path) -> Path:
    """
    创建大型 Excel 文件（用于性能测试）

    Returns:
        Excel 文件路径
    """
    file_path = temp_dir / "large.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "LargeData"

    # 添加表头
    headers = ["ID", "Name", "Value1", "Value2", "Value3"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # 添加 1000 行数据
    for row in range(2, 1002):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=f"Item_{row-1}")
        ws.cell(row=row, column=3, value=row * 10)
        ws.cell(row=row, column=4, value=row * 20)
        ws.cell(row=row, column=5, value=row * 30)

    wb.save(file_path)
    return file_path


# ============================================================================
# JSON 测试数据
# ============================================================================


@pytest.fixture
def sample_json_data() -> dict:
    """
    示例 JSON 数据（用于测试 JSON 到 Excel 的转换）

    Returns:
        JSON 数据字典
    """
    return {
        "workbook": {
            "sheets": [
                {
                    "name": "Sheet1",
                    "rows": [
                        {
                            "cells": [
                                {"value": "Name", "row": 1, "col": 1},
                                {"value": "Age", "row": 1, "col": 2},
                                {"value": "City", "row": 1, "col": 3},
                            ]
                        },
                        {
                            "cells": [
                                {"value": "Alice", "row": 2, "col": 1},
                                {"value": 25, "row": 2, "col": 2},
                                {"value": "Beijing", "row": 2, "col": 3},
                            ]
                        },
                    ],
                }
            ]
        }
    }


# ============================================================================
# Mock Fixtures
# ============================================================================


@pytest.fixture
def mock_supabase_client(monkeypatch):
    """
    Mock Supabase 客户端

    用于测试时避免实际的网络请求
    """

    class MockSupabaseClient:
        def __init__(self):
            self.storage = MockStorage()

    class MockStorage:
        def from_(self, bucket_name):
            return MockBucket(bucket_name)

    class MockBucket:
        def __init__(self, bucket_name):
            self.bucket_name = bucket_name

        def upload(self, path, file_data):
            return {"path": path, "status": "success"}

        def download(self, path):
            return b"mock file content"

        def list(self, path=""):
            return [{"name": "file1.xlsx"}, {"name": "file2.xlsx"}]

        def remove(self, paths):
            return {"status": "success"}

    return MockSupabaseClient()


@pytest.fixture
def merged_excel_file(temp_dir: Path) -> Path:
    """
    创建包含合并单元格的 Excel 文件

    Returns:
        Excel 文件路径
    """
    file_path = temp_dir / "merged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 添加数据
    ws["A1"] = "Merged Header"
    ws["A2"] = "Data1"
    ws["B2"] = "Data2"

    # 合并单元格 A1:B2
    ws.merge_cells("A1:B2")

    wb.save(file_path)
    return file_path
