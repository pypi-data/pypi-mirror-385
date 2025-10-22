"""
文件合并器模块

提供多个 Excel 文件的合并功能，支持重名处理和格式保留。
"""

from typing import Any, List, Optional, Literal

from openpyxl import load_workbook, Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet

from ..utils.errors import ValidationError
from ..utils.validator import Validator, validate_excel_file
from ..utils.logger import get_logger

logger = get_logger(__name__)


class FileMerger:
    """
    Excel 文件合并器

    提供多个 Excel 文件的合并功能，支持：
    - 保留所有格式信息（字体、颜色、边框、合并单元格等）
    - 多种重名处理策略（rename、skip、overwrite）
    - 选择性合并（指定工作表）
    """

    def __init__(self) -> None:
        """初始化 FileMerger"""
        self.validator = Validator()
        logger.debug("FileMerger 初始化完成")

    def merge_files(
        self,
        file_paths: List[str],
        output_path: str,
        handle_duplicates: Literal["rename", "skip", "overwrite"] = "rename",
        preserve_formats: bool = True,
        sheet_names: Optional[List[str]] = None,
    ) -> dict[str, Any]:
        """
        合并多个 Excel 文件

        Args:
            file_paths: 要合并的 Excel 文件路径列表
            output_path: 输出文件路径
            handle_duplicates: 重名处理策略
                - 'rename': 自动重命名（如 Sheet1 → Sheet1_2）
                - 'skip': 跳过重名工作表
                - 'overwrite': 覆盖已存在的工作表
            preserve_formats: 是否保留格式信息（默认 True）
            sheet_names: 要合并的工作表名称列表（None 表示全部）

        Returns:
            包含合并结果的字典

        Raises:
            ValidationError: 参数无效
            FileNotFoundError: 文件不存在
        """
        logger.info(f"开始合并 {len(file_paths)} 个 Excel 文件到 {output_path}")

        # 验证参数
        if not file_paths:
            raise ValidationError(
                error_code="E204",
                message="文件路径列表不能为空",
                context={"file_paths": file_paths},
                suggestion="请提供至少一个 Excel 文件路径",
            )

        if handle_duplicates not in ["rename", "skip", "overwrite"]:
            raise ValidationError(
                error_code="E204",
                message=f"无效的重名处理策略: {handle_duplicates}",
                context={"handle_duplicates": handle_duplicates},
                suggestion="请使用 'rename'、'skip' 或 'overwrite'",
            )

        # 验证所有文件存在
        for file_path in file_paths:
            validate_excel_file(file_path)

        # 创建新工作簿
        output_wb = OpenpyxlWorkbook()
        # 删除默认创建的空工作表
        if "Sheet" in output_wb.sheetnames:
            output_wb.remove(output_wb["Sheet"])

        # 统计信息
        merged_sheets = 0
        skipped_sheets = 0
        renamed_sheets = 0

        # 遍历所有文件
        for file_path in file_paths:
            logger.debug(f"处理文件: {file_path}")

            # 加载源工作簿
            source_wb = load_workbook(file_path)

            # 遍历所有工作表
            for sheet_name in source_wb.sheetnames:
                # 如果指定了工作表名称列表，只处理指定的工作表
                if sheet_names is not None and sheet_name not in sheet_names:
                    logger.debug(f"跳过工作表（未在指定列表中）: {sheet_name}")
                    continue

                # 处理重名
                target_name = sheet_name
                if sheet_name in output_wb.sheetnames:
                    if handle_duplicates == "skip":
                        logger.debug(f"跳过重名工作表: {sheet_name}")
                        skipped_sheets += 1
                        continue
                    elif handle_duplicates == "overwrite":
                        logger.debug(f"覆盖已存在的工作表: {sheet_name}")
                        output_wb.remove(output_wb[sheet_name])
                    elif handle_duplicates == "rename":
                        # 自动重命名
                        target_name = self._generate_unique_name(sheet_name, output_wb.sheetnames)
                        logger.debug(f"重命名工作表: {sheet_name} -> {target_name}")
                        renamed_sheets += 1

                # 复制工作表
                source_ws = source_wb[sheet_name]
                self._copy_worksheet(source_ws, output_wb, target_name, preserve_formats)
                merged_sheets += 1
                logger.debug(f"已合并工作表: {target_name}")

        # 保存输出文件
        output_wb.save(output_path)
        logger.info(
            f"文件合并完成: {merged_sheets} 个工作表已合并, "
            f"{skipped_sheets} 个跳过, {renamed_sheets} 个重命名"
        )

        return {
            "success": True,
            "merged_sheets": merged_sheets,
            "skipped_sheets": skipped_sheets,
            "renamed_sheets": renamed_sheets,
            "output_path": output_path,
        }

    def _generate_unique_name(self, base_name: str, existing_names: List[str]) -> str:
        """
        生成唯一的工作表名称

        Args:
            base_name: 基础名称
            existing_names: 已存在的名称列表

        Returns:
            唯一的名称（如 Sheet1_2）
        """
        if base_name not in existing_names:
            return base_name

        # 尝试添加数字后缀
        counter = 2
        while True:
            new_name = f"{base_name}_{counter}"
            if new_name not in existing_names:
                return new_name
            counter += 1

            # 防止无限循环（理论上不会发生）
            if counter > 1000:
                raise ValidationError(
                    error_code="E401",
                    message=f"无法为工作表 '{base_name}' 生成唯一名称",
                    context={"base_name": base_name, "counter": counter},
                    suggestion="请手动重命名部分工作表",
                )

    def _copy_worksheet(
        self,
        source_ws: Worksheet,
        target_wb: OpenpyxlWorkbook,
        target_name: str,
        preserve_formats: bool,
    ) -> None:
        """
        复制工作表到目标工作簿

        Args:
            source_ws: 源工作表
            target_wb: 目标工作簿
            target_name: 目标工作表名称
            preserve_formats: 是否保留格式
        """
        # 创建新工作表
        target_ws = target_wb.create_sheet(title=target_name)

        # 复制单元格数据和格式
        for row in source_ws.iter_rows():
            for cell in row:
                target_cell = target_ws[cell.coordinate]

                # 复制值
                target_cell.value = cell.value

                # 复制格式
                if preserve_formats and cell.has_style:
                    # 复制字体
                    if cell.font:
                        target_cell.font = cell.font.copy()

                    # 复制填充
                    if cell.fill:
                        target_cell.fill = cell.fill.copy()

                    # 复制边框
                    if cell.border:
                        target_cell.border = cell.border.copy()

                    # 复制对齐
                    if cell.alignment:
                        target_cell.alignment = cell.alignment.copy()

                    # 复制数字格式
                    if cell.number_format:
                        target_cell.number_format = cell.number_format

        # 复制合并单元格
        if preserve_formats:
            for merged_range in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merged_range))

        # 复制列宽
        if preserve_formats:
            for col_letter, col_dim in source_ws.column_dimensions.items():
                if col_dim.width:
                    target_ws.column_dimensions[col_letter].width = col_dim.width

        # 复制行高
        if preserve_formats:
            for row_num, row_dim in source_ws.row_dimensions.items():
                if row_dim.height:
                    target_ws.row_dimensions[row_num].height = row_dim.height

        logger.debug(f"工作表复制完成: {target_name}")
