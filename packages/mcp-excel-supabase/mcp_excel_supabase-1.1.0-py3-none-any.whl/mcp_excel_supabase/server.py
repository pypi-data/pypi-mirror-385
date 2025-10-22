"""
MCP Excel Supabase Server

提供 Excel 文件处理和 Supabase 存储管理的 MCP 服务器。
"""

from typing import Any, Dict, List, Optional

from mcp.server.fastmcp import FastMCP

from mcp_excel_supabase.excel.parser import ExcelParser
from mcp_excel_supabase.excel.generator import ExcelGenerator
from mcp_excel_supabase.excel.format_editor import FormatEditor
from mcp_excel_supabase.excel.cell_merger import CellMerger
from mcp_excel_supabase.excel.dimension_adjuster import DimensionAdjuster
from mcp_excel_supabase.excel.formula_manager import FormulaManager
from mcp_excel_supabase.excel.sheet_manager import SheetManager
from mcp_excel_supabase.excel.file_merger import FileMerger
from mcp_excel_supabase.storage.uploader import FileUploader
from mcp_excel_supabase.storage.downloader import FileDownloader
from mcp_excel_supabase.storage.manager import FileManager
from mcp_excel_supabase.tools.schemas import (
    ParseExcelOutput,
    CreateExcelOutput,
    ModifyCellFormatOutput,
    MergeCellsOutput,
    UnmergeCellsOutput,
    SetRowHeightsOutput,
    SetColumnWidthsOutput,
    ManageStorageOutput,
    SetFormulaOutput,
    RecalculateFormulasOutput,
    ManageSheetsOutput,
    MergeExcelFilesOutput,
)
from mcp_excel_supabase.utils.logger import get_logger

# 创建日志记录器
logger = get_logger(__name__)

# 创建 MCP 服务器实例
mcp = FastMCP("Excel-Supabase-Server")


# ============================================================================
# Tool 1: Parse Excel to JSON
# ============================================================================


@mcp.tool()
def parse_excel_to_json(file_path: str, extract_formats: bool = True) -> Dict[str, Any]:
    """
    解析 Excel 文件为 JSON 格式

    Args:
        file_path: Excel 文件路径（本地路径或 Supabase URL）
        extract_formats: 是否提取单元格格式信息（默认 True）

    Returns:
        包含解析结果的字典
    """
    try:
        logger.info(f"开始解析 Excel 文件: {file_path}")

        # 创建解析器
        parser = ExcelParser()

        # 解析文件
        workbook = parser.parse_file(file_path)

        # 转换为字典
        result = ParseExcelOutput(success=True, workbook=workbook.model_dump(), error=None)

        logger.info(f"Excel 文件解析成功: {file_path}")
        return result.model_dump()

    except Exception as e:
        logger.error(f"解析 Excel 文件失败: {str(e)}")
        result = ParseExcelOutput(success=False, workbook=None, error=str(e))
        return result.model_dump()


# ============================================================================
# Tool 2: Create Excel from JSON
# ============================================================================


@mcp.tool()
def create_excel_from_json(
    workbook_data: Dict[str, Any], output_path: str, apply_formats: bool = True
) -> Dict[str, Any]:
    """
    从 JSON 数据创建 Excel 文件

    Args:
        workbook_data: 工作簿数据（JSON 格式）
        output_path: 输出文件路径
        apply_formats: 是否应用单元格格式（默认 True）

    Returns:
        包含创建结果的字典
    """
    try:
        logger.info(f"开始创建 Excel 文件: {output_path}")

        # 创建生成器
        generator = ExcelGenerator()

        # 生成文件
        generator.generate_file(workbook_data, output_path)

        # 返回结果
        result = CreateExcelOutput(success=True, file_path=output_path, error=None)

        logger.info(f"Excel 文件创建成功: {output_path}")
        return result.model_dump()

    except Exception as e:
        logger.error(f"创建 Excel 文件失败: {str(e)}")
        result = CreateExcelOutput(success=False, file_path=None, error=str(e))
        return result.model_dump()


# ============================================================================
# Tool 3: Modify Cell Format
# ============================================================================


@mcp.tool()
def modify_cell_format(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    format_data: Dict[str, Any],
    output_path: Optional[str] = None,
) -> Dict[str, Any]:
    """
    修改单元格格式

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        cell_range: 单元格范围，如 'A1' 或 'A1:B10'
        format_data: 格式数据（JSON 格式）
        output_path: 输出文件路径（默认覆盖原文件）

    Returns:
        包含修改结果的字典
    """
    try:
        logger.info(f"开始修改单元格格式: {file_path}, {sheet_name}, {cell_range}")

        # 如果没有指定输出路径，使用原文件路径
        if output_path is None:
            output_path = file_path

        # 先解析Excel文件
        parser = ExcelParser()
        workbook = parser.parse_file(file_path)

        # 创建格式编辑器
        editor = FormatEditor(workbook)

        # 解析单元格范围为坐标列表，并批量修改格式
        cells = _cells_from_range(cell_range)
        editor.modify_cells_format(
            sheet_name,
            cells,
            font=format_data.get("font"),
            fill=format_data.get("fill"),
            border=format_data.get("border"),
            alignment=format_data.get("alignment"),
            number_format=format_data.get("number_format"),
        )

        # 生成修改后的文件（覆盖原文件）
        generator = ExcelGenerator()
        generator.generate_file(workbook, output_path, overwrite=True)

        # 计算修改的单元格数量
        cells_modified = _count_cells_in_range(cell_range)

        # 返回结果
        result = ModifyCellFormatOutput(
            success=True,
            file_path=output_path,
            cells_modified=cells_modified,
            error=None,
        )

        logger.info(f"单元格格式修改成功: {output_path}")
        return result.model_dump()

    except Exception as e:
        logger.error(f"修改单元格格式失败: {str(e)}")
        result = ModifyCellFormatOutput(
            success=False, file_path=None, cells_modified=None, error=str(e)
        )
        return result.model_dump()


# ============================================================================
# Tool 4: Merge Cells
# ============================================================================


@mcp.tool()
def merge_cells(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    output_path: Optional[str] = None,
) -> Dict[str, Any]:
    """
    合并单元格

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        cell_range: 要合并的单元格范围，如 'A1:B2'
        output_path: 输出文件路径（默认覆盖原文件）

    Returns:
        包含合并结果的字典
    """
    try:
        logger.info(f"开始合并单元格: {file_path}, {sheet_name}, {cell_range}")

        # 如果没有指定输出路径，使用原文件路径
        if output_path is None:
            output_path = file_path

        # 先解析Excel文件
        parser = ExcelParser()
        workbook = parser.parse_file(file_path)

        # 创建单元格合并器
        merger = CellMerger(workbook)

        # 合并单元格
        start_row, start_col, end_row, end_col = _range_start_end(cell_range)
        merger.merge_cells(sheet_name, start_row, start_col, end_row, end_col)

        # 生成修改后的文件（覆盖原文件）
        generator = ExcelGenerator()
        generator.generate_file(workbook, output_path, overwrite=True)

        # 返回结果
        result = MergeCellsOutput(
            success=True, file_path=output_path, merged_range=cell_range, error=None
        )

        logger.info(f"单元格合并成功: {output_path}")
        return result.model_dump()

    except Exception as e:
        logger.error(f"合并单元格失败: {str(e)}")
        result = MergeCellsOutput(success=False, file_path=None, merged_range=None, error=str(e))
        return result.model_dump()


# ============================================================================
# Tool 5: Unmerge Cells
# ============================================================================


@mcp.tool()
def unmerge_cells(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    output_path: Optional[str] = None,
) -> Dict[str, Any]:
    """
    取消合并单元格

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        cell_range: 要取消合并的单元格范围，如 'A1:B2'
        output_path: 输出文件路径（默认覆盖原文件）

    Returns:
        包含取消合并结果的字典
    """
    try:
        logger.info(f"开始取消合并单元格: {file_path}, {sheet_name}, {cell_range}")

        # 如果没有指定输出路径，使用原文件路径
        if output_path is None:
            output_path = file_path

        # 先解析Excel文件
        parser = ExcelParser()
        workbook = parser.parse_file(file_path)

        # 创建单元格合并器
        merger = CellMerger(workbook)

        # 取消合并单元格（使用范围左上角单元格）
        start_row, start_col, _, _ = _range_start_end(cell_range)
        merger.unmerge_cells(sheet_name, start_row, start_col)

        # 生成修改后的文件（覆盖原文件）
        generator = ExcelGenerator()
        generator.generate_file(workbook, output_path, overwrite=True)

        # 返回结果
        result = UnmergeCellsOutput(
            success=True, file_path=output_path, unmerged_range=cell_range, error=None
        )

        logger.info(f"取消合并单元格成功: {output_path}")
        return result.model_dump()

    except Exception as e:
        logger.error(f"取消合并单元格失败: {str(e)}")
        result = UnmergeCellsOutput(
            success=False, file_path=None, unmerged_range=None, error=str(e)
        )
        return result.model_dump()


# ============================================================================
# Tool 6: Set Row Heights
# ============================================================================


@mcp.tool()
def set_row_heights(
    file_path: str,
    sheet_name: str,
    row_heights: List[Dict[str, Any]],
    output_path: Optional[str] = None,
) -> Dict[str, Any]:
    """
    设置行高

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        row_heights: 行高规格列表，每项包含 row_number 和 height
        output_path: 输出文件路径（默认覆盖原文件）

    Returns:
        包含设置结果的字典
    """
    try:
        logger.info(f"开始设置行高: {file_path}, {sheet_name}")

        # 如果没有指定输出路径，使用原文件路径
        if output_path is None:
            output_path = file_path

        # 先解析Excel文件
        parser = ExcelParser()
        workbook = parser.parse_file(file_path)

        # 创建尺寸调整器
        adjuster = DimensionAdjuster(workbook)

        # 转换为字典格式
        heights_dict = {spec["row_number"]: spec["height"] for spec in row_heights}

        # 空输入直接返回成功
        if not heights_dict:
            result = SetRowHeightsOutput(
                success=True,
                file_path=output_path,
                rows_modified=0,
                error=None,
            )
            logger.info(f"行高设置成功(空列表): {output_path}")
            return result.model_dump()

        # 设置行高
        adjuster.set_row_heights(sheet_name, heights_dict)

        # 生成修改后的文件（覆盖原文件）
        generator = ExcelGenerator()
        generator.generate_file(workbook, output_path, overwrite=True)

        # 返回结果
        result = SetRowHeightsOutput(
            success=True,
            file_path=output_path,
            rows_modified=len(row_heights),
            error=None,
        )

        logger.info(f"行高设置成功: {output_path}")
        return result.model_dump()

    except Exception as e:
        logger.error(f"设置行高失败: {str(e)}")
        result = SetRowHeightsOutput(
            success=False, file_path=None, rows_modified=None, error=str(e)
        )
        return result.model_dump()


# ============================================================================
# Tool 7: Set Column Widths
# ============================================================================


@mcp.tool()
def set_column_widths(
    file_path: str,
    sheet_name: str,
    column_widths: List[Dict[str, Any]],
    output_path: Optional[str] = None,
) -> Dict[str, Any]:
    """
    设置列宽

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        column_widths: 列宽规格列表，每项包含 column_letter 和 width
        output_path: 输出文件路径（默认覆盖原文件）

    Returns:
        包含设置结果的字典
    """
    try:
        logger.info(f"开始设置列宽: {file_path}, {sheet_name}")

        # 如果没有指定输出路径，使用原文件路径
        if output_path is None:
            output_path = file_path

        # 先解析Excel文件
        parser = ExcelParser()
        workbook = parser.parse_file(file_path)

        # 创建尺寸调整器
        adjuster = DimensionAdjuster(workbook)

        # 转换为数字列号的字典格式
        widths_dict = {
            _column_letter_to_index(spec["column_letter"]): spec["width"] for spec in column_widths
        }

        # 空输入直接返回成功
        if not widths_dict:
            result = SetColumnWidthsOutput(
                success=True,
                file_path=output_path,
                columns_modified=0,
                error=None,
            )
            logger.info(f"列宽设置成功(空列表): {output_path}")
            return result.model_dump()

        # 设置列宽
        adjuster.set_column_widths(sheet_name, widths_dict)

        # 生成修改后的文件（覆盖原文件）
        generator = ExcelGenerator()
        generator.generate_file(workbook, output_path, overwrite=True)

        # 返回结果
        result = SetColumnWidthsOutput(
            success=True,
            file_path=output_path,
            columns_modified=len(column_widths),
            error=None,
        )

        logger.info(f"列宽设置成功: {output_path}")
        return result.model_dump()

    except Exception as e:
        logger.error(f"设置列宽失败: {str(e)}")
        result = SetColumnWidthsOutput(
            success=False, file_path=None, columns_modified=None, error=str(e)
        )
        return result.model_dump()


# ============================================================================
# Tool 8: Manage Storage
# ============================================================================


@mcp.tool()
def manage_storage(
    operation: str,
    file_path: Optional[str] = None,
    remote_path: Optional[str] = None,
    bucket_name: Optional[str] = None,
    search_pattern: Optional[str] = None,
    prefix: Optional[str] = None,
) -> Dict[str, Any]:
    """
    管理 Supabase 存储

    Args:
        operation: 操作类型 ('upload', 'download', 'list', 'delete', 'search')
        file_path: 本地文件路径（用于 upload/download）
        remote_path: 远程文件路径（用于 upload/download/delete）
        bucket_name: 存储桶名称
        search_pattern: 搜索模式（用于 search）
        prefix: 路径前缀（用于 list）

    Returns:
        包含操作结果的字典
    """
    try:
        logger.info(f"开始执行存储操作: {operation}")

        result_data: Any = None

        if operation == "upload":
            if not file_path or not remote_path or not bucket_name:
                raise ValueError("upload 操作需要 file_path, remote_path 和 bucket_name")
            uploader = FileUploader()
            result_data = uploader.upload_file(
                file_path=file_path,
                bucket_name=bucket_name,
                remote_path=remote_path,
            )

        elif operation == "download":
            if not file_path or not remote_path or not bucket_name:
                raise ValueError("download 操作需要 file_path, remote_path 和 bucket_name")
            downloader = FileDownloader()
            result_data = downloader.download_file(
                remote_path=remote_path,
                local_path=file_path,
                bucket_name=bucket_name,
            )

        elif operation == "list":
            if not bucket_name:
                raise ValueError("list 操作需要 bucket_name")
            manager = FileManager()
            result_data = manager.list_files(bucket_name, path=prefix or "")

        elif operation == "delete":
            if not remote_path or not bucket_name:
                raise ValueError("delete 操作需要 remote_path 和 bucket_name")
            manager = FileManager()
            result_data = manager.delete_file(remote_path, bucket_name)

        elif operation == "search":
            if not bucket_name or not search_pattern:
                raise ValueError("search 操作需要 bucket_name 和 search_pattern")
            manager = FileManager()
            result_data = manager.search_files(bucket_name, search_pattern)

        else:
            raise ValueError(f"不支持的操作类型: {operation}")

        # 返回结果
        result = ManageStorageOutput(
            success=True, operation=operation, result=result_data, error=None
        )

        logger.info(f"存储操作成功: {operation}")
        return result.model_dump()

    except Exception as e:
        logger.error(f"存储操作失败: {str(e)}")
        result = ManageStorageOutput(success=False, operation=operation, result=None, error=str(e))
        return result.model_dump()


# ============================================================================
# Helper Functions
# ============================================================================


def _count_cells_in_range(cell_range: str) -> int:
    """
    计算单元格范围中的单元格数量

    Args:
        cell_range: 单元格范围，如 'A1' 或 'A1:B10'

    Returns:
        单元格数量
    """
    if ":" not in cell_range:
        # 单个单元格
        return 1

    # 范围
    from openpyxl.utils import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return (max_col - min_col + 1) * (max_row - min_row + 1)


def _range_start_end(cell_range: str) -> tuple[int, int, int, int]:
    """将范围字符串解析为起止行列 (start_row, start_col, end_row, end_col)."""
    from openpyxl.utils import range_boundaries

    if ":" in cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        return min_row, min_col, max_row, max_col
    else:
        # 单个单元格，如 'A1'
        min_col, min_row, max_col, max_row = range_boundaries(f"{cell_range}:{cell_range}")
        return min_row, min_col, max_row, max_col


def _cells_from_range(cell_range: str) -> list[tuple[int, int]]:
    """将范围字符串展开为所有 (row, col) 元组列表。"""
    from openpyxl.utils import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(
        cell_range if ":" in cell_range else f"{cell_range}:{cell_range}"
    )
    cells: list[tuple[int, int]] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cells.append((r, c))
    return cells


def _column_letter_to_index(letter: str) -> int:
    """将列字母转换为列号（1-based）。"""
    from openpyxl.utils import column_index_from_string

    return column_index_from_string(letter)


# ============================================================================
# Tool 9: Set Formula
# ============================================================================


@mcp.tool()
def set_formula(
    file_path: str, sheet_name: str, cell: str, formula: str, save: bool = True
) -> Dict[str, Any]:
    """
    设置单元格公式

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        cell: 单元格位置（如 "A1"）
        formula: 公式字符串（如 "=SUM(A1:A10)"）
        save: 是否保存文件（默认 True）

    Returns:
        包含设置结果的字典
    """
    try:
        logger.info(f"设置公式: {file_path} - {sheet_name}!{cell} = {formula}")

        # 创建公式管理器
        manager = FormulaManager()

        # 设置公式
        result = manager.set_formula(
            file_path=file_path, sheet_name=sheet_name, cell=cell, formula=formula, save=save
        )

        # 返回结果
        output = SetFormulaOutput(
            success=True,
            cell=result.get("cell"),
            formula=result.get("formula"),
            message=result.get("message"),
            error=None,
        )

        logger.info(f"公式设置成功: {sheet_name}!{cell}")
        return output.model_dump()

    except Exception as e:
        logger.error(f"设置公式失败: {str(e)}")
        output = SetFormulaOutput(
            success=False, cell=None, formula=None, message=None, error=str(e)
        )
        return output.model_dump()


# ============================================================================
# Tool 10: Recalculate Formulas
# ============================================================================


@mcp.tool()
def recalculate_formulas(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    重新计算 Excel 文件中的公式

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称（如果为空则计算所有工作表）

    Returns:
        包含计算结果的字典
    """
    try:
        logger.info(f"重新计算公式: {file_path}" + (f" - {sheet_name}" if sheet_name else ""))

        # 创建公式管理器
        manager = FormulaManager()

        # 重新计算
        if sheet_name:
            result = manager.recalculate_sheet(file_path=file_path, sheet_name=sheet_name)
        else:
            result = manager.recalculate_all(file_path=file_path)

        # 返回结果
        output = RecalculateFormulasOutput(
            success=True,
            count=result.get("count"),
            results=result.get("results"),
            message=result.get("message"),
            error=None,
        )

        logger.info(f"公式计算成功，共 {result.get('count')} 个结果")
        return output.model_dump()

    except Exception as e:
        logger.error(f"重新计算公式失败: {str(e)}")
        output = RecalculateFormulasOutput(
            success=False, count=None, results=None, message=None, error=str(e)
        )
        return output.model_dump()


# ============================================================================
# Tool 11: Manage Sheets
# ============================================================================


@mcp.tool()
def manage_sheets(
    file_path: str,
    operation: str,
    sheet_name: Optional[str] = None,
    new_name: Optional[str] = None,
    position: Optional[int] = None,
) -> Dict[str, Any]:
    """
    管理 Excel 工作表

    Args:
        file_path: Excel 文件路径
        operation: 操作类型（'create', 'delete', 'rename', 'copy', 'move'）
        sheet_name: 工作表名称
        new_name: 新名称（用于 rename 和 copy）
        position: 位置（用于 create、copy 和 move）

    Returns:
        包含操作结果的字典
    """
    try:
        logger.info(f"管理工作表: {operation} - {sheet_name}")

        # 创建管理器
        manager = SheetManager()

        # 根据操作类型执行相应操作
        if operation == "create":
            if not sheet_name:
                raise ValueError("create 操作需要提供 sheet_name")
            _ = manager.create_sheet(file_path, sheet_name, position)
            message = f"工作表 '{sheet_name}' 创建成功"

        elif operation == "delete":
            if not sheet_name:
                raise ValueError("delete 操作需要提供 sheet_name")
            _ = manager.delete_sheet(file_path, sheet_name)
            message = f"工作表 '{sheet_name}' 删除成功"

        elif operation == "rename":
            if not sheet_name or not new_name:
                raise ValueError("rename 操作需要提供 sheet_name 和 new_name")
            _ = manager.rename_sheet(file_path, sheet_name, new_name)
            message = f"工作表 '{sheet_name}' 重命名为 '{new_name}'"

        elif operation == "copy":
            if not sheet_name or not new_name:
                raise ValueError("copy 操作需要提供 sheet_name 和 new_name")
            _ = manager.copy_sheet(file_path, sheet_name, new_name, position)
            message = f"工作表 '{sheet_name}' 复制为 '{new_name}'"

        elif operation == "move":
            if not sheet_name or position is None:
                raise ValueError("move 操作需要提供 sheet_name 和 position")
            _ = manager.move_sheet(file_path, sheet_name, position)
            message = f"工作表 '{sheet_name}' 移动到位置 {position}"

        else:
            raise ValueError(
                f"不支持的操作类型: {operation}。"
                f"支持的操作: create, delete, rename, copy, move"
            )

        # 返回结果
        output = ManageSheetsOutput(
            success=True, operation=operation, message=message, error=None
        )

        logger.info(f"工作表管理成功: {message}")
        return output.model_dump()

    except Exception as e:
        logger.error(f"管理工作表失败: {str(e)}")
        output = ManageSheetsOutput(
            success=False, operation=operation, message=None, error=str(e)
        )
        return output.model_dump()


# ============================================================================
# Tool 12: Merge Excel Files
# ============================================================================


@mcp.tool()
def merge_excel_files(
    file_paths: List[str],
    output_path: str,
    handle_duplicates: str = "rename",
    preserve_formats: bool = True,
    sheet_names: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """
    合并多个 Excel 文件

    Args:
        file_paths: 要合并的 Excel 文件路径列表
        output_path: 输出文件路径
        handle_duplicates: 重名处理策略（'rename', 'skip', 'overwrite'）
        preserve_formats: 是否保留格式信息
        sheet_names: 要合并的工作表名称列表（None 表示全部）

    Returns:
        包含合并结果的字典
    """
    try:
        logger.info(f"开始合并 {len(file_paths)} 个 Excel 文件")

        # 创建合并器
        merger = FileMerger()

        # 执行合并
        result = merger.merge_files(
            file_paths=file_paths,
            output_path=output_path,
            handle_duplicates=handle_duplicates,  # type: ignore
            preserve_formats=preserve_formats,
            sheet_names=sheet_names,
        )

        # 返回结果
        output = MergeExcelFilesOutput(
            success=True,
            merged_sheets=result["merged_sheets"],
            skipped_sheets=result["skipped_sheets"],
            renamed_sheets=result["renamed_sheets"],
            output_path=result["output_path"],
            error=None,
        )

        logger.info(
            f"文件合并成功: {result['merged_sheets']} 个工作表已合并, "
            f"{result['skipped_sheets']} 个跳过, {result['renamed_sheets']} 个重命名"
        )
        return output.model_dump()

    except Exception as e:
        logger.error(f"合并 Excel 文件失败: {str(e)}")
        output = MergeExcelFilesOutput(
            success=False,
            merged_sheets=None,
            skipped_sheets=None,
            renamed_sheets=None,
            output_path=None,
            error=str(e),
        )
        return output.model_dump()


# ============================================================================
# Main Entry Point
# ============================================================================


def main() -> None:
    """MCP 服务器主入口函数

    支持通过环境变量配置传输方式：
    - MCP_TRANSPORT: 传输方式 (stdio|http|sse)，默认 stdio
    - MCP_HOST: HTTP/SSE 服务器地址，默认 127.0.0.1
    - MCP_PORT: HTTP/SSE 服务器端口，默认 8000
    """
    import os

    # 读取传输配置
    transport = os.getenv("MCP_TRANSPORT", "stdio").lower()
    host = os.getenv("MCP_HOST", "127.0.0.1")
    port = int(os.getenv("MCP_PORT", "8000"))

    logger.info(f"启动 Excel-Supabase MCP 服务器 (传输方式: {transport})")

    # 根据传输方式启动服务器
    if transport == "stdio":
        mcp.run()
    elif transport == "http":
        logger.info(f"HTTP 服务器地址: http://{host}:{port}/mcp/")
        mcp.run(transport="http", host=host, port=port)
    elif transport == "sse":
        logger.info(f"SSE 服务器地址: http://{host}:{port}/sse")
        mcp.run(transport="sse", host=host, port=port)
    else:
        logger.error(f"不支持的传输方式: {transport}")
        logger.error("支持的传输方式: stdio, http, sse")
        raise ValueError(f"不支持的传输方式: {transport}")


if __name__ == "__main__":
    main()
