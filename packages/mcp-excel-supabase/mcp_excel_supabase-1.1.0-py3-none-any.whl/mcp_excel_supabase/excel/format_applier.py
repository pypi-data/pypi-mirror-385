"""
Excel 格式应用器

将 schemas 定义的格式模型应用到 openpyxl 的 Cell 对象上。
这是 format_extractor 的逆向操作。
"""

from typing import Optional
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color

from .schemas import (
    AlignmentFormat,
    BorderFormat,
    BorderSide,
    CellFormat,
    FillFormat,
    FontFormat,
)
from ..utils.logger import Logger

logger = Logger("format_applier")


class FormatApplier:
    """Excel 格式应用器"""

    @staticmethod
    def _hex_to_color(hex_color: Optional[str]) -> Optional[Color]:
        """
        将十六进制颜色字符串转换为 openpyxl Color 对象

        Args:
            hex_color: 十六进制颜色字符串（如 '#FF0000' 或 'FF0000'）

        Returns:
            Color: openpyxl Color 对象，如果输入为 None 则返回 None
        """
        if hex_color is None:
            return None

        # 移除 # 前缀
        color_str = hex_color.lstrip("#")

        # 确保是 6 位十六进制
        if len(color_str) != 6:
            logger.warning(f"无效的颜色格式: {hex_color}，应为 6 位十六进制")
            return None

        # openpyxl 使用 AARRGGBB 格式，添加 alpha 通道（FF 表示完全不透明）
        return Color(rgb=f"FF{color_str}")

    @staticmethod
    def apply_font_format(cell: OpenpyxlCell, font_format: Optional[FontFormat]) -> None:
        """
        应用字体格式

        Args:
            cell: openpyxl Cell 对象
            font_format: FontFormat 对象
        """
        if font_format is None:
            return

        # 转换颜色
        color = FormatApplier._hex_to_color(font_format.color)

        # 创建 Font 对象
        cell.font = Font(
            name=font_format.name,
            size=font_format.size,
            bold=font_format.bold,
            italic=font_format.italic,
            underline=font_format.underline,  # type: ignore
            color=color,
        )

        logger.debug(f"应用字体格式到单元格 {cell.coordinate}")

    @staticmethod
    def apply_fill_format(cell: OpenpyxlCell, fill_format: Optional[FillFormat]) -> None:
        """
        应用填充格式

        Args:
            cell: openpyxl Cell 对象
            fill_format: FillFormat 对象
        """
        if fill_format is None:
            return

        # 转换背景颜色
        bg_color = FormatApplier._hex_to_color(fill_format.background_color)

        # 创建 PatternFill 对象
        # 如果有背景色但没有指定图案类型，默认使用 'solid'
        pattern_type = fill_format.pattern_type or ("solid" if bg_color else None)

        if pattern_type:
            cell.fill = PatternFill(
                patternType=pattern_type,  # type: ignore
                fgColor=bg_color,  # type: ignore
            )
            logger.debug(f"应用填充格式到单元格 {cell.coordinate}")

    @staticmethod
    def apply_border_format(cell: OpenpyxlCell, border_format: Optional[BorderFormat]) -> None:
        """
        应用边框格式

        Args:
            cell: openpyxl Cell 对象
            border_format: BorderFormat 对象
        """
        if border_format is None:
            return

        # 辅助函数：将 BorderSide 转换为 openpyxl Side
        def create_side(border_side: Optional[BorderSide]) -> Optional[Side]:
            if border_side is None:
                return None

            color = FormatApplier._hex_to_color(border_side.color)
            return Side(style=border_side.style, color=color)  # type: ignore

        # 创建四个方向的边框
        top = create_side(border_format.top)
        bottom = create_side(border_format.bottom)
        left = create_side(border_format.left)
        right = create_side(border_format.right)

        # 创建 Border 对象
        cell.border = Border(
            top=top,
            bottom=bottom,
            left=left,
            right=right,
        )

        logger.debug(f"应用边框格式到单元格 {cell.coordinate}")

    @staticmethod
    def apply_alignment_format(
        cell: OpenpyxlCell, alignment_format: Optional[AlignmentFormat]
    ) -> None:
        """
        应用对齐格式

        Args:
            cell: openpyxl Cell 对象
            alignment_format: AlignmentFormat 对象
        """
        if alignment_format is None:
            return

        # 创建 Alignment 对象
        cell.alignment = Alignment(
            horizontal=alignment_format.horizontal,
            vertical=alignment_format.vertical,
            wrap_text=alignment_format.wrap_text,
        )

        logger.debug(f"应用对齐格式到单元格 {cell.coordinate}")

    @staticmethod
    def apply_number_format(cell: OpenpyxlCell, number_format: Optional[str]) -> None:
        """
        应用数字格式

        Args:
            cell: openpyxl Cell 对象
            number_format: 数字格式字符串（如 '0.00', 'yyyy-mm-dd'）
        """
        if number_format is None:
            return

        cell.number_format = number_format
        logger.debug(f"应用数字格式 '{number_format}' 到单元格 {cell.coordinate}")

    @staticmethod
    def apply_cell_format(cell: OpenpyxlCell, cell_format: Optional[CellFormat]) -> None:
        """
        应用完整的单元格格式

        Args:
            cell: openpyxl Cell 对象
            cell_format: CellFormat 对象
        """
        if cell_format is None:
            return

        logger.debug(f"开始应用格式到单元格 {cell.coordinate}")

        # 应用各种格式
        FormatApplier.apply_font_format(cell, cell_format.font)
        FormatApplier.apply_fill_format(cell, cell_format.fill)
        FormatApplier.apply_border_format(cell, cell_format.border)
        FormatApplier.apply_alignment_format(cell, cell_format.alignment)
        FormatApplier.apply_number_format(cell, cell_format.number_format)

        logger.info(f"单元格 {cell.coordinate} 格式应用完成")


# ============================================================================
# 导出
# ============================================================================

__all__ = ["FormatApplier"]
