"""
Excel 生成器

从 JSON 格式（使用 schemas 定义的数据模型）生成 Excel 文件。
这是 parser 的逆向操作。
"""

from pathlib import Path
from typing import Any, Union

from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.utils import get_column_letter

from .schemas import Workbook, Sheet, Row, Cell
from .format_applier import FormatApplier
from .data_validator import DataValidator
from ..utils.logger import Logger
from ..utils.validator import Validator
from ..utils.errors import FileWriteError

logger = Logger("generator")


class ExcelGenerator:
    """Excel 生成器"""

    def __init__(self) -> None:
        """初始化生成器"""
        self.validator = Validator()
        self.data_validator = DataValidator()
        self.format_applier = FormatApplier()

    def generate_file(
        self, workbook: Workbook, file_path: Union[str, Path], overwrite: bool = False
    ) -> Path:
        """
        生成 Excel 文件

        Args:
            workbook: Workbook 对象
            file_path: 输出文件路径
            overwrite: 是否覆盖已存在的文件

        Returns:
            Path: 生成的文件路径

        Raises:
            FileWriteError: 文件写入失败
            ValidationError: 数据验证失败
        """
        # 验证 workbook 数据
        workbook = self.data_validator.validate_workbook(workbook)

        # 转换为 Path 对象
        path = Path(file_path)

        # 检查文件是否已存在
        if path.exists() and not overwrite:
            raise FileWriteError(
                str(path),
                "文件已存在，请设置 overwrite=True 以覆盖",
            )

        # 确保父目录存在
        path.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"开始生成 Excel 文件: {path}")

        try:
            # 创建 openpyxl Workbook
            wb = self._create_workbook(workbook)

            # 保存文件
            wb.save(path)

            logger.info(f"Excel 文件生成成功: {path}")
            return path

        except Exception as e:
            logger.error(f"生成 Excel 文件失败: {e}")
            raise FileWriteError(str(path), str(e)) from e

    def _create_workbook(self, workbook: Workbook) -> OpenpyxlWorkbook:
        """
        创建 openpyxl Workbook

        Args:
            workbook: Workbook schema 对象

        Returns:
            OpenpyxlWorkbook: openpyxl Workbook 对象
        """
        # 创建新的工作簿
        wb = OpenpyxlWorkbook()

        # 删除默认创建的工作表
        if wb.active:
            wb.remove(wb.active)

        # 创建所有工作表
        for sheet_data in workbook.sheets:
            self._create_sheet(wb, sheet_data)

        logger.debug(f"创建了 {len(workbook.sheets)} 个工作表")
        return wb

    def _create_sheet(self, wb: OpenpyxlWorkbook, sheet_data: Sheet) -> None:
        """
        创建工作表

        Args:
            wb: openpyxl Workbook 对象
            sheet_data: Sheet schema 对象
        """
        # 创建工作表
        ws = wb.create_sheet(title=sheet_data.name)

        logger.debug(f"创建工作表: {sheet_data.name}")

        # 写入所有行的数据
        self._write_rows(ws, sheet_data.rows)

        # 应用合并单元格
        self._apply_merged_cells(ws, sheet_data.merged_cells)

        # 设置列宽
        self._apply_column_widths(ws, sheet_data.column_widths)

        logger.info(f"工作表 '{sheet_data.name}' 创建完成")

    def _write_rows(self, ws: Any, rows: list[Row]) -> None:
        """
        写入所有行的数据

        Args:
            ws: openpyxl Worksheet 对象
            rows: Row 对象列表
        """
        for row_data in rows:
            # 写入行中的所有单元格
            for cell_data in row_data.cells:
                self._write_cell(ws, cell_data)

            # 设置行高
            if row_data.height is not None and row_data.cells:
                # 使用第一个单元格的行号
                row_num = row_data.cells[0].row
                ws.row_dimensions[row_num].height = row_data.height

    def _write_cell(self, ws: Any, cell_data: Cell) -> None:
        """
        写入单元格数据和格式

        Args:
            ws: openpyxl Worksheet 对象
            cell_data: Cell schema 对象
        """
        # 获取 openpyxl cell 对象
        openpyxl_cell = ws.cell(row=cell_data.row, column=cell_data.column)

        # 写入值
        if cell_data.data_type == "formula" and isinstance(cell_data.value, str):
            # 公式需要以 = 开头
            formula = cell_data.value if cell_data.value.startswith("=") else f"={cell_data.value}"
            openpyxl_cell.value = formula
        elif cell_data.data_type == "null":
            openpyxl_cell.value = None
        else:
            openpyxl_cell.value = cell_data.value

        # 应用格式
        if cell_data.format is not None:
            self.format_applier.apply_cell_format(openpyxl_cell, cell_data.format)

    def _apply_merged_cells(self, ws: Any, merged_cells: list) -> None:
        """
        应用合并单元格

        Args:
            ws: openpyxl Worksheet 对象
            merged_cells: MergedCell 对象列表
        """
        for merged_cell in merged_cells:
            # 构建合并范围字符串（如 "A1:B2"）
            start_col = get_column_letter(merged_cell.start_column)
            end_col = get_column_letter(merged_cell.end_column)
            merge_range = f"{start_col}{merged_cell.start_row}:{end_col}{merged_cell.end_row}"

            # 合并单元格
            ws.merge_cells(merge_range)
            logger.debug(f"合并单元格: {merge_range}")

    def _apply_column_widths(self, ws: Any, column_widths: dict[int, float]) -> None:
        """
        应用列宽设置

        Args:
            ws: openpyxl Worksheet 对象
            column_widths: 列宽字典，键为列号（1-based），值为宽度
        """
        for col_num, width in column_widths.items():
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width
            logger.debug(f"设置列 {col_letter} 宽度: {width}")


# ============================================================================
# 导出
# ============================================================================

__all__ = ["ExcelGenerator"]
