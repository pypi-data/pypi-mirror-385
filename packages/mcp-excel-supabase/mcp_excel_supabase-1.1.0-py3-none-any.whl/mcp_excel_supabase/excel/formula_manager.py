"""
公式管理器模块

提供 Excel 文件中公式的设置、管理和批量计算功能。
"""

from typing import Any, Dict
from openpyxl import load_workbook

from .formula_engine import FormulaEngine
from ..utils.errors import (
    FormulaError,
    ValidationError,
    FileReadError,
)
from ..utils.logger import logger
from ..utils.validator import Validator, validate_excel_file


class FormulaManager:
    """
    公式管理器类

    提供 Excel 文件中公式的设置、管理和批量计算功能。
    """

    def __init__(self) -> None:
        """初始化公式管理器"""
        self.engine = FormulaEngine()
        self.validator = Validator()
        logger.info("FormulaManager 初始化完成")

    def set_formula(
        self, file_path: str, sheet_name: str, cell: str, formula: str, save: bool = True
    ) -> Dict[str, Any]:
        """
        设置单元格公式

        Args:
            file_path: Excel 文件路径
            sheet_name: 工作表名称
            cell: 单元格位置（如 "A1"）
            formula: 公式字符串（如 "=SUM(A1:A10)"）
            save: 是否保存文件（默认 True）

        Returns:
            Dict[str, Any]: 包含设置结果的字典
                {
                    "success": True,
                    "cell": "A1",
                    "formula": "=SUM(A1:A10)",
                    "message": "公式设置成功"
                }

        Raises:
            FileReadError: 文件读取失败
            ValidationError: 参数验证失败
            FormulaError: 公式设置失败
        """
        try:
            # 验证参数
            validate_excel_file(file_path)
            self.validator.validate_non_empty(sheet_name, "sheet_name")
            self.validator.validate_non_empty(cell, "cell")
            self.validator.validate_non_empty(formula, "formula")

            # 验证公式格式
            if not self.engine.is_formula(formula):
                raise FormulaError(
                    error_code="E301",
                    message=f"无效的公式格式: {formula}",
                    context={"formula": formula},
                    suggestion="公式必须以 = 开头",
                )

            # 加载工作簿
            wb = load_workbook(file_path)

            # 检查工作表是否存在
            if sheet_name not in wb.sheetnames:
                raise ValidationError(
                    error_code="E201",
                    message=f"工作表 '{sheet_name}' 不存在",
                    context={"sheet_name": sheet_name, "available_sheets": wb.sheetnames},
                )

            ws = wb[sheet_name]

            # 设置公式
            ws[cell] = formula
            logger.info(f"设置公式: {sheet_name}!{cell} = {formula}")

            # 保存文件
            if save:
                wb.save(file_path)
                logger.info(f"文件已保存: {file_path}")

            return {"success": True, "cell": cell, "formula": formula, "message": "公式设置成功"}

        except Exception as e:
            if isinstance(e, (FileReadError, ValidationError, FormulaError)):
                raise
            raise FormulaError(
                error_code="E306",
                message=f"设置公式失败: {str(e)}",
                context={
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "cell": cell,
                    "formula": formula,
                    "error": str(e),
                },
            )

    def set_formulas(
        self, file_path: str, sheet_name: str, formulas: Dict[str, str], save: bool = True
    ) -> Dict[str, Any]:
        """
        批量设置公式

        Args:
            file_path: Excel 文件路径
            sheet_name: 工作表名称
            formulas: 公式字典，键为单元格位置，值为公式
                     如 {"A1": "=SUM(B1:B10)", "A2": "=AVERAGE(B1:B10)"}
            save: 是否保存文件（默认 True）

        Returns:
            Dict[str, Any]: 包含设置结果的字典
                {
                    "success": True,
                    "count": 2,
                    "formulas": {"A1": "=SUM(B1:B10)", "A2": "=AVERAGE(B1:B10)"},
                    "message": "批量设置公式成功"
                }
        """
        try:
            # 验证参数
            validate_excel_file(file_path)
            self.validator.validate_non_empty(sheet_name, "sheet_name")

            if not formulas:
                return {"success": True, "count": 0, "formulas": {}, "message": "没有公式需要设置"}

            # 加载工作簿
            wb = load_workbook(file_path)

            # 检查工作表是否存在
            if sheet_name not in wb.sheetnames:
                raise ValidationError(
                    error_code="E201",
                    message=f"工作表 '{sheet_name}' 不存在",
                    context={"sheet_name": sheet_name, "available_sheets": wb.sheetnames},
                )

            ws = wb[sheet_name]

            # 批量设置公式
            for cell, formula in formulas.items():
                # 验证公式格式
                if not self.engine.is_formula(formula):
                    logger.warning(f"跳过无效公式: {cell} = {formula}")
                    continue

                ws[cell] = formula
                logger.debug(f"设置公式: {sheet_name}!{cell} = {formula}")

            # 保存文件
            if save:
                wb.save(file_path)
                logger.info(f"文件已保存: {file_path}")

            return {
                "success": True,
                "count": len(formulas),
                "formulas": formulas,
                "message": f"批量设置 {len(formulas)} 个公式成功",
            }

        except Exception as e:
            if isinstance(e, (FileReadError, ValidationError, FormulaError)):
                raise
            raise FormulaError(
                error_code="E307",
                message=f"批量设置公式失败: {str(e)}",
                context={
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "formulas_count": len(formulas) if formulas else 0,
                    "error": str(e),
                },
            )

    def recalculate_all(self, file_path: str) -> Dict[str, Any]:
        """
        重新计算文件中的所有公式

        Args:
            file_path: Excel 文件路径

        Returns:
            Dict[str, Any]: 计算结果字典
                {
                    "success": True,
                    "count": 10,
                    "results": {...},
                    "message": "重新计算完成"
                }
        """
        try:
            # 验证文件
            validate_excel_file(file_path)

            logger.info(f"开始重新计算所有公式: {file_path}")

            # 使用公式引擎计算
            results = self.engine.calculate_from_file(file_path)

            logger.info(f"重新计算完成，共 {len(results)} 个结果")

            return {
                "success": True,
                "count": len(results),
                "results": results,
                "message": f"重新计算完成，共 {len(results)} 个公式",
            }

        except Exception as e:
            if isinstance(e, (FileReadError, FormulaError)):
                raise
            raise FormulaError(
                error_code="E308",
                message=f"重新计算失败: {str(e)}",
                context={"file_path": file_path, "error": str(e)},
            )

    def recalculate_sheet(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        重新计算指定工作表中的公式

        Args:
            file_path: Excel 文件路径
            sheet_name: 工作表名称

        Returns:
            Dict[str, Any]: 计算结果字典
        """
        try:
            # 验证参数
            validate_excel_file(file_path)
            self.validator.validate_non_empty(sheet_name, "sheet_name")

            logger.info(f"开始重新计算工作表: {sheet_name}")

            # 计算所有公式
            all_results = self.engine.calculate_from_file(file_path)

            # 过滤出指定工作表的结果
            sheet_results = {}
            sheet_name_upper = sheet_name.upper()
            for key, value in all_results.items():
                # 键格式: '[file.xlsx]SHEET'!A1
                if sheet_name_upper in key.upper():
                    sheet_results[key] = value

            logger.info(f"工作表 {sheet_name} 重新计算完成，共 {len(sheet_results)} 个结果")

            return {
                "success": True,
                "sheet_name": sheet_name,
                "count": len(sheet_results),
                "results": sheet_results,
                "message": f"工作表 {sheet_name} 重新计算完成，共 {len(sheet_results)} 个公式",
            }

        except Exception as e:
            if isinstance(e, (FileReadError, ValidationError, FormulaError)):
                raise
            raise FormulaError(
                error_code="E309",
                message=f"重新计算工作表失败: {str(e)}",
                context={"file_path": file_path, "sheet_name": sheet_name, "error": str(e)},
            )
