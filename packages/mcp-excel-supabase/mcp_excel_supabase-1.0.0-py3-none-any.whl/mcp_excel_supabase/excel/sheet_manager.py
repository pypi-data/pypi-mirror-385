"""
Sheet 管理器模块

提供 Excel 工作表的创建、删除、重命名、复制、移动等管理功能。
"""

from typing import Optional

from openpyxl import load_workbook

from ..utils.errors import (
    SheetNotFoundError,
    SheetAlreadyExistsError,
    ValidationError,
)
from ..utils.validator import Validator, validate_excel_file
from ..utils.logger import get_logger

logger = get_logger(__name__)


class SheetManager:
    """
    Excel 工作表管理器

    提供工作表的创建、删除、重命名、复制、移动等功能。
    所有操作都会直接修改文件并保存。
    """

    def __init__(self) -> None:
        """初始化 SheetManager"""
        self.validator = Validator()
        logger.debug("SheetManager 初始化完成")

    def create_sheet(
        self, file_path: str, sheet_name: str, position: Optional[int] = None
    ) -> dict[str, bool]:
        """
        创建新工作表

        Args:
            file_path: Excel 文件路径
            sheet_name: 新工作表名称
            position: 插入位置（0-based，None 表示末尾）

        Returns:
            包含操作结果的字典

        Raises:
            SheetAlreadyExistsError: 工作表已存在
            ValidationError: 工作表名称无效
        """
        logger.info(f"创建工作表: {sheet_name} 在文件 {file_path}")

        # 验证文件
        validate_excel_file(file_path)

        # 验证工作表名称
        self.validator.validate_sheet_name(sheet_name)

        # 加载工作簿
        wb = load_workbook(file_path)

        # 检查工作表是否已存在
        if sheet_name in wb.sheetnames:
            raise SheetAlreadyExistsError(sheet_name)

        # 创建工作表
        if position is None:
            # 在末尾创建
            wb.create_sheet(title=sheet_name)
            logger.debug(f"在末尾创建工作表: {sheet_name}")
        else:
            # 在指定位置创建
            self.validator.validate_range(
                position, "position", min_val=0, max_val=len(wb.sheetnames)
            )
            wb.create_sheet(title=sheet_name, index=position)
            logger.debug(f"在位置 {position} 创建工作表: {sheet_name}")

        # 保存文件
        wb.save(file_path)
        logger.info(f"工作表 {sheet_name} 创建成功")

        return {"success": True}

    def delete_sheet(self, file_path: str, sheet_name: str) -> dict[str, bool]:
        """
        删除工作表

        Args:
            file_path: Excel 文件路径
            sheet_name: 要删除的工作表名称

        Returns:
            包含操作结果的字典

        Raises:
            SheetNotFoundError: 工作表不存在
            ValidationError: 不能删除最后一个工作表
        """
        logger.info(f"删除工作表: {sheet_name} 从文件 {file_path}")

        # 验证文件
        validate_excel_file(file_path)

        # 加载工作簿
        wb = load_workbook(file_path)

        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(sheet_name, wb.sheetnames)

        # 检查是否是最后一个工作表
        if len(wb.sheetnames) == 1:
            raise ValidationError(
                error_code="E401",
                message="不能删除最后一个工作表",
                context={"sheet_name": sheet_name},
                suggestion="Excel 文件至少需要保留一个工作表",
            )

        # 删除工作表
        ws = wb[sheet_name]
        wb.remove(ws)
        logger.debug(f"工作表 {sheet_name} 已删除")

        # 保存文件
        wb.save(file_path)
        logger.info(f"工作表 {sheet_name} 删除成功")

        return {"success": True}

    def rename_sheet(self, file_path: str, old_name: str, new_name: str) -> dict[str, bool]:
        """
        重命名工作表

        Args:
            file_path: Excel 文件路径
            old_name: 原工作表名称
            new_name: 新工作表名称

        Returns:
            包含操作结果的字典

        Raises:
            SheetNotFoundError: 原工作表不存在
            SheetAlreadyExistsError: 新名称已被使用
            ValidationError: 新名称无效
        """
        logger.info(f"重命名工作表: {old_name} -> {new_name} 在文件 {file_path}")

        # 验证文件
        validate_excel_file(file_path)

        # 验证新工作表名称
        self.validator.validate_sheet_name(new_name)

        # 加载工作簿
        wb = load_workbook(file_path)

        # 检查原工作表是否存在
        if old_name not in wb.sheetnames:
            raise SheetNotFoundError(old_name, wb.sheetnames)

        # 检查新名称是否已存在
        if new_name in wb.sheetnames:
            raise SheetAlreadyExistsError(new_name)

        # 重命名工作表
        ws = wb[old_name]
        ws.title = new_name
        logger.debug(f"工作表已重命名: {old_name} -> {new_name}")

        # 保存文件
        wb.save(file_path)
        logger.info(f"工作表重命名成功: {old_name} -> {new_name}")

        return {"success": True}

    def copy_sheet(
        self,
        file_path: str,
        source_name: str,
        target_name: str,
        position: Optional[int] = None,
    ) -> dict[str, bool]:
        """
        复制工作表

        Args:
            file_path: Excel 文件路径
            source_name: 源工作表名称
            target_name: 目标工作表名称
            position: 插入位置（0-based，None 表示末尾）

        Returns:
            包含操作结果的字典

        Raises:
            SheetNotFoundError: 源工作表不存在
            SheetAlreadyExistsError: 目标名称已存在
            ValidationError: 目标名称无效
        """
        logger.info(f"复制工作表: {source_name} -> {target_name} 在文件 {file_path}")

        # 验证文件
        validate_excel_file(file_path)

        # 验证目标工作表名称
        self.validator.validate_sheet_name(target_name)

        # 加载工作簿
        wb = load_workbook(file_path)

        # 检查源工作表是否存在
        if source_name not in wb.sheetnames:
            raise SheetNotFoundError(source_name, wb.sheetnames)

        # 检查目标名称是否已存在
        if target_name in wb.sheetnames:
            raise SheetAlreadyExistsError(target_name)

        # 复制工作表
        source_ws = wb[source_name]
        target_ws = wb.copy_worksheet(source_ws)
        target_ws.title = target_name
        logger.debug(f"工作表已复制: {source_name} -> {target_name}")

        # 如果指定了位置，移动到指定位置
        if position is not None:
            self.validator.validate_range(
                position, "position", min_val=0, max_val=len(wb.sheetnames) - 1
            )
            # 移动工作表到指定位置
            wb._sheets.remove(target_ws)  # type: ignore
            wb._sheets.insert(position, target_ws)  # type: ignore
            logger.debug(f"工作表已移动到位置 {position}")

        # 保存文件
        wb.save(file_path)
        logger.info(f"工作表复制成功: {source_name} -> {target_name}")

        return {"success": True}

    def move_sheet(self, file_path: str, sheet_name: str, position: int) -> dict[str, bool]:
        """
        移动工作表到指定位置

        Args:
            file_path: Excel 文件路径
            sheet_name: 工作表名称
            position: 目标位置（0-based）

        Returns:
            包含操作结果的字典

        Raises:
            SheetNotFoundError: 工作表不存在
            ValidationError: 位置无效
        """
        logger.info(f"移动工作表: {sheet_name} 到位置 {position} 在文件 {file_path}")

        # 验证文件
        validate_excel_file(file_path)

        # 加载工作簿
        wb = load_workbook(file_path)

        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(sheet_name, wb.sheetnames)

        # 验证位置
        self.validator.validate_range(
            position, "position", min_val=0, max_val=len(wb.sheetnames) - 1
        )

        # 移动工作表
        ws = wb[sheet_name]
        wb._sheets.remove(ws)  # type: ignore
        wb._sheets.insert(position, ws)  # type: ignore
        logger.debug(f"工作表 {sheet_name} 已移动到位置 {position}")

        # 保存文件
        wb.save(file_path)
        logger.info(f"工作表移动成功: {sheet_name} -> 位置 {position}")

        return {"success": True}
