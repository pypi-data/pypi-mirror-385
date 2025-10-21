"""
Stream Processing Module

Provides stream processing for large Excel files:
- Chunked reading for large files
- Stream writing
- Memory usage monitoring
- Progress callbacks
"""

import psutil
from typing import Any, Callable, Dict, Iterator, List, Optional, Union, cast

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from ..utils.errors import FileOperationError
from ..utils.logger import Logger
from ..utils.validator import validate_excel_file, Validator

logger = Logger("stream_processor")


class MemoryMonitor:
    """Memory usage monitor for tracking process memory consumption"""

    def __init__(self) -> None:
        """Initialize memory monitor"""
        self.process = psutil.Process()
        self.initial_memory = self.get_memory_mb()

    def get_memory_mb(self) -> float:
        """
        Get current memory usage in MB

        Returns:
            Memory usage in MB
        """
        return float(self.process.memory_info().rss / 1024 / 1024)

    def get_memory_increase_mb(self) -> float:
        """
        Get memory increase since initialization

        Returns:
            Memory increase in MB
        """
        return self.get_memory_mb() - self.initial_memory

    def check_memory_limit(self, limit_mb: float) -> bool:
        """
        Check if memory usage exceeds limit

        Args:
            limit_mb: Memory limit in MB

        Returns:
            True if limit exceeded
        """
        current = self.get_memory_mb()
        if current > limit_mb:
            logger.warning(f"Memory limit exceeded: {current:.1f}MB > {limit_mb:.1f}MB")
            return True
        return False

    def log_memory_usage(self, description: str = "") -> None:
        """
        Log current memory usage

        Args:
            description: Description for the log entry
        """
        current = self.get_memory_mb()
        increase = self.get_memory_increase_mb()
        logger.info(
            f"Memory usage {description}: current={current:.1f}MB, " f"increase={increase:.1f}MB"
        )


class StreamReader:
    """
    Stream reader for large Excel files

    Uses openpyxl's read_only mode to read files in chunks
    """

    def __init__(
        self,
        file_path: str,
        chunk_size: int = 1000,
        memory_limit_mb: Optional[float] = None,
    ) -> None:
        """
        Initialize stream reader

        Args:
            file_path: Path to Excel file
            chunk_size: Number of rows per chunk
            memory_limit_mb: Memory limit in MB, None for no limit
        """
        validate_excel_file(file_path=file_path)
        self.file_path = file_path
        self.chunk_size = chunk_size
        self.memory_limit_mb = memory_limit_mb
        self.monitor = MemoryMonitor()

    def read_chunks(
        self,
        sheet_name: Optional[str] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Iterator[List[List[Any]]]:
        """
        Read Excel file in chunks

        Args:
            sheet_name: Sheet name to read, None for active sheet
            progress_callback: Progress callback function(current_row, total_rows)

        Yields:
            List of rows (each row is a list of cell values)

        Example:
            reader = StreamReader("large.xlsx", chunk_size=1000)
            for chunk in reader.read_chunks():
                process_chunk(chunk)
        """
        logger.info(f"Starting stream read: {self.file_path}, chunk_size={self.chunk_size}")
        self.monitor.log_memory_usage("before read")

        try:
            # Open workbook in read_only mode
            wb = load_workbook(filename=self.file_path, read_only=True, data_only=True)

            # Select worksheet
            ws_temp: Optional[Union[Worksheet, ReadOnlyWorksheet]] = None
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise FileOperationError(
                        error_code="E301",
                        message=f"Sheet not found: {sheet_name}",
                        context={"available_sheets": wb.sheetnames},
                    )
                ws_temp = wb[sheet_name]  # type: ignore
            else:
                ws_temp = wb.active  # type: ignore

            if ws_temp is None:
                raise FileOperationError(
                    error_code="E301",
                    message="No active worksheet found",
                    context={},
                )
            ws = ws_temp

            # Get total rows (approximate in read_only mode)
            total_rows = ws.max_row or 0

            # Read in chunks
            chunk: List[List[Any]] = []
            current_row = 0

            for row in ws.rows:
                # Check memory limit
                if self.memory_limit_mb and self.monitor.check_memory_limit(self.memory_limit_mb):
                    raise FileOperationError(
                        error_code="E302",
                        message="Memory limit exceeded",
                        context={
                            "limit_mb": self.memory_limit_mb,
                            "current_mb": self.monitor.get_memory_mb(),
                        },
                    )

                # Extract cell values
                row_data = [cell.value for cell in row]
                chunk.append(row_data)
                current_row += 1

                # Yield chunk when size reached
                if len(chunk) >= self.chunk_size:
                    if progress_callback:
                        progress_callback(current_row, total_rows)
                    yield chunk
                    chunk = []

            # Yield remaining rows
            if chunk:
                if progress_callback:
                    progress_callback(current_row, total_rows)
                yield chunk

            wb.close()
            self.monitor.log_memory_usage("after read")
            logger.info(f"Stream read complete: {current_row} rows")

        except FileOperationError:
            raise
        except Exception as e:
            raise FileOperationError(
                error_code="E303",
                message=f"Stream read failed: {str(e)}",
                context={"file_path": self.file_path},
            ) from e

    def read_all_as_chunks(self, sheet_name: Optional[str] = None) -> List[List[List[Any]]]:
        """
        Read all data and return as list of chunks

        Args:
            sheet_name: Sheet name to read

        Returns:
            List of chunks
        """
        chunks = []
        for chunk in self.read_chunks(sheet_name=sheet_name):
            chunks.append(chunk)
        return chunks


class StreamWriter:
    """
    Stream writer for large Excel files

    Uses openpyxl's write_only mode for efficient writing
    """

    def __init__(
        self,
        file_path: str,
        memory_limit_mb: Optional[float] = None,
    ) -> None:
        """
        Initialize stream writer

        Args:
            file_path: Output file path
            memory_limit_mb: Memory limit in MB
        """
        # Validate file path (allow non-existent files for writing)
        Validator.validate_file_path(file_path=file_path, must_exist=False)
        self.file_path = file_path
        self.memory_limit_mb = memory_limit_mb
        self.monitor = MemoryMonitor()
        self.wb: Optional[Workbook] = None
        self.ws: Optional[Worksheet] = None
        self.rows_written = 0

    def __enter__(self) -> "StreamWriter":
        """Context manager entry"""
        self.open()
        return self

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        """Context manager exit"""
        self.close()

    def open(self, sheet_name: str = "Sheet1") -> None:
        """
        Open writer

        Args:
            sheet_name: Sheet name
        """
        logger.info(f"Starting stream write: {self.file_path}")
        self.monitor.log_memory_usage("before write")

        # Create workbook in write_only mode
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet(title=sheet_name)
        self.rows_written = 0

    def write_row(self, row_data: List[Any]) -> None:
        """
        Write a single row

        Args:
            row_data: Row data
        """
        if self.ws is None:
            raise FileOperationError(
                error_code="E304",
                message="Writer not opened",
                context={},
            )

        # Check memory limit
        if self.memory_limit_mb and self.monitor.check_memory_limit(self.memory_limit_mb):
            raise FileOperationError(
                error_code="E305",
                message="Memory limit exceeded",
                context={
                    "limit_mb": self.memory_limit_mb,
                    "current_mb": self.monitor.get_memory_mb(),
                },
            )

        self.ws.append(row_data)
        self.rows_written += 1

    def write_rows(
        self,
        rows: List[List[Any]],
        progress_callback: Optional[Callable[[int], None]] = None,
    ) -> None:
        """
        Write multiple rows

        Args:
            rows: List of rows
            progress_callback: Progress callback function(rows_written)
        """
        for row in rows:
            self.write_row(row)
            if progress_callback:
                progress_callback(self.rows_written)

    def close(self) -> None:
        """Close writer and save file"""
        if self.wb is not None:
            self.wb.save(self.file_path)
            self.wb.close()
            self.monitor.log_memory_usage("after write")
            logger.info(f"Stream write complete: {self.rows_written} rows")
            self.wb = None
            self.ws = None


def stream_copy(
    source_path: str,
    dest_path: str,
    chunk_size: int = 1000,
    sheet_name: Optional[str] = None,
    progress_callback: Optional[Callable[[str, int, int], None]] = None,
) -> Dict[str, Any]:
    """
    Stream copy Excel file

    Args:
        source_path: Source file path
        dest_path: Destination file path
        chunk_size: Chunk size
        sheet_name: Sheet name
        progress_callback: Progress callback(stage, current, total)

    Returns:
        Statistics dictionary

    Example:
        stats = stream_copy("large.xlsx", "copy.xlsx", chunk_size=1000)
    """
    logger.info(f"Stream copy: {source_path} -> {dest_path}")
    monitor = MemoryMonitor()

    reader = StreamReader(source_path, chunk_size=chunk_size)
    total_rows = 0

    with StreamWriter(dest_path) as writer:
        for chunk in reader.read_chunks(sheet_name=sheet_name):
            writer.write_rows(chunk)
            total_rows += len(chunk)

            if progress_callback:
                progress_callback("copying", total_rows, 0)

    stats = {
        "source_path": source_path,
        "dest_path": dest_path,
        "total_rows": total_rows,
        "chunk_size": chunk_size,
        "memory_increase_mb": monitor.get_memory_increase_mb(),
    }

    logger.info(f"Stream copy complete: {stats}")
    return stats
