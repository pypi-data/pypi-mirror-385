"""
Excel 解析器

读取 Excel 文件并转换为 JSON 格式（使用 schemas 定义的数据模型）。
"""

from pathlib import Path
from typing import Any, Dict, List, Union, Optional, cast
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell as OpenpyxlCell

from .schemas import Cell, MergedCell, Row, Sheet, Workbook
from .format_extractor import FormatExtractor
from ..utils.logger import get_logger
from ..utils.validator import Validator
from ..utils.errors import FileNotFoundError as MCPFileNotFoundError, FileReadError

logger = get_logger(__name__)


class ExcelParser:
    """Excel 解析器"""

    def __init__(self) -> None:
        """初始化解析器"""
        self.validator = Validator()
        self.format_extractor = FormatExtractor()

    def parse_file(self, file_path: Union[str, Path]) -> Workbook:
        """
        解析 Excel 文件

        Args:
            file_path: Excel 文件路径

        Returns:
            Workbook 对象

        Raises:
            FileNotFoundError: 文件不存在
            FileReadError: 文件读取失败
        """
        # 验证文件路径
        try:
            path = self.validator.validate_file_path(
                file_path, must_exist=True, extensions=[".xlsx", ".xls"]
            )
        except FileNotFoundError as e:
            raise MCPFileNotFoundError(str(path)) from e

        logger.info(f"开始解析 Excel 文件: {path}")

        try:
            # 加载工作簿（data_only=True 读取公式的计算结果）
            wb = load_workbook(filename=path, data_only=True)
        except Exception as e:
            logger.error(f"加载 Excel 文件失败: {e}")
            raise FileReadError(str(path), str(e)) from e

        # 解析工作簿
        workbook = self._parse_workbook(wb, path)

        logger.info(f"Excel 文件解析完成: {path}, 共 {len(workbook.sheets)} 个工作表")
        return workbook

    def _parse_workbook(self, wb: Any, file_path: Path) -> Workbook:
        """
        解析工作簿

        Args:
            wb: openpyxl Workbook 对象
            file_path: 文件路径

        Returns:
            Workbook 对象
        """
        sheets: List[Sheet] = []

        # 遍历所有工作表
        for ws in wb.worksheets:
            sheet = self._parse_sheet(ws)
            sheets.append(sheet)

        # 创建元数据
        metadata: Dict[str, Any] = {
            "filename": file_path.name,
            "parsed_at": datetime.now().isoformat(),
            "sheet_count": len(sheets),
        }

        return Workbook(sheets=sheets, metadata=metadata)

    def _parse_sheet(self, ws: Worksheet) -> Sheet:
        """
        解析工作表

        Args:
            ws: openpyxl Worksheet 对象

        Returns:
            Sheet 对象
        """
        logger.debug(f"解析工作表: {ws.title}")

        # 解析行数据
        rows = self._parse_rows(ws)

        # 解析合并单元格
        merged_cells = self._parse_merged_cells(ws)

        # 解析列宽
        column_widths = self._parse_column_widths(ws)

        return Sheet(
            name=ws.title,
            rows=rows,
            merged_cells=merged_cells,
            column_widths=column_widths,
        )

    def _parse_rows(self, ws: Worksheet) -> List[Row]:
        """
        解析所有行

        Args:
            ws: openpyxl Worksheet 对象

        Returns:
            Row 对象列表
        """
        rows: List[Row] = []

        # 使用 iter_rows 提高性能
        for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
            cells: List[Cell] = []

            for col_idx, openpyxl_cell in enumerate(row_cells, start=1):
                cell = self._parse_cell(cast(OpenpyxlCell, openpyxl_cell), row_idx, col_idx)
                cells.append(cell)

            # 获取行高
            row_height = None
            if row_idx in ws.row_dimensions:
                row_height = ws.row_dimensions[row_idx].height

            rows.append(Row(cells=cells, height=row_height))

        return rows

    def _parse_cell(self, openpyxl_cell: OpenpyxlCell, row: int, column: int) -> Cell:
        """
        解析单元格

        Args:
            openpyxl_cell: openpyxl Cell 对象
            row: 行号（1-based）
            column: 列号（1-based）

        Returns:
            Cell 对象
        """
        # 获取单元格值
        value = openpyxl_cell.value

        # 识别数据类型
        data_type = self._identify_data_type(openpyxl_cell, value)

        # 将值转换为允许的类型
        safe_value: Optional[Union[str, int, float, bool]]
        if value is None or isinstance(value, (str, int, float, bool)):
            safe_value = value
        else:
            safe_value = str(value)

        # 提取格式
        cell_format = self.format_extractor.extract_cell_format(openpyxl_cell)

        return Cell(
            value=safe_value,
            data_type=data_type,
            format=cell_format,
            row=row,
            column=column,
        )

    def _identify_data_type(self, openpyxl_cell: OpenpyxlCell, value: Any) -> str:
        """
        识别单元格数据类型

        Args:
            openpyxl_cell: openpyxl Cell 对象
            value: 单元格值

        Returns:
            数据类型字符串：string, number, boolean, formula, date, null
        """
        if value is None:
            return "null"

        # openpyxl 的 data_type 属性
        # 's' = string, 'n' = number, 'b' = boolean, 'f' = formula, 'd' = date
        cell_type = openpyxl_cell.data_type

        if cell_type == "f":
            return "formula"
        elif cell_type == "b":
            return "boolean"
        elif cell_type == "n":
            # 检查是否是日期
            if openpyxl_cell.is_date:
                return "date"
            return "number"
        elif cell_type == "s":
            return "string"
        else:
            # 根据值的类型判断
            if isinstance(value, bool):
                return "boolean"
            elif isinstance(value, (int, float)):
                return "number"
            elif isinstance(value, str):
                return "string"
            else:
                return "null"

    def _parse_merged_cells(self, ws: Worksheet) -> List[MergedCell]:
        """
        解析合并单元格

        Args:
            ws: openpyxl Worksheet 对象

        Returns:
            MergedCell 对象列表
        """
        merged_cells: List[MergedCell] = []

        for merged_range in ws.merged_cells.ranges:
            # merged_range 格式: 'A1:B2'
            # 获取边界
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            max_row = merged_range.max_row
            max_col = merged_range.max_col

            merged_cells.append(
                MergedCell(
                    start_row=min_row,
                    start_column=min_col,
                    end_row=max_row,
                    end_column=max_col,
                )
            )

        return merged_cells

    def _parse_column_widths(self, ws: Worksheet) -> Dict[int, float]:
        """
        解析列宽

        Args:
            ws: openpyxl Worksheet 对象

        Returns:
            列宽字典，键为列号（1-based），值为宽度
        """
        column_widths: Dict[int, float] = {}

        for col_letter, dimension in ws.column_dimensions.items():
            if dimension.width is not None:
                # 将列字母转换为列号
                col_idx = self._column_letter_to_index(col_letter)
                column_widths[col_idx] = dimension.width

        return column_widths

    @staticmethod
    def _column_letter_to_index(column_letter: str) -> int:
        """
        将列字母转换为列号（1-based）

        Args:
            column_letter: 列字母（如 'A', 'B', 'AA'）

        Returns:
            列号（1-based）
        """
        from openpyxl.utils import column_index_from_string

        return int(column_index_from_string(column_letter))


# ============================================================================
# 导出
# ============================================================================

__all__ = ["ExcelParser"]
