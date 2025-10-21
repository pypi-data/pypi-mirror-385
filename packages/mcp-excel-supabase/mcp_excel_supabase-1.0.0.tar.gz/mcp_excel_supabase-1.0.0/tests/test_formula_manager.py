"""
测试公式管理器模块
"""

import pytest
import tempfile
import os
from openpyxl import Workbook, load_workbook

from mcp_excel_supabase.excel.formula_manager import FormulaManager
from mcp_excel_supabase.utils.errors import FormulaError, ValidationError


class TestFormulaManager:
    """测试 FormulaManager 类"""

    @pytest.fixture
    def manager(self):
        """创建公式管理器实例"""
        return FormulaManager()

    @pytest.fixture
    def simple_excel(self):
        """创建简单的 Excel 文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # 设置数据
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = 30
        ws["B1"] = 5
        ws["B2"] = 15
        ws["B3"] = 25

        # 保存到临时文件
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        wb.save(temp_file.name)

        yield temp_file.name

        # 清理
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)

    def test_set_formula_success(self, manager, simple_excel):
        """测试设置公式成功"""
        result = manager.set_formula(
            file_path=simple_excel, sheet_name="Data", cell="C1", formula="=SUM(A1:A3)"
        )

        # 验证返回结果
        assert result["success"] is True
        assert result["cell"] == "C1"
        assert result["formula"] == "=SUM(A1:A3)"

        # 验证文件中的公式
        wb = load_workbook(simple_excel)
        ws = wb["Data"]
        assert ws["C1"].value == "=SUM(A1:A3)"

    def test_set_formula_invalid_formula(self, manager, simple_excel):
        """测试设置无效公式"""
        with pytest.raises(FormulaError) as exc_info:
            manager.set_formula(
                file_path=simple_excel, sheet_name="Data", cell="C1", formula="SUM(A1:A3)"  # 缺少 =
            )
        assert exc_info.value.error_code == "E301"

    def test_set_formula_invalid_sheet(self, manager, simple_excel):
        """测试设置公式到不存在的工作表"""
        with pytest.raises(ValidationError) as exc_info:
            manager.set_formula(
                file_path=simple_excel, sheet_name="NonExistent", cell="C1", formula="=SUM(A1:A3)"
            )
        assert exc_info.value.error_code == "E201"

    def test_set_formulas_success(self, manager, simple_excel):
        """测试批量设置公式成功"""
        formulas = {"C1": "=SUM(A1:A3)", "C2": "=AVERAGE(A1:A3)", "C3": "=MAX(A1:A3)"}

        result = manager.set_formulas(file_path=simple_excel, sheet_name="Data", formulas=formulas)

        # 验证返回结果
        assert result["success"] is True
        assert result["count"] == 3

        # 验证文件中的公式
        wb = load_workbook(simple_excel)
        ws = wb["Data"]
        assert ws["C1"].value == "=SUM(A1:A3)"
        assert ws["C2"].value == "=AVERAGE(A1:A3)"
        assert ws["C3"].value == "=MAX(A1:A3)"

    def test_set_formulas_empty(self, manager, simple_excel):
        """测试批量设置空公式列表"""
        result = manager.set_formulas(file_path=simple_excel, sheet_name="Data", formulas={})

        assert result["success"] is True
        assert result["count"] == 0

    def test_recalculate_all(self, manager, simple_excel):
        """测试重新计算所有公式"""
        # 先设置一些公式
        manager.set_formula(simple_excel, "Data", "C1", "=SUM(A1:A3)")
        manager.set_formula(simple_excel, "Data", "C2", "=AVERAGE(A1:A3)")

        # 重新计算
        result = manager.recalculate_all(simple_excel)

        # 验证结果
        assert result["success"] is True
        assert result["count"] > 0
        assert "results" in result

        # 验证计算结果
        results = result["results"]
        sum_found = False
        avg_found = False

        for key, value in results.items():
            if "C1" in key:
                assert value == 60.0  # 10 + 20 + 30
                sum_found = True
            elif "C2" in key:
                assert value == 20.0  # (10 + 20 + 30) / 3
                avg_found = True

        assert sum_found, "未找到 SUM 公式的计算结果"
        assert avg_found, "未找到 AVERAGE 公式的计算结果"

    def test_recalculate_sheet(self, manager, simple_excel):
        """测试重新计算指定工作表"""
        # 先设置公式
        manager.set_formula(simple_excel, "Data", "C1", "=SUM(A1:A3)")

        # 重新计算工作表
        result = manager.recalculate_sheet(simple_excel, "Data")

        # 验证结果
        assert result["success"] is True
        assert result["sheet_name"] == "Data"
        assert result["count"] > 0
        assert "results" in result

    def test_recalculate_sheet_invalid(self, manager, simple_excel):
        """测试重新计算不存在的工作表"""
        # 注意：recalculate_sheet 会计算所有公式然后过滤
        # 如果工作表不存在，可能返回空结果而不是错误
        # 这取决于实现细节
        result = manager.recalculate_sheet(simple_excel, "NonExistent")

        # 应该返回成功但结果为空
        assert result["success"] is True
        assert result["count"] == 0

    def test_set_formula_no_save(self, manager, simple_excel):
        """测试设置公式但不保存"""
        result = manager.set_formula(
            file_path=simple_excel, sheet_name="Data", cell="C1", formula="=SUM(A1:A3)", save=False
        )

        # 验证返回结果
        assert result["success"] is True

        # 重新加载文件，公式不应该存在（因为没保存）
        wb = load_workbook(simple_excel)
        ws = wb["Data"]
        # C1 应该是空的或原始值
        assert ws["C1"].value != "=SUM(A1:A3)"

    def test_set_formulas_no_save(self, manager, simple_excel):
        """测试批量设置公式但不保存"""
        formulas = {"C1": "=SUM(A1:A3)", "C2": "=AVERAGE(A1:A3)"}

        result = manager.set_formulas(
            file_path=simple_excel, sheet_name="Data", formulas=formulas, save=False
        )

        # 验证返回结果
        assert result["success"] is True
        assert result["count"] == 2

        # 重新加载文件，公式不应该存在
        wb = load_workbook(simple_excel)
        ws = wb["Data"]
        assert ws["C1"].value != "=SUM(A1:A3)"
