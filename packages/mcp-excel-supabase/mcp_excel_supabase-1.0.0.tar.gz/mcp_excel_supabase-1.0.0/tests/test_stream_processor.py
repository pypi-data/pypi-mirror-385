"""
流式处理模块测试
"""

import os
import tempfile

import pytest
from openpyxl import Workbook

from mcp_excel_supabase.excel.stream_processor import (
    MemoryMonitor,
    StreamReader,
    StreamWriter,
    stream_copy,
)
from mcp_excel_supabase.utils.errors import FileOperationError


@pytest.fixture
def temp_dir():
    """创建临时目录"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir


@pytest.fixture
def sample_excel_file(temp_dir):
    """创建示例Excel文件"""
    file_path = os.path.join(temp_dir, "sample.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入100行数据
    for i in range(100):
        ws.append([f"A{i}", f"B{i}", i, i * 2, i * 3])

    wb.save(file_path)
    return file_path


@pytest.fixture
def large_excel_file(temp_dir):
    """创建大型Excel文件"""
    file_path = os.path.join(temp_dir, "large.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # 写入1000行数据
    for i in range(1000):
        ws.append([f"Col{j}_{i}" for j in range(10)])

    wb.save(file_path)
    return file_path


class TestMemoryMonitor:
    """测试MemoryMonitor类"""

    def test_get_memory_mb(self):
        """测试获取内存使用量"""
        monitor = MemoryMonitor()
        memory = monitor.get_memory_mb()
        assert memory > 0
        assert isinstance(memory, float)

    def test_get_memory_increase(self):
        """测试获取内存增量"""
        monitor = MemoryMonitor()

        # 分配一些内存
        data = [i for i in range(100000)]

        increase = monitor.get_memory_increase_mb()
        # 增量应该是float类型
        assert isinstance(increase, float)

        # 清理
        del data

    def test_check_memory_limit(self):
        """测试内存限制检查"""
        monitor = MemoryMonitor()
        current = monitor.get_memory_mb()

        # 设置一个很高的限制，应该不会超限
        assert not monitor.check_memory_limit(current + 1000)

        # 设置一个很低的限制，应该超限
        assert monitor.check_memory_limit(1)

    def test_log_memory_usage(self):
        """测试记录内存使用"""
        monitor = MemoryMonitor()
        # 应该不抛出异常
        monitor.log_memory_usage("test")


class TestStreamReader:
    """测试StreamReader类"""

    def test_init_with_valid_file(self, sample_excel_file):
        """测试使用有效文件初始化"""
        reader = StreamReader(sample_excel_file, chunk_size=10)
        assert reader.file_path == sample_excel_file
        assert reader.chunk_size == 10

    def test_init_with_invalid_file(self):
        """测试使用无效文件初始化"""
        with pytest.raises(FileOperationError):
            StreamReader("nonexistent.xlsx")

    def test_read_chunks_basic(self, sample_excel_file):
        """测试基本分块读取"""
        reader = StreamReader(sample_excel_file, chunk_size=10)
        chunks = list(reader.read_chunks())

        # 100行数据，每批10行，应该有10批
        assert len(chunks) == 10
        assert all(len(chunk) == 10 for chunk in chunks)

    def test_read_chunks_with_remainder(self, sample_excel_file):
        """测试有余数的分块读取"""
        reader = StreamReader(sample_excel_file, chunk_size=15)
        chunks = list(reader.read_chunks())

        # 100行数据，每批15行，应该有7批
        assert len(chunks) == 7
        assert len(chunks[-1]) == 10  # 最后一批只有10行

    def test_read_chunks_with_progress_callback(self, sample_excel_file):
        """测试带进度回调的分块读取"""
        progress_calls = []

        def progress_callback(current, total):
            progress_calls.append((current, total))

        reader = StreamReader(sample_excel_file, chunk_size=20)
        list(reader.read_chunks(progress_callback=progress_callback))

        # 应该有5次进度回调
        assert len(progress_calls) == 5
        assert progress_calls[-1][0] == 100  # 最后一次应该是100行

    def test_read_chunks_specific_sheet(self, temp_dir):
        """测试读取指定工作表"""
        file_path = os.path.join(temp_dir, "multi_sheet.xlsx")
        wb = Workbook()

        # 创建两个工作表
        ws1 = wb.active
        ws1.title = "Sheet1"
        for i in range(10):
            ws1.append([f"S1_{i}"])

        ws2 = wb.create_sheet("Sheet2")
        for i in range(20):
            ws2.append([f"S2_{i}"])

        wb.save(file_path)

        # 读取Sheet2
        reader = StreamReader(file_path, chunk_size=5)
        chunks = list(reader.read_chunks(sheet_name="Sheet2"))

        assert len(chunks) == 4  # 20行，每批5行
        assert chunks[0][0][0] == "S2_0"

    def test_read_chunks_nonexistent_sheet(self, sample_excel_file):
        """测试读取不存在的工作表"""
        reader = StreamReader(sample_excel_file, chunk_size=10)
        with pytest.raises(FileOperationError) as exc_info:
            list(reader.read_chunks(sheet_name="NonExistent"))
        assert exc_info.value.error_code == "E301"

    def test_read_all_as_chunks(self, sample_excel_file):
        """测试读取所有数据为分块列表"""
        reader = StreamReader(sample_excel_file, chunk_size=25)
        chunks = reader.read_all_as_chunks()

        assert len(chunks) == 4  # 100行，每批25行
        assert isinstance(chunks, list)
        assert all(isinstance(chunk, list) for chunk in chunks)

    def test_memory_limit(self, sample_excel_file):
        """测试内存限制"""
        # 设置一个极低的内存限制，应该触发异常
        reader = StreamReader(sample_excel_file, chunk_size=10, memory_limit_mb=1)
        with pytest.raises(FileOperationError) as exc_info:
            list(reader.read_chunks())
        assert exc_info.value.error_code == "E302"


class TestStreamWriter:
    """测试StreamWriter类"""

    def test_context_manager(self, temp_dir):
        """测试上下文管理器"""
        file_path = os.path.join(temp_dir, "output.xlsx")

        with StreamWriter(file_path) as writer:
            writer.write_row(["A", "B", "C"])
            writer.write_row([1, 2, 3])

        # 验证文件已创建
        assert os.path.exists(file_path)

    def test_write_row(self, temp_dir):
        """测试写入单行"""
        file_path = os.path.join(temp_dir, "output.xlsx")

        writer = StreamWriter(file_path)
        writer.open()
        writer.write_row(["Header1", "Header2", "Header3"])
        writer.write_row([1, 2, 3])
        writer.write_row([4, 5, 6])
        writer.close()

        assert writer.rows_written == 3
        assert os.path.exists(file_path)

    def test_write_rows(self, temp_dir):
        """测试批量写入"""
        file_path = os.path.join(temp_dir, "output.xlsx")

        rows = [
            ["A", "B", "C"],
            [1, 2, 3],
            [4, 5, 6],
            [7, 8, 9],
        ]

        with StreamWriter(file_path) as writer:
            writer.write_rows(rows)

        assert writer.rows_written == 4

    def test_write_rows_with_progress(self, temp_dir):
        """测试带进度回调的批量写入"""
        file_path = os.path.join(temp_dir, "output.xlsx")
        progress_calls = []

        def progress_callback(rows_written):
            progress_calls.append(rows_written)

        rows = [[i] for i in range(50)]

        with StreamWriter(file_path) as writer:
            writer.write_rows(rows, progress_callback=progress_callback)

        assert len(progress_calls) == 50
        assert progress_calls[-1] == 50

    def test_write_without_open(self, temp_dir):
        """测试未打开就写入"""
        file_path = os.path.join(temp_dir, "output.xlsx")
        writer = StreamWriter(file_path)

        with pytest.raises(FileOperationError) as exc_info:
            writer.write_row([1, 2, 3])
        assert exc_info.value.error_code == "E304"


class TestStreamCopy:
    """测试stream_copy函数"""

    def test_basic_copy(self, sample_excel_file, temp_dir):
        """测试基本复制"""
        dest_path = os.path.join(temp_dir, "copy.xlsx")
        stats = stream_copy(sample_excel_file, dest_path, chunk_size=20)

        assert os.path.exists(dest_path)
        assert stats["total_rows"] == 100
        assert stats["chunk_size"] == 20
        assert "memory_increase_mb" in stats

    def test_copy_with_progress(self, sample_excel_file, temp_dir):
        """测试带进度回调的复制"""
        dest_path = os.path.join(temp_dir, "copy.xlsx")
        progress_calls = []

        def progress_callback(stage, current, total):
            progress_calls.append((stage, current, total))

        stream_copy(
            sample_excel_file,
            dest_path,
            chunk_size=25,
            progress_callback=progress_callback,
        )

        assert len(progress_calls) == 4  # 100行，每批25行
        assert all(stage == "copying" for stage, _, _ in progress_calls)

    def test_copy_large_file(self, large_excel_file, temp_dir):
        """测试复制大文件"""
        dest_path = os.path.join(temp_dir, "large_copy.xlsx")
        stats = stream_copy(large_excel_file, dest_path, chunk_size=100)

        assert os.path.exists(dest_path)
        assert stats["total_rows"] == 1000


class TestIntegration:
    """集成测试"""

    def test_read_write_cycle(self, sample_excel_file, temp_dir):
        """测试读写循环"""
        output_path = os.path.join(temp_dir, "output.xlsx")

        # 读取并写入
        reader = StreamReader(sample_excel_file, chunk_size=20)
        with StreamWriter(output_path) as writer:
            for chunk in reader.read_chunks():
                writer.write_rows(chunk)

        # 验证输出文件
        assert os.path.exists(output_path)

        # 读取输出文件验证数据
        reader2 = StreamReader(output_path, chunk_size=100)
        chunks = list(reader2.read_chunks())
        assert len(chunks) == 1
        assert len(chunks[0]) == 100
