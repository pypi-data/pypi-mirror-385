"""
测试格式应用器模块
"""

import pytest
from openpyxl import Workbook
from openpyxl.styles import Color

from src.mcp_excel_supabase.excel.format_applier import FormatApplier
from src.mcp_excel_supabase.excel.schemas import (
    FontFormat,
    FillFormat,
    BorderFormat,
    BorderSide,
    AlignmentFormat,
    CellFormat,
)


class TestFormatApplier:
    """测试 FormatApplier 类"""

    @pytest.fixture
    def workbook(self):
        """创建测试用的工作簿"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        return wb, ws

    def test_hex_to_color_valid(self):
        """测试十六进制颜色转换（有效输入）"""
        # 带 # 前缀
        color = FormatApplier._hex_to_color("#FF0000")
        assert isinstance(color, Color)
        assert color.rgb == "FFFF0000"  # AARRGGBB 格式

        # 不带 # 前缀
        color = FormatApplier._hex_to_color("00FF00")
        assert isinstance(color, Color)
        assert color.rgb == "FF00FF00"

    def test_hex_to_color_none(self):
        """测试十六进制颜色转换（None 输入）"""
        color = FormatApplier._hex_to_color(None)
        assert color is None

    def test_hex_to_color_invalid(self):
        """测试十六进制颜色转换（无效输入）"""
        color = FormatApplier._hex_to_color("#FFF")  # 长度不是 6
        assert color is None

    def test_apply_font_format(self, workbook):
        """测试应用字体格式"""
        _, ws = workbook
        cell = ws["A1"]

        font_format = FontFormat(
            name="Arial",
            size=12,
            bold=True,
            italic=True,
            underline="single",
            color="#FF0000",
        )

        FormatApplier.apply_font_format(cell, font_format)

        assert cell.font.name == "Arial"
        assert cell.font.size == 12
        assert cell.font.bold is True
        assert cell.font.italic is True
        assert cell.font.underline == "single"
        assert cell.font.color.rgb == "FFFF0000"

    def test_apply_font_format_none(self, workbook):
        """测试应用 None 字体格式"""
        _, ws = workbook
        cell = ws["A1"]

        # 应该不抛出异常
        FormatApplier.apply_font_format(cell, None)

    def test_apply_fill_format(self, workbook):
        """测试应用填充格式"""
        _, ws = workbook
        cell = ws["A1"]

        fill_format = FillFormat(
            background_color="#FFFF00",
            pattern_type="solid",
        )

        FormatApplier.apply_fill_format(cell, fill_format)

        assert cell.fill.patternType == "solid"
        assert cell.fill.fgColor.rgb == "FFFFFF00"

    def test_apply_fill_format_default_pattern(self, workbook):
        """测试应用填充格式（默认图案类型）"""
        _, ws = workbook
        cell = ws["A1"]

        # 只指定背景色，不指定图案类型
        fill_format = FillFormat(background_color="#00FF00")

        FormatApplier.apply_fill_format(cell, fill_format)

        # 应该默认使用 solid
        assert cell.fill.patternType == "solid"
        assert cell.fill.fgColor.rgb == "FF00FF00"

    def test_apply_fill_format_none(self, workbook):
        """测试应用 None 填充格式"""
        _, ws = workbook
        cell = ws["A1"]

        FormatApplier.apply_fill_format(cell, None)

    def test_apply_border_format(self, workbook):
        """测试应用边框格式"""
        _, ws = workbook
        cell = ws["A1"]

        border_format = BorderFormat(
            top=BorderSide(style="thin", color="#000000"),
            bottom=BorderSide(style="medium", color="#FF0000"),
            left=BorderSide(style="thick", color="#00FF00"),
            right=BorderSide(style="thin", color="#0000FF"),
        )

        FormatApplier.apply_border_format(cell, border_format)

        assert cell.border.top.style == "thin"
        assert cell.border.top.color.rgb == "FF000000"
        assert cell.border.bottom.style == "medium"
        assert cell.border.bottom.color.rgb == "FFFF0000"
        assert cell.border.left.style == "thick"
        assert cell.border.left.color.rgb == "FF00FF00"
        assert cell.border.right.style == "thin"
        assert cell.border.right.color.rgb == "FF0000FF"

    def test_apply_border_format_partial(self, workbook):
        """测试应用部分边框格式"""
        _, ws = workbook
        cell = ws["A1"]

        # 只设置上边框和下边框
        border_format = BorderFormat(
            top=BorderSide(style="thin"),
            bottom=BorderSide(style="thin"),
        )

        FormatApplier.apply_border_format(cell, border_format)

        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style == "thin"

    def test_apply_border_format_none(self, workbook):
        """测试应用 None 边框格式"""
        _, ws = workbook
        cell = ws["A1"]

        FormatApplier.apply_border_format(cell, None)

    def test_apply_alignment_format(self, workbook):
        """测试应用对齐格式"""
        _, ws = workbook
        cell = ws["A1"]

        alignment_format = AlignmentFormat(
            horizontal="center",
            vertical="top",
            wrap_text=True,
        )

        FormatApplier.apply_alignment_format(cell, alignment_format)

        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "top"
        assert cell.alignment.wrap_text is True

    def test_apply_alignment_format_none(self, workbook):
        """测试应用 None 对齐格式"""
        _, ws = workbook
        cell = ws["A1"]

        FormatApplier.apply_alignment_format(cell, None)

    def test_apply_number_format(self, workbook):
        """测试应用数字格式"""
        _, ws = workbook
        cell = ws["A1"]

        FormatApplier.apply_number_format(cell, "0.00")

        assert cell.number_format == "0.00"

    def test_apply_number_format_date(self, workbook):
        """测试应用日期格式"""
        _, ws = workbook
        cell = ws["A1"]

        FormatApplier.apply_number_format(cell, "yyyy-mm-dd")

        assert cell.number_format == "yyyy-mm-dd"

    def test_apply_number_format_none(self, workbook):
        """测试应用 None 数字格式"""
        _, ws = workbook
        cell = ws["A1"]

        FormatApplier.apply_number_format(cell, None)

    def test_apply_cell_format_complete(self, workbook):
        """测试应用完整的单元格格式"""
        _, ws = workbook
        cell = ws["A1"]

        cell_format = CellFormat(
            font=FontFormat(name="Arial", size=12, bold=True, color="#FF0000"),
            fill=FillFormat(background_color="#FFFF00", pattern_type="solid"),
            border=BorderFormat(
                top=BorderSide(style="thin", color="#000000"),
                bottom=BorderSide(style="thin", color="#000000"),
            ),
            alignment=AlignmentFormat(horizontal="center", vertical="center"),
            number_format="0.00",
        )

        FormatApplier.apply_cell_format(cell, cell_format)

        # 验证所有格式都已应用
        assert cell.font.name == "Arial"
        assert cell.font.size == 12
        assert cell.font.bold is True
        assert cell.fill.patternType == "solid"
        assert cell.border.top.style == "thin"
        assert cell.alignment.horizontal == "center"
        assert cell.number_format == "0.00"

    def test_apply_cell_format_none(self, workbook):
        """测试应用 None 单元格格式"""
        _, ws = workbook
        cell = ws["A1"]

        FormatApplier.apply_cell_format(cell, None)
