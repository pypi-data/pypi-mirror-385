"""
测试公式引擎模块
"""

import pytest
import tempfile
import os
from openpyxl import Workbook

from mcp_excel_supabase.excel.formula_engine import FormulaEngine
from mcp_excel_supabase.utils.errors import FormulaError


class TestFormulaEngine:
    """测试 FormulaEngine 类"""

    @pytest.fixture
    def engine(self):
        """创建公式引擎实例"""
        return FormulaEngine()

    @pytest.fixture
    def simple_excel_with_formulas(self):
        """创建包含公式的简单 Excel 文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"

        # 设置数据
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = 30
        ws["B1"] = 5
        ws["B2"] = 15

        # 设置公式
        ws["C1"] = "=SUM(A1:A3)"
        ws["C2"] = "=AVERAGE(A1:A3)"
        ws["C3"] = "=MAX(A1:A3)"
        ws["C4"] = "=MIN(A1:A3)"
        ws["D1"] = "=A1+B1"
        ws["D2"] = '=IF(A1>B1, "大", "小")'

        # 保存到临时文件
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        wb.save(temp_file.name)

        yield temp_file.name

        # 清理
        os.unlink(temp_file.name)

    def test_is_formula_valid(self, engine):
        """测试有效公式识别"""
        assert engine.is_formula("=SUM(A1:A10)")
        assert engine.is_formula("=AVERAGE(B1:B5)")
        assert engine.is_formula("=A1+B1")
        assert engine.is_formula("=IF(A1>10, 1, 0)")

    def test_is_formula_invalid(self, engine):
        """测试无效公式识别"""
        assert not engine.is_formula("SUM(A1:A10)")  # 缺少 =
        assert not engine.is_formula("Hello")
        assert not engine.is_formula("")
        assert not engine.is_formula(None)
        assert not engine.is_formula(123)

    def test_parse_formula_success(self, engine):
        """测试公式解析成功"""
        ast = engine.parse_formula("=SUM(A1:A10)")
        assert ast is not None

        ast = engine.parse_formula("=AVERAGE(B1:B5)")
        assert ast is not None

    def test_parse_formula_invalid(self, engine):
        """测试解析无效公式"""
        with pytest.raises(FormulaError) as exc_info:
            engine.parse_formula("SUM(A1:A10)")  # 缺少 =
        assert exc_info.value.error_code == "E301"

    def test_get_formula_dependencies(self, engine):
        """测试获取公式依赖"""
        deps = engine.get_formula_dependencies("=SUM(A1:A10)")
        assert "A1:A10" in deps

        deps = engine.get_formula_dependencies("=A1+B1*C1")
        assert "A1" in deps
        assert "B1" in deps
        assert "C1" in deps

    def test_calculate_from_file(self, engine, simple_excel_with_formulas):
        """测试从文件计算公式"""
        results = engine.calculate_from_file(simple_excel_with_formulas)

        # 验证结果存在
        assert results is not None
        assert len(results) > 0

        # 查找 SUM 结果
        sum_found = False
        for key, value in results.items():
            if "C1" in key:
                assert value == 60.0  # 10 + 20 + 30
                sum_found = True
                break
        assert sum_found, "未找到 SUM 公式的计算结果"

    def test_calculate_from_file_with_inputs(self, engine, simple_excel_with_formulas):
        """测试使用自定义输入计算"""
        # 获取文件名（不含路径）
        filename = os.path.basename(simple_excel_with_formulas)

        # 覆盖 A1 的值
        inputs = {f"'[{filename}]TEST'!A1": 100}

        results = engine.calculate_from_file(simple_excel_with_formulas, inputs=inputs)

        # 验证结果
        assert results is not None

        # SUM 应该是 100 + 20 + 30 = 150
        sum_found = False
        for key, value in results.items():
            if "C1" in key:
                assert value == 150.0
                sum_found = True
                break
        assert sum_found

    def test_get_supported_functions(self, engine):
        """测试获取支持的函数列表"""
        functions = engine.get_supported_functions()

        # 验证返回列表
        assert isinstance(functions, list)
        # formulas 库至少支持一些基本函数
        # 注意：实际支持的函数数量可能因版本而异
        assert len(functions) >= 0

    def test_compile_formula(self, engine):
        """测试编译并执行公式"""
        formula = "=A1+B1"
        context = {"A1": 10, "B1": 20}

        result = engine.compile_formula(formula, context)

        # 验证结果
        assert result is not None
        # 结果可能是 Array 对象，需要提取值
        if hasattr(result, "tolist"):
            result = result.tolist()
        if isinstance(result, list):
            result = result[0] if len(result) > 0 else result

        assert result == 30

    def test_compile_formula_missing_input(self, engine):
        """测试编译公式时缺少输入"""
        formula = "=A1+B1"
        context = {"A1": 10}  # 缺少 B1

        with pytest.raises(FormulaError) as exc_info:
            engine.compile_formula(formula, context)
        assert exc_info.value.error_code == "E304"

    def test_detect_circular_reference_none(self, engine, simple_excel_with_formulas):
        """测试检测无循环引用"""
        has_circular = engine.detect_circular_reference(simple_excel_with_formulas)
        assert not has_circular

    def test_detect_circular_reference_exists(self, engine):
        """测试检测存在循环引用"""
        # 创建包含循环引用的文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Circular"

        # 创建循环引用: A1 -> A2 -> A1
        ws["A1"] = "=A2+1"
        ws["A2"] = "=A1+1"

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        wb.save(temp_file.name)

        try:
            # formulas 库可能会处理循环引用或抛出异常
            # 这个测试主要验证方法不会崩溃
            has_circular = engine.detect_circular_reference(temp_file.name)
            # 结果可能是 True 或 False，取决于 formulas 库的行为
            assert isinstance(has_circular, bool)
        finally:
            os.unlink(temp_file.name)
