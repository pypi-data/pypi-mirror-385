"""
FileMerger 单元测试
"""

import pytest
from openpyxl import load_workbook

from mcp_excel_supabase.excel.file_merger import FileMerger
from mcp_excel_supabase.utils.errors import ValidationError


class TestFileMerger:
    """FileMerger 测试类"""

    def test_merge_two_files(self, simple_excel_file, multi_sheet_excel_file, tmp_path):
        """测试合并两个文件"""
        merger = FileMerger()
        output_path = str(tmp_path / "merged.xlsx")

        # 合并文件
        result = merger.merge_files(
            file_paths=[simple_excel_file, multi_sheet_excel_file],
            output_path=output_path,
        )

        assert result["success"] is True
        assert result["merged_sheets"] == 4  # 1 + 3
        assert result["skipped_sheets"] == 0
        assert result["renamed_sheets"] == 0

        # 验证输出文件
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) == 4

    def test_merge_with_rename_strategy(self, tmp_path):
        """测试使用 rename 策略合并"""
        merger = FileMerger()

        # 创建两个有相同工作表名称的文件
        from openpyxl import Workbook

        file1 = str(tmp_path / "file1.xlsx")
        wb1 = Workbook()
        wb1.active.title = "Sheet1"  # type: ignore
        wb1.save(file1)

        file2 = str(tmp_path / "file2.xlsx")
        wb2 = Workbook()
        wb2.active.title = "Sheet1"  # type: ignore
        wb2.save(file2)

        output_path = str(tmp_path / "merged.xlsx")

        # 合并文件（rename 策略）
        result = merger.merge_files(
            file_paths=[file1, file2],
            output_path=output_path,
            handle_duplicates="rename",
        )

        assert result["success"] is True
        assert result["merged_sheets"] == 2
        assert result["renamed_sheets"] == 1

        # 验证工作表名称
        wb = load_workbook(output_path)
        assert "Sheet1" in wb.sheetnames
        assert "Sheet1_2" in wb.sheetnames

    def test_merge_with_skip_strategy(self, tmp_path):
        """测试使用 skip 策略合并"""
        merger = FileMerger()

        # 创建两个有相同工作表名称的文件
        from openpyxl import Workbook

        file1 = str(tmp_path / "file1.xlsx")
        wb1 = Workbook()
        wb1.active.title = "Sheet1"  # type: ignore
        wb1.save(file1)

        file2 = str(tmp_path / "file2.xlsx")
        wb2 = Workbook()
        wb2.active.title = "Sheet1"  # type: ignore
        wb2.save(file2)

        output_path = str(tmp_path / "merged.xlsx")

        # 合并文件（skip 策略）
        result = merger.merge_files(
            file_paths=[file1, file2],
            output_path=output_path,
            handle_duplicates="skip",
        )

        assert result["success"] is True
        assert result["merged_sheets"] == 1
        assert result["skipped_sheets"] == 1

        # 验证只有一个工作表
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) == 1
        assert "Sheet1" in wb.sheetnames

    def test_merge_with_overwrite_strategy(self, tmp_path):
        """测试使用 overwrite 策略合并"""
        merger = FileMerger()

        # 创建两个有相同工作表名称的文件
        from openpyxl import Workbook

        file1 = str(tmp_path / "file1.xlsx")
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Sheet1"  # type: ignore
        ws1["A1"] = "File1"  # type: ignore
        wb1.save(file1)

        file2 = str(tmp_path / "file2.xlsx")
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sheet1"  # type: ignore
        ws2["A1"] = "File2"  # type: ignore
        wb2.save(file2)

        output_path = str(tmp_path / "merged.xlsx")

        # 合并文件（overwrite 策略）
        result = merger.merge_files(
            file_paths=[file1, file2],
            output_path=output_path,
            handle_duplicates="overwrite",
        )

        assert result["success"] is True
        assert result["merged_sheets"] == 2

        # 验证第二个文件覆盖了第一个
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) == 1
        assert wb["Sheet1"]["A1"].value == "File2"

    def test_merge_with_sheet_names_filter(self, multi_sheet_excel_file, tmp_path):
        """测试选择性合并指定工作表"""
        merger = FileMerger()
        output_path = str(tmp_path / "merged.xlsx")

        # 只合并 Sales 和 Summary 工作表
        result = merger.merge_files(
            file_paths=[multi_sheet_excel_file],
            output_path=output_path,
            sheet_names=["Sales", "Summary"],
        )

        assert result["success"] is True
        assert result["merged_sheets"] == 2

        # 验证只有指定的工作表
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) == 2
        assert "Sales" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert "Expenses" not in wb.sheetnames

    def test_merge_empty_file_list(self, tmp_path):
        """测试合并空文件列表"""
        merger = FileMerger()
        output_path = str(tmp_path / "merged.xlsx")

        with pytest.raises(ValidationError) as exc_info:
            merger.merge_files(file_paths=[], output_path=output_path)

        assert "文件路径列表不能为空" in str(exc_info.value)

    def test_merge_invalid_duplicate_strategy(self, simple_excel_file, tmp_path):
        """测试无效的重名处理策略"""
        merger = FileMerger()
        output_path = str(tmp_path / "merged.xlsx")

        with pytest.raises(ValidationError) as exc_info:
            merger.merge_files(
                file_paths=[simple_excel_file],
                output_path=output_path,
                handle_duplicates="invalid",  # type: ignore
            )

        assert "无效的重名处理策略" in str(exc_info.value)

    def test_merge_preserves_formats(self, formatted_excel_file, tmp_path):
        """测试合并保留格式"""
        merger = FileMerger()
        output_path = str(tmp_path / "merged.xlsx")

        # 获取原始格式
        wb_orig = load_workbook(formatted_excel_file)
        ws_orig = wb_orig.active
        orig_font = ws_orig["A1"].font  # type: ignore

        # 合并文件（保留格式）
        result = merger.merge_files(
            file_paths=[formatted_excel_file],
            output_path=output_path,
            preserve_formats=True,
        )

        assert result["success"] is True

        # 验证格式已保留
        wb_merged = load_workbook(output_path)
        ws_merged = wb_merged.active
        merged_font = ws_merged["A1"].font  # type: ignore

        assert merged_font.name == orig_font.name
        assert merged_font.size == orig_font.size
        assert merged_font.bold == orig_font.bold

    def test_merge_without_formats(self, formatted_excel_file, tmp_path):
        """测试合并不保留格式"""
        merger = FileMerger()
        output_path = str(tmp_path / "merged.xlsx")

        # 合并文件（不保留格式）
        result = merger.merge_files(
            file_paths=[formatted_excel_file],
            output_path=output_path,
            preserve_formats=False,
        )

        assert result["success"] is True

        # 验证数据已复制但格式可能不同
        wb_merged = load_workbook(output_path)
        ws_merged = wb_merged.active
        assert ws_merged["A1"].value is not None  # type: ignore

    def test_generate_unique_name(self):
        """测试生成唯一名称"""
        merger = FileMerger()

        # 测试基础情况
        name = merger._generate_unique_name("Sheet1", [])
        assert name == "Sheet1"

        # 测试已存在的情况
        name = merger._generate_unique_name("Sheet1", ["Sheet1"])
        assert name == "Sheet1_2"

        # 测试多次重复
        name = merger._generate_unique_name("Sheet1", ["Sheet1", "Sheet1_2", "Sheet1_3"])
        assert name == "Sheet1_4"

    def test_merge_multiple_files(
        self, simple_excel_file, multi_sheet_excel_file, formatted_excel_file, tmp_path
    ):
        """测试合并多个文件"""
        merger = FileMerger()
        output_path = str(tmp_path / "merged.xlsx")

        # 合并三个文件
        result = merger.merge_files(
            file_paths=[simple_excel_file, multi_sheet_excel_file, formatted_excel_file],
            output_path=output_path,
        )

        assert result["success"] is True
        assert result["merged_sheets"] >= 3  # 至少有3个工作表

        # 验证输出文件
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) >= 3
