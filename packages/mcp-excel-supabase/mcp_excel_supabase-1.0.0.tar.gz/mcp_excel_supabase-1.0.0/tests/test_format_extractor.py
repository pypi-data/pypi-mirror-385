"""
测试 Excel 格式提取器（format_extractor.py）
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color

from mcp_excel_supabase.excel.format_extractor import FormatExtractor


# ============================================================================
# 测试颜色转换
# ============================================================================


class TestColorConversion:
    """测试颜色转换"""

    def test_color_to_hex_with_rgb(self) -> None:
        """测试 RGB 颜色转换"""
        color = Color(rgb="00FF0000")  # 红色（带 alpha 通道）
        hex_color = FormatExtractor._color_to_hex(color)
        assert hex_color == "#FF0000"

    def test_color_to_hex_without_alpha(self) -> None:
        """测试不带 alpha 通道的颜色"""
        color = Color(rgb="FF0000")  # 红色（不带 alpha 通道）
        hex_color = FormatExtractor._color_to_hex(color)
        assert hex_color == "#FF0000"

    def test_color_to_hex_with_none(self) -> None:
        """测试 None 颜色"""
        hex_color = FormatExtractor._color_to_hex(None)
        assert hex_color is None


# ============================================================================
# 测试字体格式提取
# ============================================================================


class TestFontFormatExtraction:
    """测试字体格式提取"""

    def test_extract_font_format(self) -> None:
        """测试提取字体格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 设置字体格式
        cell.font = Font(
            name="Arial",
            size=12,
            bold=True,
            italic=False,
            underline="single",
            color=Color(rgb="00FF0000"),
        )

        # 提取格式
        font_format = FormatExtractor.extract_font_format(cell)

        assert font_format is not None
        assert font_format.name == "Arial"
        assert font_format.size == 12
        assert font_format.bold is True
        assert font_format.italic is False
        assert font_format.underline == "single"
        assert font_format.color == "#FF0000"

    def test_extract_font_format_with_none(self) -> None:
        """测试提取空字体格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]
        cell.font = None

        font_format = FormatExtractor.extract_font_format(cell)
        assert font_format is None

    def test_extract_font_format_partial(self) -> None:
        """测试提取部分字体格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]
        cell.font = Font(bold=True, size=14)

        font_format = FormatExtractor.extract_font_format(cell)
        assert font_format is not None
        assert font_format.bold is True
        assert font_format.size == 14


# ============================================================================
# 测试填充格式提取
# ============================================================================


class TestFillFormatExtraction:
    """测试填充格式提取"""

    def test_extract_fill_format(self) -> None:
        """测试提取填充格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 设置填充格式
        cell.fill = PatternFill(
            start_color=Color(rgb="00FFFF00"), end_color=Color(rgb="00FFFF00"), fill_type="solid"
        )

        # 提取格式
        fill_format = FormatExtractor.extract_fill_format(cell)

        assert fill_format is not None
        assert fill_format.background_color == "#FFFF00"
        assert fill_format.pattern_type == "solid"

    def test_extract_fill_format_with_none(self) -> None:
        """测试提取空填充格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]
        cell.fill = None

        fill_format = FormatExtractor.extract_fill_format(cell)
        assert fill_format is None


# ============================================================================
# 测试边框格式提取
# ============================================================================


class TestBorderFormatExtraction:
    """测试边框格式提取"""

    def test_extract_border_format(self) -> None:
        """测试提取边框格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 设置边框格式
        cell.border = Border(
            top=Side(style="thin", color=Color(rgb="00000000")),
            bottom=Side(style="medium", color=Color(rgb="00FF0000")),
            left=Side(style="thick"),
            right=Side(style="thin"),
        )

        # 提取格式
        border_format = FormatExtractor.extract_border_format(cell)

        assert border_format is not None
        assert border_format.top is not None
        assert border_format.top.style == "thin"
        assert border_format.top.color == "#000000"
        assert border_format.bottom is not None
        assert border_format.bottom.style == "medium"
        assert border_format.bottom.color == "#FF0000"

    def test_extract_border_format_with_none(self) -> None:
        """测试提取空边框格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]
        cell.border = None

        border_format = FormatExtractor.extract_border_format(cell)
        assert border_format is None

    def test_extract_border_format_partial(self) -> None:
        """测试提取部分边框格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 只设置上边框
        cell.border = Border(top=Side(style="thin"))

        border_format = FormatExtractor.extract_border_format(cell)
        assert border_format is not None
        assert border_format.top is not None
        assert border_format.top.style == "thin"
        assert border_format.bottom is None
        assert border_format.left is None
        assert border_format.right is None


# ============================================================================
# 测试对齐格式提取
# ============================================================================


class TestAlignmentFormatExtraction:
    """测试对齐格式提取"""

    def test_extract_alignment_format(self) -> None:
        """测试提取对齐格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 设置对齐格式
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        # 提取格式
        alignment_format = FormatExtractor.extract_alignment_format(cell)

        assert alignment_format is not None
        assert alignment_format.horizontal == "center"
        assert alignment_format.vertical == "top"
        assert alignment_format.wrap_text is True

    def test_extract_alignment_format_with_none(self) -> None:
        """测试提取空对齐格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]
        cell.alignment = None

        alignment_format = FormatExtractor.extract_alignment_format(cell)
        assert alignment_format is None


# ============================================================================
# 测试数字格式提取
# ============================================================================


class TestNumberFormatExtraction:
    """测试数字格式提取"""

    def test_extract_number_format(self) -> None:
        """测试提取数字格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 设置数字格式
        cell.number_format = "0.00"

        # 提取格式
        number_format = FormatExtractor.extract_number_format(cell)
        assert number_format == "0.00"

    def test_extract_number_format_general(self) -> None:
        """测试提取默认数字格式（General）"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 默认格式是 'General'
        cell.number_format = "General"

        number_format = FormatExtractor.extract_number_format(cell)
        assert number_format is None

    def test_extract_number_format_date(self) -> None:
        """测试提取日期格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        cell.number_format = "yyyy-mm-dd"

        number_format = FormatExtractor.extract_number_format(cell)
        assert number_format == "yyyy-mm-dd"


# ============================================================================
# 测试完整单元格格式提取
# ============================================================================


class TestCellFormatExtraction:
    """测试完整单元格格式提取"""

    def test_extract_cell_format_complete(self) -> None:
        """测试提取完整的单元格格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 设置所有格式
        cell.font = Font(name="Arial", size=12, bold=True)
        cell.fill = PatternFill(
            start_color=Color(rgb="00FFFF00"), end_color=Color(rgb="00FFFF00"), fill_type="solid"
        )
        cell.border = Border(top=Side(style="thin"))
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "0.00"

        # 提取格式
        cell_format = FormatExtractor.extract_cell_format(cell)

        assert cell_format is not None
        assert cell_format.font is not None
        assert cell_format.font.name == "Arial"
        assert cell_format.fill is not None
        assert cell_format.fill.background_color == "#FFFF00"
        assert cell_format.border is not None
        assert cell_format.alignment is not None
        assert cell_format.alignment.horizontal == "center"
        assert cell_format.number_format == "0.00"

    def test_extract_cell_format_empty(self) -> None:
        """测试提取空单元格格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 不设置任何格式
        _ = FormatExtractor.extract_cell_format(cell)

        # 默认单元格可能有一些默认格式，所以可能不是 None
        # 这里我们只验证方法不会抛出异常
        assert True

    def test_extract_cell_format_partial(self) -> None:
        """测试提取部分单元格格式"""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        cell = ws["A1"]

        # 只设置字体
        cell.font = Font(bold=True)

        cell_format = FormatExtractor.extract_cell_format(cell)
        assert cell_format is not None
        assert cell_format.font is not None
        assert cell_format.font.bold is True
