import os, re, argparse, pdb
from colorama import Fore, Style
from datetime import datetime
from copy import deepcopy
from typing import Any
from collections import OrderedDict
from warnings import warn

import openpyxl.styles
import pandas as pd
import numpy as np
from mmengine.config import Config



MODELS = ['DATransUNet', 'DconnNet', 'LM_Net', 
          'MedNext', 'Resnet50', 'SwinTransformer', 
          'SwinUMamba', 'VMamba',]


def parser():
    argparser = argparse.ArgumentParser(description="CONCLUDE")
    argparser.add_argument('--exp-name', type=str, nargs='+', default=None)
    argparser.add_argument('--test-dir', type=str, default=os.environ.get("mm_testdir"))
    argparser.add_argument('--out-dir', type=str, default=os.environ.get("mm_concdir"))
    args = argparser.parse_args()
    return args


def VersionToFullExpName(exp_name, search_dir):
    if exp_name[-1] == ".":
        raise AttributeError(f"Target experiment name must not end with '.': {exp_name}")
    exp_list = os.listdir(search_dir)
    for exp in exp_list:
        if exp == exp_name:
            print(f"Found experiment: {exp_name} <-> {exp}")
            return exp
        elif exp.startswith(exp_name):
            pattern = r'\.[a-zA-Z]'    # Regex to find the first occurrence of '.' followed by a letter
            match = re.search(pattern, exp)
            if match is None:
                raise FileNotFoundError(f"No matching experiment name found for {exp_name}")
            if exp[:match.start()] == exp_name:
                print(f"Found experiment by prefix: {exp_name} -> {exp}")
                return exp
    raise RuntimeError(f"No matching experiment name found for '{exp_name}'")


class ExcelFormatter:
    @staticmethod
    def read_to_dict(exp_root:str, exps:list[str]
        ) -> dict[str,dict[str,dict[str,pd.DataFrame]]]:
        
        print(f"Reading csvs from {exp_root}")
        all_results = {}
        for exp in exps:
            round_root = os.path.join(exp_root, exp)
            round_results = {}
            for round in os.listdir(round_root):
                model_root = os.path.join(round_root, round)
                models_results = {}
                for model in os.listdir(model_root):
                    run_root = os.path.join(model_root, model)
                    
                    # save configs, only one config file should exist based on mmengine design.
                    py_cfgs = [file for file in os.listdir(run_root) if file.endswith('.py')]
                    if len(py_cfgs) > 1:
                        raise RuntimeError(f"Found more than 1 py config file of one model：{run_root}")
                    elif len(py_cfgs) == 1:
                        cfg_file_path = os.path.join(run_root, py_cfgs[0])
                        model_cfg = Config.fromfile(cfg_file_path)
                    
                    csvs = []
                    for run in os.listdir(run_root):
                        if run[:8].isdigit():
                            csv_root = os.path.join(run_root, run)
                            csvs += [os.path.join(csv_root, f) \
                                    for f in os.listdir(csv_root) 
                                    if f.endswith('.csv') and f.startswith('PerClassResult_')]
                    
                    if len(csvs) > 1:
                        raise RuntimeError(f"Num of csv file illegal, got {len(csvs)}: {run_root}")
                    elif len(csvs) == 0:
                        warn(Fore.YELLOW + f"No csv file found: {run_root}" + Style.RESET_ALL, UserWarning)
                    elif len(csvs) == 1:
                        csv_file_path = csvs[0]
                        pd_csv_dataframe = pd.read_csv(csv_file_path, index_col=0)
                        # drop the last row, which is mean row
                        pd_csv_dataframe = pd_csv_dataframe.drop(pd_csv_dataframe.index[-1])
                        models_results[model] = {
                                "csv": pd_csv_dataframe,
                                "config": model_cfg
                            }
                    else:
                        raise RuntimeError(f"Unexpected Error, Num of csv file illegal, got {len(csvs)}: {run_root}")
                
                # if there exists at least one log in a round, register it.
                if len(models_results) != 0:
                    round_results[round] = models_results
            
            # if there exists at least one log in a experiment, register it.
            if len(round_results) != 0:
                all_results[exp] = round_results

        return all_results

    @classmethod
    def center_all_cells(cls, xlsx_writer:pd.ExcelWriter) -> pd.ExcelWriter:
        align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        if not hasattr(cls, 'SHEET_NAME'):
           raise NotImplementedError
        worksheet = xlsx_writer.sheets[cls.SHEET_NAME] # type: ignore
        for row in worksheet.rows:
            for cell in row:
                cell.alignment = align
        return xlsx_writer
    
    @staticmethod
    def create_xlsx_writer(save_path) -> pd.ExcelWriter:
        if not os.path.exists(os.path.dirname(save_path)):
            os.makedirs(os.path.dirname(save_path))
        return pd.ExcelWriter(save_path)

    @staticmethod
    def done(xlsx_writer:pd.ExcelWriter) -> None:
        xlsx_writer.close()


    def convert_to_xlsx(self, all_csvs:dict, xlsx_writer:pd.ExcelWriter) -> pd.ExcelWriter:
        raise NotImplementedError


class DetailFormatter(ExcelFormatter):
    SHEET_NAME = 'Detail'
    
    def __init__(self):
        self.num_columns_ahead_value = 3
        self.dataset_heads:dict[str, dict[str, Any]] = OrderedDict()


    def _find_head_first_empty_column_index(self):
        if len(self.dataset_heads) == 0:
            return self.num_columns_ahead_value
        else:
            last_dataset = list(self.dataset_heads.keys())[-1]
            last_column_index_in_use = self.dataset_heads[last_dataset]['xlsx_column_range'][1]
            return last_column_index_in_use + 1


    def get_dataset_head(self, modality:str, dataset:str, class_names:list[str]):
        assert isinstance(class_names, list)
        head_name = f"{modality} - {dataset}"
        
        # Create New Dataset Head and Set Index Range.
        if self.dataset_heads.get(head_name) is None:
            next_column_index = self._find_head_first_empty_column_index()
            self.dataset_heads[head_name] = {
                'head_name': head_name,
                'class_names': class_names,
                'xlsx_column_range': [next_column_index,
                                      next_column_index + len(class_names) - 1]
            }
        # this is used for locate the start column index for a dataset
        return self.dataset_heads[head_name]


    def format_experiment_rows(self, all_csvs:dict[str,dict[str,dict[str,pd.DataFrame]]]):
        experiment_rows = []
        for exp_name, rounds in all_csvs.items():
            exp_temp = {'dataset_modality': [], 
                        'dataset_name': [],
                        'models_results': {},}
            exp_temp.values()
            for round_name, models in rounds.items():
                for model_name, run_dict in models.items():
                    exp_temp['dataset_modality'].append(run_dict['config'].get('modality'))
                    exp_temp['dataset_name'].append(run_dict['config'].get('dataset_source'))

                    if exp_temp['models_results'].get(model_name) is None:
                        exp_temp['models_results'][model_name] = [run_dict['csv']]
                    else:
                        exp_temp['models_results'][model_name].append(run_dict['csv'])
            
            # Check dataset consistency in each exp.
            exp_temp['dataset_modality'] = set(exp_temp['dataset_modality']) # type:ignore
            exp_temp['dataset_name'] = set(exp_temp['dataset_name']) # type:ignore
            assert len(exp_temp['dataset_modality']) == len(exp_temp['dataset_name']) == 1, \
                (f"Expect all runs using the same dataset in each exps, "
                f"but exp {exp_name} has not only one: {exp_temp['dataset_name']}.")
            exp_temp['dataset_modality'] = exp_temp['dataset_modality'].pop()
            exp_temp['dataset_name'] = exp_temp['dataset_name'].pop()

            # Register Dataset to formatter cache
            example_model_name = list(exp_temp['models_results'].keys())[0]
            example_array = exp_temp['models_results'][example_model_name][0] # type:ignore
            register_result = self.get_dataset_head(
                exp_temp['dataset_modality'], 
                exp_temp['dataset_name'],
                example_array.index.to_list()
                )
            (column_start_index, _) = register_result['xlsx_column_range']

            # calculate average for each model
            for model_name, arrays in exp_temp['models_results'].items():
                if model_name in ['dataset_modality', 'dataset_name']:
                    continue
                # calculate the dataframes average on each element
                df_values = np.mean([df.values for df in arrays], axis=0)
                # set_trace()
                averaged_df:pd.DataFrame = deepcopy(arrays[0])
                averaged_df.loc[:] = df_values # type:ignore
                # insert a row to xlsx. exp name, model name and metric name are ahead of the value
                empty_space = [None] * (column_start_index-3)
                for metric_name in averaged_df.columns:
                    one_row = [exp_name, model_name, metric_name, *empty_space] + averaged_df[metric_name].tolist()
                    experiment_rows.append(one_row)

        return experiment_rows


    def format_final(self, experiment_rows:list) -> list[list]:
        modality_and_dataset = []
        for name, info in self.dataset_heads.items():
            modality_and_dataset += [info['head_name']] * len(info['class_names'])
        
        class_name_list = []
        for info in self.dataset_heads.values():
            class_name_list += info['class_names']
        head_rows = [['exp', 'model', None    , *modality_and_dataset],
                     ['exp', 'model', 'Class' , *class_name_list]]
        all_rows = head_rows + experiment_rows
        
        return all_rows


    def merge_cells(self, xlsx_writer:pd.ExcelWriter) -> pd.ExcelWriter:
        # scan the entire column and merge the continuous cells with the same value
        worksheet = xlsx_writer.sheets[self.SHEET_NAME]
        for col_num in range(self.num_columns_ahead_value):
            prev_cell_value = None
            start_row = None
            for row_num in range(worksheet.max_row):
                cell_value = worksheet.cell(row=row_num+1, column=col_num+1).value
                if cell_value == prev_cell_value:
                    if start_row is None:
                        start_row = row_num + 1
                else:
                    if start_row is not None:
                        worksheet.merge_cells(
                            start_row=start_row-1,
                            start_column=col_num+1,
                            end_row=row_num,
                            end_column=col_num+1
                        )
                    start_row = None
                prev_cell_value = cell_value
            if start_row is not None:
                worksheet.merge_cells(
                    start_row=start_row-1,
                    start_column=col_num+1,
                    end_row=worksheet.max_row,
                    end_column=col_num+1
                )
        
        # scan the first row and merge the continuous cells with the same value
        prev_cell_value = None
        start_column = None
        for col_num in range(worksheet.max_column):
            cell_value = worksheet.cell(row=1, column=col_num+1).value
            if prev_cell_value == cell_value:
                if start_column is None:
                    start_column = col_num + 1
            else:
                if start_column is not None:
                    worksheet.merge_cells(
                        start_row=1,
                        start_column=start_column-1,
                        end_row=1,
                        end_column=col_num
                    )
                start_column = None
            prev_cell_value = cell_value
        if start_column is not None:
            worksheet.merge_cells(
                            start_row=1,
                            start_column=start_column-1,
                            end_row=1,
                            end_column=worksheet.max_column
                        )
            
        return xlsx_writer


    def convert_to_xlsx(self, all_csvs:dict, xlsx_writer:pd.ExcelWriter) -> pd.ExcelWriter:
        print(f"Formatting as detailed sheet...")
        experiment_rows = self.format_experiment_rows(all_csvs)
        context = self.format_final(experiment_rows)

        print(f"Formatting Done, Send to writer.")
        pd.DataFrame(context).to_excel(xlsx_writer, index=False, header=False, sheet_name=self.SHEET_NAME)
        
        self.merge_cells(xlsx_writer)
        self.center_all_cells(xlsx_writer)
        
        return xlsx_writer


class ExpCompareFormatter(ExcelFormatter):
    SHEET_NAME = 'ExpCompare'
    METRICS = ('Dice', 'Precision', 'Recall') # 'Acc', 'Fscore'
    
    def format_experiment_rows(self, all_csvs:dict[str,dict[str,dict[str,pd.DataFrame]]]):
        experiment_rows = []
        for exp_name, rounds in all_csvs.items():
            exp_temp = {'dataset_modality': [], 
                        'dataset_name': [],
                        'model_results': {}}
            for round_name, models in rounds.items():
                for model_name, run_dict in models.items():
                    exp_temp['dataset_modality'].append(run_dict['config'].get('modality'))
                    exp_temp['dataset_name'].append(run_dict['config'].get('dataset_source'))

                    if exp_temp.get(model_name) is None:
                        exp_temp['model_results'][model_name] = [run_dict['csv']]
                    else:
                        exp_temp['model_results'][model_name].append(run_dict['csv'])
            
            # Check dataset consistency in each exp.
            exp_temp['dataset_modality'] = set(exp_temp['dataset_modality']) # type:ignore
            exp_temp['dataset_name'] = set(exp_temp['dataset_name']) # type:ignore
            assert len(exp_temp['dataset_modality']) == len(exp_temp['dataset_name']) == 1, \
                f"Expect all runs using the same dataset in each exps, \
                but exp {exp_name} has more than one: {exp_temp['dataset_name']}."
            exp_temp['dataset_modality'] = exp_temp['dataset_modality'].pop()
            exp_temp['dataset_name'] = exp_temp['dataset_name'].pop()

            one_row = [exp_temp['dataset_modality'], exp_temp['dataset_name'], exp_name]
            
            # calculate average for each model, class-wise result is averaged
            WriteStartColumnIndex = 4
            for model_name in MODELS:
                arrays = exp_temp['model_results'].get(model_name, None)
                if arrays is None:
                    for _ in self.METRICS:
                        one_row.append(None)
                    continue
                
                # calculate the dataframes average on each element
                # This is experiment round(trial) wise
                df_values = np.mean([df.values for df in arrays], axis=0)   # [Class, Metrics]

                # Use copy and overwrite to get a dataframe containing the same index and headers
                averaged_df:pd.DataFrame = deepcopy(arrays[0])
                averaged_df.loc[:] = df_values # type:ignore
                mean_metric_df = averaged_df.mean(axis=0)
                
                # insert a row to xlsx. exp name, model name and metric name are ahead of the value
                for metric in self.METRICS:
                    one_row.append(mean_metric_df[metric])
                
            experiment_rows.append(one_row)

        return experiment_rows


    def merge_cells(self, xlsx_writer:pd.ExcelWriter) -> None:
        # Merge A1:A2 B1:B2 C1:C2
        worksheet = xlsx_writer.sheets[self.SHEET_NAME]
        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        worksheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
        
        for i in range(0, len(MODELS)):
            worksheet.merge_cells(start_row=1, start_column=4+len(self.METRICS)*i, end_row=1, end_column=4+len(self.METRICS)*(i+1)-1)


    def convert_to_xlsx(self, all_csvs: dict, xlsx_writer: pd.ExcelWriter) -> pd.ExcelWriter:
        print('Formatting as a comparable sheet...')
        rows_to_be_writen = self.format_experiment_rows(all_csvs)

        first_column = ['Modality', 'Dataset', 'Exp Name']
        second_column = [None, None, None]
        for model_name in MODELS:
            first_column  += [model_name]
            first_column  += [None] * (len(self.METRICS)-1)
            second_column += list(self.METRICS)
        
        all_rows = [first_column, second_column] + rows_to_be_writen
        pd.DataFrame(data=all_rows).to_excel(
            xlsx_writer, index=False, header=False, sheet_name=self.SHEET_NAME)
        self.merge_cells(xlsx_writer)
        self.center_all_cells(xlsx_writer)
        
        return xlsx_writer


def main():
    args = parser()
    if args.exp_name is None:
        exps = [folder for folder in os.listdir(args.test_dir) \
                if os.path.isdir(os.path.join(args.test_dir,folder))]
    else:
        exps = [VersionToFullExpName(exp, args.test_dir) for exp in args.exp_name]
    
    # prepare
    result_dict = ExcelFormatter.read_to_dict(args.test_dir, sorted(exps))
    time_name = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    save_path = os.path.join(args.out_dir, f"TestConclude_{time_name}.xlsx")
    writer = ExcelFormatter.create_xlsx_writer(save_path)
    
    try:
        # Format a detailed sheet
        Sheet1 = DetailFormatter()
        writer = Sheet1.convert_to_xlsx(result_dict, writer)
        
        # Format a sheet which is easy to compare across different experiments
        Sheet2 = ExpCompareFormatter()
        writer = Sheet2.convert_to_xlsx(result_dict, writer)
    
    except:
        ExcelFormatter.done(writer)
        os.remove(save_path)
    
    else:
        ExcelFormatter.done(writer)



if __name__ == "__main__":
    main()