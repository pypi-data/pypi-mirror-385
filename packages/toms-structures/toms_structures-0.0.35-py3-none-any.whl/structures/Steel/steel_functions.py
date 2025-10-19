from openpyxl import load_workbook
import math

wb = load_workbook(filename="Steel Design Calculator.xlsx")


def steel_data(string, wb):
    SectionSize = []
    for j in wb[string]:
        SectionSize.append(j[0].value)
    SectionSize = list(filter(lambda item: item is not None, SectionSize))
    for j1 in range(4):
        SectionSize.pop(0)
    return SectionSize


def import_steel_data():
    wb = load_workbook(filename="Steel Design Calculator.xlsx")
    # SectionType = wb.sheetnames
    return wb


class SteelBeamFactory:
    def get_beam(self, values):
        if (
            values["SectionType"] == "Universal_Beam"
            or values["SectionType"] == "Universal_Column"
        ):
            return ISection(values)
        if values["SectionType"] == "PFC" or values["SectionType"] == "T-section":
            return ChannelSection(values)
        elif values["SectionType"] in ["RHS", "SHS"]:
            return SquareSection(values)
        elif values["SectionType"] in ["CHS"]:
            return CircularSection(values)


class SteelBeam:
    def __init__(self, values):
        self.values = values
        self.section_properties = {}

    def calculate(self, values):
        section = self.fetch_section_properties(
            values["SectionType"], values["SectionSize"], 0, 0, 0
        )
        self.compact(
            section,
            values["SectionType"],
            float(values["segment length"]),
            float(values["alpha_m"]),
            values["restraint"],
            values["load_height_position"],
            values["longitudinal_position"],
            values["ends_with_restraint"],
        )
        self.section_moment_capacity()
        self.member_moment_capacity()
        self.shear(section, values["SectionType"])
        self.shear_moment(section, 0)
        self.axial_compression(
            section, values["SectionType"], float(values["segment length"])
        )
        self.uniaxial_bending_compression(section, float(values["N*"]))
        self.out_of_plane_compression(section, float(values["N*"]))
        return section

    def fetch_section_properties(self, SectionType, SectionSize, b, d, t):
        data = [
            "d",
            "bf",
            "tf",
            "tw",
            "Ix",
            "Zx",
            "Sx",
            "rx",
            "Iy",
            "Iw",
            "J",
            "Zy",
            "Sy",
            "fyf",
            "fyw",
            "kf",
            "Ag",
            "ry",
        ]
        self.section_properties = {}
        if SectionType == "RHS" or SectionType == "SHS":
            data = [
                "d",
                "bf",
                "tf",
                "tw",
                "Ix",
                "Zx",
                "Sx",
                "rx",
                "Iy",
                "Iw",
                "J",
                "Zy",
                "Sy",
                "Zex",
                "Zey",
                "fyf",
                "fyw",
                "kf",
                "Ag",
                "ry",
            ]
        if SectionType == "T-section":
            self.section_properties["d"] = 1
        if SectionType == "CHS":
            data = ["Ag", "I", "Z", "S", "r", "J", "Ze", "fy", "kf"]
        sheet_ranges = wb[SectionType]
        # use format x = sheet_ranges[A6].value
        count4 = 0
        count3 = 0
        for row1 in sheet_ranges:
            count3 += 1
            for cell1 in row1:
                if SectionSize == cell1.value:
                    count4 = count3
                    break
        count = 0
        for row in sheet_ranges:
            count += 1
            count1 = 0
            for cell in row:
                count1 += 1
                for x in range(len(data)):
                    if cell.value == data[x]:
                        data1 = sheet_ranges.cell(row=count4, column=count1)
                        self.section_properties[cell.value] = data1.value
        if "fyw" in self.section_properties:
            self.section_properties["fy"] = min(
                self.section_properties["fyw"], self.section_properties["fyf"]
            )
        print(self.section_properties)
        return self.section_properties

    def axial_compression(self, section_properties, SectionType, L):
        self.axial_compression_formula(self.section_properties, L, "x", SectionType)
        self.axial_compression_formula(self.section_properties, L, "y", SectionType)
        return self.section_properties

    def get_alpha_b(self):
        return 1

    def axial_compression_formula(self, section_properties, L, xy, SectionType):
        self.section_properties["Ns" + xy] = (
            self.section_properties["kf"]
            * self.section_properties["Ag"]
            * self.section_properties["fy"]
            / 1000
        )
        self.section_properties["PhiNs" + xy] = 0.9 * self.section_properties["Ns" + xy]
        ke = 1
        Le = L * ke * 1000
        lambdan = (
            (Le / self.section_properties["r" + xy])
            * math.sqrt(self.section_properties["kf"])
            * math.sqrt(self.section_properties["fy"] / 250)
        )
        alpha_a = (2100 * (lambdan - 13.5)) / (lambdan**2 - 15.3 * lambdan + 2050)
        alpha_b = self.get_alpha_b()
        lambda1 = lambdan + alpha_a * alpha_b
        n = max(0.00326 * (lambda1 - 13.5), 0)
        curly = ((lambda1 / 90) ** 2 + 1 + n) / (2 * (lambda1 / 90) ** 2)
        alpha_c = curly * (1 - math.sqrt(1 - (90 / (curly * lambda1)) ** 2))
        self.section_properties["Nc" + xy] = min(
            alpha_c * self.section_properties["Ns" + xy],
            self.section_properties["Ns" + xy],
        )
        self.section_properties["PhiNc" + xy] = 0.9 * self.section_properties["Nc" + xy]
        return self.section_properties

    def get_kt(self):
        self.values["segment length"] = float(self.values["segment length"])
        if (
            self.values["SectionType"] == "Universal_Beam"
            or self.values["SectionType"] == "Universal_Column"
            or self.values["SectionType"] == "PFC"
            or self.values["SectionType"] == "T-section"
        ):
            if (
                self.values["restraint"] == "FF"
                or self.values["restraint"] == "FL"
                or self.values["restraint"] == "LL"
                or self.values["restraint"] == "FU"
            ):
                kt = 1
            elif (
                self.values["restraint"] == "FP"
                or self.values["restraint"] == "PL"
                or self.values["restraint"] == "PU"
            ):
                kt = 1 + (
                    (
                        (
                            self.section_properties["d"]
                            - self.section_properties["tf"] * 2
                        )
                        / self.values["segment length"]
                        / 1000
                    )
                    * (
                        self.section_properties["tf"]
                        / (2 * self.section_properties["tw"])
                    )
                    ** 3
                )
                print(kt, "THis is L")
            elif self.values["restraint"] == "PP":
                kt = 1 + (
                    2
                    * (
                        (
                            self.section_properties["d"]
                            - self.section_properties["tf"] * 2
                        )
                        / self.values["segment length"]
                        / 1000
                    )
                    * (
                        self.section_properties["tf"]
                        / (2 * self.section_properties["tw"])
                    )
                    ** 3
                )
        self.values["kt"] = kt

    # TODO: effective length is using segment length not overall length, this is a mistake.
    def get_Le(self):
        Table_5_6_3 = wb["Table 5.6.3"]
        count1 = 0
        for row in Table_5_6_3:
            count = 0
            if (
                self.values["restraint"] in row[0].value
                and self.values["ends_with_restraint"] in row[1].value
            ):
                row2 = count1
            if (
                row[0].value == self.values["longitudinal_position"]
                and self.values["restraint"] in row[1].value
            ):
                row1 = count1
            for cell in row:
                if self.values["load_height_position"] == cell.value:
                    column = count
                count += 1
            count1 += 1
        kl = Table_5_6_3.cell(row=row1 + 1, column=column + 1).value
        kr = Table_5_6_3.cell(row=row2 + 1, column=3).value
        if (
            self.values["SectionType"] == "Universal_Beam"
            or self.values["SectionType"] == "Universal_Column"
            or self.values["SectionType"] == "PFC"
            or self.values["SectionType"] == "T-section"
            or self.values["SectionType"] == "RHS"
        ):
            Le = self.values["segment length"] * self.values["kt"] * kl * kr
        else:
            Le = self.values["segment length"]
        self.values["Le"] = Le
        return

    # This function will determine if the section is compact or not and then calculate the sectional moment capacity
    def compact(
        self,
        section_properties,
        SectionType,
        L,
        alpha_m,
        restraint,
        load_height_position,
        longitudinal_position,
        ends_with_restraint,
    ):
        self.get_kt()
        self.get_Le()
        # Determine if section is compact from section properties
        # The 0.5 applies because the flange extends from each side of the web
        if (
            SectionType == "Universal_Beam"
            or SectionType == "Universal_Column"
            or SectionType == "Welded_Beam"
            or SectionType == "T-section"
        ):
            f_lambda_e = (
                0.5
                * (self.section_properties["bf"] / self.section_properties["tf"])
                * math.sqrt(self.section_properties["fyf"] / 250)
            )
        elif SectionType == "PFC":
            f_lambda_e = (
                self.section_properties["bf"] / self.section_properties["tf"]
            ) * math.sqrt(self.section_properties["fyf"] / 250)
        # Use Table 5.2 to find lambda values
        table_5_2 = wb["Table 5.2"]
        if (
            SectionType == "Universal_Beam"
            or SectionType == "Universal_Column"
            or SectionType == "Welded_Beam"
            or SectionType == "PFC"
        ):
            for row in table_5_2:
                if row[0].value == "UCOneSR":
                    f_lambda_ey = row[6].value
                    f_lambda_ep = row[5].value
                    break
            for row in table_5_2:
                if row[0].value == "CTOneSR":
                    f_lambda_ey_OoP = row[6].value
                    f_lambda_ep_OoP = row[5].value
                    break
        else:
            f_lambda_e = 2
            f_lambda_ey = 1
            f_lambda_ep = 1
            f_lambda_ey_OoP = 1
            f_lambda_ep_OoP = 1

        w_lambda_e = (
            (self.section_properties["d"] - 2 * self.section_properties["tf"])
            / self.section_properties["tw"]
        ) * math.sqrt(self.section_properties["fyw"] / 250)
        # Use Table 5.2 to find lambda values
        table_5_2 = wb["Table 5.2"]
        for row in table_5_2:
            if row[0].value == "CTBothSR":
                w_lambda_ey = row[6].value
                w_lambda_ep = row[5].value
                break
        if f_lambda_e / f_lambda_ey > w_lambda_e / w_lambda_ey:
            lambda_s = f_lambda_e
            lambda_sy = f_lambda_ey
            lambda_sp = f_lambda_ep
        else:
            lambda_s = w_lambda_e
            lambda_sy = w_lambda_ey
            lambda_sp = w_lambda_ep
        if SectionType != "RHS" and SectionType != "SHS" and SectionType != "CHS":
            if lambda_s < lambda_sp:
                self.section_properties["Zex"] = min(
                    1.5 * self.section_properties["Zx"], self.section_properties["Sx"]
                )
                self.section_properties["compactness"] = "compact"
                print(self.section_properties["Zex"])
            elif lambda_s > lambda_sp and lambda_s < lambda_sy:
                self.section_properties["Zex"] = self.section_properties["Zx"] + (
                    ((lambda_sy - lambda_s) / (lambda_sy - lambda_sp))
                    * (
                        min(
                            1.5 * self.section_properties["Zx"],
                            self.section_properties["Sx"],
                        )
                        - self.section_properties["Zx"]
                    )
                )
                self.section_properties["compactness"] = "non-compact"
            elif lambda_s > lambda_sy:
                self.section_properties["Zex"] = self.section_properties["Zx"] * (
                    lambda_sy / lambda_s
                )
                self.section_properties["compactness"] = "slender"
            if f_lambda_e < f_lambda_ep_OoP:
                self.section_properties["Zey"] = min(
                    1.5 * self.section_properties["Zy"], self.section_properties["Sy"]
                )
                self.section_properties["compactness OoP"] = "compact"
                print(self.section_properties["Zey"])
            elif f_lambda_e > f_lambda_ep_OoP and f_lambda_e < f_lambda_ey_OoP:
                self.section_properties["Zey"] = self.section_properties["Zy"] + (
                    (
                        (f_lambda_ey_OoP - f_lambda_e)
                        / (f_lambda_ey_OoP - f_lambda_ep_OoP)
                    )
                    * (
                        min(
                            1.5 * self.section_properties["Zy"],
                            self.section_properties["Sy"],
                        )
                        - self.section_properties["Zy"]
                    )
                )
                self.section_properties["compactness OoP"] = "non-compact"
            elif f_lambda_e > f_lambda_ey_OoP:
                self.section_properties["Zey"] = self.section_properties["Zy"] * (
                    f_lambda_ey_OoP / f_lambda_e
                )
                self.section_properties["compactness OoP"] = "slender"

    def section_moment_capacity(self):
        self.section_properties["Msx"] = (
            min(self.section_properties["fyf"], self.section_properties["fyw"])
            * self.section_properties["Zex"]
            / 1000000
        )
        self.section_properties["Msy"] = (
            min(self.section_properties["fyf"], self.section_properties["fyw"])
            * self.section_properties["Zey"]
            / 1000000
        )
        self.section_properties["PhiMsx"] = 0.9 * self.section_properties["Msx"]
        self.section_properties["PhiMsy"] = 0.9 * self.section_properties["Msy"]

    def member_moment_capacity(self):
        Moa = (
            math.sqrt(
                (
                    (
                        math.pi**2
                        * 200
                        * 10**9
                        * self.section_properties["Iy"]
                        * 10 ** (-12)
                    )
                    / (self.values["Le"] ** 2)
                )
                * (
                    80 * 10**9 * self.section_properties["J"] * 10 ** (-12)
                    + (
                        (
                            math.pi**2
                            * 200
                            * 10**9
                            * self.section_properties["Iw"]
                            * 10 ** (-18)
                        )
                        / (self.values["Le"] ** 2)
                    )
                )
            )
            / 1000
        )
        Moa_OoP = (
            math.sqrt(
                (
                    (
                        math.pi**2
                        * 200
                        * 10**9
                        * self.section_properties["Ix"]
                        * 10 ** (-12)
                    )
                    / (self.values["Le"] ** 2)
                )
                * (
                    80 * 10**9 * self.section_properties["J"] * 10 ** (-12)
                    + (
                        (
                            math.pi**2
                            * 200
                            * 10**9
                            * self.section_properties["Iw"]
                            * 10 ** (-18)
                        )
                        / (self.values["Le"] ** 2)
                    )
                )
            )
            / 1000
        )

        alpha_s = 0.6 * (
            math.sqrt((self.section_properties["Msx"] / Moa) ** 2 + 3)
            - (self.section_properties["Msx"] / Moa)
        )
        alpha_s_OoP = 0.6 * (
            math.sqrt((self.section_properties["Msy"] / Moa_OoP) ** 2 + 3)
            - (self.section_properties["Msy"] / Moa_OoP)
        )
        self.values["alpha_m"] = float(self.values["alpha_m"])
        self.section_properties["PhiMbx"] = min(
            0.9 * self.values["alpha_m"] * alpha_s * self.section_properties["Msx"],
            self.section_properties["PhiMsx"],
        )
        self.section_properties["Mbx"] = min(
            self.values["alpha_m"] * alpha_s * self.section_properties["Msx"],
            self.section_properties["Msx"],
        )
        self.section_properties["PhiMby"] = min(
            0.9 * self.values["alpha_m"] * alpha_s_OoP * self.section_properties["Msy"],
            self.section_properties["PhiMsy"],
        )
        self.section_properties["Mby"] = min(
            self.values["alpha_m"] * alpha_s_OoP * self.section_properties["Msy"],
            self.section_properties["Msy"],
        )
        self.section_properties["Moa"] = Moa
        self.section_properties["alpha_m"] = self.values["alpha_m"]
        self.section_properties["alpha_s"] = alpha_s
        return self.section_properties

    def shear(self, section_properties, SectionType):
        self.section_properties["Vw"] = (
            0.6
            * self.section_properties["fyw"]
            * (self.section_properties["d"] - self.section_properties["tf"] * 2)
            * self.section_properties["tw"]
            / 1000
        )
        if (
            self.section_properties["d"] - self.section_properties["tf"]
        ) / self.section_properties["tw"] <= 82 / math.sqrt(
            self.section_properties["fyw"]
        ):
            self.section_properties["Vu"] = self.section_properties["Vw"]
        elif (
            self.section_properties["d"] - self.section_properties["tf"]
        ) / self.section_properties["tw"] >= 82 / math.sqrt(
            self.section_properties["fyw"]
        ):
            self.section_properties["alpha_v"] = (
                82
                / (
                    (
                        (
                            self.section_properties["d"]
                            - self.section_properties["tf"] * 2
                        )
                        / self.section_properties["tw"]
                    )
                    * math.sqrt(self.section_properties["fyw"] / 250)
                )
            ) ** 2
            self.section_properties["Vu"] = min(
                self.section_properties["Vw"] * self.section_properties["alpha_v"],
                self.section_properties["Vw"],
            )
        if SectionType == "SHS" or SectionType == "RHS":
            self.section_properties["Vu"] = 2 * self.section_properties["Vu"]
        self.section_properties["PhiVu"] = 0.9 * self.section_properties["Vu"]
        return self.section_properties

    def shear_moment(self, section_properties, M):
        if M <= 0.75 * self.section_properties["PhiMsx"]:
            self.section_properties["Vvm"] = self.section_properties["Vu"]
            self.section_properties["PhiVvm"] = 0.9 * self.section_properties["Vvm"]
        elif (
            M > 0.75 * self.section_properties["PhiMsx"]
            and M < self.section_properties["PhiMsx"]
        ):
            self.section_properties["Vvm"] = self.section_properties["Vu"] * (
                2.2 - ((1.6 * M) / self.section_properties["PhiMsx"])
            )
            self.section_properties["PhiVvm"] = 0.9 * self.section_properties["Vvm"]
        else:
            self.section_properties["Vvm"] = "M* exceeds PhiMsx"
        return self.section_properties

    # AS 4100:2020 Cl 8.4.2.2
    def uniaxial_bending_compression(self, section_properties, N):
        phi = 0.9
        self.section_properties["Mix"] = self.section_properties["Msx"] * (
            1 - N / self.section_properties["PhiNcy"]
        )
        self.section_properties["PhiMix"] = phi * self.section_properties["Mix"]
        self.section_properties["Miy"] = self.section_properties["Msy"] * (
            1 - N / self.section_properties["PhiNcy"]
        )
        self.section_properties["PhiMiy"] = phi * self.section_properties["Miy"]
        return self.section_properties

    # AS 4100:2020 Cl 8.4.4.1
    def out_of_plane_compression(self, section_properties, N):
        phi = 0.9
        self.section_properties["Mox"] = self.section_properties["Mbx"] * (
            1 - N / phi / self.section_properties["Ncy"]
        )
        self.section_properties["PhiMox"] = phi * self.section_properties["Mox"]
        self.section_properties["Mox"] = self.section_properties["Mbx"] * (
            1 - N / phi / self.section_properties["Ncy"]
        )
        self.section_properties["PhiMox"] = phi * self.section_properties["Mox"]
        return self.section_properties


class ISection(SteelBeam):
    def get_alpha_b(self):
        if self.section_properties["tf"] <= 40:
            return 0
        else:
            return 1


class SquareSection(SteelBeam):
    def get_kt(self):
        self.values["segment length"] = float(self.values["segment length"])
        if (
            self.values["restraint"] == "FF"
            or self.values["restraint"] == "FL"
            or self.values["restraint"] == "LL"
            or self.values["restraint"] == "FU"
        ):
            kt = 1
        elif (
            self.values["restraint"] == "FP"
            or self.values["restraint"] == "PL"
            or self.values["restraint"] == "PU"
        ):
            kt = (
                1
                + (
                    (
                        (
                            self.section_properties["d"]
                            - self.section_properties["tf"] * 2
                        )
                        / self.values["segment length"]
                        / 1000
                    )
                    * (
                        self.section_properties["tf"]
                        / (2 * self.section_properties["tw"])
                    )
                    ** 3
                )
                / 2
            )
        elif self.values["restraint"] == "PP":
            kt = (
                1
                + (
                    2
                    * (
                        (
                            self.section_properties["d"]
                            - self.section_properties["tf"] * 2
                        )
                        / self.values["segment length"]
                        / 1000
                    )
                    * (
                        self.section_properties["tf"]
                        / (2 * self.section_properties["tw"])
                    )
                    ** 3
                )
                / 2
            )
        self.values["kt"] = kt

    def get_alpha_b(self):
        return -0.5

    def compact(
        self,
        section_properties,
        SectionType,
        L,
        alpha_m,
        restraint,
        load_height_position,
        longitudinal_position,
        ends_with_restraint,
    ):
        self.get_kt()
        self.get_Le()


class CircularSection(SteelBeam):
    def axial_compression(self, section_properties, SectionType, L):
        self.axial_compression_formula(self.section_properties, L, "", SectionType)
        self.section_properties["PhiNsx"] = self.section_properties["PhiNs"]
        self.section_properties["PhiNsy"] = self.section_properties["PhiNs"]
        self.section_properties["PhiNcx"] = self.section_properties["PhiNc"]
        self.section_properties["PhiNcy"] = self.section_properties["PhiNc"]
        print("mark", self.section_properties)
        return self.section_properties

    def shear(self, section_properties, SectionType):
        self.section_properties["Vw"] = (
            0.36 * self.section_properties["fy"] * self.section_properties["Ag"]
        )
        self.section_properties["Vu"] = self.section_properties["Vw"]
        self.section_properties["PhiVu"] = 0.9 * self.section_properties["Vu"]
        return self.section_properties

    def compact(
        self,
        section_properties,
        SectionType,
        L,
        alpha_m,
        restraint,
        load_height_position,
        longitudinal_position,
        ends_with_restraint,
    ):
        pass

    def section_moment_capacity(self):
        self.section_properties["Msx"] = (
            self.section_properties["fy"] * self.section_properties["Ze"] / 1e6
        )
        self.section_properties["Msy"] = (
            self.section_properties["fy"] * self.section_properties["Ze"] / 1e6
        )
        self.section_properties["PhiMsx"] = 0.9 * self.section_properties["Msx"]
        self.section_properties["PhiMsy"] = 0.9 * self.section_properties["Msy"]

    def member_moment_capacity(self):
        self.section_properties["PhiMbx"] = self.section_properties["PhiMsx"]
        self.section_properties["PhiMby"] = self.section_properties["PhiMsy"]
        return self.section_properties

    def get_alpha_b(self):
        return -0.5

    def out_of_plane_compression(self, section_properties, N):
        pass

    def uniaxial_bending_compression(self, section_properties, N):
        pass


class ChannelSection(SteelBeam):
    def get_alpha_b(self):
        return 0.5


wb.close()
