from typing import Dict, Any, Optional, Tuple, List
from pathlib import Path
from dataclasses import dataclass

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, Session

from ...database import Base, Project, ParameterCategory, Parameter, ParameterValue
from ...io_formats.registry import register_importer


@dataclass
class ImportResult:
    project_name_en: str
    output_db_path: str
    categories_count: int
    parameters_count: int
    values_count: int


@register_importer("excel_rich")
class RichExcelImporter:
    """富格式Excel导入器，支持excel_rich导出格式"""
    
    def __init__(self):
        self.export_dir = Path("exports")
        self.export_dir.mkdir(exist_ok=True)

    def import_project(self, excel_path: str, output_db_path: str) -> ImportResult:
        """创建新库导入项目"""
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
        
        db_path = Path(output_db_path)
        if db_path.exists():
            db_path.unlink()
        
        engine = create_engine(f"sqlite:///{db_path}")
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        Base.metadata.create_all(bind=engine)
        db: Session = SessionLocal()
        
        try:
            wb = load_workbook(str(excel_file))
            
            # 解析项目概览
            if "项目概览" not in wb.sheetnames:
                raise ValueError("Excel文件缺少'项目概览'工作表")
            
            overview = wb["项目概览"]
            project_info = self._parse_project_overview(overview)
            
            # 创建项目
            project = Project(
                name=project_info["name"],
                name_en=project_info["name_en"],
                description=project_info.get("description"),
                time_horizon=project_info["time_horizon"],
                start_year=project_info["start_year"],
                year_step=project_info["year_step"],
                end_year=project_info["end_year"],
            )
            db.add(project)
            db.flush()
            
            categories_count = 0
            parameters_count = 0
            values_count = 0
            
            # 解析参数目录
            if "参数目录" in wb.sheetnames:
                catalog = wb["参数目录"]
                categories_info = self._parse_catalog(catalog)
            else:
                # 如果没有参数目录，从其他工作表推断
                categories_info = self._infer_categories_from_sheets(wb)
            
            # 创建分类和参数
            for cat_info in categories_info:
                # 创建分类
                category = ParameterCategory(
                    project_id=project.id,
                    name=cat_info["name"],
                    name_en=cat_info["name_en"],
                    description=cat_info.get("description"),
                )
                db.add(category)
                db.flush()
                categories_count += 1
                
                # 创建参数
                for param_info in cat_info["parameters"]:
                    parameter = Parameter(
                        category_id=category.id,
                        name=param_info["name"],
                        name_en=param_info["name_en"],
                        param_type=param_info["param_type"],
                        unit=param_info.get("unit"),
                        description=param_info.get("description"),
                        is_list=param_info["is_list"],
                        is_year_related=param_info["is_year_related"],
                        list_length=param_info.get("list_length")
                    )
                    db.add(parameter)
                    db.flush()
                    parameters_count += 1
                    
                    # 设置参数值
                    if param_info.get("values"):
                        for value_info in param_info["values"]:
                            value = ParameterValue(
                                parameter_id=parameter.id,
                                value=value_info["value"],
                                list_index=value_info.get("list_index")
                            )
                            db.add(value)
                            values_count += 1
            
            db.commit()
            
            return ImportResult(
                project_name_en=project.name_en,
                output_db_path=str(db_path),
                categories_count=categories_count,
                parameters_count=parameters_count,
                values_count=values_count,
            )
            
        finally:
            db.close()

    def import_into_current_db(self, db: Session, excel_path: str) -> Project:
        """导入到当前数据库（处理重名）"""
        wb = load_workbook(str(excel_path))
        
        # 解析项目概览
        if "项目概览" not in wb.sheetnames:
            raise ValueError("Excel文件缺少'项目概览'工作表")
        
        overview = wb["项目概览"]
        project_info = self._parse_project_overview(overview)
        
        # 处理重名
        base_name = project_info["name"]
        base_name_en = project_info["name_en"]
        new_name = base_name
        new_name_en = base_name_en
        
        suffix = 0
        while db.query(Project).filter(Project.name_en == new_name_en).first() is not None:
            suffix += 1
            new_name = f"{base_name} 副本{suffix}"
            new_name_en = f"{base_name_en}_copy{suffix}"
        
        # 创建项目
        project = Project(
            name=new_name,
            name_en=new_name_en,
            description=project_info.get("description"),
            time_horizon=project_info["time_horizon"],
            start_year=project_info["start_year"],
            year_step=project_info["year_step"],
            end_year=project_info["end_year"],
        )
        db.add(project)
        db.flush()
        
        # 解析参数目录
        if "参数目录" in wb.sheetnames:
            catalog = wb["参数目录"]
            categories_info = self._parse_catalog(catalog)
        else:
            categories_info = self._infer_categories_from_sheets(wb)
        
        # 从各个分类工作表中解析详细的参数定义
        categories_info = self._parse_detailed_parameters_from_sheets(wb, categories_info)
        
        # 解析参数值
        values_data = self._parse_values_from_sheets(wb, project_info)
        
        # 创建分类和参数
        for cat_info in categories_info:
            # 创建分类
            category = ParameterCategory(
                project_id=project.id,
                name=cat_info["name"],
                name_en=cat_info["name_en"],
                description=cat_info.get("description"),
            )
            db.add(category)
            db.flush()
            
            # 创建参数
            for param_info in cat_info["parameters"]:
                # 计算list_length
                list_length = None
                if param_info["is_list"]:
                    list_length = project.time_horizon
                
                parameter = Parameter(
                    category_id=category.id,
                    name=param_info["name"],
                    name_en=param_info["name_en"],
                    param_type=param_info["param_type"],
                    unit=param_info.get("unit"),
                    description=param_info.get("description"),
                    is_list=param_info["is_list"],
                    is_year_related=param_info["is_year_related"],
                    list_length=list_length
                )
                db.add(parameter)
                db.flush()
                
                # 设置参数值
                param_key = f"{cat_info['name_en']}.{param_info['name_en']}"
                if param_key in values_data:
                    for value_info in values_data[param_key]:
                        value = ParameterValue(
                            parameter_id=parameter.id,
                            value=value_info["value"],
                            list_index=value_info.get("list_index")
                        )
                        db.add(value)
        
        db.commit()
        return project

    def _parse_project_overview(self, ws) -> Dict[str, Any]:
        """解析项目概览工作表"""
        info = {
            "name": None,
            "name_en": None,
            "description": None,
            "time_horizon": None,
            "start_year": None,
            "year_step": None,
            "end_year": None,
        }
        
        # 查找项目信息
        for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
            if len(row) < 2:
                continue
            k, v = row[0], row[1]
            if not k:
                continue
                
            if k == "项目名称":
                info["name"] = v
            elif k == "项目英文名":
                info["name_en"] = v
            elif k == "项目描述":
                info["description"] = v
            elif k == "时间长度":
                try:
                    info["time_horizon"] = int(str(v).split()[0])
                except Exception:
                    info["time_horizon"] = int(v)
            elif k == "起始年份":
                info["start_year"] = int(v)
            elif k == "年份步长":
                info["year_step"] = int(v)
            elif k == "结束年份":
                info["end_year"] = int(v)
        
        # 检查必需字段
        missing = [k for k, v in info.items() if v is None]
        if missing:
            raise ValueError(f"项目概览缺少字段: {missing}")
        
        return info

    def _parse_catalog(self, ws) -> List[Dict[str, Any]]:
        """解析参数目录工作表"""
        categories = {}
        
        # 查找表头
        header_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "分类":
                header_row = row
                break
        
        if not header_row:
            raise ValueError("参数目录工作表格式错误：找不到表头")
        
        # 解析数据
        for row in range(header_row + 1, ws.max_row + 1):
            cat_name = ws.cell(row=row, column=1).value
            cat_name_en = ws.cell(row=row, column=2).value
            param_name = ws.cell(row=row, column=3).value
            param_name_en = ws.cell(row=row, column=4).value
            param_desc = ws.cell(row=row, column=5).value
            
            if not cat_name or not cat_name_en or not param_name or not param_name_en:
                continue
            
            # 创建或获取分类
            if cat_name_en not in categories:
                categories[cat_name_en] = {
                    "name": cat_name,
                    "name_en": cat_name_en,
                    "description": None,
                    "parameters": []
                }
            
            # 添加参数
            categories[cat_name_en]["parameters"].append({
                "name": param_name,
                "name_en": param_name_en,
                "description": param_desc,
                "param_type": "string",  # 默认类型，后续会从值工作表更新
                "unit": None,
                "is_list": False,
                "is_year_related": False,
                "list_length": None,
                "values": []
            })
        
        return list(categories.values())

    def _infer_categories_from_sheets(self, wb) -> List[Dict[str, Any]]:
        """从工作表推断分类信息"""
        categories = {}
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ["项目概览", "参数目录", "值_单值", "值_列表_年份", "值_列表_非年份"]:
                continue
            
            ws = wb[sheet_name]
            
            # 解析分类标题
            title = ws.cell(row=1, column=1).value
            if not title or not isinstance(title, str):
                continue
            
            # 提取分类名称
            if "(" in title and ")" in title:
                try:
                    cat_name = title.split(" (")[0].strip()
                    cat_name_en = title.split(" (")[1][:-1].strip()
                except Exception:
                    cat_name = sheet_name
                    cat_name_en = sheet_name
            else:
                cat_name = sheet_name
                cat_name_en = sheet_name
            
            # 解析描述
            desc_cell = ws.cell(row=2, column=1).value
            description = None
            if desc_cell and isinstance(desc_cell, str) and desc_cell.startswith("描述:"):
                description = desc_cell[3:].strip()
            
            # 解析参数
            parameters = self._parse_parameters_from_sheet(ws)
            
            categories[cat_name_en] = {
                "name": cat_name,
                "name_en": cat_name_en,
                "description": description,
                "parameters": parameters
            }
        
        return list(categories.values())

    def _parse_detailed_parameters_from_sheets(self, wb, categories_info: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """从各个分类工作表中解析详细的参数定义"""
        # 创建分类映射
        category_map = {cat["name_en"]: cat for cat in categories_info}
        
        # 遍历所有工作表，查找分类工作表
        for sheet_name in wb.sheetnames:
            if sheet_name in ["项目概览", "参数目录", "值_单值", "值_列表_年份", "值_列表_非年份"]:
                continue
            
            # 检查是否是分类工作表
            if sheet_name in category_map:
                ws = wb[sheet_name]
                
                # 解析分类描述
                description = self._parse_category_description(ws)
                if description:
                    category_map[sheet_name]["description"] = description
                
                # 解析参数定义
                parameters = self._parse_parameters_from_sheet(ws)
                category_map[sheet_name]["parameters"] = parameters
        
        return list(category_map.values())

    def _parse_category_description(self, ws) -> Optional[str]:
        """从分类工作表中解析分类描述"""
        # 检查第2行是否有描述信息
        desc_cell = ws.cell(row=2, column=1).value
        if desc_cell and isinstance(desc_cell, str) and desc_cell.startswith("描述:"):
            return desc_cell[3:].strip()
        return None

    def _parse_parameters_from_sheet(self, ws) -> List[Dict[str, Any]]:
        """从工作表解析参数信息"""
        parameters = []
        
        # 查找参数定义表头
        header_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "参数名称":
                header_row = row
                break
        
        if not header_row:
            return parameters
        
        # 解析参数定义
        for row in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            name_en = ws.cell(row=row, column=2).value
            param_type = ws.cell(row=row, column=3).value
            unit = ws.cell(row=row, column=4).value
            is_list = ws.cell(row=row, column=5).value
            is_year_related = ws.cell(row=row, column=6).value
            description = ws.cell(row=row, column=7).value
            
            if not name or not name_en:
                continue
            
            parameters.append({
                "name": name,
                "name_en": name_en,
                "param_type": param_type or "string",
                "unit": unit,
                "description": description,
                "is_list": is_list == "是",
                "is_year_related": is_year_related == "是",
                "list_length": None,
                "values": []
            })
        
        return parameters

    def _parse_values_from_sheets(self, wb, project_info: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
        """从值工作表解析参数值"""
        values = {}
        
        # 解析单值参数
        if "值_单值" in wb.sheetnames:
            single_values = self._parse_single_values(wb["值_单值"])
            values.update(single_values)
        
        # 解析年份相关列表参数
        if "值_列表_年份" in wb.sheetnames:
            year_values = self._parse_year_list_values(wb["值_列表_年份"], project_info)
            values.update(year_values)
        
        # 解析非年份列表参数
        if "值_列表_非年份" in wb.sheetnames:
            noyear_values = self._parse_noyear_list_values(wb["值_列表_非年份"])
            values.update(noyear_values)
        
        return values

    def _parse_single_values(self, ws) -> Dict[str, List[Dict[str, Any]]]:
        """解析单值参数"""
        values = {}
        
        # 查找表头
        header_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "参数名称":
                header_row = row
                break
        
        if not header_row:
            return values
        
        # 解析数据
        for row in range(header_row + 1, ws.max_row + 1):
            param_name_en = ws.cell(row=row, column=2).value
            value = ws.cell(row=row, column=3).value
            
            if not param_name_en or value is None:
                continue
            
            # 使用分类.参数名作为键
            # 需要从分类分隔符中推断分类名
            category_name_en = self._infer_category_from_single_values(ws, row)
            param_key = f"{category_name_en}.{param_name_en}"
            
            if param_key not in values:
                values[param_key] = []
            
            values[param_key].append({
                "value": str(value),
                "list_index": None
            })
        
        return values

    def _infer_category_from_single_values(self, ws, row: int) -> str:
        """从单值工作表中推断分类名"""
        # 向上查找分类分隔符
        for r in range(row, 0, -1):
            cell_value = ws.cell(row=r, column=1).value
            if cell_value and isinstance(cell_value, str) and cell_value.startswith("【"):
                # 提取分类名，格式：【分类名 / 分类英文名】
                content = cell_value[1:-1]  # 去掉【】
                if " / " in content:
                    return content.split(" / ")[1].strip()
        return "unknown"

    def _parse_year_list_values(self, ws, project_info: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
        """解析年份相关列表参数"""
        values = {}
        
        # 生成年份列表
        years = [project_info["start_year"] + i * project_info["year_step"] 
                for i in range(project_info["time_horizon"])]
        
        # 查找表头
        header_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "参数名称":
                header_row = row
                break
        
        if not header_row:
            return values
        
        # 解析数据
        for row in range(header_row + 1, ws.max_row + 1):
            param_name_en = ws.cell(row=row, column=2).value
            
            if not param_name_en:
                continue
            
            # 使用分类.参数名作为键
            category_name_en = self._infer_category_from_single_values(ws, row)
            param_key = f"{category_name_en}.{param_name_en}"
            
            if param_key not in values:
                values[param_key] = []
            
            # 读取年份对应的值
            for i, year in enumerate(years):
                value = ws.cell(row=row, column=3 + i).value
                if value is not None:
                    values[param_key].append({
                        "value": str(value),
                        "list_index": i
                    })
        
        return values

    def _parse_noyear_list_values(self, ws) -> Dict[str, List[Dict[str, Any]]]:
        """解析非年份列表参数"""
        values = {}
        
        # 查找表头
        header_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "参数名称":
                header_row = row
                break
        
        if not header_row:
            return values
        
        # 解析数据
        for row in range(header_row + 1, ws.max_row + 1):
            param_name_en = ws.cell(row=row, column=2).value
            
            if not param_name_en:
                continue
            
            # 使用分类.参数名作为键
            category_name_en = self._infer_category_from_single_values(ws, row)
            param_key = f"{category_name_en}.{param_name_en}"
            
            if param_key not in values:
                values[param_key] = []
            
            # 读取列表值
            col = 3
            index = 0
            while True:
                value = ws.cell(row=row, column=col).value
                if value is None:
                    break
                
                values[param_key].append({
                    "value": str(value),
                    "list_index": index
                })
                
                col += 1
                index += 1
        
        return values
