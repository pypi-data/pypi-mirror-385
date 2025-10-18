from typing import List, Dict, Any, Tuple
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ...crud import get_project_by_name_en, get_parameter_categories, get_parameters_with_values
from ...io_formats.registry import register_exporter


@register_exporter("excel_rich")
class RichExcelExporter:
    def __init__(self):
        self.export_dir = Path("exports")
        self.export_dir.mkdir(exist_ok=True)
        self.font_normal = Font(name="微软雅黑", size=10)
        self.font_small_italic_gray = Font(name="微软雅黑", size=9, italic=True, color="666666")
        self.header_font_dark = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        self.header_fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_fill_green = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        self.header_fill_teal = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        self.title_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
        self.title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.center = Alignment(horizontal="center", vertical="center")
        self.left = Alignment(horizontal="left", vertical="center")
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 统一接口
    def export(self, project_name_en: str, db_session) -> str:
        project = get_project_by_name_en(db_session, project_name_en)
        if not project:
            raise ValueError(f"项目 '{project_name_en}' 不存在")
        categories = get_parameter_categories(db_session, project.id)
        if not categories:
            raise ValueError(f"项目 '{project_name_en}' 没有参数分类")
        cat_to_params: Dict[int, List[Dict[str, Any]]] = {}
        for cat in categories:
            cat_to_params[cat.id] = get_parameters_with_values(db_session, cat.id)
        wb = Workbook()
        wb.remove(wb.active)
        self._sheet_project_overview(wb, project, categories)
        catalog_ws = wb.create_sheet("参数目录", 1)
        category_sheet_refs: Dict[Tuple[str, str], Tuple[str, int]] = {}
        for cat in categories:
            sheet_name = self._safe_sheet_name(cat.name_en)
            ws = wb.create_sheet(sheet_name)
            self._sheet_category_definition(ws, cat, cat_to_params[cat.id], category_sheet_refs)
        values_single_refs: Dict[Tuple[str, str], int] = {}
        self._sheet_values_single(wb.create_sheet("值_单值"), categories, cat_to_params, values_single_refs)
        values_year_refs: Dict[Tuple[str, str], int] = {}
        self._sheet_values_list_year(wb.create_sheet("值_列表_年份"), project, categories, cat_to_params, values_year_refs)
        values_noyear_refs: Dict[Tuple[str, str], int] = {}
        self._sheet_values_list_noyear(wb.create_sheet("值_列表_非年份"), categories, cat_to_params, values_noyear_refs)
        self._sheet_catalog(catalog_ws, categories, cat_to_params, category_sheet_refs)
        self._update_param_definition_links(wb, categories, values_single_refs, values_year_refs, values_noyear_refs)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{project.name_en}_parameters_rich_{ts}.xlsx"
        path = self.export_dir / filename
        wb.save(path)
        return str(path)

    def get_export_formats(self) -> List[Dict[str, str]]:
        return [
            {"format": "excel_rich", "name": "富格式Excel文件", "extension": ".xlsx", "description": "多sheet、超链接、隐藏元数据的Excel文件"}
        ]

    # ---------- helpers (从原 rich 导出迁移) ----------
    def _sheet_project_overview(self, wb: Workbook, project, categories: List[Any]):
        ws = wb.create_sheet("项目概览", 0)
        ws.merge_cells('A1:D1')
        ws['A1'] = f"项目参数导出 - {project.name}"
        ws['A1'].font = self.title_font
        ws['A1'].fill = self.title_fill
        ws['A1'].alignment = self.center
        info = [["项目名称", project.name],["项目英文名", project.name_en],["项目描述", project.description or "无"],["时间长度", f"{project.time_horizon} 年"],["起始年份", project.start_year],["年份步长", project.year_step],["结束年份", project.end_year],["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],]
        row = 3
        for label, val in info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = val
            ws[f'A{row}'].font = Font(name="微软雅黑", size=11, bold=True)
            ws[f'B{row}'].font = Font(name="微软雅黑", size=11)
            row += 1
        ws[f'A{row+1}'] = "分类名称"; ws[f'B{row+1}'] = "分类英文名"; ws[f'C{row+1}'] = "参数数量"; ws[f'D{row+1}'] = "描述"
        for col in ['A','B','C','D']:
            cell = ws[f'{col}{row+1}']
            cell.font = self.header_font_dark
            cell.fill = self.header_fill_blue
            cell.alignment = self.center
            cell.border = self.thin_border
        r = row + 2
        for cat in categories:
            ws[f'A{r}'] = cat.name
            ws[f'B{r}'] = cat.name_en
            ws[f'C{r}'] = len(getattr(cat, 'parameters', []) or [])
            ws[f'D{r}'] = cat.description or "无"
            for col in ['A','B','C','D']:
                ws[f'{col}{r}'].border = self.thin_border
            ws[f'C{r}'].alignment = self.center
            r += 1
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30

    def _sheet_catalog(self, ws, categories, cat_to_params, refs):
        headers = ["分类", "分类英文名", "参数名称", "参数英文名", "描述"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = self.header_font_dark
            c.fill = self.header_fill_green
            c.alignment = self.center
            c.border = self.thin_border
        row = 2
        for cat in categories:
            params = cat_to_params[cat.id]
            params_sorted = sorted(params, key=lambda p: (p['name'] or '', p['name_en'] or ''))
            if params_sorted:
                start_row = row
                end_row = row + len(params_sorted) - 1
                if start_row == end_row:
                    cat_cell = ws.cell(row=row, column=1, value=cat.name)
                    cat_cell.font = self.font_normal
                    cat_cell.alignment = Alignment(horizontal="center", vertical="center")
                    cat_en_cell = ws.cell(row=row, column=2, value=cat.name_en)
                    cat_en_cell.font = self.font_normal
                    cat_en_cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    ws.merge_cells(f'A{start_row}:A{end_row}')
                    ws.merge_cells(f'B{start_row}:B{end_row}')
                    cat_cell = ws[f'A{start_row}']
                    cat_cell.value = cat.name
                    cat_cell.font = self.font_normal
                    cat_cell.alignment = Alignment(horizontal="center", vertical="center")
                    cat_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cat_en_cell = ws[f'B{start_row}']
                    cat_en_cell.value = cat.name_en
                    cat_en_cell.font = self.font_normal
                    cat_en_cell.alignment = Alignment(horizontal="center", vertical="center")
                    cat_en_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    for r in range(start_row + 1, end_row + 1):
                        ws.cell(row=r, column=1).border = Border(left=Side(style='thin'), right=Side(style='thin'))
                        ws.cell(row=r, column=2).border = Border(left=Side(style='thin'), right=Side(style='thin'))
                        ws.cell(row=r, column=3).border = Border(left=Side(style='thin'), right=Side(style='thin'))
                        ws.cell(row=r, column=4).border = Border(left=Side(style='thin'), right=Side(style='thin'))
                        ws.cell(row=r, column=5).border = Border(left=Side(style='thin'), right=Side(style='thin'))
                for p in params_sorted:
                    name_cell = ws.cell(row=row, column=3, value=p['name'])
                    name_cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
                    key = (cat.name_en, p['name_en'])
                    if key in refs:
                        sheet_name, target_row = refs[key]
                        name_cell.hyperlink = f"#{sheet_name}!A{target_row}"
                        name_cell.style = "Hyperlink"
                    ws.cell(row=row, column=4, value=p['name_en']).font = self.font_normal
                    ws.cell(row=row, column=5, value=p.get('description') or "").font = self.font_normal
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).border = self.thin_border
                    row += 1
                if params_sorted:
                    last_row = row - 1
                    for col in range(1, 6):
                        cell = ws.cell(row=last_row, column=col)
                        cb = cell.border
                        cell.border = Border(left=cb.left, right=cb.right, top=cb.top, bottom=Side(style='thin'))
        widths = [16, 20, 18, 22, 40]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    def _sheet_category_definition(self, ws, category, params: List[Dict[str, Any]], refs: Dict[Tuple[str, str], Tuple[str, int]]):
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{category.name} ({category.name_en})"
        ws['A1'].font = self.title_font
        ws['A1'].fill = self.title_fill
        ws['A1'].alignment = self.center
        info_row = 2
        if category.description:
            ws.merge_cells('A2:G2')
            ws['A2'] = f"描述: {category.description}"
            ws['A2'].font = Font(name="微软雅黑", size=10, italic=True)
            ws['A2'].alignment = self.left
            info_row = 3
        header_row = info_row + 1
        headers = ["参数名称", "参数英文名", "参数类型", "单位", "是否列表", "是否关联年份", "描述"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=header_row, column=i, value=h)
            c.font = self.header_font_dark
            c.fill = self.header_fill_green
            c.alignment = self.center
            c.border = self.thin_border
        row = header_row + 1
        def keyf(p):
            if not p['is_list']:
                g = 0
            elif p['is_list'] and not p['is_year_related']:
                g = 1
            else:
                g = 2
            return (g, p['name'] or '')
        for p in sorted(params, key=keyf):
            name_cell = ws.cell(row=row, column=1, value=p['name'])
            name_cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
            ws.cell(row=row, column=2, value=p['name_en']).font = self.font_normal
            ws.cell(row=row, column=3, value=p['param_type']).font = self.font_normal
            ws.cell(row=row, column=4, value=p.get('unit') or "").font = self.font_normal
            ws.cell(row=row, column=5, value="是" if p['is_list'] else "否").font = self.font_normal
            ws.cell(row=row, column=6, value="是" if p['is_year_related'] else "否").font = self.font_normal
            ws.cell(row=row, column=7, value=p.get('description') or "").font = self.font_normal
            refs[(category.name_en, p['name_en'])] = (ws.title, row)
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.thin_border
            row += 1
        widths = [16, 20, 12, 10, 10, 12, 40]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    def _append_category_separator(self, ws, row: int, title: str, max_col: int = None) -> int:
        if max_col is None:
            max_col = ws.max_column if ws.max_column > 0 else 8
        end_col = get_column_letter(max_col)
        ws.merge_cells(f'A{row}:{end_col}{row}')
        c = ws[f'A{row}']
        c.value = f"【{title}】"
        c.font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        c.alignment = self.center
        return row + 1

    def _sheet_values_single(self, ws, categories, cat_to_params, refs: Dict[Tuple[str, str], int] = None):
        headers = ["参数名称", "参数英文名", "值", "单位"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = self.header_font_dark
            c.fill = self.header_fill_blue
            c.alignment = self.center
            c.border = self.thin_border
        row = 2
        for cat in categories:
            row = self._append_category_separator(ws, row, f"{cat.name} / {cat.name_en}", 4)
            for p in cat_to_params[cat.id]:
                if p['is_list']:
                    continue
                if refs is not None:
                    refs[(cat.name_en, p['name_en'])] = row
                ws.cell(row=row, column=1, value=p['name']).font = self.font_normal
                ws.cell(row=row, column=2, value=p['name_en']).font = self.font_normal
                val = p.get('current_value')
                ws.cell(row=row, column=3, value=("" if val is None else str(val))).font = self.font_normal
                ws.cell(row=row, column=4, value=p.get('unit') or "").font = self.font_normal
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.thin_border
                row += 1
        widths = [18, 22, 16, 10]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    def _sheet_values_list_year(self, ws, project, categories, cat_to_params, refs: Dict[Tuple[str, str], int] = None):
        years = [project.start_year + i * project.year_step for i in range(project.time_horizon)]
        headers = ["参数名称", "参数英文名", *years, "单位"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = self.header_font_dark
            c.fill = self.header_fill_teal
            c.alignment = self.center
            c.border = self.thin_border
        row = 2
        for cat in categories:
            row = self._append_category_separator(ws, row, f"{cat.name} / {cat.name_en}", len(years) + 3)
            for p in cat_to_params[cat.id]:
                if not (p['is_list'] and p['is_year_related']):
                    continue
                if refs is not None:
                    refs[(cat.name_en, p['name_en'])] = row
                ws.cell(row=row, column=1, value=p['name']).font = self.font_normal
                ws.cell(row=row, column=2, value=p['name_en']).font = self.font_normal
                values = p.get('current_values') or []
                for i, y in enumerate(years):
                    val = values[i] if i < len(values) else None
                    ws.cell(row=row, column=3 + i, value=(None if val in (None, "") else val)).font = self.font_normal
                ws.cell(row=row, column=3 + len(years), value=p.get('unit') or "").font = self.font_normal
                for col in range(1, 4 + len(years)):
                    ws.cell(row=row, column=col).border = self.thin_border
                row += 1
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        for idx in range(3, 3 + len(years)):
            ws.column_dimensions[get_column_letter(idx)].width = 10
        ws.column_dimensions[get_column_letter(3 + len(years))].width = 10

    def _sheet_values_list_noyear(self, ws, categories, cat_to_params, refs: Dict[Tuple[str, str], int] = None):
        for i, h in enumerate(["参数名称", "参数英文名"], 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = self.header_font_dark
            c.fill = self.header_fill_green
            c.alignment = self.center
            c.border = self.thin_border
        max_len = 0
        for cat in categories:
            for p in cat_to_params[cat.id]:
                if p['is_list'] and not p['is_year_related']:
                    max_len = max(max_len, int(p.get('list_length') or len(p.get('current_values') or []) or 0))
        col = 3
        for idx in range(1, max_len + 1):
            c = ws.cell(row=1, column=col, value=idx)
            c.font = self.header_font_dark
            c.fill = self.header_fill_green
            c.alignment = self.center
            c.border = self.thin_border
            col += 1
        c = ws.cell(row=1, column=col, value="单位")
        c.font = self.header_font_dark
        c.fill = self.header_fill_green
        c.alignment = self.center
        c.border = self.thin_border
        row = 2
        for cat in categories:
            row = self._append_category_separator(ws, row, f"{cat.name} / {cat.name_en}", max_len + 3)
            for p in cat_to_params[cat.id]:
                if not (p['is_list'] and not p['is_year_related']):
                    continue
                if refs is not None:
                    refs[(cat.name_en, p['name_en'])] = row
                ws.cell(row=row, column=1, value=p['name']).font = self.font_normal
                ws.cell(row=row, column=2, value=p['name_en']).font = self.font_normal
                values = p.get('current_values') or []
                for i in range(max_len):
                    val = values[i] if i < len(values) else None
                    ws.cell(row=row, column=3 + i, value=(None if val in (None, "") else val)).font = self.font_normal
                ws.cell(row=row, column=3 + max_len, value=p.get('unit') or "").font = self.font_normal
                for ccol in range(1, 4 + max_len):
                    ws.cell(row=row, column=ccol).border = self.thin_border
                row += 1
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        for idx in range(3, 3 + max_len):
            ws.column_dimensions[get_column_letter(idx)].width = 10
        ws.column_dimensions[get_column_letter(3 + max_len)].width = 10

    def _update_param_definition_links(self, wb, categories, values_single_refs, values_year_refs, values_noyear_refs):
        for cat in categories:
            sheet_name = self._safe_sheet_name(cat.name_en)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in range(2, ws.max_row + 1):
                    param_name_en_cell = ws.cell(row=row, column=2)
                    if param_name_en_cell.value:
                        param_name_en = param_name_en_cell.value
                        key = (cat.name_en, param_name_en)
                        is_list_cell = ws.cell(row=row, column=5)
                        is_year_related_cell = ws.cell(row=row, column=6)
                        if is_list_cell.value == "否":
                            if key in values_single_refs:
                                target_row = values_single_refs[key]
                                name_cell = ws.cell(row=row, column=1)
                                name_cell.hyperlink = f"#值_单值!A{target_row}"
                                name_cell.style = "Hyperlink"
                        elif is_year_related_cell.value == "是":
                            if key in values_year_refs:
                                target_row = values_year_refs[key]
                                name_cell = ws.cell(row=row, column=1)
                                name_cell.hyperlink = f"#值_列表_年份!A{target_row}"
                                name_cell.style = "Hyperlink"
                        else:
                            if key in values_noyear_refs:
                                target_row = values_noyear_refs[key]
                                name_cell = ws.cell(row=row, column=1)
                                name_cell.hyperlink = f"#值_列表_非年份!A{target_row}"
                                name_cell.style = "Hyperlink"

    def _safe_sheet_name(self, name: str) -> str:
        name = (name or "").strip() or "Sheet"
        invalid = set('[]:*?/\\')
        safe = ''.join(ch for ch in name if ch not in invalid)
        return safe[:31]
