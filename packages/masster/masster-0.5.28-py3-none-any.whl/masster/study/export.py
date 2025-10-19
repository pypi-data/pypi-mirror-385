from __future__ import annotations

import os

from datetime import datetime

import numpy as np
import pandas as pd
import polars as pl

from tqdm import tqdm

from masster.spectrum import combine_peaks
from masster.study.defaults import export_mgf_defaults
from masster._version import get_version


def _get_mgf_df(self, **kwargs):
    """
    Generate MGF data as a Polars DataFrame.

    This is the core data generation function used by export_mgf().

    Parameters:
        **kwargs: Keyword arguments for export parameters. Same as export_mgf()
                 except return_data is not relevant here.

    Returns:
        pl.DataFrame: DataFrame with columns:
            - mgf_index: MGF index
            - title: MGF title string
            - feature_id: Consensus feature ID
            - feature_uid: Consensus feature UID
            - charge: Charge state
            - pepmass: Precursor m/z
            - rtinseconds: Retention time in seconds
            - mslevel: MS level
            - type: Spectrum type (e.g., "MS2")
            - energy: Collision energy (if available)
            - spec_len: Number of peaks in spectrum
            - spec_mz: List of spectrum m/z values
            - spec_int: List of spectrum intensity values
    """
    # parameters initialization
    params = export_mgf_defaults()
    for key, value in kwargs.items():
        if isinstance(value, export_mgf_defaults):
            params = value
            self.logger.debug("Using provided export_defaults parameters")
        else:
            if hasattr(params, key):
                if params.set(key, value, validate=True):
                    self.logger.debug(f"Updated parameter {key} = {value}")
                else:
                    self.logger.warning(
                        f"Failed to set parameter {key} = {value} (validation failed)",
                    )
            else:
                self.logger.debug(f"Unknown parameter {key} ignored")
    # end of parameter initialization

    # Store parameters in the Study object
    self.update_history(["get_mgf"], params.to_dict())
    self.logger.debug("Parameters stored to get_mgf")

    # Get parameter values for use in the method
    selection = params.get("selection")
    split_energy = params.get("split_energy")
    merge = params.get("merge")
    mz_start = params.get("mz_start")
    mz_end = params.get("mz_end")
    rt_start = params.get("rt_start")
    rt_end = params.get("rt_end")
    centroid = params.get("centroid")
    inty_min = params.get("inty_min")
    deisotope = params.get("deisotope")

    if self.consensus_df is None:
        self.logger.error("No consensus map found. Please run merge() first.")
        return None

    # MS2 data is optional - we can generate MS1 data without it
    ms2_available = self.consensus_ms2 is not None and not self.consensus_ms2.is_empty()
    if not ms2_available:
        self.logger.info("No consensus MS2 data found. Generating MS1-only MGF data.")

    # Convert to pandas for merge operation only if we have MS2 data
    if ms2_available:
        consensus_df_pd = self.consensus_df.to_pandas()
        consensus_ms2_pd = self.consensus_ms2.to_pandas()

        features = pd.merge(
            consensus_df_pd,
            consensus_ms2_pd,
            how="right",
            on="consensus_uid",
        )
        if len(features) == 0:
            self.logger.warning("No MS2 features found.")
            grouped = {}  # Empty groupby result
        else:
            # Pre-group by consensus_uid for fast access
            grouped = features.groupby("consensus_uid")
    else:
        grouped = {}  # No MS2 data available

    def filter_peaks(spec, inty_min=None):
        spec = spec.copy()
        length = len(spec.mz)
        mask = np.ones(length, dtype=bool)
        if inty_min is not None and inty_min > 0:
            mask = mask & (spec.inty >= inty_min)
        for attr in spec.__dict__:
            arr = getattr(spec, attr)
            if isinstance(arr, list | np.ndarray) and hasattr(arr, "__len__") and len(arr) == length:
                setattr(spec, attr, np.array(arr)[mask])
        return spec

    def safe_charge(charge_value):
        """Safely convert charge value to integer, handling NaN and None"""
        if charge_value is None or (isinstance(charge_value, float) and np.isnan(charge_value)):
            return 1
        return int(round(charge_value))

    def create_ion_dict(title, id, uid, mz, rt, charge, spect, mgf_id):
        """Create a dictionary representing an ion for the DataFrame."""
        if spect is None:
            return None

        # Prepare spectrum data
        spectrum_mz = spect.mz.tolist() if hasattr(spect.mz, "tolist") else list(spect.mz)
        spectrum_inty = spect.inty.tolist() if hasattr(spect.inty, "tolist") else list(spect.inty)

        # Determine MS level
        ms_level = spect.ms_level if spect.ms_level is not None else 1

        # Get energy if available
        energy = getattr(spect, "energy", None)

        # Determine spectrum type based on MS level
        spec_type = f"MS{ms_level}" if ms_level > 1 else "MS1"

        # Calculate spectrum length
        spec_len = len(spectrum_mz)

        return {
            "mgf_index": mgf_id,
            "title": title,
            "feature_id": id,
            "feature_uid": uid,
            "charge": charge,
            "pepmass": mz,
            "rtinseconds": rt,
            "mslevel": ms_level,
            "type": spec_type,
            "energy": energy,
            "spec_len": spec_len,
            "spec_mz": spectrum_mz,
            "spec_int": spectrum_inty,
        }

    # Collect all ion data
    ion_data = []
    skip = 0
    mgf_counter = 0
    self.logger.debug(f"Generating MGF data for {len(self.consensus_df)} consensus features...")

    # First, generate MS1 spectra for all consensus features using isotope data
    self.logger.debug("Generating MS1 spectra from isotope data...")
    for row in self.consensus_df.iter_rows(named=True):
        # Apply filtering at individual feature level for MS1 data
        consensus_uid = row["consensus_uid"]
        consensus_mz = row["mz"]
        consensus_rt = row["rt"]
        consensus_inty_mean = row.get("inty_mean", 0)

        if mz_start is not None and consensus_mz < mz_start:
            continue
        if mz_end is not None and consensus_mz > mz_end:
            continue
        if rt_start is not None and consensus_rt < rt_start:
            continue
        if rt_end is not None and consensus_rt > rt_end:
            continue

        # Create MS1 spectrum using isotope data
        iso_data = row.get("iso", None)

        if iso_data is not None and len(iso_data) > 0:
            # Use isotope data for spectrum
            spectrum_mz = [float(peak[0]) for peak in iso_data]
            spectrum_inty = [float(peak[1]) for peak in iso_data]
        else:
            # Use consensus mz and inty_mean as single peak
            spectrum_mz = [float(consensus_mz)]
            spectrum_inty = [float(consensus_inty_mean)]

        # Apply intensity minimum filter if specified
        if inty_min is not None and inty_min > 0:
            filtered_pairs = [
                (mz, inty) for mz, inty in zip(spectrum_mz, spectrum_inty, strict=False) if inty >= inty_min
            ]
            if filtered_pairs:
                spectrum_mz, spectrum_inty = zip(*filtered_pairs, strict=False)
                spectrum_mz = list(spectrum_mz)
                spectrum_inty = list(spectrum_inty)
            else:
                # If all peaks are below threshold, skip this feature
                continue

        mgf_counter += 1

        # Create MS1 spectrum object to use with create_ion_dict
        class SimpleSpectrum:
            def __init__(self, mz_list, inty_list):
                self.mz = np.array(mz_list)
                self.inty = np.array(inty_list)
                self.ms_level = 1
                self.energy = None

        ms1_spectrum = SimpleSpectrum(spectrum_mz, spectrum_inty)

        # Use create_ion_dict to ensure consistent schema
        ion_dict = create_ion_dict(
            f"uid:{consensus_uid}, rt:{consensus_rt:.2f}, mz:{consensus_mz:.4f}, MS1",
            row["consensus_id"],
            consensus_uid,
            consensus_mz,
            consensus_rt,
            safe_charge(row.get("charge_mean")),
            ms1_spectrum,
            mgf_counter,
        )

        if ion_dict is not None:
            ion_data.append(ion_dict)

    self.logger.debug(f"Generated {len(ion_data)} MS1 spectra from isotope data")

    # Now generate MS2 spectra if available
    if ms2_available and len(grouped) > 0:
        self.logger.debug(f"Processing MS2 data for {len(grouped)} consensus features with MS2...")
        tdqm_disable = self.log_level not in ["TRACE", "DEBUG", "INFO"]
        for _consensus_uid, cons_ms2 in tqdm(
            grouped,
            total=len(grouped),
            desc=f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]} | INFO     | {self.log_label}Feature",
            disable=tdqm_disable,
        ):
            # Use the first row for feature-level info
            row = cons_ms2.iloc[0]
            if mz_start is not None and row["mz"] < mz_start:
                continue
            if mz_end is not None and row["mz"] > mz_end:
                continue
            if rt_start is not None and row["rt"] < rt_start:
                continue
            if rt_end is not None and row["rt"] > rt_end:
                continue
            if len(cons_ms2) == 0:
                skip += 1
                continue

            if split_energy:
                energies = cons_ms2["energy"].unique()
                for e in energies:
                    cons_ms2_e = cons_ms2[cons_ms2["energy"] == e]
                    if selection == "best":
                        # Check if the filtered DataFrame is empty
                        if len(cons_ms2_e) == 0:
                            continue
                        idx = cons_ms2_e["prec_inty"].idxmax()
                        cons_ms2_e_row = cons_ms2_e.loc[idx]
                        spect = cons_ms2_e_row["spec"]
                        if spect is None:
                            skip += 1
                            continue
                        if centroid:
                            spect = spect.centroid()
                        if deisotope:
                            spect = spect.deisotope()
                        spect = filter_peaks(spect, inty_min=inty_min)
                        mgf_counter += 1
                        ion_dict = create_ion_dict(
                            f"uid:{cons_ms2_e_row['consensus_uid']}, rt:{cons_ms2_e_row['rt']:.2f}, mz:{cons_ms2_e_row['mz']:.4f}, energy:{e}, sample_uid:{cons_ms2_e_row['sample_uid']}, scan_id:{cons_ms2_e_row['scan_id']}",
                            cons_ms2_e_row["consensus_id"],
                            cons_ms2_e_row["consensus_uid"],
                            cons_ms2_e_row["mz"],
                            cons_ms2_e_row["rt"],
                            safe_charge(cons_ms2_e_row["charge_mean"]),
                            spect,
                            mgf_counter,
                        )
                        if ion_dict is not None:
                            ion_data.append(ion_dict)
                    else:
                        for row_e in cons_ms2_e.iter_rows(named=True):
                            spect = row_e["spec"]
                            if spect is None:
                                continue
                            if centroid:
                                spect = spect.centroid()
                            if deisotope:
                                spect = spect.deisotope()
                            spect = filter_peaks(spect, inty_min=inty_min)
                            mgf_counter += 1
                            ion_dict = create_ion_dict(
                                f"uid:{row_e['consensus_uid']}, rt:{row_e['rt']:.2f}, mz:{row_e['mz']:.4f}, energy:{e}, sample_uid:{row_e['sample_uid']}, scanid:{row_e['scan_id']}",
                                row_e["consensus_id"],
                                row_e["consensus_uid"],
                                row_e["mz"],
                                row_e["rt"],
                                safe_charge(row_e["charge_mean"]),
                                spect,
                                mgf_counter,
                            )
                            if ion_dict is not None:
                                ion_data.append(ion_dict)
            else:
                if selection == "best":
                    idx = cons_ms2["prec_inty"].idxmax()
                    cons_ms2_e_row = cons_ms2.loc[idx]
                    spect = cons_ms2_e_row["spec"]
                    if spect is None:
                        continue
                    if centroid:
                        spect = spect.centroid()
                    if deisotope:
                        spect = spect.deisotope()
                    spect = filter_peaks(spect, inty_min=inty_min)
                    mgf_counter += 1
                    ion_dict = create_ion_dict(
                        f"uid:{cons_ms2_e_row['consensus_uid']}, rt:{cons_ms2_e_row['rt']:.2f}, mz:{cons_ms2_e_row['mz']:.4f}, energy:{cons_ms2_e_row['energy']}, sample_uid:{cons_ms2_e_row['sample_uid']}, scan_id:{cons_ms2_e_row['scan_id']}",
                        cons_ms2_e_row["consensus_id"],
                        cons_ms2_e_row["consensus_uid"],
                        cons_ms2_e_row["mz"],
                        cons_ms2_e_row["rt"],
                        safe_charge(cons_ms2_e_row["charge_mean"]),
                        spect,
                        mgf_counter,
                    )
                    if ion_dict is not None:
                        ion_data.append(ion_dict)

                elif selection == "all":
                    if merge:
                        specs = [row_e["spec"] for row_e in cons_ms2.iter_rows(named=True) if row_e["spec"] is not None]
                        if not specs:
                            continue
                        spect = combine_peaks(specs)
                        if centroid:
                            spect = spect.denoise()
                            spect = spect.centroid()
                        if deisotope:
                            spect = spect.deisotope()
                        spect = filter_peaks(spect, inty_min=inty_min)
                        mgf_counter += 1
                        ion_dict = create_ion_dict(
                            f"uid:{row['consensus_uid']}, rt:{row['rt']:.2f}, mz:{row['mz']:.4f}, sample_uid:{row['sample_uid']}, scan_id:{row['scan_id']}",
                            row["consensus_id"],
                            row["consensus_uid"],
                            row["mz"],
                            row["rt"],
                            safe_charge(row["charge_mean"]),
                            spect,
                            mgf_counter,
                        )
                        if ion_dict is not None:
                            ion_data.append(ion_dict)
                    else:
                        for row_e in cons_ms2.iter_rows(named=True):
                            spect = row_e["spec"]
                            if spect is None:
                                continue
                            if centroid:
                                spect = spect.centroid()
                            if deisotope:
                                spect = spect.deisotope()
                            spect = filter_peaks(spect, inty_min=inty_min)
                            mgf_counter += 1
                            ion_dict = create_ion_dict(
                                f"uid:{row_e['consensus_uid']}, rt:{row_e['rt']:.2f}, mz:{row_e['mz']:.4f}, energy:{row_e['energy']}, sample_uid:{row_e['sample_uid']}, scan_id:{row_e['scan_id']}",
                                row_e["consensus_id"],
                                row_e["consensus_uid"],
                                row_e["mz"],
                                row_e["rt"],
                                safe_charge(row_e["charge_mean"]),
                                spect,
                                mgf_counter,
                            )
                            if ion_dict is not None:
                                ion_data.append(ion_dict)
    else:
        self.logger.info("Skipping MS2 data generation - no MS2 data available")

    self.logger.debug(f"Generated MGF data for {len(ion_data)} spectra (MS1 + MS2)")
    self.logger.debug(f"Skipped {skip} MS2 features due to missing data.")

    # Convert to Polars DataFrame
    if not ion_data:
        return pl.DataFrame()

    return pl.DataFrame(ion_data, infer_schema_length=None)


def export_mgf(self, **kwargs):
    """
    Export consensus features as MGF format for database searching.

    Parameters:
        **kwargs: Keyword arguments for export parameters. Can include:
            - An export_defaults instance to set all parameters at once
            - Individual parameter names and values (see export_defaults for details)

    Key Parameters:
        filename (str): Output MGF file name (default: "features.mgf").
        selection (str): "best" for first scan, "all" for every scan (default: "best").
        split_energy (bool): Process MS2 scans by unique energy (default: True).
        merge (bool): If selection="all", merge MS2 scans into one spectrum (default: False).
        mz_start (float): Minimum m/z for feature selection (default: None).
        mz_end (float): Maximum m/z for feature selection (default: None).
        rt_start (float): Minimum RT for feature selection (default: None).
        rt_end (float): Maximum RT for feature selection (default: None).
        centroid (bool): Apply centroiding to spectra (default: True).
        inty_min (float): Minimum intensity threshold (default: None).
        deisotope (bool): Apply deisotoping to spectra (default: True).
        verbose (bool): Enable verbose logging (default: False).
        precursor_trim (float): Precursor trimming value (default: -10).
        centroid_algo (str): Centroiding algorithm (default: "lmp").

    Returns:
        None: Writes MGF file to disk.
    """
    # Get mgf data as DataFrame
    from masster.study.export import _get_mgf_df

    mgf_data = _get_mgf_df(self, **kwargs)

    if mgf_data is None or len(mgf_data) == 0:
        self.logger.warning("No MGF data generated.")
        return

    # Get filename from parameters
    params = export_mgf_defaults()
    for key, value in kwargs.items():
        if isinstance(value, export_mgf_defaults):
            params = value
        else:
            if hasattr(params, key):
                params.set(key, value, validate=True)

    filename = params.get("filename")

    # Prepare output path
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    # Write MGF file
    with open(filename, "w", encoding="utf-8") as f:
        for row in mgf_data.iter_rows(named=True):
            # Write BEGIN IONS
            f.write("BEGIN IONS\n")

            # Write metadata
            if row["mgf_index"] is not None:
                f.write(f"INDEX={row['mgf_index']}\n")
            f.write(f"TITLE={row['title']}\n")
            f.write(f"FEATURE_ID={row['feature_id']}\n")
            f.write(f"FEATURE_UID={row['feature_uid']}\n")
            f.write(f"CHARGE={row['charge']}\n")
            f.write(f"PEPMASS={row['pepmass']}\n")
            f.write(f"RTINSECONDS={row['rtinseconds']}\n")
            f.write(f"MSLEVEL={row['mslevel']}\n")

            if row["energy"] is not None:
                f.write(f"ENERGY={row['energy']}\n")

            # Write spectrum data
            spectrum_mz = row["spec_mz"]
            spectrum_inty = row["spec_int"]
            for mz_val, inty in zip(spectrum_mz, spectrum_inty, strict=False):
                f.write(f"{mz_val:.5f} {inty:.0f}\n")

            # Write END IONS
            f.write("END IONS\n\n")

    self.logger.success(f"Exported {len(mgf_data)} spectra to {filename}")


def export_mztab(self, filename: str | None = None, include_mgf=True, **kwargs) -> None:
    """
    Export the study as a fully compliant mzTab-M file.

    Args:
        filename (str, optional): Path to the output mzTab-M file.
        title (str, optional): Human-readable title for the file.
        description (str, optional): Human-readable description.
        **kwargs: Additional metadata or export options.
    """

    def safe_str(value, default="null"):
        """Convert value to string, replacing empty strings with 'null'"""
        if value is None:
            return default
        str_val = str(value)
        return str_val if str_val.strip() != "" else default

    if filename is None:
        filename = "study.mztab"
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    # Get identification data if available
    id_data = None
    top_id_data = None
    full_id_data = None
    try:
        # Import here to avoid circular imports
        from masster.study.id import get_id

        # Get full enriched identification data for SOME section
        full_id_data = get_id(self)
        if full_id_data is not None and not full_id_data.is_empty():
            # Get top scoring identification for each consensus_uid for SML section
            top_id_data = (
                full_id_data.group_by("consensus_uid")
                .agg(pl.all().sort_by("score", descending=True).first())
                .sort("consensus_uid")
            )
            # Keep raw id_data for backward compatibility (if needed elsewhere)
            id_data = self.id_df if hasattr(self, "id_df") and self.id_df is not None else None
        else:
            self.logger.info("No identification data available for mzTab export")
    except Exception as e:
        self.logger.debug(f"Could not retrieve identification data: {e}")
        id_data = None
        top_id_data = None
        full_id_data = None

    # get mgf data only if requested
    mgf_data = None
    mgf_mapping: dict[str, list[int]] = {}
    if include_mgf:
        from masster.study.export import _get_mgf_df

        mgf_data = _get_mgf_df(self, **kwargs)
        # Create mapping from feature_uid to MGF indexes
        if mgf_data is not None and len(mgf_data) > 0:
            for row in mgf_data.iter_rows(named=True):
                feature_uid = row["feature_uid"]
                mgf_index = row["mgf_index"]
                if feature_uid not in mgf_mapping:
                    mgf_mapping[feature_uid] = []
                mgf_mapping[feature_uid].append(mgf_index)

    # --- Prepare MTD (metadata) section ---
    mtd_lines = []
    mtd_lines.append(
        f"COM\tfile generated by MASSter {get_version()} on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    )
    mtd_lines.append("\nMTD\tmzTab-version\t2.2.0-M")
    id = self.label if self.label else self.folder
    mtd_lines.append(f"MTD\tmzTab-id\t{id}")
    mtd_lines.append("")
    mtd_lines.append("MTD\tcv[1]-label\tMS")
    mtd_lines.append("MTD\tcv[1]-full_name\tPSI-MS controlled vocabulary")
    mtd_lines.append("MTD\tcv[1]-version\t4.1.199")
    mtd_lines.append(
        "MTD\tcv[1]-uri\thttps://raw.githubusercontent.com/HUPO-PSI/psi-ms-CV/master/psi-ms.obo",
    )
    mtd_lines.append("")
    mtd_lines.append(
        "MTD\tsmall_molecule-quantification_unit\t[MS, MS:1001844, MS1 feature area, ]",
    )
    mtd_lines.append(
        "MTD\tsmall_molecule_feature-quantification_unit\t[MS, MS:1001844, MS1 feature area, ]",
    )
    mtd_lines.append(
        "MTD\tsmall_molecule-identification_reliability\t[MS, MS:1002955, hr-ms compound identification confidence level, ]",
    )

    # Add identification confidence measures if identification data is available
    if full_id_data is not None:
        mtd_lines.append(
            "MTD\tid_confidence_measure[1]\t[MS, MS:1002888, small molecule confidence measure, ]",
        )
    else:
        mtd_lines.append(
            "MTD\tid_confidence_measure[1]\t[MS, MS:1002888, small molecule confidence measure, ]",
        )

    mtd_lines.append("")
    mtd_lines.append("MTD\tsoftware[1]\t[MS, MS:1003430, OpenMS, unknown]")
    mtd_lines.append(f"MTD\tsoftware[2]\t[MS, MS:1002878, MASSter, {get_version()}]")
    mtd_lines.append(
        "MTD\tquantification_method\t[MS, MS:1001834, LC-MS label-free quantitation analysis, ]",
    )
    mtd_lines.append("")

    # Database information - updated based on identification data
    if full_id_data is not None and hasattr(self, "lib_df") and self.lib_df is not None and not self.lib_df.is_empty():
        mtd_lines.append('MTD\tdatabase[1]\t[, , "compound library", ]')
        mtd_lines.append("MTD\tdatabase[1]-prefix\tcmpd")
        mtd_lines.append("MTD\tdatabase[1]-version\tUnknown")
        mtd_lines.append("MTD\tdatabase[1]-uri\thttps://pubchem.ncbi.nlm.nih.gov/")
    else:
        mtd_lines.append('MTD\tdatabase[1]\t[, , "PubChem", ]')
        mtd_lines.append("MTD\tdatabase[1]-prefix\tCID")
        mtd_lines.append("MTD\tdatabase[1]-version\tUnknown")
        mtd_lines.append("MTD\tdatabase[1]-uri\thttps://pubchem.ncbi.nlm.nih.gov/")

    # Get abundance matrix to determine the number of assays needed
    abundance_matrix = self.get_consensus_matrix()

    # Get sample columns (excluding consensus_uid)
    sample_columns = [col for col in abundance_matrix.columns if col != "consensus_uid"]
    n_assays = len(sample_columns)

    # Define samples, ms_runs, and assays based on the abundance matrix columns
    # Determine scan polarity based on study polarity
    study_polarity = getattr(self, "polarity", "positive")
    if study_polarity in ["negative", "neg"]:
        scan_polarity_cv = "[MS, MS:1000129, negative scan, ]"
    else:
        scan_polarity_cv = "[MS, MS:1000130, positive scan, ]"

    for i, sample_col in enumerate(sample_columns, 1):
        mtd_lines.append(f"\nMTD\tsample[{i}]\t{sample_col}")
        mtd_lines.append(f"MTD\tsample[{i}]-description\t{sample_col}")
        mtd_lines.append(f"MTD\tms_run[{i}]-location\tfile://unknown")
        mtd_lines.append(f"MTD\tms_run[{i}]-scan_polarity\t{scan_polarity_cv}")
        mtd_lines.append(f"MTD\tassay[{i}]\tAssay_{i}")
        mtd_lines.append(f"MTD\tassay[{i}]-sample_ref\tsample[{i}]")
        mtd_lines.append(f"MTD\tassay[{i}]-ms_run_ref\tms_run[{i}]")
    mtd_lines.append("")
    mtd_lines.append("MTD\tstudy_variable[1]\tundefined")
    mtd_lines.append("MTD\tstudy_variable[1]_refs\tundefined")
    # assay_refs = '|'.join([f"assay[{i}]" for i in range(1, len(self.samples_df)+1)])
    # mtd_lines.append(f"MTD\tstudy_variable[1]-assay_refs\t{assay_refs}")
    # mtd_lines.append("MTD\tstudy_variable[1]-description\tAll assays grouped (default)")
    with open(filename, "w", encoding="utf-8") as f:
        for line in mtd_lines:
            f.write(line + "\n")

    # --- SML (Small Molecule) table ---
    sml_lines = []
    sml_header = [
        "SMH",
        "SML_ID",
        "SMF_ID_REFS",
        "database_identifier",
        "chemical_formula",
        "smiles",
        "inchi",
        "chemical_name",
        "uri",
        "theoretical_neutral_mass",
        "adduct_ions",
        "reliability",
        "best_id_confidence_measure",
        "best_id_confidence_value",
        "opt_global_mgf_index",
    ]

    # round to int - handle both Polars and Pandas DataFrames
    if hasattr(abundance_matrix, "with_columns"):
        # Polars DataFrame
        numeric_cols = [col for col in abundance_matrix.columns if abundance_matrix[col].dtype.is_numeric()]
        abundance_matrix = abundance_matrix.with_columns(
            [abundance_matrix[col].round(0) for col in numeric_cols],
        )
    else:
        # Pandas DataFrame
        abundance_matrix = abundance_matrix.round(0)

    # Use the n_assays already calculated from abundance matrix columns
    sml_header += [f"abundance_assay[{i}]" for i in range(1, n_assays + 1)]
    sml_header += [
        "abundance_study_variable[1]",
        "abundance_variation_study_variable[1]",
    ]
    sml_lines.append("\t".join(sml_header))

    # get adducts from consensus_df['adduct_top'] - use the top-ranked adduct directly
    adduct_list = []
    for idx, row in enumerate(self.consensus_df.iter_rows(named=True), 1):
        adduct = "null"
        # Use adduct_top if available, otherwise fall back to null
        if "adduct_top" in row and row["adduct_top"] is not None:
            adduct = str(row["adduct_top"])
            # Replace ? with H for better mzTab compatibility
            adduct = adduct.replace("?", "H")

        adduct_list.append(adduct)

    for idx, row in enumerate(self.consensus_df.iter_rows(named=True), 1):
        # Get identification information for this consensus_uid if available
        consensus_uid = row["consensus_uid"]
        id_info = None
        if top_id_data is not None:
            id_matches = top_id_data.filter(pl.col("consensus_uid") == consensus_uid)
            if id_matches.height > 0:
                id_info = id_matches.row(0, named=True)

        # Populate identification fields
        database_identifier = "null"
        chemical_formula = "null"
        smiles_val = "null"
        inchi_val = "null"
        chemical_name = "null"
        best_id_confidence_measure = "null"
        best_id_confidence_value = "null"
        reliability = "4"  # Default: unknown compound
        theoretical_neutral_mass = "null"  # Only set when we have database identification

        if id_info:
            # Use cmpd_uid as database identifier with prefix
            if id_info.get("cmpd_uid") is not None:
                database_identifier = f"cmpd:{id_info['cmpd_uid']}"

            # Chemical formula
            if id_info.get("formula") is not None and id_info["formula"] != "":
                chemical_formula = safe_str(id_info["formula"])

            # SMILES
            if id_info.get("smiles") is not None and id_info["smiles"] != "":
                smiles_val = safe_str(id_info["smiles"])

            # InChI
            if id_info.get("inchi") is not None and id_info["inchi"] != "":
                inchi_val = safe_str(id_info["inchi"])

            # Chemical name
            if id_info.get("name") is not None and id_info["name"] != "":
                chemical_name = safe_str(id_info["name"])

            # Theoretical neutral mass - only from identification data, not consensus_df
            if id_info.get("neutral_mass") is not None:
                theoretical_neutral_mass = safe_str(id_info["neutral_mass"])
            elif id_info.get("mass") is not None:
                theoretical_neutral_mass = safe_str(id_info["mass"])

            # Identification confidence
            if id_info.get("matcher") is not None:
                best_id_confidence_measure = f"[MS, MS:1002888, {id_info['matcher']}, ]"

            if id_info.get("score") is not None:
                best_id_confidence_value = safe_str(id_info["score"])

            # Set reliability based on identification quality
            # Using mzTab-M hr-ms identification levels: 2a=compound match, 2b=library spectrum match, 3=compound class, 4=unknown
            if id_info.get("score", 0) >= 0.8:
                reliability = "2a"  # High confidence compound match
            elif id_info.get("score", 0) >= 0.5:
                reliability = "2b"  # Moderate confidence match
            elif id_info.get("score", 0) >= 0.2:
                reliability = "3"  # Compound class level
            else:
                reliability = "4"  # Unknown compound

        # Get MGF indexes for this consensus feature
        mgf_indexes = mgf_mapping.get(row["consensus_uid"], [])

        sml_row = [
            "SML",
            str(idx),
            str(idx),
            database_identifier,
            chemical_formula,
            smiles_val,
            inchi_val,
            chemical_name,
            safe_str(row.get("uri", "null")),
            theoretical_neutral_mass,  # Only set when database_identifier is not null
            adduct_list[idx - 1],
            reliability,
            best_id_confidence_measure,
            best_id_confidence_value,
            ",".join(map(str, mgf_indexes)) if mgf_indexes else "null",
        ]
        # Add abundance values for each assay
        consensus_uid = row["consensus_uid"]
        # Check if consensus_uid exists in the abundance_matrix (Polars)
        filtered_matrix = abundance_matrix.filter(
            pl.col("consensus_uid") == consensus_uid,
        )
        if filtered_matrix.height > 0:
            # Get the first (and should be only) matching row
            abundance_row = filtered_matrix.row(0, named=True)
            # Extract values excluding the consensus_uid column
            abundance_values = [abundance_row[col] for col in abundance_matrix.columns if col != "consensus_uid"]
            sml_row += [safe_str(val) if val is not None else "null" for val in abundance_values]

            # Calculate study variable statistics
            non_null_values = [val for val in abundance_values if val is not None]
            if non_null_values:
                abundance_study_variable = sum(non_null_values) / len(non_null_values)
                abundance_variation_study_variable = (
                    (sum((x - abundance_study_variable) ** 2 for x in non_null_values) / len(non_null_values)) ** 0.5
                    if len(non_null_values) > 1
                    else 0
                )
            else:
                abundance_study_variable = "null"
                abundance_variation_study_variable = "null"

            sml_row += [
                safe_str(abundance_study_variable),
                safe_str(abundance_variation_study_variable),
            ]
        else:
            sml_row += ["null"] * n_assays
            sml_row += ["null", "null"]  # Study variable columns
        sml_lines.append("\t".join(sml_row))
    with open(filename, "a", encoding="utf-8") as f:
        f.write("\n")
        for line in sml_lines:
            f.write(line + "\n")

    # --- SMF (Small Molecule Feature) table ---
    smf_lines = []
    smf_header = [
        "SFH",
        "SMF_ID",
        "SOME_ID_REFS",
        "SOME_ID_REF_ambiguity_code",
        "adduct_ion",
        "isotopomer",
        "exp_mass_to_charge",
        "charge",
        "retention_time_in_seconds",
        "retention_time_in_seconds_start",
        "retention_time_in_seconds_end",
    ]
    smf_header += [f"abundance_assay[{i}]" for i in range(1, n_assays + 1)]
    smf_header += [
        "abundance_study_variable[1]",
        "abundance_variation_study_variable[1]",
    ]
    smf_lines.append("\t".join(smf_header))

    # SMF table uses the same consensus features as SML, just different metadata
    for idx, row in enumerate(self.consensus_df.iter_rows(named=True), 1):
        # References to SOME entries - each SMF can reference multiple SOME entries for the same consensus_uid
        some_refs = "null"
        some_ambiguity = "null"
        consensus_uid = row["consensus_uid"]

        if full_id_data is not None:
            # Find all SOME entries for this consensus_uid
            some_matches = full_id_data.filter(pl.col("consensus_uid") == consensus_uid)
            if some_matches.height > 0:
                # Generate SOME IDs - we'll create a mapping in the SOME section
                # For now, use a simple approach based on consensus_uid and lib_uid
                some_ids = []
                for i, some_row in enumerate(some_matches.iter_rows(named=True)):
                    # Create a unique SOME ID based on consensus_uid and position
                    some_id_base = consensus_uid * 1000  # Ensure uniqueness across consensus features
                    some_id = some_id_base + i + 1
                    some_ids.append(str(some_id))

                if some_ids:
                    some_refs = "|".join(some_ids)
                    # Set ambiguity code: 1=ambiguous identification, 2=multiple evidence same molecule, 3=both
                    if len(some_ids) > 1:
                        # Check if all identifications point to the same compound
                        unique_cmpds = {
                            match["cmpd_uid"]
                            for match in some_matches.iter_rows(named=True)
                            if match.get("cmpd_uid") is not None
                        }
                        if len(unique_cmpds) > 1:
                            some_ambiguity = "1"  # Ambiguous identification
                        else:
                            some_ambiguity = "2"  # Multiple evidence for same molecule
                    else:
                        some_ambiguity = "null"

        # Format isotopomer according to mzTab-M specification
        iso_value = row.get("iso_mean", 0)
        if iso_value is not None and round(iso_value) != 0:
            isotopomer = f'[MS,MS:1002957,"isotopomer MS peak","+{round(iso_value)}"]'
        else:
            isotopomer = "null"

        smf_row = [
            "SMF",
            str(idx),
            some_refs,
            some_ambiguity,
            adduct_list[idx - 1],  # adduct_ion
            isotopomer,  # isotopomer formatted according to mzTab-M specification
            safe_str(row.get("mz", "null")),  # exp_mass_to_charge
            safe_str(
                row.get("adduct_charge_top", "null"),
            ),  # Use top-ranked adduct charge
            safe_str(row.get("rt", "null")),  # retention_time_in_seconds
            safe_str(row.get("retention_time_in_seconds_start", "null")),
            safe_str(row.get("retention_time_in_seconds_end", "null")),
        ]
        # Add abundance values for each assay - same as SML (Polars)
        consensus_uid = row["consensus_uid"]
        filtered_matrix = abundance_matrix.filter(
            pl.col("consensus_uid") == consensus_uid,
        )
        if filtered_matrix.height > 0:
            # Get the first (and should be only) matching row
            abundance_row = filtered_matrix.row(0, named=True)
            # Extract values excluding the consensus_uid column
            abundance_values = [abundance_row[col] for col in abundance_matrix.columns if col != "consensus_uid"]
            abundance_strings = [safe_str(val) if val is not None else "null" for val in abundance_values]
            smf_row += abundance_strings

            # Calculate study variable statistics (same as in SML section)
            non_null_values = [val for val in abundance_values if val is not None]
            if non_null_values:
                abundance_study_variable = sum(non_null_values) / len(non_null_values)
                abundance_variation_study_variable = (
                    (sum((x - abundance_study_variable) ** 2 for x in non_null_values) / len(non_null_values)) ** 0.5
                    if len(non_null_values) > 1
                    else 0
                )
            else:
                abundance_study_variable = "null"
                abundance_variation_study_variable = "null"

            smf_row += [
                safe_str(abundance_study_variable),
                safe_str(abundance_variation_study_variable),
            ]
        else:
            smf_row += ["null"] * n_assays
            smf_row += ["null", "null"]  # Study variable columns
        smf_lines.append("\t".join(smf_row))
    with open(filename, "a", encoding="utf-8") as f:
        f.write("\n")
        for line in smf_lines:
            f.write(line + "\n")

    # --- SOME (Small Molecule Evidence) table ---
    if full_id_data is not None and not full_id_data.is_empty():
        some_lines = []
        # Add comment about spectra_ref being dummy placeholders
        some_lines.append(
            "COM\tThe spectra_ref are dummy placeholders, as the annotation was based on aggregated data",
        )
        some_header = [
            "SHE",
            "SOME_ID",
            "evidence_input_id",
            "database_identifier",
            "chemical_formula",
            "smiles",
            "inchi",
            "chemical_name",
            "uri",
            "derivatized_form",
            "adduct_ion",
            "exp_mass_to_charge",
            "charge",
            "theoretical_mass_to_charge",
            "spectra_ref",
            "identification_method",
            "ms_level",
            "id_confidence_measure[1]",
            "rank",
        ]
        some_lines.append("\t".join(some_header))

        # Create SOME entries for all identification results using enriched data
        for consensus_uid in self.consensus_df.select("consensus_uid").to_series().unique():
            # Get consensus feature data for this consensus_uid
            consensus_feature_data = self.consensus_df.filter(
                pl.col("consensus_uid") == consensus_uid,
            )
            if consensus_feature_data.height == 0:
                continue
            consensus_row = consensus_feature_data.row(0, named=True)

            # Get all identification results for this consensus feature from enriched data
            some_matches = full_id_data.filter(pl.col("consensus_uid") == consensus_uid)

            if some_matches.height > 0:
                # Sort by score descending to maintain rank order
                some_matches = some_matches.sort("score", descending=True)

                for i, some_row in enumerate(some_matches.iter_rows(named=True)):
                    # Generate unique SOME_ID
                    some_id_base = consensus_uid * 1000
                    some_id = some_id_base + i + 1

                    # Create evidence input ID using consensus_uid:mz:rt format
                    consensus_mz = consensus_row.get("mz", 0)
                    consensus_rt = consensus_row.get("rt", 0)
                    evidence_id = f"consensus_uid={consensus_uid}:mz={consensus_mz:.4f}:rt={consensus_rt:.2f}"

                    # Database identifier - use db_id if available, otherwise fallback to cmpd_uid
                    db_id = "null"
                    if some_row.get("db_id") is not None and some_row["db_id"] != "":
                        db_id = safe_str(some_row["db_id"])
                    elif some_row.get("cmpd_uid") is not None:
                        db_id = f"cmpd:{some_row['cmpd_uid']}"

                    # Get adduct information
                    adduct_ion = "null"
                    if some_row.get("adduct") is not None and some_row["adduct"] != "":
                        adduct_ion = safe_str(some_row["adduct"])
                        # Replace ? with H for better mzTab compatibility
                        adduct_ion = adduct_ion.replace("?", "H")

                    # Spectra reference - reference to first ms_run with spectrum index 0
                    spectra_ref = "ms_run[1]:spectrum=0"

                    # Identification method
                    id_method = "[MS, MS:1002888, small molecule confidence measure, ]"
                    if some_row.get("matcher") is not None:
                        id_method = f"[MS, MS:1002888, {some_row['matcher']}, ]"

                    # MS level - assume MS1 for now
                    ms_level = "[MS, MS:1000511, ms level, 1]"

                    # Experimental mass-to-charge from consensus feature
                    exp_mz = safe_str(consensus_mz)

                    # Theoretical mass-to-charge from lib_df
                    theoretical_mz = "null"
                    if some_row.get("mz") is not None:  # This comes from lib_df via get_id() join
                        theoretical_mz = safe_str(some_row["mz"])

                    some_line = [
                        "SOME",
                        str(some_id),
                        evidence_id,
                        db_id,
                        safe_str(some_row.get("formula", "null")),
                        safe_str(some_row.get("smiles", "null")),
                        safe_str(some_row.get("inchi", "null")),
                        safe_str(some_row.get("name", "null")),
                        "null",  # uri - not available in current data
                        "null",  # derivatized_form
                        adduct_ion,
                        exp_mz,  # experimental m/z from consensus feature
                        safe_str(
                            consensus_row.get("adduct_charge_top", "1"),
                        ),  # Use consensus feature's top adduct charge
                        theoretical_mz,  # theoretical m/z from lib_df
                        spectra_ref,
                        id_method,
                        ms_level,
                        safe_str(some_row.get("score", "null")),
                        str(i + 1),  # rank within this consensus feature
                    ]
                    some_lines.append("\t".join(some_line))

        # Write SOME table
        with open(filename, "a", encoding="utf-8") as f:
            f.write("\n")
            for line in some_lines:
                f.write(line + "\n")

    # --- MGF table ---
    if include_mgf and mgf_data is not None and len(mgf_data) > 0:
        mgf_lines = []
        # Header
        mgf_header = [
            "COM",
            "MGH",
            "mgf_id",
            "prec_id",
            "prec_rt",
            "prec_mz",
            "prec_int",
            "energy",
            "level",
            "title",
            "spec_tic",
            "spec_len",
            "spec_mz",
            "spec_int",
        ]
        mgf_lines.append("\t".join(mgf_header))

        # Data rows
        for row in mgf_data.iter_rows(named=True):
            # Calculate spectrum TIC (total ion current) from the spectrum data
            spectrum_mz = row["spec_mz"]
            spectrum_inty = row["spec_int"]
            spec_tic = sum(spectrum_inty) if spectrum_inty else 0
            spec_len = row["spec_len"] if row["spec_len"] is not None else 0

            # Format spectrum data as pipe-separated strings
            spec_mz_str = "|".join([f"{mz:.4f}" for mz in spectrum_mz]) if spectrum_mz else ""
            spec_int_str = "|".join([f"{int(inty)}" for inty in spectrum_inty]) if spectrum_inty else ""

            mgf_row = [
                "COM",
                "MGF",
                str(row["mgf_index"]) if row["mgf_index"] is not None else "null",
                str(row["feature_id"]) if row["feature_id"] is not None else "null",
                f"{row['rtinseconds']:.2f}" if row["rtinseconds"] is not None else "null",
                f"{row['pepmass']:.4f}" if row["pepmass"] is not None else "null",
                "null",  # prec_int - not available in current data
                str(row["energy"]) if row["energy"] is not None else "null",
                str(row["mslevel"]) if row["mslevel"] is not None else "null",
                str(row["title"]) if row["title"] is not None else "null",
                f"{int(spec_tic)}" if spec_tic > 0 else "null",
                str(spec_len) if spec_len > 0 else "null",
                spec_mz_str if spec_mz_str else "null",
                spec_int_str if spec_int_str else "null",
            ]
            mgf_lines.append("\t".join(mgf_row))

        # Write MGF table
        with open(filename, "a", encoding="utf-8") as f:
            f.write("\n")
            for line in mgf_lines:
                f.write(line + "\n")

    self.logger.success(f"Exported mzTab-M to {filename}")


def export_xlsx(self, filename: str | None = None) -> None:
    """
    Export the study data to an Excel workbook with multiple worksheets.

    The Excel file contains five worksheets:
    - samples: Samples dataframe
    - consensus: Consensus features dataframe
    - identification: Identification results with library annotations (get_id)
    - gaps: Gaps matrix showing filled vs non-filled features (get_gaps_matrix)
    - matrix: Consensus matrix with samples as columns (get_consensus_matrix)

    Args:
        filename (str, optional): Path to the output Excel file. Defaults to "study.xlsx"
                                in the study folder.
    """
    try:
        import openpyxl
    except ImportError:
        self.logger.error(
            "openpyxl package is required for Excel export. Install with: pip install openpyxl",
        )
        return

    # Set default filename
    if filename is None:
        filename = "study.xlsx"

    # Make filename absolute if not already
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    self.logger.debug("Exporting study to Excel...")

    # Prepare data for export in the desired order
    from collections import OrderedDict

    worksheets = OrderedDict()

    # 1. Samples dataframe (first worksheet)
    if self.samples_df is not None and not self.samples_df.is_empty():
        samples_pandas = self.samples_df.to_pandas()
        worksheets["samples"] = samples_pandas
        self.logger.debug(f"Added samples worksheet with {len(samples_pandas)} rows")
    else:
        self.logger.warning("samples_df is empty or None, skipping worksheet")

    # 2. Consensus dataframe (renamed to 'consensus')
    if self.consensus_df is not None and not self.consensus_df.is_empty():
        consensus_pandas = self.consensus_df.to_pandas()
        worksheets["consensus"] = consensus_pandas
        self.logger.debug(
            f"Added consensus worksheet with {len(consensus_pandas)} rows",
        )
    else:
        self.logger.warning("consensus_df is empty or None, skipping worksheet")

    # 3. Identification results
    try:
        from masster.study.id import get_id

        id_df = get_id(self)
        if id_df is not None and not id_df.is_empty():
            id_pandas = id_df.to_pandas()
            worksheets["identification"] = id_pandas
            self.logger.debug(
                f"Added identification worksheet with {len(id_pandas)} rows",
            )
        else:
            self.logger.warning(
                "get_id() returned empty data, skipping identification worksheet",
            )
    except Exception as e:
        self.logger.debug(
            f"Error getting identification data: {e}. Skipping identification worksheet.",
        )

    # 4. Gaps matrix (filled vs non-filled features)
    try:
        gaps_df = self.get_gaps_matrix()
        if gaps_df is not None and not gaps_df.is_empty():
            gaps_pandas = gaps_df.to_pandas()
            worksheets["gaps"] = gaps_pandas
            self.logger.debug(
                f"Added gaps worksheet with {len(gaps_pandas)} rows",
            )
        else:
            self.logger.warning(
                "get_gaps_matrix() returned empty data, skipping gaps worksheet",
            )
    except Exception as e:
        self.logger.debug(
            f"Error getting gaps data: {e}. Skipping gaps worksheet.",
        )

    # 5. Consensus matrix (last worksheet)
    try:
        matrix_df = self.get_consensus_matrix()
        if matrix_df is not None and not matrix_df.is_empty():
            matrix_pandas = matrix_df.to_pandas()
            worksheets["matrix"] = matrix_pandas
            self.logger.debug(f"Added matrix worksheet with {len(matrix_pandas)} rows")
        else:
            self.logger.warning(
                "get_consensus_matrix() returned empty data, skipping matrix worksheet",
            )
    except Exception as e:
        self.logger.error(f"Error getting consensus matrix: {e}")

    # Check if we have any data to export
    if not worksheets:
        self.logger.error("No data available to export to Excel")
        return

    # Write to Excel file
    try:
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            for sheet_name, data in worksheets.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)
                self.logger.debug(
                    f"Written worksheet '{sheet_name}' with shape {data.shape}",
                )

        self.logger.success(f"Study exported to {filename}")

    except Exception as e:
        self.logger.error(f"Error writing Excel file: {e}")


def export_parquet(self, filename: str | None = None) -> None:
    """
    Export the study data to multiple Parquet files with different suffixes.

    The export creates separate Parquet files for each dataset:
    - <filename>_samples.parquet: Samples dataframe
    - <filename>_consensus.parquet: Consensus features dataframe
    - <filename>_identification.parquet: Identification results with library annotations
    - <filename>_matrix.parquet: Consensus matrix with samples as columns

    Args:
        filename (str, optional): Base name for the output files. Defaults to "study"
                                 in the study folder.
    """
    # Set default filename
    if filename is None:
        filename = "study"

    # Make filename absolute path if not already (without extension)
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    self.logger.debug(f"Exporting study to Parquet files with filename: {filename}")

    exported_files = []

    # 1. Samples dataframe
    if self.samples_df is not None and not self.samples_df.is_empty():
        samples_file = f"{filename}_samples.parquet"
        try:
            self.samples_df.write_parquet(samples_file)
            exported_files.append(samples_file)
            self.logger.debug(
                f"Exported samples to {samples_file} ({self.samples_df.height} rows)",
            )
        except Exception as e:
            self.logger.error(f"Error writing samples parquet file: {e}")
    else:
        self.logger.warning(
            "samples_df is empty or None, skipping samples parquet file",
        )

    # 2. Consensus dataframe
    if self.consensus_df is not None and not self.consensus_df.is_empty():
        consensus_file = f"{filename}_consensus.parquet"
        try:
            # Create a copy of consensus_df for parquet export
            consensus_export_df = self.consensus_df.clone()
            
            # Handle Object dtype columns that can't be serialized to parquet
            for col in consensus_export_df.columns:
                if consensus_export_df[col].dtype == pl.Object:
                    if col == "iso":
                        # Convert numpy arrays to string representation for parquet compatibility
                        # This preserves the data while making it parquet-serializable
                        consensus_export_df = consensus_export_df.with_columns([
                            pl.col("iso").map_elements(
                                lambda x: str(x.tolist()) if x is not None else None,
                                return_dtype=pl.String
                            ).alias("iso")
                        ])
                    else:
                        # For other Object columns, convert to string representation
                        consensus_export_df = consensus_export_df.with_columns([
                            pl.col(col).map_elements(
                                lambda x: str(x) if x is not None else None,
                                return_dtype=pl.String
                            ).alias(col)
                        ])
            
            consensus_export_df.write_parquet(consensus_file)
            exported_files.append(consensus_file)
            self.logger.debug(
                f"Exported consensus to {consensus_file} ({consensus_export_df.height} rows)",
            )
        except Exception as e:
            self.logger.error(f"Error writing consensus parquet file: {e}")
    else:
        self.logger.warning(
            "consensus_df is empty or None, skipping consensus parquet file",
        )

    # 3. Identification results
    try:
        from masster.study.id import get_id

        id_df = get_id(self)
        if id_df is not None and not id_df.is_empty():
            identification_file = f"{filename}_identification.parquet"
            try:
                id_df.write_parquet(identification_file)
                exported_files.append(identification_file)
                self.logger.debug(
                    f"Exported identification to {identification_file} ({id_df.height} rows)",
                )
            except Exception as e:
                self.logger.error(f"Error writing identification parquet file: {e}")
        else:
            self.logger.warning(
                "get_id() returned empty data, skipping identification parquet file",
            )
    except Exception as e:
        self.logger.warning(
            f"Error getting identification data: {e}. Skipping identification parquet file.",
        )

    # 4. Consensus matrix
    try:
        matrix_df = self.get_consensus_matrix()
        if matrix_df is not None and not matrix_df.is_empty():
            matrix_file = f"{filename}_matrix.parquet"
            try:
                matrix_df.write_parquet(matrix_file)
                exported_files.append(matrix_file)
                self.logger.debug(
                    f"Exported matrix to {matrix_file} ({matrix_df.height} rows)",
                )
            except Exception as e:
                self.logger.error(f"Error writing matrix parquet file: {e}")
        else:
            self.logger.warning(
                "get_consensus_matrix() returned empty data, skipping matrix parquet file",
            )
    except Exception as e:
        self.logger.error(f"Error getting consensus matrix: {e}")

    # Report results
    if exported_files:
        self.logger.success(f"Study exported to {len(exported_files)} Parquet files.")
    else:
        self.logger.error("No Parquet files were created - no data available to export")
