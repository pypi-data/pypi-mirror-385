import traceback
import requests
import json
import csv
import logging
import os
import time
import threading
import concurrent.futures
import subprocess
import sys
import importlib.metadata
import tempfile
import textwrap
import signal
import re
import gzip
import shutil
import builtins
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pathlib import Path
from datetime import datetime, timedelta
from functools import lru_cache
from tqdm import tqdm
from queue import Queue
from argparse import ArgumentParser, RawDescriptionHelpFormatter
from jinja2 import Environment, FileSystemLoader
from collections import Counter, defaultdict
from packaging import version as pkg_version
from vuln_checker import __version__

# Global log file handle
log_file = None

# Custom print function
def print(*args, **kwargs):
    builtins.print(*args, file=log_file, **kwargs)
    if log_file is not None:
        log_file.flush()

# Redirect uncaught exceptions to log file
def log_exception_handler(exc_type, exc_value, exc_traceback):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    builtins.print("Uncaught exception:", file=log_file)
    traceback.print_exception(exc_type, exc_value, exc_traceback, file=log_file)
    log_file.flush()

sys.excepthook = log_exception_handler

logging.basicConfig(level=logging.WARNING)
try:
    import tomllib
except ModuleNotFoundError:
    import tomli as tomllib

def signal_handler(sig, frame):
    print("\n⛔ Received Ctrl+C, exiting gracefully...")
    sys.exit(1)

signal.signal(signal.SIGINT, signal_handler)

# ==== NVD FEED DOWNLOAD/UPDATE FUNCTIONS ====

def download_nvdcve_year(year, feed_dir="nvd_feeds", force_update=False):
    url = f"https://nvd.nist.gov/feeds/json/cve/2.0/nvdcve-2.0-{year}.json.gz"
    feed_path = Path(feed_dir)
    feed_path.mkdir(exist_ok=True)
    gz_file = feed_path / f"nvdcve-2.0-{year}.json.gz"
    json_file = feed_path / f"nvdcve-2.0-{year}.json"
    
    if json_file.exists() and not force_update:
        file_modified_time = datetime.fromtimestamp(json_file.stat().st_mtime)
        if (datetime.now() - file_modified_time) < timedelta(hours=24):
            print(f"✅ JSON feed for {year} is fresh (less than 24h old), skipping download")
            return True
        else:
            print(f"⚠️ Feed for {year} is older than 24h, refreshing download")
    elif force_update:
        print(f"⚠️ Force update enabled: re-downloading feed for {year}")

    print(f"📥 Downloading {url} ...")
    try:
        response = requests.get(url, stream=True)
        if response.status_code != 200:
            print(f"❌ Failed to download feed for {year}: HTTP {response.status_code}")
            return False
        
        with open(gz_file, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        
        with gzip.open(gz_file, 'rb') as f_in, open(json_file, 'wb') as f_out:
            shutil.copyfileobj(f_in, f_out)
        
        gz_file.unlink()
        print(f"✅ Decompressed and saved feed for {year}")
        return True
    except Exception as e:
        print(f"❌ Error downloading {year} feed: {e}")
        return False

def download_all_nvd_feeds(feed_dir="nvd_feeds", force_update=False):
    current_year = datetime.now().year
    for year in range(2002, current_year + 1):
        success = download_nvdcve_year(year, feed_dir, force_update=force_update)
        if not success:
            print(f"⚠️ Skipping {year} due to download error")

def nvd_feeds_need_update(feed_dir="nvd_feeds", max_age_hours=24):
    feed_path = Path(feed_dir)
    feed_files = list(feed_path.glob("nvdcve-2.0-*.json"))
    if not feed_files:
        print("📥 No feed files found. Download required.")
        return True
    now = datetime.now()
    for f in feed_files:
        last_mod = datetime.fromtimestamp(f.stat().st_mtime)
        if now - last_mod > timedelta(hours=max_age_hours):
            print(f"⚠️ {f.name} last updated at {last_mod.strftime('%Y-%m-%d %H:%M:%S')}. Needs refresh.")
            return True
    print("✅ NVD feed files are fresh.")
    return False

def load_excluded_cpes(filename="excluded_cpes.txt"):
    # Get directory of the current script
    base_path = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_path, filename)

    cpes = set()
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                cpe = line.strip()
                if cpe and not cpe.startswith("#"):
                    cpes.add(cpe)
    except FileNotFoundError:
        print(f"Warning: Excluded CPE file not found: {file_path}")
    except Exception as e:
        print(f"Error reading excluded CPE file: {e}")
    return cpes

# Usage
EXCLUDED_CPES = load_excluded_cpes()

def is_excluded_cpe(cpe_uri: str) -> bool:
    cpe_uri = cpe_uri.strip()
    if cpe_uri in EXCLUDED_CPES:
        return True

    for excluded in EXCLUDED_CPES:
        excluded = excluded.strip()
        if excluded.endswith("*"):
            prefix = excluded[:-1]
            if cpe_uri.startswith(prefix):
                return True
    return False

# ==== LOCAL FEED MANAGER ====

class LocalFeedManager:
    def __init__(self, feed_dir="nvd_feeds"):
        self.feed_dir = Path(feed_dir)
        self.index = defaultdict(list)  # (vendor, product) -> list of CVEs
        self.loaded = False

    def load_feeds(self):
        if self.loaded:
            return
        self.index.clear()
        json_files = sorted(self.feed_dir.glob("nvdcve-2.0-*.json"))
        if not json_files:
            print("❌ No NVD feed files found. Please run with --update-feeds first.")
            return
        
        total_cves = 0
        for json_file in json_files:
            print(f"📥 Loading NVD feed file: {json_file.name}")
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    vulns = data.get("vulnerabilities", [])
                    for vuln in vulns:
                        vuln_with_feed = dict(vuln)
                        vuln_with_feed['nvd_feed_file'] = json_file.name
                        cve_id = vuln_with_feed.get("cve", {}).get("id", "unknown")
                        configs = vuln_with_feed.get("cve", {}).get("configurations", [])
                        for config in configs:
                            for node in config.get("nodes", []):
                                for cpe_match in node.get("cpeMatch", []):
                                    if not cpe_match.get("vulnerable") or not cpe_match.get("criteria"):
                                        continue
                                    cpe_uri = cpe_match.get("criteria")
                                    cpe_parts = cpe_uri.split(":")
                                    if len(cpe_parts) >= 6:
                                        vendor = cpe_parts[3].lower()
                                        product = cpe_parts[4].lower()
                                        self.index[(vendor, product)].append({
                                            "vuln": vuln_with_feed,
                                            "cpe_match": cpe_match
                                        })
                        total_cves += 1
            except Exception as e:
                print(f"❌ Error loading {json_file}: {e}")
        
        print(f"✅ Loaded {total_cves} CVEs from {len(json_files)} local NVD feeds, indexed {len(self.index)} vendor-product pairs")
        self.loaded = True

    def search_cves_for_cpe(self, target_cpe, severity=None):
        if not self.loaded:
            return []
        
        # ADD THIS CHECK AT THE START
        if is_excluded_cpe(target_cpe):
            print(f"⚠️  Skipping excluded CPE: {target_cpe}")
            return []
        
        matched_cves = []
        target_vendor, target_product, target_version = None, None, None
        parts = target_cpe.split(":")
        if len(parts) >= 6:
            target_vendor = parts[3].lower()
            target_product = parts[4].lower()
            target_version = parts[5]
        else:
            print(f"Invalid CPE format: {target_cpe}")
            return []

        # Get relevant CVEs from index
        cve_entries = self.index.get((target_vendor, target_product), [])
        # print(f"[DEBUG] Found {len(cve_entries)} potential CVEs for {target_vendor}:{target_product}")

        for entry in cve_entries:
            vuln = entry["vuln"]
            cpe_match = entry["cpe_match"]
            cve_obj = vuln.get("cve", {})
            cve_id = cve_obj.get("id", "unknown")
            
            # Severity filter
            if severity:
                severity_match = False
                metrics = cve_obj.get("metrics", {})
                for metric_key in ["cvssMetricV31", "cvssMetricV30", "cvssMetricV2"]:
                    metric_list = metrics.get(metric_key, [])
                    for metric in metric_list:
                        cvss_data = metric.get("cvssData", {})
                        base_severity = cvss_data.get("baseSeverity", "").upper()
                        if base_severity == severity.upper():
                            severity_match = True
                            break
                    if severity_match:
                        break
                if not severity_match:
                    # print(f"[DEBUG] CVE {cve_id} skipped due to severity mismatch (wanted {severity})")
                    continue
            
            # Version validation
            is_version_vulnerable = False
            cpe_uri = cpe_match.get("criteria")
            cpe_parts = cpe_uri.split(":")
            cpe_version = cpe_parts[5] if len(cpe_parts) >= 6 else None
            
            # Initialize version range variables
            start_v = None
            start_ex_v = None
            end_v = None
            end_ex_v = None
            
            try:
                # Clean and parse target version
                clean_target_version = re.sub(r'[-_].*$', '', target_version)  # e.g., "33.0.0-jre" -> "33.0.0"
                target_v = pkg_version.parse(clean_target_version)
                # print(f"[DEBUG] Parsed target version {target_version} as {clean_target_version} for CVE {cve_id}")
                
                # Parse CVE version and ranges
                if cpe_version and cpe_version != "*":
                    try:
                        cpe_v = pkg_version.parse(cpe_version)
                        if target_v == cpe_v:
                            is_version_vulnerable = True
                            # print(f"[DEBUG] Exact version match: {clean_target_version} == {cpe_version} for CVE {cve_id}")
                        # else:
                        #     print(f"[DEBUG] No exact version match: {clean_target_version} != {cpe_version} for CVE {cve_id}")
                    except (pkg_version.InvalidVersion, ValueError):
                        # print(f"[DEBUG] Failed to parse CPE version {cpe_version} for CVE {cve_id} - checking ranges")
                        continue
                
                # Parse version ranges
                version_start = cpe_match.get("versionStartIncluding")
                version_start_ex = cpe_match.get("versionStartExcluding")
                version_end = cpe_match.get("versionEndIncluding")
                version_end_ex = cpe_match.get("versionEndExcluding")
                
                try:
                    if version_start:
                        start_v = pkg_version.parse(version_start)
                    if version_start_ex:
                        start_ex_v = pkg_version.parse(version_start_ex)
                    if version_end:
                        end_v = pkg_version.parse(version_end)
                    if version_end_ex:
                        end_ex_v = pkg_version.parse(version_end_ex)
                except (pkg_version.InvalidVersion, ValueError) as e:
                    # print(f"[DEBUG] Failed to parse version range for CVE {cve_id}: {e} - skipping")
                    continue
                
                # Range-based checks
                lower_match = True
                upper_match = True
                
                if start_v:
                    lower_match = target_v >= start_v
                elif start_ex_v:
                    lower_match = target_v > start_ex_v
                
                if end_v:
                    upper_match = target_v <= end_v
                elif end_ex_v:
                    upper_match = target_v < end_ex_v
                
                if start_v or start_ex_v or end_v or end_ex_v:
                    is_version_vulnerable = lower_match and upper_match
                    if not is_version_vulnerable:
                        range_str = f"{version_start or version_start_ex or 'N/A'}–{version_end or version_end_ex or 'N/A'}"
                        # print(f"[DEBUG] Version {clean_target_version} outside range {range_str} for CVE {cve_id} - skipping")
                        continue
                elif cpe_version == "*":
                    # print(f"[DEBUG] Wildcard CPE with no ranges for CVE {cve_id} - assuming NOT vulnerable")
                    continue
                
            except (pkg_version.InvalidVersion, ValueError) as e:
                # print(f"[DEBUG] Version parsing failed for {target_version} in CVE {cve_id}: {e} - skipping")
                continue
            
            if is_version_vulnerable:
                feed_file = vuln.get('nvd_feed_file', 'unknown')
                print(f"[DEBUG] ✅ Matched CVE {cve_id} from {feed_file} for {target_cpe} (version {clean_target_version})")
                matched_cves.append(vuln)
        
        return matched_cves

# Global feed manager
local_feed_manager = None

class CPECache:
    def __init__(self, cache_file="cve_cache.json"):
        self.cache_file = cache_file
        self.lock = threading.Lock()
        self.cache = self.load_cache()

    def load_cache(self):
        try:
            with open(self.cache_file, 'r') as f:
                return json.load(f)
        except Exception:
            return {}

    def save_cache(self):
        with self.lock:
            cache_copy = dict(self.cache)
            with open(self.cache_file, 'w') as f:
                json.dump(cache_copy, f, indent=2)

    def get_cves(self, cpe_uri, severity=None):
        key = f"{cpe_uri}_{severity or 'ALL'}"
        return self.cache.get(key)

    def set_cves(self, cpe_uri, severity, cves):
        key = f"{cpe_uri}_{severity or 'ALL'}"
        cached_cves = []
        for cve in cves:
            cached_cve = dict(cve)
            if 'nvd_feed_file' not in cached_cve:
                cached_cve['nvd_feed_file'] = 'unknown'
            cve_id = cached_cve.get('cve', {}).get('id', 'unknown')
            print(f"[CACHING] CVE {cve_id} with nvd_feed_file = {cached_cve['nvd_feed_file']}")
            cached_cves.append(cached_cve)
        
        self.cache[key] = {
            'cves': cached_cves,
            'timestamp': time.time(),
            'cpe': cpe_uri,
            'nvd_feed_file': [cve.get('nvd_feed_file', 'unknown') for cve in cached_cves]
        }
        self.save_cache()

    def is_fresh(self, entry, max_age_hours=24):
        return (time.time() - entry['timestamp']) < (max_age_hours * 3600)

cpe_cache = CPECache()

def fetch_cves_from_local_feeds(cpe_uri, severity=None):
    if not local_feed_manager or not local_feed_manager.loaded:
        print("❌ Local feeds not loaded. Cannot fetch CVEs.")
        return []
    
    # print(f"🔍 Searching local feeds for: {cpe_uri}")
    results = local_feed_manager.search_cves_for_cpe(cpe_uri, severity)
    for cve in results:
        cve_id = cve.get('cve', {}).get('id', 'unknown')
        # print(f"[DEBUG FETCH] CVE {cve_id} has nvd_feed_file = {cve.get('nvd_feed_file', 'unknown')}")
    return results

def fetch_cves_cached(cpe_uri, severity=None, bypass_cache=False):
    if not bypass_cache:
        cached_entry = cpe_cache.get_cves(cpe_uri, severity)
        if cached_entry and cpe_cache.is_fresh(cached_entry):
            print(f"[CACHE] Using cached CVEs for {cpe_uri} (severity: {severity or 'ALL'})")
            return cached_entry['cves']
    
    cves = fetch_cves_from_local_feeds(cpe_uri, severity)
    cpe_cache.set_cves(cpe_uri, severity, cves)
    return cves

def fetch_cves_cached_with_enrichment(cpe_uri, severity=None, component_name="", component_version="", purl="", parent_jar=None):
    cached_entry = cpe_cache.get_cves(cpe_uri, severity)
    if cached_entry and cpe_cache.is_fresh(cached_entry):
        # print(f"[CACHE] Using cached enriched CVEs for {cpe_uri} (severity: {severity or 'ALL'})")
        return cached_entry['cves']

    raw_cves = fetch_cves_from_local_feeds(cpe_uri, severity)
    final_cves = []
    for cve in raw_cves:
        cve_id = cve.get('cve', {}).get('id', 'unknown')
        feed_file = cve.get('nvd_feed_file', 'unknown')
        # print(f"[DEBUG RAW] CVE {cve_id} from feed file = {feed_file}")
        
        product_field = f"{component_name}:{component_version}"
        if parent_jar and parent_jar not in (component_name, "", component_version):
            product_field += f" ({parent_jar})"
        
        enriched_cve = {
            "cve": cve.get("cve", cve),
            "product": product_field,
            "sbom": True,
            "purl": purl or "N/A",
            "cpe_used": cpe_uri,
            "cpe_source": "local_feeds",
            "nvd_feed_file": feed_file
        }
        # print(f"[DEBUG ENRICHED] CVE {cve_id} with feed file = {enriched_cve['nvd_feed_file']}")
        final_cves.append(enriched_cve)
    
    cpe_cache.set_cves(cpe_uri, severity, final_cves)
    return final_cves

def fetch_cves_for_sbom_cpe(cpe_uri, severity, component_name, component_version, purl, parent_jar):
    return fetch_cves_cached_with_enrichment(
        cpe_uri, severity, component_name, component_version, purl, parent_jar
    )

# ==== SBOM PARSING ====

def parse_cyclonedx_sbom_enhanced(sbom_file_path):
    try:
        with open(sbom_file_path, 'r', encoding='utf-8') as f:
            sbom_data = json.load(f)
        if sbom_data.get('bomFormat', '').lower() != 'cyclonedx':
            print(f"❌ File {sbom_file_path} is not a valid CycloneDX SBOM")
            return []
        components = []
        sbom_components = sbom_data.get('components', [])
        print(f"📋 Found {len(sbom_components)} components in SBOM")
        for component in sbom_components:
            name = component.get('name', 'unknown')
            version = component.get('version', 'unknown')
            purl = component.get('purl', '')
            primary_cpe = component.get('cpe', '')
            additional_cpes = []
            parent_jar = None
            for prop in component.get('properties', []):
                if prop.get('name') == 'syft:cpe23':
                    val = prop.get('value', '')
                    if val and val.startswith('cpe:2.3:a:'):
                        additional_cpes.append(val)
                if 'syft:metadata:virtualPath' in prop.get('name'):
                    path_val = prop.get('value', '')
                    jar_name = path_val.split("/")[-1].replace('.jar', '')
                    parent_jar = jar_name
            component_type = component.get('type', 'library')
            if component_type in ['library', 'application', 'framework']:
                components.append((name, version, primary_cpe, purl, additional_cpes, parent_jar))
        return components
    except Exception as e:
        print(f"❌ Error parsing SBOM file: {e}")
        return []

def lookup_cpe_from_txt(product, version, cpes_file):
    product = product.lower()
    version = version.lower()
    if not os.path.exists(cpes_file):
        return None
    with open(cpes_file, "r", encoding="utf-8") as f:
        for line in f:
            cpe = line.strip()
            if not cpe.startswith("cpe:2.3:a:"):
                continue
            parts = cpe.split(":")
            if len(parts) >= 6:
                cpe_product = parts[4].lower()
                cpe_version = parts[5].lower()
                if cpe_product == product and cpe_version == version:
                    return cpe
    return None

def process_single_component(component_data, args):
    component_name, component_version, primary_cpe, purl, additional_cpes, parent_jar = component_data
    print(f"🔍 Processing component: {component_name}:{component_version}")
    all_component_cves = []
    severities = [s.strip().upper() for s in args.severity.split(",")] if args.severity else [None]
    
    cpe_list = [c for c in [primary_cpe] + additional_cpes if c and c.startswith('cpe:2.3:a:')]
    local_cpe = lookup_cpe_from_txt(component_name, component_version, args.cpes_file or "cpes_list.txt")
    if local_cpe:
        cpe_list.append(local_cpe)
    
    used_cpe = None
    for cpe in cpe_list:
        cves = fetch_cves_for_sbom_cpe(cpe, None, component_name, component_version, purl, parent_jar)
        if cves:
            used_cpe = cpe
            filtered_cves = []
            for cve in cves:
                cve_severity = cve.get("cve", {}).get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseSeverity", "UNKNOWN").upper()
                if severities != [None] and cve_severity not in severities:
                    # print(f"[DEBUG] CVE {cve.get('cve', {}).get('id')} skipped due to severity filter ({cve_severity} not in {severities})")
                    continue
                filtered_cves.append(cve)
            all_component_cves.extend(filtered_cves)
            break
    
    if all_component_cves:
        print(f"🔴 Found {len(all_component_cves)} CVEs for component {component_name}:{component_version}")
    else:
        print(f"✅ No CVEs found for component {component_name}:{component_version}")
    
    return all_component_cves

def process_sbom_vulnerabilities_parallel(args, max_workers=5):
    if not args.sbom:
        return []
    sbom_file_path = args.sbom
    if not os.path.exists(sbom_file_path):
        print(f"❌ SBOM file not found: {sbom_file_path}")
        return []
    print(f"🔍 Loading SBOM file: {sbom_file_path}")
    components = parse_cyclonedx_sbom_enhanced(sbom_file_path)

    if getattr(args, "skip_search", False):
        components = [c for c in components if c[2] or c[4]]
    if getattr(args, "max_components", None):
        components = components[:args.max_components]

    print(f"🚦 {len(components)} components to analyze after filtering")

    global cpe_cache
    if hasattr(args, "cache_file") and args.cache_file:
        cpe_cache = CPECache(cache_file=args.cache_file)
    else:
        cpe_cache = CPECache()

    workers = max_workers if getattr(args, "fast", False) else 1
    all_cves = []
    lock = threading.Lock()

    def worker(component):
        local_cves = process_single_component(component, args)
        with lock:
            all_cves.extend(local_cves)

    print(f"🚀 Starting parallel processing with {workers} workers")
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [executor.submit(worker, c) for c in components]
            for _ in tqdm(concurrent.futures.as_completed(futures), total=len(components), desc="Processing SBOM components", unit="component"):
                pass
    except KeyboardInterrupt:
        print("\n⛔ Scan aborted by user")
        sys.exit(1)

    print(f"\n📊 SBOM analysis complete: {len(components)} components processed, {len(all_cves)} total CVEs found.")
    return all_cves

# ==== OUTPUT FUNCTIONS ====

def output_results_sbom_enhanced(cves, output_format="json", output_file=None):
    if not cves:
        print("⚠️ No CVEs found.")
        return
    if output_format == "json":
        enriched = []
        for item in cves:
            cve = item["cve"]
            enriched.append({
                "product": item.get("product"),
                "purl": item.get("purl"),
                "cpe_used": item.get("cpe_used"),
                "cpe_source": item.get("cpe_source"),
                "nvd_feed_file": item.get("nvd_feed_file", "unknown"),
                "id": {"value": cve["id"], "url": f"https://nvd.nist.gov/vuln/detail/{cve['id']}"},
                "published": cve.get("published"),
                "lastModified": cve.get("lastModified"),
                "cvssScore": cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseScore", "N/A"),
                "severity": cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseSeverity", "N/A"),
                "description": cve.get("descriptions", [{}])[0].get("value", "N/A")
            })
        with open(output_file or "sbom_vulnerabilities_enhanced.json", "w", encoding="utf-8") as f:
            json.dump(enriched, f, indent=2)
        print(f"✅ JSON report written to {output_file or 'sbom_vulnerabilities_enhanced.json'}")
    elif output_format == "csv":
        keys = ["product", "purl", "cpe_used", "cpe_source", "nvd_feed_file", "id", "published", "lastModified", "cvssScore", "severity", "description"]
        with open(output_file or "sbom_vulnerabilities_enhanced.csv", "w", newline='', encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            for item in cves:
                cve = item["cve"]
                writer.writerow({
                    "product": item.get("product"),
                    "purl": item.get("purl"),
                    "cpe_used": item.get("cpe_used"),
                    "cpe_source": item.get("cpe_source"),
                    "nvd_feed_file": item.get("nvd_feed_file", "unknown"),
                    "id": f'=HYPERLINK("https://nvd.nist.gov/vuln/detail/{cve["id"]}", "{cve["id"]}")',
                    "published": cve.get("published"),
                    "lastModified": cve.get("lastModified"),
                    "cvssScore": cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseScore", "N/A"),
                    "severity": cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseSeverity", "N/A"),
                    "description": cve.get("descriptions", [{}])[0].get("value", "N/A")
                })
        print(f"✅ CSV report written to {output_file or 'sbom_vulnerabilities_enhanced.csv'}")

def output_results(cves, output_format="json", output_file=None):
    if not cves:
        print("⚠️ No CVEs found.")
        return
    if output_format == "json":
        enriched = []
        for item in cves:
            cve = item.get("cve", item)
            enriched.append({
                "product": item.get("product", "Unknown"),
                "cpe_used": item.get("cpe_used", "N/A"),
                "cpe_source": item.get("cpe_source", "N/A"),
                "nvd_feed_file": item.get("nvd_feed_file", "unknown"),
                "id": {"value": cve["id"], "url": f"https://nvd.nist.gov/vuln/detail/{cve['id']}"},
                "published": cve.get("published"),
                "lastModified": cve.get("lastModified"),
                "cvssScore": cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseScore", "N/A"),
                "severity": cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseSeverity", "N/A"),
                "description": cve.get("descriptions", [{}])[0].get("value", "N/A")
            })
        with open(output_file or "vulnerabilities.json", "w", encoding="utf-8") as f:
            json.dump(enriched, f, indent=2)
        print(f"✅ JSON report written to {output_file or 'vulnerabilities.json'}")
    elif output_format == "csv":
        keys = ["product", "cpe_used", "cpe_source", "nvd_feed_file", "id", "published", "lastModified", "cvssScore", "severity", "description"]
        with open(output_file or "vulnerabilities.csv", "w", newline='', encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            for item in cves:
                cve = item.get("cve", item)
                writer.writerow({
                    "product": item.get("product", "Unknown"),
                    "cpe_used": item.get("cpe_used", "N/A"),
                    "cpe_source": item.get("cpe_source", "N/A"),
                    "nvd_feed_file": item.get("nvd_feed_file", "unknown"),
                    "id": f'=HYPERLINK("https://nvd.nist.gov/vuln/detail/{cve["id"]}", "{cve["id"]}")',
                    "published": cve.get("published"),
                    "lastModified": cve.get("lastModified"),
                    "cvssScore": cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseScore", "N/A"),
                    "severity": cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseSeverity", "N/A"),
                    "description": cve.get("descriptions", [{}])[0].get("value", "N/A")
                })
        print(f"✅ CSV report written to {output_file or 'vulnerabilities.csv'}")

def generate_html_report(cves, output_file="cve_report.html"):
    script_dir = Path(__file__).parent
    template_dir = script_dir / "templates"
    template_path = template_dir / "template.html"

    if not template_path.is_file():
        print(f"❌ Error: 'template.html' not found in {template_dir}")
        sys.exit(1)

    env = Environment(loader=FileSystemLoader(str(template_dir)))
    try:
        template = env.get_template("template.html")
    except Exception as e:
        print(f"❌ Error loading template.html: {e}")
        sys.exit(1)

    rows = []
    severity_counter = Counter()
    for item in cves:
        cve = item["cve"]
        product = item.get("product", "Unknown")
        cve_id = cve["id"]
        url = f"https://nvd.nist.gov/vuln/detail/{cve_id}"
        metrics = cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {})
        severity = metrics.get("baseSeverity", "UNKNOWN")
        score = metrics.get("baseScore", "N/A")
        description = cve.get("descriptions", [{}])[0].get("value", "N/A")
        published = cve.get("published", "N/A")
        severity_counter[severity] += 1
        rows.append({
            "product": product,
            "id": cve_id,
            "url": url,
            "severity": severity,
            "score": score,
            "description": description,
            "published": published,
            "nvd_feed_file": item.get("nvd_feed_file", "unknown")
        })
    html = template.render(cves=rows, severity_counts=severity_counter)
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"📄 HTML report written to {output_file}")

def output_results_excel(cves, output_file="vulnerabilities.xlsx"):
    if not cves:
        print("No CVEs to output.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    headers = [
        "Vuln ID", "Severity", "Source",
        "Component Name", "Library Version",
        "Description", "Dev Comments"
    ]
    ws.append(headers)

    # Set header font bold
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_num).font = Font(bold=True)

    for idx, item in enumerate(cves, start=2):
        cve = item.get("cve", item)
        cve_id = cve.get("id", "")

        severity = cve.get("metrics", {}).get("cvssMetricV31", [{}])[0].get("cvssData", {}).get("baseSeverity", "")
        source = item.get("cpe_source", "")
        component_name = item.get("product", "")
        library_version = ""
        if ":" in component_name:
            # Assuming product is "name:version"
            parts = component_name.split(":")
            if len(parts) >= 2:
                component_name = parts[0]
                library_version = parts[1]

        description = cve.get("descriptions", [{}])[0].get("value", "")
        dev_comments = ""  # You can populate from your data if available

        # Vuln ID with hyperlink
        cve_cell = ws.cell(row=idx, column=1, value=cve_id)
        if cve_id:
            url = f"https://nvd.nist.gov/vuln/detail/{cve_id}"
            cve_cell.hyperlink = url
            cve_cell.style = "Hyperlink"

        ws.cell(row=idx, column=2, value=severity)
        ws.cell(row=idx, column=3, value=source)
        ws.cell(row=idx, column=4, value=component_name)
        ws.cell(row=idx, column=5, value=library_version)
        ws.cell(row=idx, column=6, value=description)
        ws.cell(row=idx, column=7, value=dev_comments)

    # Adjust column widths for readability
    for col_num in range(1, len(headers) + 1):
        max_length = max(len(str(cell.value or "")) for cell in ws[get_column_letter(col_num)])
        ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 5, 50)

    wb.save(output_file)
    print(f"Excel report written to {output_file}")

def get_project_version():
    toml_path = Path(__file__).parent / "pyproject.toml"
    try:
        with open(toml_path, "rb") as f:
            data = tomllib.load(f)
            return data["project"]["version"]
    except Exception:
        return __version__

def get_installed_version(package_name="vuln-checker"):
    try:
        return importlib.metadata.version(package_name)
    except importlib.metadata.PackageNotFoundError:
        return None

def get_latest_version_from_pypi(package_name="vuln-checker"):
    try:
        response = requests.get(f"https://pypi.org/pypi/{package_name}/json", timeout=5)
        if response.status_code == 200:
            return response.json()["info"]["version"]
    except Exception as e:
        print(f"⚠️ Failed to fetch version from PyPI: {e}")
    return None

def check_for_upgrade(package_name="vuln-checker", auto_confirm=False):
    installed = get_installed_version(package_name)
    latest = get_latest_version_from_pypi(package_name)
    if not installed:
        print(f"⚠️ '{package_name}' not installed via pip. Cannot upgrade.")
        return
    if not latest:
        print("❌ Could not retrieve latest version from PyPI.")
        return
    v_installed = pkg_version.parse(installed)
    v_latest = pkg_version.parse(latest)
    if v_installed > v_latest:
        print(f"🛑 You have a newer version ({v_installed}) than PyPI ({v_latest}). Skipping upgrade.")
        return
    elif v_installed == v_latest:
        print(f"✅ You're already using the latest version: {installed}")
        return
    print(f"🚀 New version available: {latest} (current: {installed})")
    if auto_confirm or input("Do you want to upgrade? [y/N]: ").strip().lower() == 'y':
        print("🔁 Exiting and upgrading in a subprocess...")
        upgrade_script = textwrap.dedent(f"""
            import subprocess, time, sys
            time.sleep(2)
            subprocess.run([
                r"{sys.executable}", "-m", "pip", "install",
                "--upgrade", "--force-reinstall", "--no-cache-dir", "--user", "{package_name}"
            ], check=True)
        """)
        with tempfile.NamedTemporaryFile('w', suffix='.py', delete=False) as tmp:
            tmp.write(upgrade_script)
            script_path = tmp.name
        subprocess.Popen([sys.executable, script_path], creationflags=subprocess.DETACHED_PROCESS)
        print("✅ Upgrade process started in background. Please re-run the tool after upgrade completes.")
        sys.exit(0)

def main():
    global log_file
    with open("vuln-checker.log", "w", encoding="utf-8") as f:
        f.write("")
    log_file = open("vuln-checker.log", "a", encoding="utf-8")
    print(f"vuln-checker version: {__version__}")

    start_time = time.time()
    parser = ArgumentParser(
        description="""\
🔍 vuln-checker: Search CVEs by CPE product/version with Enhanced SBOM Support

Features:
- Uses local NVD JSON feeds (no API dependency)
- Auto-updates feeds if older than 24 hours
- Parse CycloneDX-JSON SBOMs and detect vulnerabilities
- Fetch matching CPEs using product & versions
- Batch mode to scan multiple product,versions via CSV
- Export results in JSON, CSV, or HTML
- Supports excluding false-positive CPEs via an external excluded_cpes.txt file placed alongside the main script.

Examples:
    vuln-checker --update-feeds
    vuln-checker --input-csv products.csv --severity High,Critical --format html --output report.html
    vuln-checker --products "jquery:1.11.3 lodash:3.5.0" --format csv --output output.csv
    vuln-checker --products "jquery:1.11.3" --format json
    vuln-checker --products "jquery:1.11.3,1.11.5" --format json
    vuln-checker --products "jquery:1.11.3,1.11.5 lodash:3.5.0" --format json
    vuln-checker --cpes-file cpes_list.txt --format json
    vuln-checker --cpes-file cpes_list.txt --severity High --format csv --output high_cves.csv
    vuln-checker --cpes-file cpes_list.txt --severity Critical,High --format html --output report.html
    vuln-checker --sbom myfile.json --format html
    vuln-checker --update-feeds --sbom myfile.json --format json
    vuln-checker --sbom myfile.json --fast --max-workers 5 --severity Critical,High --format csv
""",
        formatter_class=RawDescriptionHelpFormatter
    )
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--input-csv", help="CSV file of products and versions")
    group.add_argument("--products", help="Products and versions string")
    group.add_argument("--cpes-file", help="Path to file of CPEs")
    group.add_argument("--sbom", help="Path to CycloneDX SBOM file (JSON)")
    
    parser.add_argument("--fast", action="store_true", help="Fast mode: parallel processing with caching")
    parser.add_argument("--max-workers", type=int, default=5, help="Number of parallel workers")
    parser.add_argument("--max-components", type=int, help="Limit number of components to process")
    parser.add_argument("--skip-search", action="store_true", help="Skip components without existing CPEs")
    parser.add_argument('--update-feeds', action='store_true', help='Download the latest NVD JSON feeds')
    parser.add_argument('--force-update', action='store_true', help='Force re-download all feeds even if fresh')
    parser.add_argument("--feed-dir", default="nvd_feeds", help="Directory for NVD feeds (default: nvd_feeds)")
    parser.add_argument("--cache-file", default="cve_cache.json", help="CVE cache file location")
    parser.add_argument("--no-clear-cache", action="store_true", help="Prevent clearing the CVE cache")
    parser.add_argument("--migrate-cache", action="store_true", help="Migrate and normalize existing CVE cache")
    parser.add_argument("--severity", help="Severity filter (comma separated)")
    parser.add_argument("--format", choices=["json", "csv", "html", "excel"], default="json")
    parser.add_argument("--output", help="Report output filename")
    parser.add_argument("--upgrade", action="store_true", help="Upgrade vuln-checker")
    parser.add_argument("--yes", action="store_true", help="Auto-confirm prompts")
    parser.add_argument('--version', action='version', version=f'%(prog)s {__version__}')

    args = parser.parse_args()

    if args.upgrade:
        check_for_upgrade(auto_confirm=args.yes)
        return
    
    if args.update_feeds or nvd_feeds_need_update(args.feed_dir, max_age_hours=24):
        print("🔄 Updating NVD feeds...")
        download_all_nvd_feeds(feed_dir=args.feed_dir, force_update=args.force_update)
        print("✅ Feeds updated.")

    global local_feed_manager
    local_feed_manager = LocalFeedManager(feed_dir=args.feed_dir)
    local_feed_manager.load_feeds()
    
    if not local_feed_manager.loaded:
        print("❌ Could not load NVD feeds. Please check feed directory or run --update-feeds")
        return

    global cpe_cache
    if args.cache_file:
        cpe_cache = CPECache(cache_file=args.cache_file)
    
    if args.migrate_cache:
        print("Cache migration not yet implemented")
        return

    if not args.no_clear_cache:
        if os.path.exists(args.cache_file or "cve_cache.json"):
            os.remove(args.cache_file or "cve_cache.json")
            print("✅ CVE cache cleared.")

    if not args.input_csv and not args.products and not args.cpes_file and not args.sbom:
        parser.error("One of --input-csv, --products, --cpes-file, or --sbom is required.")

    all_cves = []

    if args.sbom:
        workers = args.max_workers if args.fast else 1
        all_cves = process_sbom_vulnerabilities_parallel(args, max_workers=workers)
    elif args.input_csv:
        all_cves = process_csv_vulnerabilities(args)
    elif args.products:
        all_cves = process_products_vulnerabilities(args)
    elif args.cpes_file:
        all_cves = process_cpes_file_vulnerabilities(args)

    if args.format == 'html':
        generate_html_report(all_cves, args.output or 'cve_report.html')
    elif args.format == 'excel':
        output_results_excel(all_cves, args.output or 'vulnerabilities.xlsx')
    else:
        if args.sbom:
            output_results_sbom_enhanced(all_cves, args.format, args.output)
        else:
            output_results(all_cves, args.format, args.output)

    elapsed = time.time() - start_time
    elapsed_minutes = elapsed / 60
    print(f"\n⏱️ Script completed in {elapsed:.2f} seconds ({elapsed_minutes:.2f} minutes)")
    log_file.flush()

def process_csv_vulnerabilities(args):
    all_cves = []
    try:
        with open(args.input_csv, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                product = (row.get('product') or row.get('name') or row.get('package') or '').strip()
                versions = (row.get('versions') or row.get('version') or '').strip()
                
                if not product:
                    continue
                
                version_list = [v.strip() for v in versions.split(',') if v.strip()]
                if not version_list:
                    continue
                
                for version in version_list:
                    cpe = f"cpe:2.3:a:{product}:{product}:{version}:*:*:*:*:*:*:*"
                    severities = [s.strip().upper() for s in args.severity.split(",")] if args.severity else [None]
                    
                    for sev in severities:
                        cves = fetch_cves_cached(cpe, sev)
                        for cve in cves:
                            cve["product"] = f"{product}:{version}"
                            cve["cpe_used"] = cpe
                            cve["cpe_source"] = "csv_input"
                            cve["nvd_feed_file"] = cve.get("nvd_feed_file", "unknown")
                        all_cves.extend(cves)
                    print(f"📦 {product}:{version} - {len(cves)} CVEs found")
    except Exception as e:
        print(f"❌ Error processing CSV file: {e}")
    return all_cves

def process_products_vulnerabilities(args):
    all_cves = []
    product_groups = args.products.split()
    
    for group in product_groups:
        parts = group.split(':')
        if len(parts) < 2:
            continue
        product = parts[0]
        versions = ':'.join(parts[1:]).split(',')
        
        for version in versions:
            version = version.strip()
            if not version:
                continue
            cpe = f"cpe:2.3:a:{product}:{product}:{version}:*:*:*:*:*:*:*"
            severities = [s.strip().upper() for s in args.severity.split(",")] if args.severity else [None]
            
            for sev in severities:
                cves = fetch_cves_cached(cpe, sev)
                for cve in cves:
                    cve["product"] = f"{product}:{version}"
                    cve["cpe_used"] = cpe
                    cve["cpe_source"] = "products_input"
                    cve["nvd_feed_file"] = cve.get("nvd_feed_file", "unknown")
                all_cves.extend(cves)
            print(f"📦 {product}:{version} - {len(cves)} CVEs found")
    return all_cves

def process_cpes_file_vulnerabilities(args):
    all_cves = []
    try:
        with open(args.cpes_file, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                cpe = line.strip()
                if not cpe or not cpe.startswith('cpe:2.3:'):
                    continue
                severities = [s.strip().upper() for s in args.severity.split(",")] if args.severity else [None]
                
                for sev in severities:
                    cves = fetch_cves_cached(cpe, sev)
                    for cve in cves:
                        cve["product"] = cpe
                        cve["cpe_used"] = cpe
                        cve["cpe_source"] = "cpes_file"
                        cve["nvd_feed_file"] = cve.get("nvd_feed_file", "unknown")
                    all_cves.extend(cves)
                print(f"📦 Line {line_num}: {cpe} - {len(cves)} CVEs found")
    except Exception as e:
        print(f"❌ Error processing CPEs file: {e}")
    return all_cves

if __name__ == "__main__":
    try:
        main()
    finally:
        if log_file is not None:
            log_file.close()