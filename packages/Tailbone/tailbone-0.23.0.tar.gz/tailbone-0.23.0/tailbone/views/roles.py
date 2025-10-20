# -*- coding: utf-8; -*-
################################################################################
#
#  Rattail -- Retail Software Framework
#  Copyright © 2010-2024 Lance Edgar
#
#  This file is part of Rattail.
#
#  Rattail is free software: you can redistribute it and/or modify it under the
#  terms of the GNU General Public License as published by the Free Software
#  Foundation, either version 3 of the License, or (at your option) any later
#  version.
#
#  Rattail is distributed in the hope that it will be useful, but WITHOUT ANY
#  WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
#  FOR A PARTICULAR PURPOSE.  See the GNU General Public License for more
#  details.
#
#  You should have received a copy of the GNU General Public License along with
#  Rattail.  If not, see <http://www.gnu.org/licenses/>.
#
################################################################################
"""
Role Views
"""

import os

from sqlalchemy import orm
from openpyxl.styles import Font, PatternFill

from rattail.db.model import Role
from rattail.excel import ExcelWriter

import colander
from deform import widget as dfwidget
from webhelpers2.html import HTML

from tailbone import grids
from tailbone.db import Session
from tailbone.views.principal import PrincipalMasterView, PermissionsRenderer


class RoleView(PrincipalMasterView):
    """
    Master view for the Role model.
    """
    model_class = Role
    has_versions = True
    touchable = True

    labels = {
        'adminish': "Admin-ish",
        'sync_me': "Sync Attrs & Perms",
    }

    grid_columns = [
        'name',
        'session_timeout',
        'sync_me',
        'sync_users',
        'node_type',
        'notes',
    ]

    form_fields = [
        'name',
        'adminish',
        'session_timeout',
        'notes',
        'sync_me',
        'sync_users',
        'node_type',
        'users',
        'permissions',
    ]

    def configure_grid(self, g):
        super().configure_grid(g)

        # name
        g.filters['name'].default_active = True
        g.filters['name'].default_verb = 'contains'
        g.set_sort_defaults('name')
        g.set_link('name')

        # notes
        g.set_renderer('notes', self.render_short_notes)

    def render_short_notes(self, role, field):
        value = getattr(role, field)
        if value is None:
            return ""
        if len(value) < 100:
            return value
        return HTML.tag('span', title=value,
                        c="{}...".format(value[:100]))

    def editable_instance(self, role):
        """
        We must prevent edit for certain built-in roles etc., depending on
        current user's permissions.
        """
        # role with node type specified, can only be edited from a
        # node of the same type
        if role.node_type and role.node_type != self.rattail_config.node_type():
            return False

        app = self.get_rattail_app()
        auth = app.get_auth_handler()

        # only "root" can edit Administrator
        if role is auth.get_role_administrator(self.Session()):
            return self.request.is_root

        # only "admin" can edit "admin-ish" roles
        if role.adminish:
            return self.request.is_admin

        # can edit Authenticated only if user has permission
        if role is auth.get_role_authenticated(self.Session()):
            return self.has_perm('edit_authenticated')

        # can edit Guest only if user has permission
        if role is auth.get_role_anonymous(self.Session()):
            return self.has_perm('edit_guest')

        # current user can edit their own roles, only if they have permission
        user = self.request.user
        if user and role in user.roles:
            return self.has_perm('edit_my')

        return True

    def deletable_instance(self, role):
        """
        We must prevent deletion for all built-in roles.
        """
        # role with node type specified, can only be edited from a
        # node of the same type
        if role.node_type and role.node_type != self.rattail_config.node_type():
            return False

        app = self.get_rattail_app()
        auth = app.get_auth_handler()

        if role is auth.get_role_administrator(self.Session()):
            return False
        if role is auth.get_role_authenticated(self.Session()):
            return False
        if role is auth.get_role_anonymous(self.Session()):
            return False

        # only "admin" can delete "admin-ish" roles
        if role.adminish:
            return self.request.is_admin

        # current user can delete their own roles, only if they have permission
        user = self.request.user
        if user and role in user.roles:
            return self.has_perm('edit_my')

        return True

    def unique_name(self, node, value):
        model = self.model
        query = self.Session.query(model.Role)\
                            .filter(model.Role.name == value)
        if self.editing:
            role = self.get_instance()
            query = query.filter(model.Role.uuid != role.uuid)
        if query.count():
            raise colander.Invalid(node, "Name must be unique")

    def configure_form(self, f):
        super().configure_form(f)
        role = f.model_instance
        app = self.get_rattail_app()
        auth = app.get_auth_handler()

        # name
        f.set_validator('name', self.unique_name)

        # adminish
        if self.request.is_admin:
            f.set_helptext('adminish',
                           "If checked, only Administrators may add/remove "
                           "users for the role.")
        else:
            f.remove('adminish')

        # session_timeout
        f.set_renderer('session_timeout', self.render_session_timeout)
        if self.editing and role is auth.get_role_anonymous(self.Session()):
            f.set_readonly('session_timeout')

        # sync_me, node_type
        if not self.creating:
            include = True
            if role is auth.get_role_administrator(self.Session()):
                include = False
            elif role is auth.get_role_authenticated(self.Session()):
                include = False
            elif role is auth.get_role_anonymous(self.Session()):
                include = False
            if not include:
                f.remove('sync_me', 'sync_users', 'node_type')
            else:
                if not self.has_perm('edit_node_sync'):
                    f.set_readonly('sync_me')
                    f.set_readonly('sync_users')
                    f.set_readonly('node_type')

        # notes
        f.set_type('notes', 'text_wrapped')

        # users
        if self.viewing:
            f.set_renderer('users', self.render_users)
        else:
            f.remove('users')

        # permissions
        self.tailbone_permissions = self.get_available_permissions()
        f.set_renderer('permissions', PermissionsRenderer(request=self.request,
                                                          permissions=self.tailbone_permissions))
        f.set_node('permissions', colander.Set())
        f.set_widget('permissions', PermissionsWidget(
            permissions=self.tailbone_permissions))
        if self.editing:
            granted = []
            for groupkey in self.tailbone_permissions:
                for key in self.tailbone_permissions[groupkey]['perms']:
                    if auth.has_permission(self.Session(), role, key,
                                           include_anonymous=False,
                                           include_authenticated=False):
                        granted.append(key)
            f.set_default('permissions', granted)
        elif self.deleting:
            f.remove_field('permissions')

    def render_users(self, role, field):
        app = self.get_rattail_app()
        auth = app.get_auth_handler()

        if role is auth.get_role_anonymous(self.Session()):
            return ("The guest role is implied for all anonymous users, "
                    "i.e. when not logged in.")

        if role is auth.get_role_authenticated(self.Session()):
            return ("The authenticated role is implied for all users, "
                    "but only when logged in.")

        route_prefix = self.get_route_prefix()
        permission_prefix = self.get_permission_prefix()
        factory = self.get_grid_factory()
        g = factory(
            self.request,
            key=f'{route_prefix}.users',
            data=[],
            columns=[
                'full_name',
                'username',
                'active',
            ],
            sortable=True,
            sorters={'full_name': True, 'username': True, 'active': True},
            default_sortkey='full_name',
        )

        if self.request.has_perm('users.view'):
            g.actions.append(self.make_action('view', icon='eye'))
        if self.request.has_perm('users.edit'):
            g.actions.append(self.make_action('edit', icon='edit'))

        return HTML.literal(
            g.render_table_element(data_prop='usersData'))

    def get_available_permissions(self):
        """
        Should return a dictionary with all "available" permissions.  The
        result of this will vary depending on the current user, because a user
        is only allowed to "manage" permissions which they themselves have been
        granted.

        In other words the return value will be the "full set" of permissions,
        if the current user is an admin; otherwise it will be the "subset" of
        permissions which the current user has been granted.
        """
        # get all known permissions from settings cache
        permissions = self.request.registry.settings.get('wutta_permissions', {})

        # admin user gets to manage all permissions
        if self.request.is_admin:
            return permissions

        # when viewing, we allow all permissions to be exposed for all users
        if self.viewing:
            return permissions

        # non-admin user can only manage permissions they've been granted
        # TODO: it seems a bit ugly, to "rebuild" permission groups like this,
        # but not sure if there's a better way?
        available = {}
        for gkey, group in permissions.items():
            for pkey, perm in group['perms'].items():
                if self.request.has_perm(pkey):
                    if gkey not in available:
                        available[gkey] = {
                            'key': gkey,
                            'label': group['label'],
                            'perms': {},
                        }
                    available[gkey]['perms'][pkey] = perm
        return available

    def render_session_timeout(self, role, field):
        app = self.get_rattail_app()
        auth = app.get_auth_handler()
        if role is auth.get_role_anonymous(self.Session()):
            return "(not applicable)"
        if role.session_timeout is None:
            return ""
        return str(role.session_timeout)

    def objectify(self, form, data=None):
        """
        Supplements the default logic, as follows:

        The role is updated as per usual, and then we also invoke
        :meth:`update_permissions()` in order to correctly handle that part,
        i.e. ensure the user can't modify permissions which they do not have.
        """
        if data is None:
            data = form.validated
        role = super().objectify(form, data)
        self.update_permissions(role, data['permissions'])
        return role

    def update_permissions(self, role, permissions):
        """
        Update the given role's permissions, according to those specified.
        Note that this will not simply "clobber" the role's existing
        permissions, but rather each "available" permission (depends on current
        user) will be examined individually, and updated as needed.
        """
        app = self.get_rattail_app()
        auth = app.get_auth_handler()

        available = self.tailbone_permissions
        for gkey, group in available.items():
            for pkey, perm in group['perms'].items():
                if pkey in permissions:
                    auth.grant_permission(role, pkey)
                else:
                    auth.revoke_permission(role, pkey)

    def template_kwargs_view(self, **kwargs):
        app = self.get_rattail_app()
        auth = app.get_auth_handler()
        model = self.model
        role = kwargs['instance']
        if role.users:
            users = sorted(role.users, key=lambda u: u.username)
            actions = [
                self.make_action('view', icon='zoomin',
                                 url=lambda r, i: self.request.route_url('users.view', uuid=r.uuid))
            ]
            kwargs['users'] = grids.Grid(self.request,
                                         data=users,
                                         columns=['username', 'active'],
                                         model_class=model.User,
                                         actions=actions)
        else:
            kwargs['users'] = None

        kwargs['guest_role'] = auth.get_role_anonymous(self.Session())
        kwargs['authenticated_role'] = auth.get_role_authenticated(self.Session())

        role = kwargs['instance']
        if role not in (kwargs['guest_role'], kwargs['authenticated_role']):
            users_data = []
            for user in role.users:
                users_data.append({
                    'uuid': user.uuid,
                    'full_name': user.display_name,
                    'username': user.username,
                    'active': "Yes" if user.active else "No",
                    '_action_url_view': self.request.route_url('users.view',
                                                               uuid=user.uuid),
                    '_action_url_edit': self.request.route_url('users.edit',
                                                               uuid=user.uuid),
                })
            kwargs['users_data'] = users_data

        return kwargs

    def before_delete(self, role):
        app = self.get_rattail_app()
        auth = app.get_auth_handler()
        admin = auth.get_role_administrator(self.Session())
        guest = auth.get_role_anonymous(self.Session())
        authenticated = auth.get_role_authenticated(self.Session())
        if role in (admin, guest, authenticated):
            self.request.session.flash("You may not delete the {} role.".format(role.name), 'error')
            return self.redirect(self.request.get_referrer(default=self.request.route_url('roles')))

    def find_principals_with_permission(self, session, permission):
        app = self.get_rattail_app()
        model = self.model
        auth = app.get_auth_handler()

        # TODO: this should search Permission table instead, and work backward to Role?
        all_roles = session.query(model.Role)\
                           .order_by(model.Role.name)\
                           .options(orm.joinedload(model.Role._permissions))
        roles = []
        for role in all_roles:
            if auth.has_permission(session, role, permission, include_anonymous=False):
                roles.append(role)
        return roles

    def find_by_perm_configure_results_grid(self, g):
        g.append('name')
        g.set_link('name')

    def find_by_perm_normalize(self, role):
        data = super().find_by_perm_normalize(role)

        data['name'] = role.name

        return data

    def download_permissions_matrix(self):
        """
        View which renders the complete role / permissions matrix data into an
        Excel spreadsheet, and returns that file.
        """
        app = self.get_rattail_app()
        model = self.model
        auth = app.get_auth_handler()

        roles = self.Session.query(model.Role)\
                            .order_by(model.Role.name)\
                            .all()

        permissions = self.get_available_permissions()

        # prep the excel writer
        path = os.path.join(self.rattail_config.workdir(),
                            'permissions-matrix.xlsx')
        writer = ExcelWriter(path, None)
        sheet = writer.sheet

        # write header
        sheet.cell(row=1, column=1, value="")
        for i, role in enumerate(roles, 2):
            sheet.cell(row=1, column=i, value=role.name)

        # font and fill pattern for permission group rows
        bold = Font(bold=True)
        group_fill = PatternFill(patternType='solid',
                                 fgColor='d9d9d9',
                                 bgColor='d9d9d9')

        # now we'll write the rows
        writing_row = 2
        for groupkey in sorted(permissions, key=lambda k: permissions[k]['label'].lower()):
            group = permissions[groupkey]

            # group label is bold, with fill pattern
            cell = sheet.cell(row=writing_row, column=1, value=group['label'])
            cell.font = bold
            cell.fill = group_fill

            # continue fill pattern for rest of group row
            for col, role in enumerate(roles, 2):
                cell = sheet.cell(row=writing_row, column=col)
                cell.fill = group_fill

            # okay, that row is done
            writing_row += 1

            # now we list each perm in the group
            perms = group['perms']
            for key in sorted(perms, key=lambda p: perms[p]['label'].lower()):
                sheet.cell(row=writing_row, column=1, value=perms[key]['label'])

                # and show an 'X' for any role which has this perm
                for col, role in enumerate(roles, 2):
                    if auth.has_permission(self.Session(), role, key,
                                           include_anonymous=False):
                        sheet.cell(row=writing_row, column=col, value="X")

                writing_row += 1

        writer.auto_resize()
        writer.auto_freeze(column=2)
        writer.save()
        return self.file_response(path)

    @classmethod
    def defaults(cls, config):
        cls._principal_defaults(config)
        cls._role_defaults(config)
        cls._defaults(config)

    @classmethod
    def _role_defaults(cls, config):
        route_prefix = cls.get_route_prefix()
        url_prefix = cls.get_url_prefix()
        permission_prefix = cls.get_permission_prefix()
        model_title = cls.get_model_title()

        # extra permissions for editing built-in roles etc.
        config.add_tailbone_permission(permission_prefix, '{}.edit_authenticated'.format(permission_prefix),
                                       "Edit the \"Authenticated\" Role")
        config.add_tailbone_permission(permission_prefix, '{}.edit_guest'.format(permission_prefix),
                                       "Edit the \"Guest\" Role")
        config.add_tailbone_permission(permission_prefix, '{}.edit_my'.format(permission_prefix),
                                       "Edit Role(s) to which current user belongs")
        config.add_tailbone_permission(permission_prefix,
                                       '{}.edit_node_sync'.format(permission_prefix),
                                       "Edit the Node Type and Sync flags for a {}".format(model_title))

        # download permissions matrix
        config.add_tailbone_permission(permission_prefix, '{}.download_permissions_matrix'.format(permission_prefix),
                                       "Download complete Role/Permissions matrix")
        config.add_route('{}.download_permissions_matrix'.format(route_prefix), '{}/permissions-matrix'.format(url_prefix),
                         request_method='GET')
        config.add_view(cls, attr='download_permissions_matrix', route_name='{}.download_permissions_matrix'.format(route_prefix),
                        permission='{}.download_permissions_matrix'.format(permission_prefix))

# TODO: deprecate / remove this
RolesView = RoleView


class PermissionsWidget(dfwidget.Widget):
    template = 'permissions'
    permissions = None
    true_val = 'true'

    def deserialize(self, field, pstruct):
        return [key for key, val in pstruct.items()
                if val == self.true_val]

    def get_checked_value(self, cstruct, value):
        if cstruct is colander.null:
            return False
        return value in cstruct

    def serialize(self, field, cstruct, **kw):
        kw.setdefault('permissions', self.permissions)
        template = self.template
        values = self.get_template_values(field, cstruct, kw)
        return field.renderer(template, **values)


def defaults(config, **kwargs):
    base = globals()

    RoleView = kwargs.get('RoleView', base['RoleView'])
    RoleView.defaults(config)


def includeme(config):
    defaults(config)
